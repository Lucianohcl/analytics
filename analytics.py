"""
NetExame Analytics BI — v8
Rodar: streamlit run analytics_bi.py
pip install streamlit pandas numpy scikit-learn statsmodels plotly openpyxl pdfplumber requests anthropic openai
pip install prophet  (opcional)
"""
import streamlit as st
import streamlit.components.v1 as components_v1
import pandas as pd
import base64
import numpy as np
import json, os, re, io, requests, time, warnings
import concurrent.futures
from io import BytesIO
from datetime import datetime
import plotly.graph_objects as go
from plotly.subplots import make_subplots

warnings.filterwarnings("ignore")

try:
    from openai import OpenAI; OPENAI_OK = True
except Exception as _e_openai_import: OPENAI_OK = False; OPENAI_ERR = str(_e_openai_import)
try:
    import anthropic; ANTHROPIC_OK = True
except Exception as _e_anthropic_import: ANTHROPIC_OK = False; ANTHROPIC_ERR = str(_e_anthropic_import)
try:
    import pdfplumber; PDF_OK = True
except: PDF_OK = False
try:
    from statsmodels.tsa.arima.model import ARIMA
    from statsmodels.tsa.holtwinters import ExponentialSmoothing, Holt
    from statsmodels.tsa.statespace.sarimax import SARIMAX
    from sklearn.metrics import mean_squared_error
    STATS_OK = True
except: STATS_OK = False
try:
    from prophet import Prophet; PROPHET_OK = True
except: PROPHET_OK = False

# ═══════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════
st.set_page_config(
    page_title="NetExame Analytics BI",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)


st.markdown("""<style>
.stApp{background:#FFFCF7;}
section[data-testid="stSidebar"]{background:linear-gradient(180deg,#0F6E56 0%,#085041 100%)!important;
  border-right:6px solid #A9762F!important;box-shadow:2px 0 10px rgba(169,118,47,.35)!important;}
section[data-testid="stSidebar"] button{background:rgba(255,255,255,.06)!important;
  color:#F2EDE1!important;border:1px solid rgba(169,118,47,.35)!important;}
section[data-testid="stSidebar"] button:hover{background:rgba(169,118,47,.25)!important;
  border-color:#A9762F!important;}
section[data-testid="stSidebar"] .stCaption,section[data-testid="stSidebar"] p{color:#A9A290!important;}
.mc{background:white;border:1px solid #E8ECF0;border-radius:12px;padding:10px 18px;
  text-align:center;box-shadow:0 1px 4px rgba(0,0,0,.06);}
.mc-lbl{font-size:.62rem;text-transform:uppercase;letter-spacing:.1em;
  color:#9CA3AF;font-weight:600;margin-bottom:3px;}
.mc-val{font-size:1.35rem;font-weight:700;color:#111827;}
.mc-val.g{color:#059669;}.mc-val.r{color:#DC2626;}.mc-val.y{color:#D97706;}.mc-val.b{color:#2563EB;}
.mc-sub{font-size:.7rem;color:#9CA3AF;margin-top:3px;}
.phdr{background:white;border:1px solid #E8ECF0;border-radius:14px;
  padding:22px 28px;margin-bottom:18px;box-shadow:0 1px 4px rgba(0,0,0,.06);}
.phdr h1{color:#111827!important;font-size:1.6rem;margin:0;}
.phdr p{color:#6B7280;margin:0;font-size:.86rem;}
.sec{font-size:.95rem;font-weight:700;color:#111827;
  padding-left:11px;border-left:3px solid #2563EB;margin:20px 0 12px;}
.al-d{background:#FEF2F2;border:1px solid #FECACA;border-radius:9px;
  padding:10px 14px;color:#DC2626;font-size:.83rem;margin:4px 0;}
.al-w{background:#FFFBEB;border:1px solid #FDE68A;border-radius:9px;
  padding:10px 14px;color:#D97706;font-size:.83rem;margin:4px 0;}
.al-s{background:#ECFDF5;border:1px solid #A7F3D0;border-radius:9px;
  padding:10px 14px;color:#059669;font-size:.83rem;margin:4px 0;}
.al-i{background:#EFF6FF;border:1px solid #BFDBFE;border-radius:9px;
  padding:10px 14px;color:#2563EB;font-size:.83rem;margin:4px 0;}
.kz-s{background:#ECFDF5;border:1px solid #A7F3D0;border-radius:10px;padding:12px 16px;}
.kz-w{background:#FFFBEB;border:1px solid #FDE68A;border-radius:10px;padding:12px 16px;}
.kz-d{background:#FEF2F2;border:1px solid #FECACA;border-radius:10px;padding:12px 16px;}
.dre-wrap{overflow-x:auto;border-radius:10px;border:1px solid #E8ECF0;margin-top:8px;}
table.dre{width:100%;border-collapse:collapse;font-size:.76rem;}
table.dre th{background:#F3F4F6;color:#6B7280;padding:7px 10px;text-align:right;
  font-weight:600;border-bottom:1px solid #E8ECF0;white-space:nowrap;}
table.dre th:first-child{text-align:left;min-width:180px;}
table.dre td{padding:5px 10px;text-align:right;color:#374151;white-space:nowrap;}
table.dre td:first-child{text-align:left;color:#111827;}
table.dre tr.cat td{background:#F9FAFB;color:#6B7280!important;
  font-size:.67rem;text-transform:uppercase;font-weight:600;}
table.dre tr.sub td{background:#FAFAFA;}
table.dre tr.tot td{background:#F3F4F6;font-weight:700;}
table.dre tr.tot td:first-child{color:#2563EB!important;}
table.dre tr.pct td:not(:first-child){color:#D97706!important;}
table.dre td.pos{color:#059669!important;}
table.dre td.neg{color:#DC2626!important;}
table.dre td.neu{color:#9CA3AF!important;}
.mc-val.pos{color:#059669!important;}
.mc-val.neg{color:#DC2626!important;}
.mc-val.neu{color:#6B7280!important;}
table.dre tr:hover td{background:#F9FAFB!important;}

/* esconde a barra de ferramentas do Plotly (camera/zoom) em telas pequenas,
   pra nao sobrepor o titulo dos graficos no celular */
@media (max-width: 600px){
  .js-plotly-plot .modebar{ display:none !important; }
}

/* modo escuro via filtro CSS (inverte a tela inteira; imagens/graficos
   sao invertidos de volta pra nao ficar com cor estranha) */
html{filter:invert(1) hue-rotate(180deg);}
img, iframe{filter:invert(1) hue-rotate(180deg);}
.wc-hero{background:linear-gradient(135deg,#0F6E56 0%,#085041 100%);border-radius:16px;
  padding:36px 40px;margin-bottom:26px;box-shadow:0 4px 14px rgba(0,0,0,.15);
  text-align:center;border:3px solid #A9762F;}
.wc-hero h1{font-family:Georgia,serif;font-weight:600;color:#fff;font-size:1.5rem;margin:0 0 8px;}
.wc-hero p{color:#9FE1CB;font-size:.92rem;letter-spacing:.01em;margin:0;}
.wc-back{background:linear-gradient(135deg,#0F6E56 0%,#085041 100%);border-radius:16px;
  padding:48px 40px;text-align:center;margin-bottom:26px;box-shadow:0 4px 14px rgba(0,0,0,.15);
  border:3px solid #A9762F;}
.wc-back-emoji{font-size:2.6rem;margin-bottom:12px;}
.wc-back h2{font-family:Georgia,serif;font-weight:600;color:#fff;font-size:1.5rem;margin:0 0 8px;}
.wc-back p{color:#9FE1CB;font-size:.92rem;margin:0;}
.wc-steps{position:relative;margin-bottom:32px;display:flex;}
.wc-steps::before{content:"";position:absolute;top:22px;left:16.5%;right:16.5%;height:2px;
  background:#0F6E56;opacity:.35;z-index:0;}
.wc-step{flex:1;text-align:center;padding:0 10px;position:relative;z-index:1;}
.wc-step-n{width:44px;height:44px;border-radius:50%;background:#0F6E56;color:#9FE1CB;
  display:flex;align-items:center;justify-content:center;font-family:Georgia,serif;
  font-weight:700;font-size:1.05rem;margin:0 auto 12px;border:2px solid rgba(169,118,47,.6);}
.wc-step-t{font-weight:700;color:#14243B;font-size:.9rem;margin-bottom:3px;}
.wc-step-s{color:#9CA3AF;font-size:.76rem;}
.wc-card-exp{background:#0F6E56;color:#fff;border-radius:12px;padding:16px 20px;
  margin-bottom:16px;border:1px solid rgba(169,118,47,.55);cursor:pointer;}
.wc-card-exp summary{font-family:Georgia,serif;font-weight:700;font-size:.85rem;
  letter-spacing:.03em;list-style:none;cursor:pointer;color:#fff;}
.wc-card-exp summary::-webkit-details-marker{display:none;}
.wc-card-exp summary::after{content:"▾";float:right;color:#9FE1CB;}
.wc-card-exp[open] summary::after{content:"▴";}
.wc-card-exp-content{margin-top:14px;padding-top:12px;border-top:1px solid rgba(255,255,255,.2);}
.wc-card-exp ul{list-style:none;margin:0;padding:0;font-size:.8rem;line-height:1.95;color:#F0FBF7;}
.wc-card-exp li{margin-bottom:4px;}
</style>""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════
# CONSTANTES
# ═══════════════════════════════════════════════════
PASTA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "clientes_bi")
CFG   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
os.makedirs(PASTA, exist_ok=True)

# Log simples de acessos de visitantes (ignora acessos com ?owner=1 na URL)
if "log_registrado" not in st.session_state:
    st.session_state.log_registrado = True
    try:
        eh_owner = st.query_params.get("owner") == "1"
        if not eh_owner:
            path_acessos = os.path.join(PASTA, "acessos.log")
            with open(path_acessos, "a", encoding="utf-8") as f:
                from datetime import timedelta
                agora_brasilia=datetime.utcnow()-timedelta(hours=3)
                f.write(f"{agora_brasilia.strftime('%Y-%m-%d %H:%M:%S')}\n")
    except Exception:
        pass

MESES   = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]
MES_NUM = {m:str(i+1).zfill(2) for i,m in enumerate(MESES)}

DEMONSTRACOES_CAMPOS = {
    "DRE": ["receita bruta de vendas","impostos sobre vendas","devoluções de vendas",
            "CMV (custo da mercadoria vendida)","despesas comerciais","despesas administrativas",
            "despesas financeiras líquidas","despesas com depreciações e amortizações",
            "receitas não operacionais","despesas não operacionais",
            "provisão para imposto de renda","provisão para contribuição social",
            "numero de vendas","pró-labore/distribuição de lucro"],
    "Balanço": ["disponibilidades saldo","contas a receber saldo",
            "estoque inicial do mês de mercadorias para revenda saldo",
            "estoque final do mês de mercadorias para revenda saldo","Outros AC","Ativo NC",
            "contas a pagar de fornecedores saldo","Passivos Financeiros","Outros PC","Passivo NC","Aporte"],
    "Fluxo": ["Disponibilidades entradas","Disponibilidades Saida",
            "Centro de Custos Entradas 1","Centro de Custos Entradas 2",
            "Centro de Custos Entradas 3","Centro de Custos Entradas 4",
            "Centro de Custos Saidas 1","Centro de Custos Saidas 2",
            "Centro de Custos Saidas 3","Centro de Custos Saidas 4"],
}

CAMPOS_DRE = ["receita bruta de vendas","impostos sobre vendas","devoluções de vendas",
    "CMV (custo da mercadoria vendida)","estoque inicial do mês de mercadorias para revenda saldo",
    "estoque final do mês de mercadorias para revenda saldo","despesas comerciais",
    "despesas administrativas","despesas financeiras líquidas",
    "despesas com depreciações e amortizações","receitas não operacionais",
    "despesas não operacionais","provisão para imposto de renda",
    "provisão para contribuição social","numero de vendas","pró-labore/distribuição de lucro"]
CAMPOS_BAL = ["disponibilidades saldo","contas a receber saldo",
    "estoque final do mês de mercadorias para revenda saldo","Outros AC","Ativo NC",
    "contas a pagar de fornecedores saldo","Passivos Financeiros","Outros PC","Passivo NC","Aporte"]
CAMPOS_FLUXO = ["Disponibilidades entradas","Disponibilidades Saida",
    "Centro de Custos Entradas 1","Centro de Custos Entradas 2",
    "Centro de Custos Entradas 3","Centro de Custos Entradas 4",
    "Centro de Custos Saidas 1","Centro de Custos Saidas 2",
    "Centro de Custos Saidas 3","Centro de Custos Saidas 4"]
TODOS = list(dict.fromkeys(CAMPOS_DRE + CAMPOS_BAL + CAMPOS_FLUXO))
MODELOS_ML = {"ARIMA":STATS_OK,"ExponentialSmoothing":STATS_OK,
              "SARIMAX":STATS_OK,"Holt":STATS_OK,"Prophet":PROPHET_OK,
              "Croston":True,"TSB":True,"Ensemble":True}
# Executor único, reaproveitado em todos os treinos de modelo (em vez de criar
# um novo a cada chamada). max_workers=4 (não 1) de propósito: se um modelo
# travar e estourar o timeout, a thread presa não bloqueia os próximos treinos,
# porque sobram outras vagas livres no pool.
_ML_EXECUTOR = concurrent.futures.ThreadPoolExecutor(max_workers=4)
CORES = ["#2176FF","#00D4AA","#FFB627","#F85149","#A371F7","#F78166","#79C0FF","#56D364"]

# ═══════════════════════════════════════════════════
# UTILITÁRIOS
# ═══════════════════════════════════════════════════
def fmt(v, t="brl"):
    try:
        f=float(v)
        if t=="brl":
            s="-" if f<0 else ""; f=abs(f)
            if f>=1e6: return f"{s}R$ {f/1e6:.1f}M"
            if f>=1e3: return f"{s}R$ {f/1e3:.0f}K"
            return f"{s}R$ {f:.0f}"
        if t=="pct": return f"{f:.1f}%"
        if t=="x":   return f"{f:.2f}x"
        if t=="d":   return f"{f:.0f}d"
    except: pass
    return "—"

def safe(a,b,d=0.):
    try: a,b=float(a),float(b); return a/b if b else d
    except: return d

def gid(n):
    import unicodedata
    s=unicodedata.normalize("NFKD",n.lower())
    s="".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+","_",s).strip("_")[:40]

def cor(v, inv=False):
    if abs(v)<0.5: return "neu"
    if inv: return "neg" if v>0 else "pos"
    return "pos" if v>0 else "neg"

def cn(df,c):
    if c in df.columns: return pd.to_numeric(df[c],errors="coerce").fillna(0.)
    return pd.Series(0.,index=df.index)

def cm_(df): return next((c for c in df.columns if c.lower().strip() in ["mês","mes"]),None)
def ca_(df): return next((c for c in df.columns if c.lower().strip() in ["ano","year"]),None)

# ═══════════════════════════════════════════════════
# PERSISTÊNCIA
# ═══════════════════════════════════════════════════
def load_cfg():
    try:
        with open(CFG,encoding="utf-8") as f: return json.load(f)
    except: return {}

def save_cfg(d):
    try:
        with open(CFG,"w",encoding="utf-8") as f: json.dump(d,f,ensure_ascii=False,indent=2)
    except: pass

def path_cli(cid): return os.path.join(PASTA,f"{gid(cid)}.json")

def salvar_cli(cid,d):
    d.setdefault("meta",{})["at"]=datetime.now().isoformat()
    with open(path_cli(cid),"w",encoding="utf-8") as f: json.dump(d,f,ensure_ascii=False,indent=2)

def load_cli(cid):
    p=path_cli(cid)
    if not os.path.exists(p): return None
    with open(p,encoding="utf-8") as f: return json.load(f)

SENHA_MASTER = "netexame2024"  # Troque para sua senha

def verificar_pin(cid, pin):
    p = load_cli(cid)
    if not p: return False
    pin_salvo = p.get("pin","")
    if not pin_salvo: return True  # sem pin = acesso livre
    return str(pin) == str(pin_salvo)

def ls_cli():
    out=[]
    for a in os.listdir(PASTA):
        if a.endswith(".json") and not a.endswith("_cenarios_ml.json"):
            try:
                with open(os.path.join(PASTA,a),encoding="utf-8") as f: d=json.load(f)
                if "nome" not in d: continue
                out.append({"id":a[:-5],"nome":d.get("nome","?"),"at":d.get("meta",{}).get("at","")[:10]})
            except: pass
    return sorted(out,key=lambda x:x["nome"].lower())

def pre_carregar_cliente(cid):
    if not cid: return
    _res,_cal,_morto=load_resultado_compras(cid)
    if _res is not None:
        st.session_state["compras_resultado"]=_res
        st.session_state["compras_calendario"]=_cal
        st.session_state["compras_morto"]=_morto
    _est=load_estoque_compras(cid)
    if _est is not None: st.session_state["compras_df_estoque"]=_est
    _cfg_ml=load_cfgml_resultado(cid)
    if _cfg_ml is not None: st.session_state["cfgml_resultado_atual"]=_cfg_ml
    _val_full=load_validacao_full(cid)
    if _val_full is not None and _val_full.get("df_comp") is not None:
        st.session_state["cfgml_df_comp_bruto"]=_val_full["df_comp"]
        st.session_state["cfgml_df_escopo_val"]=_val_full["df_escopo"]
        st.session_state["cfgml_produto_col_val_usado"]=_val_full["produto_col"]
        st.session_state["cfgml_col_data_val_usado"]=_val_full["col_data"]
        st.session_state["cfgml_metrica_val_usado"]=_val_full["metrica"]
    _ff_salvo=load_fluxo_financeiro(cid)
    if _ff_salvo:
        st.session_state["ff_parcelas_pagar"]=_ff_salvo.get("ff_parcelas_pagar") or [{"pct":100,"dias":30}]
        st.session_state["ff_parcelas_receber"]=_ff_salvo.get("ff_parcelas_receber") or [{"pct":100,"dias":30}]
        st.session_state["ff_encargos_pct"]=_ff_salvo.get("ff_encargos_pct") or 0.0
        st.session_state["ff_saldo_inicial"]=_ff_salvo.get("ff_saldo_inicial") or 0.0
        st.session_state["ff_incluir_contas_pr"]=_ff_salvo.get("ff_incluir_contas_pr") or False
        st.session_state["ff_config_carregada"]=True
    _cp=load_contas_pr(cid,"pagar")
    if _cp is not None: st.session_state["ff_contas_pagar_df"]=_cp
    _cr=load_contas_pr(cid,"receber")
    if _cr is not None: st.session_state["ff_contas_receber_df"]=_cr
    _sc_forn=load_scorecard_forn(cid)
    if _sc_forn is not None: st.session_state["gs_scorecard_forn"]=_sc_forn
    _cfg_ml_config=load_config_ml(cid)
    if _cfg_ml_config:
        for _k,_v in _cfg_ml_config.items():
            st.session_state[_k]=_v
        st.session_state["cfgml_config_carregada"]=True
    _cfg_compras=load_config_compras(cid)
    if _cfg_compras:
        for _k,_v in _cfg_compras.items():
            st.session_state[_k]=_v
    _lt_compras=load_leadtime_compras(cid)
    if _lt_compras:
        for _k,_v in _lt_compras.items():
            st.session_state[_k]=_v
    _cfg_mlp=load_config_mlp(cid)
    if _cfg_mlp:
        for _k,_v in _cfg_mlp.items():
            if _v is not None:
                st.session_state[_k]=_v
    _pv=load_snap(cid,"pareto_visao")
    if _pv: st.session_state["pareto_visao"]=_pv.get("visao","")
    # O resultado do Pareto (pareto_resultado_atual) não é mais pré-carregado aqui —
    # a própria página de Pareto já recarrega do disco sozinha quando necessário,
    # de forma correta e ciente de Filial (evita o "achou que você trocou de visão" à toa).    

def renderizar_card_cliente(c):
    c1,c2,c3,c4=st.columns([4,2,1,1])
    p2=load_cli(c["id"]); tem_pin=bool(p2.get("pin","")) if p2 else False
    c1.markdown(f'<div style="color:#111827;font-weight:600;padding:10px 0">'
               f'{c["nome"]} {"🔒" if tem_pin else "🔓"}</div>',unsafe_allow_html=True)
    c2.markdown(f'<div style="color:#9CA3AF;font-size:.82rem;padding:12px 0">🕐 {c["at"]}</div>',unsafe_allow_html=True)
    if c3.button("📂 Abrir",key=f"ab_{c['id']}",use_container_width=True):
        if tem_pin:
            st.session_state.pin_pendente=c["id"]
        else:
            limpar_sessao_cliente()
            p_open=load_cli(c["id"])
            st.session_state.cid=c["id"]
            st.session_state.df_raw=load_df(c["id"])
            st.session_state.projecoes={}
            st.session_state.saldo_ini=float(p_open.get("saldo_ini",0)) if p_open else 0.
            st.session_state.entradas_vista=float(p_open.get("entradas_vista",0)) if p_open else 0.
            st.session_state.freq_fluxo=p_open.get("freq_fluxo","Mensal") if p_open else "Mensal"
            st.session_state["_limpar_busca_cliente"]=True
            pre_carregar_cliente(c["id"])
            addlog(f"'{c['nome']}' aberto"); ir("boas_vindas")
    if c4.button("🗑",key=f"dl_{c['id']}",use_container_width=True):
        os.remove(path_cli(c["id"]))
        if st.session_state.cid==c["id"]: st.session_state.cid=None; st.session_state.df_raw=None
        st.rerun()
    st.divider()

def save_df(cid,df):
    df.to_csv(os.path.join(PASTA,f"{gid(cid)}_dados.csv"),sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_df(cid):
    p=os.path.join(PASTA,f"{gid(cid)}_dados.csv")
    if not os.path.exists(p): return None
    try: return pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
    except: return None

def path_rateio_contas(cid): return os.path.join(PASTA,f"{gid(cid)}_rateio_contas.json")

def path_projecoes_ml(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_projecoes_ml__{_sufixo_filial(filial)}.json")

def save_projecoes_ml(cid,projs,filial=None):
    with open(path_projecoes_ml(cid,filial),"w",encoding="utf-8") as f:
        json.dump(projs,f,ensure_ascii=False,default=str)

def load_projecoes_ml(cid,filial=None):
    p=path_projecoes_ml(cid,filial)
    if not os.path.exists(p): return None
    try:
        with open(p,"r",encoding="utf-8") as f: return json.load(f)
    except: return None

def save_rateio_contas(cid,rateio_dict):
    with open(path_rateio_contas(cid),"w",encoding="utf-8") as f:
        json.dump(rateio_dict,f,ensure_ascii=False,indent=2)

def load_rateio_contas(cid):
    p=path_rateio_contas(cid)
    if not os.path.exists(p): return {}
    try:
        with open(p,encoding="utf-8") as f: return json.load(f)
    except: return {}

# ═══════════════════════════════════════════════════
# LEITURA
# ═══════════════════════════════════════════════════
def limpar(df):
    cols,cnt=[],{}
    for c in df.columns:
        s=str(c).strip()
        if s.lower() in ["nan","none",""]: s=f"__c{len(cols)}"
        if s in cnt: cnt[s]+=1; s=f"{s}_{cnt[s]}"
        else: cnt[s]=0
        cols.append(s)
    df.columns=cols
    df=df.dropna(how="all").dropna(axis=1,how="all")
    df=df[[c for c in df.columns if not c.startswith("__c")]]
    return df.reset_index(drop=True)

def _converter_valor_balancete(v):
    """Converte string de valor contábil (com vírgula decimal e sinal D/C/-) para float."""
    import re
    v=v.strip()
    negativo = v.startswith("-") or v.endswith("-") or v.upper().endswith("C")
    v_limpo = re.sub(r'[^\d,]', '', v)
    v_limpo = v_limpo.replace(".", "").replace(",", ".")
    try:
        num = float(v_limpo)
    except ValueError:
        return 0.0
    return -num if negativo else num

def parser_balancete_texto(texto):
    """Fallback para PDFs de balancete sem estrutura de tabela detectável —
    interpreta linhas de texto corrido no formato: codigo [codigo-formatado] nome valor1 valor2 valor3 valor4"""
    import re
    # Tenta identificar o período no cabeçalho, em 2 formatos possíveis:
    # "Período: 12/2019 a 12/2019" (mês/ano) ou "Periodo: 01/01/2021 a 31/03/2021" (dia/mês/ano)
    # Sempre usa a DATA FINAL do range, pois o Saldo Atual representa a posição naquele momento
    ano_periodo, mes_periodo = None, None
    meses_num = {"01":"jan","02":"fev","03":"mar","04":"abr","05":"mai","06":"jun",
                 "07":"jul","08":"ago","09":"set","10":"out","11":"nov","12":"dez"}

    m_periodo_longo = re.search(r'Per[ií]odo:\s*\d{2}/\d{2}/\d{4}\s*a\s*\d{2}/(\d{2})/(\d{4})', texto)
    if m_periodo_longo:
        mes_periodo = meses_num.get(m_periodo_longo.group(1))
        ano_periodo = m_periodo_longo.group(2)
    else:
        m_periodo_curto = re.findall(r'Per[ií]odo:\s*(\d{2})/(\d{4})(?:\s*a\s*(\d{2})/(\d{4}))?', texto)
        if m_periodo_curto:
            grupo = m_periodo_curto[0]
            if grupo[2] and grupo[3]:
                mes_periodo = meses_num.get(grupo[2])
                ano_periodo = grupo[3]
            else:
                mes_periodo = meses_num.get(grupo[0])
                ano_periodo = grupo[1]

    linhas_extraidas=[]
    padrao_valor=r'-?[\d\.]+,\d{2}-?[DC]?'
    # Aceita opcionalmente um segundo código formatado (ex: "11110-1") entre o código numérico e o nome
    padrao_linha=re.compile(
        r'^\s*(\d[\d\.]*)\s+(?:[\d]+-[\d]+\s+)?(.+?)\s+('
        +padrao_valor+r')\s+('+padrao_valor+r')\s+('+padrao_valor+r')\s+('+padrao_valor+r')\s*$'
    )
    for linha in texto.split("\n"):
        linha=linha.strip()
        if not linha or "Saldo Anterior" in linha or "Página" in linha: continue
        m=padrao_linha.match(linha)
        if m:
            codigo,nome,saldo_ant,debito,credito,saldo_atual=m.groups()
            linhas_extraidas.append({
                "Código":codigo.strip(),"Nome":nome.strip(),
                "Saldo Anterior":_converter_valor_balancete(saldo_ant),
                "Débito":_converter_valor_balancete(debito),
                "Crédito":_converter_valor_balancete(credito),
                "Saldo Atual":_converter_valor_balancete(saldo_atual),
                "Ano":ano_periodo,"Mês":mes_periodo,
            })
    # Proteção: garante que TODAS as linhas usem o mesmo ano/mês (evita inconsistência
    # quando o cabeçalho de período se repete de forma diferente em páginas distintas do PDF)
    if linhas_extraidas and (not ano_periodo or not mes_periodo):
        # Não conseguiu identificar no cabeçalho — sem fallback seguro, aborta
        return None
    for item in linhas_extraidas:
        item["Ano"]=ano_periodo
        item["Mês"]=mes_periodo

    if not linhas_extraidas: return None

    # Identifica contas-folha: uma conta só é processada se NENHUMA outra conta
    # tiver um código que comece com o dela + separador (evita duplicação hierárquica)
    todos_codigos=[item["Código"].replace(".","") for item in linhas_extraidas]
    def eh_folha(codigo):
        cod=codigo.replace(".","")
        for outro in todos_codigos:
            if outro!=cod and outro.startswith(cod) and len(outro)>len(cod):
                return False
        return True

    # Mapa de DRE (prefixo 3=Receita, 4=Despesa) e Balanço (prefixo 1=Ativo, 2=Passivo)
    mapa_balancete={
        "caixa":"disponibilidades saldo","banco":"disponibilidades saldo",
        "disponib":"disponibilidades saldo","aplicaç":"disponibilidades saldo",
        "duplicatas a receber":"contas a receber saldo","cliente":"contas a receber saldo",
        "cartoes a receber":"contas a receber saldo","cartões a receber":"contas a receber saldo",
        "cheques a receber":"contas a receber saldo",
        "estoque":"estoque final do mês de mercadorias para revenda saldo",
        "mercadorias p/revenda":"estoque final do mês de mercadorias para revenda saldo",
        "mercadorias para revenda":"estoque final do mês de mercadorias para revenda saldo",
        "custo da mercadoria":"CMV (custo da mercadoria vendida)",
        "fornecedores de mercadoria":"contas a pagar de fornecedores saldo",
        "impostos a recuperar":"Outros AC","icms a recuperar":"Outros AC",
        "adiantamento":"Outros AC","provisao":"Outros AC","provisão":"Outros AC",
        "imobilizado":"Ativo NC","veiculo":"Ativo NC","veículo":"Ativo NC",
        "maquinas":"Ativo NC","máquinas":"Ativo NC","moveis":"Ativo NC","móveis":"Ativo NC",
        "equipamento":"Ativo NC","depreciaç":"Ativo NC","depreciac":"Ativo NC",
        "edifica":"Ativo NC","bens tangiveis":"Ativo NC","bens tangíveis":"Ativo NC",
        "despesas antecipadas":"Outros AC","despesas a apropriar":"Outros AC",
        "impostos e taxas a a":"Outros AC",
        "fornecedor":"contas a pagar de fornecedores saldo",
        "emprestimo":"Passivos Financeiros","empréstimo":"Passivos Financeiros",
        "financiamento":"Passivos Financeiros",
        "salari":"Outros PC","salário":"Outros PC","encargo":"Outros PC",
        "impostos a recolher":"Outros PC","tributos a pagar":"Outros PC",
        "capital social":"PL","reservas":"PL","lucros acumulados":"PL",
        # DRE - Receitas (prefixo 3)
        "venda":"receita bruta de vendas","faturamento":"receita bruta de vendas",
        "receita":"receita bruta de vendas","prestaç":"receita bruta de vendas",
        # DRE - Despesas (prefixo 4)
        "despesas administrativas":"despesas administrativas",
        "honorarios diretoria":"despesas administrativas","honorarios contabeis":"despesas administrativas",
        "honorarios juridicos":"despesas administrativas","material de consumo":"despesas administrativas",
        "material de escritorio":"despesas administrativas","telefone":"despesas administrativas",
        "agua":"despesas administrativas","energia":"despesas administrativas",
        "manutencao":"despesas administrativas","combustiveis":"despesas administrativas",
        "seguros":"despesas administrativas","consertos":"despesas administrativas",
        "graficas":"despesas administrativas","depreciacao":"despesas com depreciações e amortizações",
        "limpeza":"despesas administrativas","copa e cozinha":"despesas administrativas",
        "mensalidades":"despesas administrativas","brindes":"despesas comerciais",
        "jornais e revistas":"despesas administrativas","despesas medicas":"despesas administrativas",
        "comiss":"despesas comerciais","marketing":"despesas comerciais","publicidade":"despesas comerciais",
        "salarios":"despesas administrativas","encargos sociais":"despesas administrativas",
        "juros":"despesas financeiras líquidas","tarifas bancarias":"despesas financeiras líquidas",
        "despesas financeiras":"despesas financeiras líquidas",
        "provisao de deved duvid":"despesas administrativas",
    }
    totais_evitar_balancete=["ativo","passivo","ativo circulante","passivo circulante",
                              "ativo não-circulante","ativo nao-circulante",
                              "receitas","despesas","despesas operacionais"]

    detalhamento_balancete=[]
    acumulado_bal={}
    detalhamento_balancete=[]
    acumulado_fluxo={}
    detalhamento_fluxo=[]
    for item in linhas_extraidas:
        if not eh_folha(item["Código"]): continue
        nome_l=item["Nome"].lower()
        if any(t==nome_l.strip() for t in totais_evitar_balancete): continue
        campo_dest=None
        primeiro_digito_chk=item["Código"].replace(".","")[0] if item["Código"] else ""
        for k,v in mapa_balancete.items():
            if k in nome_l:
                if primeiro_digito_chk in ("3","4") and v not in ("receita bruta de vendas","faturamento","despesas administrativas","despesas comerciais","despesas financeiras líquidas","despesas com depreciações e amortizações","CMV (custo da mercadoria vendida)"):
                    continue
                if primeiro_digito_chk in ("1","2") and v in ("receita bruta de vendas","despesas administrativas","despesas comerciais","despesas financeiras líquidas","despesas com depreciações e amortizações","CMV (custo da mercadoria vendida)"):
                    continue
                campo_dest=v; break
        if not campo_dest: continue
        primeiro_digito=item["Código"].replace(".","")[0] if item["Código"] else ""
        if primeiro_digito=="3":
            valor=abs(item["Crédito"])
        elif primeiro_digito=="4":
            valor=abs(item["Débito"])
        else:
            valor=item["Saldo Atual"]
        if valor==0: continue
        chave=(item["Ano"],item["Mês"],campo_dest)
        acumulado_bal[chave]=acumulado_bal.get(chave,0)+valor
        detalhamento_balancete.append({"ano":item["Ano"],"mes":item["Mês"],
                                       "campo_pai":campo_dest,"subconta":item["Nome"],"valor":valor})

        if campo_dest=="disponibilidades saldo":
            entrada_v=abs(item["Débito"])
            saida_v=abs(item["Crédito"])
            if entrada_v>0:
                chave_ent=(item["Ano"],item["Mês"],"Disponibilidades entradas")
                acumulado_fluxo[chave_ent]=acumulado_fluxo.get(chave_ent,0)+entrada_v
                detalhamento_fluxo.append({"ano":item["Ano"],"mes":item["Mês"],
                                           "campo_pai":"Disponibilidades entradas","subconta":item["Nome"],"valor":entrada_v})
            if saida_v>0:
                chave_sai=(item["Ano"],item["Mês"],"Disponibilidades Saida")
                acumulado_fluxo[chave_sai]=acumulado_fluxo.get(chave_sai,0)+saida_v
                detalhamento_fluxo.append({"ano":item["Ano"],"mes":item["Mês"],
                                           "campo_pai":"Disponibilidades Saida","subconta":item["Nome"],"valor":saida_v})

    campos_dre_balancete={"receita bruta de vendas","despesas administrativas","despesas comerciais",
                          "despesas financeiras líquidas","despesas com depreciações e amortizações"}
    celulas_balancete=[]
    for (a,m,c),v in acumulado_bal.items():
        v_final=v if c in campos_dre_balancete else max(v,0)
        celulas_balancete.append({"ano":a,"mes":m,"campo":c,"valor":v_final})
    for (a,m,c),v in acumulado_fluxo.items():
        celulas_balancete.append({"ano":a,"mes":m,"campo":c,"valor":v})

    detalhamento_total=detalhamento_balancete+detalhamento_fluxo
    if detalhamento_total:
        celulas_balancete.append({"_detalhamento":detalhamento_total})
    return celulas_balancete if celulas_balancete else None

def col_filial(df):
    return next((c for c in df.columns if c.strip().lower() in
        ["filial","loja","unidade","ponto de venda","pdv","codigo loja","código loja","cnpj"]),None)    

def ler(b,nome):
    n=nome.lower()
    try:
        if n.endswith(".pdf"):
            if not PDF_OK: return None,"pip install pdfplumber"
            dfs=[]
            texto_completo=""
            with pdfplumber.open(io.BytesIO(b)) as pdf:
                for pg in pdf.pages:
                    for t in pg.extract_tables() or []:
                        if t and len(t)>1:
                            try: dfs.append(pd.DataFrame(t[1:],columns=t[0]))
                            except: pass
                    txt_pg=pg.extract_text() or ""
                    texto_completo+=txt_pg+"\n"
            if dfs:
                return limpar(pd.concat(dfs,ignore_index=True)),"PDF lido"
            # Fallback: sem tabela detectada — tenta interpretar como balancete em texto corrido
            celulas_balancete=parser_balancete_texto(texto_completo)
            if celulas_balancete:
                return ("_CELULAS_BALANCETE_",celulas_balancete),"PDF lido (formato balancete — texto)"
            return None,"Sem tabelas no PDF"
        elif n.endswith((".xlsx",".xls",".xlsm")):
            xls=pd.read_excel(io.BytesIO(b),sheet_name=None)
            dfs=[(s,df) for s,df in xls.items() if not df.dropna(how="all").empty]
            if not dfs: return None,"Excel vazio"
            dfs.sort(key=lambda x:len(x[1]),reverse=True)
            s,df=dfs[0]; return limpar(df),f"Excel — '{s}'"
        else:
            for enc in ["utf-8-sig","utf-8","latin1","cp1252"]:
                for sep in [";",",","\t","|"]:
                    try:
                        df=pd.read_csv(io.BytesIO(b),sep=sep,encoding=enc,on_bad_lines="skip",decimal=",")
                        if df.shape[1]>=2: return limpar(df),"CSV lido"
                    except: pass
            return None,"Não foi possível ler"
    except Exception as e: return None,str(e)

# ═══════════════════════════════════════════════════
# IA
# ═══════════════════════════════════════════════════
def eh_formato_longo(df):
    """Detecta se o arquivo está no formato vertical: uma coluna Ano, uma Mês, uma Conta, uma Valor."""
    cols_lower=[str(c).lower().strip() for c in df.columns]
    tem_ano="ano" in cols_lower or "year" in cols_lower
    tem_mes="mês" in cols_lower or "mes" in cols_lower or "month" in cols_lower
    tem_conta=any(c in cols_lower for c in ["conta","descrição","descricao","item"])
    tem_valor=any(c in cols_lower for c in ["valor","value","montante"])
    return tem_ano and tem_mes and tem_conta and tem_valor

def parser_formato_longo(df, tipo):
    """Parser determinístico para formato vertical (1 linha = 1 valor).
    Sempre confiável, não depende de IA, não trunca, sempre converte valores para positivo."""
    cols_lower={str(c).lower().strip():c for c in df.columns}
    col_ano=cols_lower.get("ano") or cols_lower.get("year")
    col_mes=cols_lower.get("mês") or cols_lower.get("mes") or cols_lower.get("month")
    col_conta=cols_lower.get("conta") or cols_lower.get("descrição") or cols_lower.get("descricao") or cols_lower.get("item")
    col_valor=cols_lower.get("valor") or cols_lower.get("value") or cols_lower.get("montante")
    col_fil=col_filial(df)
    if not all([col_ano,col_mes,col_conta,col_valor]):
        return []

    meses_pt={"jan":"jan","fev":"fev","mar":"mar","abr":"abr","mai":"mai","jun":"jun",
               "jul":"jul","ago":"ago","set":"set","out":"out","nov":"nov","dez":"dez",
               "janeiro":"jan","fevereiro":"fev","março":"mar","abril":"abr","maio":"mai",
               "junho":"jun","julho":"jul","agosto":"ago","setembro":"set",
               "outubro":"out","novembro":"nov","dezembro":"dez"}

    if tipo=="DRE":
        mapa={
            "receita bruta":"receita bruta de vendas","faturamento":"receita bruta de vendas",
            "receita de vendas":"receita bruta de vendas","receita de serviço":"receita bruta de vendas",
            "serviços recorrentes":"receita bruta de vendas","serviços avulsos":"receita bruta de vendas",
            "venda de mercadorias":"receita bruta de vendas","vendas brutas":"receita bruta de vendas",
            "outras receitas operacionais":"receita bruta de vendas",
            "imposto sobre vendas":"impostos sobre vendas","deduç":"impostos sobre vendas",
            "devoluç":"devoluções de vendas","desconto concedido":"devoluções de vendas",
            "cmv":"CMV (custo da mercadoria vendida)","custo da mercadoria":"CMV (custo da mercadoria vendida)",
            "custo do produto":"CMV (custo da mercadoria vendida)","custo variável":"CMV (custo da mercadoria vendida)",
            "mão de obra direta":"CMV (custo da mercadoria vendida)","materiais e insumos":"CMV (custo da mercadoria vendida)",
            "fretes de produção":"CMV (custo da mercadoria vendida)",
            "despesas comerciais":"despesas comerciais","salários comerciais":"despesas comerciais",
            "comissões":"despesas comerciais","marketing":"despesas comerciais","viagens comerciais":"despesas comerciais",
            "despesas administrativas":"despesas administrativas","salários administrativos":"despesas administrativas",
            "pró-labore":"despesas administrativas","aluguel":"despesas administrativas",
            "energia elétrica":"despesas administrativas","internet e telefonia":"despesas administrativas",
            "contabilidade":"despesas administrativas","sistemas e licenças":"despesas administrativas",
            "material de escritório":"despesas administrativas","manutenção":"despesas administrativas",
            "seguros":"despesas administrativas","treinamentos":"despesas administrativas",
            "despesas financeiras":"despesas financeiras líquidas","juros bancários":"despesas financeiras líquidas",
            "tarifas bancárias":"despesas financeiras líquidas",
            "receitas financeiras":"receitas não operacionais","juros ativos":"receitas não operacionais",
            "deprecia":"despesas com depreciações e amortizações",
            "irpj":"provisão para imposto de renda","csll":"provisão para contribuição social",
        }
    elif tipo=="BALANCO":
        mapa={
            "caixa":"disponibilidades saldo","banco":"disponibilidades saldo","disponib":"disponibilidades saldo",
            "aplicaç":"disponibilidades saldo","cliente":"contas a receber saldo","receber":"contas a receber saldo",
            "estoque":"estoque final do mês de mercadorias para revenda saldo",
            "impostos a recuperar":"Outros AC","outros ac":"Outros AC",
            "imobilizado":"Ativo NC","ativo nc":"Ativo NC",
            "fornecedor":"contas a pagar de fornecedores saldo",
            "empréstimos cp":"Passivos Financeiros","empréstimo cp":"Passivos Financeiros",
            "financiamentos cp":"Passivos Financeiros","financiamento":"Passivos Financeiros",
            "salários e encargos":"Outros PC","impostos a recolher":"Outros PC","outros pc":"Outros PC",
            "empréstimos lp":"Passivo NC","empréstimo lp":"Passivo NC",
            "financiamentos lp":"Passivo NC","passivo nc":"Passivo NC",
            "capital social":"PL","reservas":"PL",
            "lucros acumulados":"PL","resultado exercício":"PL","resultado do exercício":"PL",
        }
    elif tipo=="FLUXO":
        mapa={
            "receita serviç":"Centro de Custos Entradas 1","receita de serviç":"Centro de Custos Entradas 1",
            "venda de serviç":"Centro de Custos Entradas 1","prestação de serviç":"Centro de Custos Entradas 1",
            "receita produt":"Centro de Custos Entradas 2","venda de produt":"Centro de Custos Entradas 2",
            "venda de mercadoria":"Centro de Custos Entradas 2","faturamento":"Centro de Custos Entradas 2",
            "receb":"Centro de Custos Entradas 3","cobrança":"Centro de Custos Entradas 3",
            "boleto":"Centro de Custos Entradas 3","cartão":"Centro de Custos Entradas 3",
            "pix recebido":"Centro de Custos Entradas 3",
            "receita financ":"Centro de Custos Entradas 4","juros recebidos":"Centro de Custos Entradas 4",
            "rendimento":"Centro de Custos Entradas 4","aplicação financeira":"Centro de Custos Entradas 4",
            "outras receitas":"Centro de Custos Entradas 4","outras entradas":"Centro de Custos Entradas 4",
            "comiss":"Centro de Custos Entradas 4","repasse":"Centro de Custos Entradas 4",
            "aporte":"Centro de Custos Entradas 4","empréstimo recebido":"Centro de Custos Entradas 4",
            "folha":"Centro de Custos Saidas 1","salári":"Centro de Custos Saidas 1",
            "pró-labore":"Centro de Custos Saidas 1","encargos":"Centro de Custos Saidas 1",
            "benefícios":"Centro de Custos Saidas 1","13º":"Centro de Custos Saidas 1",
            "férias":"Centro de Custos Saidas 1","rescis":"Centro de Custos Saidas 1",
            "fornecedor":"Centro de Custos Saidas 2","compra de mercadoria":"Centro de Custos Saidas 2",
            "compra de insumo":"Centro de Custos Saidas 2","matéria-prima":"Centro de Custos Saidas 2",
            "pagamento a fornecedor":"Centro de Custos Saidas 2",
            "impostos":"Centro de Custos Saidas 3","tributos":"Centro de Custos Saidas 3",
            "desp. operac":"Centro de Custos Saidas 3","despesas operac":"Centro de Custos Saidas 3",
            "aluguel":"Centro de Custos Saidas 3","energia":"Centro de Custos Saidas 3",
            "água":"Centro de Custos Saidas 3","internet":"Centro de Custos Saidas 3",
            "telefon":"Centro de Custos Saidas 3","contabilidade":"Centro de Custos Saidas 3",
            "manutenção":"Centro de Custos Saidas 3","marketing":"Centro de Custos Saidas 3",
            "publicidade":"Centro de Custos Saidas 3","honorári":"Centro de Custos Saidas 3",
            "taxa bancária":"Centro de Custos Saidas 3","tarifa bancária":"Centro de Custos Saidas 3",
            "investimentos":"Centro de Custos Saidas 4","empréstimo pago":"Centro de Custos Saidas 4",
            "financiamento pago":"Centro de Custos Saidas 4","amortização":"Centro de Custos Saidas 4",
            "compra de ativo":"Centro de Custos Saidas 4","imobilizado":"Centro de Custos Saidas 4",
        }
    else:
        return []

    acumulado={}
    detalhamento=[]
    for _,row in df.iterrows():
        conta_raw=str(row.get(col_conta,"")).strip()
        conta=conta_raw.lower()
        ano=str(row.get(col_ano,"")).strip()
        mes_raw=str(row.get(col_mes,"")).lower().strip()[:3]
        mes=meses_pt.get(mes_raw,mes_raw)
        filial=str(row.get(col_fil,"")).strip() if col_fil else None
        if not conta or not ano or not mes: continue
        campo_dest=None
        for k,v in mapa.items():
            if k in conta: campo_dest=v; break
        if not campo_dest: continue
        try:
            v=abs(float(row.get(col_valor,0)))
            chave=(ano,mes,filial,campo_dest)
            acumulado[chave]=acumulado.get(chave,0)+v
            detalhamento.append({"ano":ano,"mes":mes,"filial":filial,"campo_pai":campo_dest,
                                 "subconta":conta_raw,"valor":v})
        except: pass

    if tipo=="FLUXO":
        campos_entrada_fl=["Centro de Custos Entradas 1","Centro de Custos Entradas 2",
                           "Centro de Custos Entradas 3","Centro de Custos Entradas 4"]
        campos_saida_fl=["Centro de Custos Saidas 1","Centro de Custos Saidas 2",
                         "Centro de Custos Saidas 3","Centro de Custos Saidas 4"]
        periodos_vistos_fl=set((a,m,f) for (a,m,f,c) in acumulado.keys())
        for (ano_fl,mes_fl,fil_fl) in periodos_vistos_fl:
            soma_ent_fl=sum(acumulado.get((ano_fl,mes_fl,fil_fl,c),0) for c in campos_entrada_fl)
            soma_sai_fl=sum(acumulado.get((ano_fl,mes_fl,fil_fl,c),0) for c in campos_saida_fl)
            acumulado[(ano_fl,mes_fl,fil_fl,"Disponibilidades entradas")]=soma_ent_fl
            acumulado[(ano_fl,mes_fl,fil_fl,"Disponibilidades Saida")]=soma_sai_fl

    celulas=[{"ano":a,"mes":m,"filial":fl,"campo":cp,"valor":v} for (a,m,fl,cp),v in acumulado.items()]
    if detalhamento:
        celulas.append({"_detalhamento":detalhamento})
    return celulas

def detectar_tipo(df, nome_arquivo=None):
    cols_str=" ".join(str(c).lower() for c in df.columns)
    vals_str=" ".join(str(v).lower() for v in df.iloc[:,0].dropna().astype(str).tolist()[:20])
    # Se existir coluna "Conta" (formato longo), usa o conteúdo dela também — tem mais sinal
    col_conta=next((c for c in df.columns if str(c).lower().strip() in ["conta","descrição","descricao"]),None)
    vals_extra=""
    if col_conta is not None:
        vals_extra=" ".join(str(v).lower() for v in df[col_conta].dropna().astype(str).unique().tolist()[:60])
    txt=cols_str+" "+vals_str+" "+vals_extra

    # Pontuação por evidência: termos FORTES são exclusivos de uma demonstração e valem mais;
    # termos FRACOS são mais genéricos e podem aparecer em mais de uma (ex: "fornecedor" e "cliente"
    # existem tanto como saldo no Balanço quanto como pagamento/recebimento no Fluxo — por isso
    # esses dois termos ambíguos saíram da lista do Balanço, para não colidir com Fluxo).
    termos_dre_fortes=["receita bruta","cmv","lucro bruto","lucro líquido","ebitda",
                        " dre","dre -","dre gerencial","receita líquida","receita liquida"]
    termos_dre_fracos=["faturamento","custo da mercadoria","despesas comerciais"]

    termos_balanco_fortes=["patrimônio líquido","ativo circulante","passivo circulante",
                            "ativo não circulante","passivo não circulante","posição patrimonial",
                            "saldo inicial","saldo final","balanço patrimonial"]
    termos_balanco_fracos=["ativo","passivo","caixa e banco","estoque","imobilizado","capital social"]

    termos_fluxo_fortes=["fluxo de caixa","recebimento de client","pagamento a fornecedor",
                          "entrada de caixa","saída de caixa","saida de caixa","desembolso",
                          "recebimentos","pagamentos","saldo acumulado"]
    termos_fluxo_fracos=["entradas","saídas","saidas","total entradas","total saídas",
                          "recebimento","pagamento"]

    def _pontuar(fortes,fracos):
        return sum(2 for p in fortes if p in txt)+sum(1 for p in fracos if p in txt)

    pontos_dre=_pontuar(termos_dre_fortes,termos_dre_fracos)
    pontos_balanco=_pontuar(termos_balanco_fortes,termos_balanco_fracos)
    pontos_fluxo=_pontuar(termos_fluxo_fortes,termos_fluxo_fracos)

    # Sinal adicional (não exclusivo) — nome do arquivo, quando disponível, reforça a pontuação
    # de conteúdo, mas nunca decide sozinho (arquivo sem sinal de nome cai só no conteúdo, como antes)
    if nome_arquivo:
        nome_l=str(nome_arquivo).lower()
        if "fluxo" in nome_l or "caixa" in nome_l: pontos_fluxo+=2
        if "balan" in nome_l: pontos_balanco+=2
        if "dre" in nome_l or "resultado" in nome_l: pontos_dre+=2

    pontos={"DRE":pontos_dre,"BALANCO":pontos_balanco,"FLUXO":pontos_fluxo}
    melhor=max(pontos,key=pontos.get)
    if pontos[melhor]==0: return "DESCONHECIDO"
    return melhor

def extrair_periodos_colunas(df):
    import re
    meses_pt={"jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,
               "jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12,
               "janeiro":1,"fevereiro":2,"março":3,"abril":4,"maio":5,"junho":6,
               "julho":7,"agosto":8,"setembro":9,"outubro":10,"novembro":11,"dezembro":12}
    periodos={}
    for col in df.columns:
        s=str(col).lower().strip()
        m=re.match(r'([a-zçã]+)[/\-\s](\d{2,4})',s)
        if m:
            mes_nome=m.group(1)[:3]
            ano_s=m.group(2)
            if mes_nome in meses_pt:
                ano=int(ano_s) if len(ano_s)==4 else 2000+int(ano_s)
                periodos[col]=(str(ano),mes_nome)
                continue
        for nome in meses_pt:
            if s.startswith(nome[:3]):
                periodos[col]=("?",nome[:3]); break
    return periodos

def parser_dre(df):
    periodos=extrair_periodos_colunas(df)
    if not periodos: return []
    import re as re_local
    ano_detectado=None
    for col in list(df.columns)[:3]:
        m=re_local.search(r'(20\d{2})',str(col))
        if m: ano_detectado=m.group(1); break
    if not ano_detectado:
        for val in df.iloc[:3].values.flatten():
            m=re_local.search(r'(20\d{2})',str(val))
            if m: ano_detectado=m.group(1); break
    if ano_detectado:
        periodos={col:(ano_detectado,mes) for col,(ano,mes) in periodos.items()}
    mapa={
        # Receitas
        "receita bruta":"receita bruta de vendas",
        "faturamento":"receita bruta de vendas",
        "receita da venda":"receita bruta de vendas",
        "serviços recorrentes":"receita bruta de vendas",
        "serviços avulsos":"receita bruta de vendas",
        "venda de mercadorias":"receita bruta de vendas",
        # Deduções
        "imposto sobre vendas":"impostos sobre vendas",
        "imposto":"impostos sobre vendas",
        "deduç":"impostos sobre vendas",
        "devoluç":"devoluções de vendas",
        "abatimento":"devoluções de vendas",
        # CMV / Custo Variável (soma de subcontas)
        "cmv":"CMV (custo da mercadoria vendida)",
        "custo da mercadoria":"CMV (custo da mercadoria vendida)",
        "custo dos serviços":"CMV (custo da mercadoria vendida)",
        "custo do produto":"CMV (custo da mercadoria vendida)",
        "custo variável":"CMV (custo da mercadoria vendida)",
        "insumos":"CMV (custo da mercadoria vendida)",
        "embalagem":"CMV (custo da mercadoria vendida)",
        "frete":"CMV (custo da mercadoria vendida)",
        # Despesas Comerciais
        "despesas comerciais":"despesas comerciais",
        "desp comerci":"despesas comerciais",
        "brindes":"despesas comerciais",
        "feiras e eventos":"despesas comerciais",
        "marketing":"despesas comerciais",
        # Despesas Administrativas (Pessoal + Estruturais + Administrativas)
        "despesas com pessoal":"despesas administrativas",
        "salários":"despesas administrativas",
        "encargos":"despesas administrativas",
        "benefícios":"despesas administrativas",
        "adicionais":"despesas administrativas",
        "despesas estruturais":"despesas administrativas",
        "aluguel":"despesas administrativas",
        "manutenção":"despesas administrativas",
        "infra tech":"despesas administrativas",
        "despesas administrativas":"despesas administrativas",
        "serviços adm":"despesas administrativas",
        "mat. escritorio":"despesas administrativas",
        "viagens e locomoção":"despesas administrativas",
        "desp admin":"despesas administrativas",
        # Despesas Financeiras
        "desp financ":"despesas financeiras líquidas",
        "juros pagos":"despesas financeiras líquidas",
        "juros recebidos":"receitas não operacionais",
        # Depreciação
        "deprecia":"despesas com depreciações e amortizações",
        # Totais (apenas para referência/validação, não obrigatórios)
        "lucro bruto":"lucro bruto",
        "margem bruta":"lucro bruto",
        "margem de contribuição":"lucro bruto",
        "lucro operacional":"lucro operacional",
        "ebitda":"EBITDA",
        "lucro líquido":"lucro líquido",
        "resultado líquido":"lucro líquido",
        # Não operacionais
        "rec/desp.não operacionais":"receitas não operacionais",
        # IR/CSLL
        "imposto de renda":"provisão para imposto de renda",
        "csll":"provisão para contribuição social",
        # Distribuição
        "distribuição de lucro":"pró-labore/distribuição de lucro",
        "aportes":"pró-labore/distribuição de lucro",
    }
    # Detecta automaticamente qual coluna tem o texto das contas
    # (a que tiver mais valores de texto não-numéricos)
    col_desc=df.columns[0]
    melhor_score=0
    for col in df.columns[:3]:
        try:
            vals=df[col].dropna().astype(str)
            score=sum(1 for v in vals if len(v)>3 and not v.replace(".","").replace(",","").replace("-","").isdigit())
            if score>melhor_score:
                melhor_score=score; col_desc=col
        except: pass

    # Acumula valores por (ano,mes,campo) para somar subcontas que mapeiam pro mesmo destino
    acumulado={}
    detalhamento=[]  # guarda subcontas originais para drill-down
    totais_evitar=["receita bruta de vendas","lucro bruto","lucro operacional","EBITDA",
                   "lucro líquido","CMV (custo da mercadoria vendida)"]
    # Linhas que são SUBcontas (detalhamento) somam; linhas TOTAIS (maiúsculas) não somam, usam direto
    import sys
    for _,row in df.iterrows():
        desc_raw=str(row.get(col_desc,"")).strip()
        desc=desc_raw.lower()
        if not desc or desc=="nan": continue
        campo_dest=None
        for k,v in mapa.items():
            if k in desc: campo_dest=v; break
        if "devoluç" in desc:
            print(f"DEBUG ACHEI LINHA: desc={desc!r} campo_dest={campo_dest!r}",file=sys.stderr)
        if not campo_dest: continue
        e_total=desc_raw.isupper() and len(desc_raw)>3
        for col,(ano,mes) in periodos.items():
            try:
                val=row.get(col,None)
                if val is None or str(val).strip() in ["","nan","-"]: continue
                v=float(str(val).replace(",",".").replace("%","").replace(" ",""))
                if v!=0 and abs(v)<2: continue
                chave=(ano,mes,campo_dest)
                if e_total and campo_dest in totais_evitar:
                    acumulado[chave]=abs(v)
                else:
                    acumulado[chave]=acumulado.get(chave,0)+abs(v)
                    if not e_total:
                        detalhamento.append({"ano":ano,"mes":mes,"campo_pai":campo_dest,
                                             "subconta":desc_raw,"valor":abs(v)})
            except: pass

    import sys
    n_dev=sum(1 for (a,m,c) in acumulado.keys() if c=="devoluções de vendas")
    print(f"DEBUG: chaves de devolucoes no acumulado = {n_dev}", file=sys.stderr)
    celulas=[{"ano":a,"mes":m,"campo":c,"valor":v} for (a,m,c),v in acumulado.items()]
    if detalhamento:
        celulas.append({"_detalhamento":detalhamento})
    return celulas

def parser_balanco(df, capturar_originais=False):
    periodos=extrair_periodos_colunas(df)
    if not periodos:
        if capturar_originais: return [], {}
        return []
    mapa={
        "caixa e bancos":"disponibilidades saldo",
        "caixa":"disponibilidades saldo",
        "banco":"disponibilidades saldo",
        "disponib":"disponibilidades saldo",
        "aplicaç":"disponibilidades saldo",
        "cliente":"contas a receber saldo",
        "receber":"contas a receber saldo",
        "duplicatas a receber":"contas a receber saldo",
        "estoque":"estoque final do mês de mercadorias para revenda saldo",
        "impostos a recuperar":"Outros AC",
        "tributos a recuperar":"Outros AC",
        "outros ac":"Outros AC",
        "imobilizado":"Ativo NC",
        "intangível":"Ativo NC",
        "investimentos":"Ativo NC",
        "ativo nc":"Ativo NC",
        "fornecedor":"contas a pagar de fornecedores saldo",
        "pagar a fornecedores":"contas a pagar de fornecedores saldo",
        "empréstimos cp":"Passivos Financeiros",
        "empréstimo cp":"Passivos Financeiros",
        "financiamentos cp":"Passivos Financeiros",
        "financiamento":"Passivos Financeiros",
        "salários e encargos":"Outros PC",
        "impostos a recolher":"Outros PC",
        "tributos a pagar":"Outros PC",
        "outros pc":"Outros PC",
        "empréstimos lp":"Passivo NC",
        "empréstimo lp":"Passivo NC",
        "financiamentos lp":"Passivo NC",
        "passivo nc":"Passivo NC",
        "capital social":"PL",
        "reservas":"PL",
        "lucros acumulados":"PL",
        "resultado exercício":"PL",
        "resultado do exercício":"PL",
        "ativo total":"ativo total saldo",
        "passivo total":"passivo total saldo",
    }
    col_desc=df.columns[0]
    melhor_score=0
    for col in df.columns[:3]:
        try:
            vals=df[col].dropna().astype(str)
            score=sum(1 for v in vals if len(v)>3 and not v.replace(".","").replace(",","").replace("-","").isdigit())
            if score>melhor_score:
                melhor_score=score; col_desc=col
        except: pass

    acumulado={}
    originais={}
    detalhamento=[]
    totais_evitar=["ativo total saldo","passivo total saldo"]

    for _,row in df.iterrows():
        desc_raw=str(row.get(col_desc,"")).strip()
        desc=desc_raw.lower()
        if not desc or desc=="nan": continue
        campo_dest=None
        for k,v in mapa.items():
            if k in desc: campo_dest=v; break
        if not campo_dest: continue
        e_total=desc_raw.isupper() and len(desc_raw)>3
        for col,(ano,mes) in periodos.items():
            try:
                val=row.get(col,None)
                if val is None or str(val).strip() in ["","nan","-"]: continue
                v=float(str(val).replace(",","."))
                chave=(ano,mes,campo_dest)
                if e_total and campo_dest in totais_evitar:
                    acumulado[chave]=v
                    if capturar_originais:
                        originais[chave]={"linha":desc_raw,"coluna":str(col),"valor_celula":val}
                else:
                    acumulado[chave]=acumulado.get(chave,0)+v
                    if not e_total:
                        detalhamento.append({"ano":ano,"mes":mes,"campo_pai":campo_dest,
                                             "subconta":desc_raw,"valor":v})
            except: pass

    celulas=[{"ano":a,"mes":m,"campo":c,"valor":v} for (a,m,c),v in acumulado.items()]
    if detalhamento:
        celulas.append({"_detalhamento":detalhamento})
    if capturar_originais:
        return celulas, originais
    return celulas

def parser_fluxo(df, capturar_originais=False):
    meses_pt={"jan":"jan","fev":"fev","mar":"mar","abr":"abr","mai":"mai","jun":"jun",
               "jul":"jul","ago":"ago","set":"set","out":"out","nov":"nov","dez":"dez",
               "janeiro":"jan","fevereiro":"fev","março":"mar","abril":"abr","maio":"mai",
               "junho":"jun","julho":"jul","agosto":"ago","setembro":"set",
               "outubro":"out","novembro":"nov","dezembro":"dez"}
    mapa={
        # Entradas - Receita de serviços/produtos
        "receita serviç":"Centro de Custos Entradas 1","receita de serviç":"Centro de Custos Entradas 1",
        "venda de serviç":"Centro de Custos Entradas 1","prestação de serviç":"Centro de Custos Entradas 1",
        "receita produt":"Centro de Custos Entradas 2","venda de produt":"Centro de Custos Entradas 2",
        "venda de mercadoria":"Centro de Custos Entradas 2","faturamento":"Centro de Custos Entradas 2",
        # Entradas - Recebimentos
        "receb":"Centro de Custos Entradas 3","cobrança":"Centro de Custos Entradas 3",
        "boleto":"Centro de Custos Entradas 3","cartão":"Centro de Custos Entradas 3",
        "pix recebido":"Centro de Custos Entradas 3",
        # Entradas - Financeiras/Outras
        "receita financ":"Centro de Custos Entradas 4","juros recebidos":"Centro de Custos Entradas 4",
        "rendimento":"Centro de Custos Entradas 4","aplicação financeira":"Centro de Custos Entradas 4",
        "outras receitas":"Centro de Custos Entradas 4","outras entradas":"Centro de Custos Entradas 4",
        "comiss":"Centro de Custos Entradas 4","repasse":"Centro de Custos Entradas 4",
        "aporte":"Centro de Custos Entradas 4","empréstimo recebido":"Centro de Custos Entradas 4",
        # Saídas - Folha
        "folha":"Centro de Custos Saidas 1","salári":"Centro de Custos Saidas 1",
        "pró-labore":"Centro de Custos Saidas 1","encargos":"Centro de Custos Saidas 1",
        "benefícios":"Centro de Custos Saidas 1","13º":"Centro de Custos Saidas 1",
        "férias":"Centro de Custos Saidas 1","rescis":"Centro de Custos Saidas 1",
        # Saídas - Fornecedores
        "fornecedor":"Centro de Custos Saidas 2","compra de mercadoria":"Centro de Custos Saidas 2",
        "compra de insumo":"Centro de Custos Saidas 2","matéria-prima":"Centro de Custos Saidas 2",
        "pagamento a fornecedor":"Centro de Custos Saidas 2",
        # Saídas - Operacionais/Impostos
        "impostos":"Centro de Custos Saidas 3","tributos":"Centro de Custos Saidas 3",
        "desp. operac":"Centro de Custos Saidas 3","despesas operac":"Centro de Custos Saidas 3",
        "aluguel":"Centro de Custos Saidas 3","energia":"Centro de Custos Saidas 3",
        "água":"Centro de Custos Saidas 3","internet":"Centro de Custos Saidas 3",
        "telefon":"Centro de Custos Saidas 3","contabilidade":"Centro de Custos Saidas 3",
        "manutenção":"Centro de Custos Saidas 3","marketing":"Centro de Custos Saidas 3",
        "publicidade":"Centro de Custos Saidas 3","honorári":"Centro de Custos Saidas 3",
        "taxa bancária":"Centro de Custos Saidas 3","tarifa bancária":"Centro de Custos Saidas 3",
        # Saídas - Investimentos
        "investimentos":"Centro de Custos Saidas 4","empréstimo pago":"Centro de Custos Saidas 4",
        "financiamento pago":"Centro de Custos Saidas 4","amortização":"Centro de Custos Saidas 4",
        "compra de ativo":"Centro de Custos Saidas 4","imobilizado":"Centro de Custos Saidas 4",
    }
    totais_mapa={
        "total entradas":"Disponibilidades entradas",
        "total saídas":"Disponibilidades Saida","total saidas":"Disponibilidades Saida",
    }
    # Palavras que indicam claramente ENTRADA ou SAÍDA, usadas só pra decidir o destino
    # de colunas não reconhecidas pelo mapa principal (evita perder dinheiro de vista)
    pistas_entrada=["receita","entrada","recebiment","venda","faturamento"]
    pistas_saida=["despesa","saida","saída","pagamento","custo","gasto"]

    col_ano=next((c for c in df.columns if str(c).lower() in ["ano","year"]),None)
    col_mes=next((c for c in df.columns if str(c).lower() in ["mês","mes","month"]),None)
    if not col_ano or not col_mes:
        if capturar_originais: return [], {}
        return []

    acumulado={}
    originais={}
    nao_reconhecidas=[]  # guarda {coluna, ano, mes, valor, destino} pra aviso na tela
    for _,row in df.iterrows():
        ano=str(row.get(col_ano,"")).strip()
        mes_raw=str(row.get(col_mes,"")).lower().strip()[:3]
        mes=meses_pt.get(mes_raw,mes_raw)
        if not ano or not mes or ano=="nan": continue
        for col in df.columns:
            if col in [col_ano,col_mes]: continue
            col_l=str(col).lower().strip()
            try:
                v=float(str(row.get(col,0)).replace(",","."))
            except:
                continue
            if v==0: continue
            # Verifica se é um TOTAL do arquivo (usa direto, não soma) — só para auditoria/referência
            campo_total=None
            for k,dest in totais_mapa.items():
                if k in col_l: campo_total=dest; break
            if campo_total:
                chave=(ano,mes,campo_total)
                if capturar_originais:
                    originais[chave]={"linha":col_l,"coluna":str(col),"valor_celula":v}
                continue
            # Senão, é uma categoria — soma no campo de destino
            campo_dest=None
            for k,dest in mapa.items():
                if k in col_l: campo_dest=dest; break
            if not campo_dest:
                # Coluna não reconhecida — NUNCA descarta o dinheiro.
                # Decide se é Entrada ou Saída por pista textual; se não achar pista, assume Entrada
                # quando valor positivo (mais comum) e avisa para revisão.
                eh_saida=any(p in col_l for p in pistas_saida)
                campo_dest="Centro de Custos Saidas 4" if eh_saida else "Centro de Custos Entradas 4"
                nao_reconhecidas.append({"coluna":str(col),"ano":ano,"mes":mes,
                                         "valor":v,"destino":"Saída" if eh_saida else "Entrada"})
            chave=(ano,mes,campo_dest)
            acumulado[chave]=acumulado.get(chave,0)+v

    # Calcula os totais (Disponibilidades entradas/Saida) somando as categorias mapeadas
    totais_calculados={}
    campos_entrada=["Centro de Custos Entradas 1","Centro de Custos Entradas 2",
                    "Centro de Custos Entradas 3","Centro de Custos Entradas 4"]
    campos_saida=["Centro de Custos Saidas 1","Centro de Custos Saidas 2",
                  "Centro de Custos Saidas 3","Centro de Custos Saidas 4"]
    periodos_vistos=set((a,m) for (a,m,c) in acumulado.keys())
    for (ano,mes) in periodos_vistos:
        soma_ent=sum(acumulado.get((ano,mes,c),0) for c in campos_entrada)
        soma_sai=sum(acumulado.get((ano,mes,c),0) for c in campos_saida)
        totais_calculados[(ano,mes,"Disponibilidades entradas")]=soma_ent
        totais_calculados[(ano,mes,"Disponibilidades Saida")]=soma_sai

    # Detalhamento para drill-down: cada categoria mapeada vira subconta do campo consolidado
    nomes_legiveis={
        "Centro de Custos Entradas 1":"Receita Serviços","Centro de Custos Entradas 2":"Receita Produtos",
        "Centro de Custos Entradas 3":"Recebimento Clientes","Centro de Custos Entradas 4":"Receita Financeira/Outras",
        "Centro de Custos Saidas 1":"Folha","Centro de Custos Saidas 2":"Fornecedores",
        "Centro de Custos Saidas 3":"Impostos/Operacionais","Centro de Custos Saidas 4":"Investimentos",
    }
    detalhamento=[]
    for (ano,mes,campo),valor in acumulado.items():
        if campo in campos_entrada:
            detalhamento.append({"ano":ano,"mes":mes,"campo_pai":"Disponibilidades entradas",
                                 "subconta":nomes_legiveis.get(campo,campo),"valor":valor})
        elif campo in campos_saida:
            detalhamento.append({"ano":ano,"mes":mes,"campo_pai":"Disponibilidades Saida",
                                 "subconta":nomes_legiveis.get(campo,campo),"valor":valor})

    celulas=[{"ano":a,"mes":m,"campo":c,"valor":v} for (a,m,c),v in acumulado.items()]
    celulas+=[{"ano":a,"mes":m,"campo":c,"valor":v} for (a,m,c),v in totais_calculados.items()]
    if detalhamento:
        celulas.append({"_detalhamento":detalhamento})
    if nao_reconhecidas:
        celulas.append({"_nao_reconhecidas":nao_reconhecidas})
    if capturar_originais:
        return celulas, originais
    return celulas

def mapear_celulas_brutas(celulas_brutas, tipo):
    """Recebe células BRUTAS no formato {ano, mes, conta(texto original), valor}
    e aplica o motor de regras determinístico (mesmo mapa do parser_dre/balanco)
    para somar subcontas no campo de destino padrão do sistema.
    Isso garante que a soma de subcontas SEMPRE seja feita por regra, nunca por IA."""

    if tipo=="DRE":
        mapa={
            "receita bruta":"receita bruta de vendas","faturamento":"receita bruta de vendas",
            "receita da venda":"receita bruta de vendas","receita de vendas":"receita bruta de vendas",
            "receita de serviços":"receita bruta de vendas","receita de serviço":"receita bruta de vendas",
            "serviços recorrentes":"receita bruta de vendas","serviços avulsos":"receita bruta de vendas",
            "venda de mercadorias":"receita bruta de vendas","vendas brutas":"receita bruta de vendas",
            "receita total":"receita bruta de vendas","outras receitas operacionais":"receita bruta de vendas",
            "imposto sobre vendas":"impostos sobre vendas","imposto":"impostos sobre vendas",
            "deduç":"impostos sobre vendas",
            "devoluç":"devoluções de vendas","abatimento":"devoluções de vendas",
            "desconto concedido":"devoluções de vendas","descontos concedidos":"devoluções de vendas",
            "cmv":"CMV (custo da mercadoria vendida)","custo da mercadoria":"CMV (custo da mercadoria vendida)",
            "custo dos serviços":"CMV (custo da mercadoria vendida)","custo do produto":"CMV (custo da mercadoria vendida)",
            "custo variável":"CMV (custo da mercadoria vendida)","insumos":"CMV (custo da mercadoria vendida)",
            "embalagem":"CMV (custo da mercadoria vendida)","frete":"CMV (custo da mercadoria vendida)",
            "mão de obra direta":"CMV (custo da mercadoria vendida)","materiais e insumos":"CMV (custo da mercadoria vendida)",
            "fretes de produção":"CMV (custo da mercadoria vendida)",
            "despesas comerciais":"despesas comerciais","desp comerci":"despesas comerciais",
            "brindes":"despesas comerciais","feiras e eventos":"despesas comerciais","marketing":"despesas comerciais",
            "salários comerciais":"despesas comerciais","comissões":"despesas comerciais",
            "viagens comerciais":"despesas comerciais","marketing e publicidade":"despesas comerciais",
            "despesas com pessoal":"despesas administrativas","salários":"despesas administrativas",
            "encargos":"despesas administrativas","benefícios":"despesas administrativas",
            "aluguel":"despesas administrativas","manutenção":"despesas administrativas",
            "despesas administrativas":"despesas administrativas","serviços adm":"despesas administrativas",
            "salários administrativos":"despesas administrativas","pró-labore":"despesas administrativas",
            "energia elétrica":"despesas administrativas","internet e telefonia":"despesas administrativas",
            "contabilidade":"despesas administrativas","sistemas e licenças":"despesas administrativas",
            "material de escritório":"despesas administrativas","seguros":"despesas administrativas",
            "treinamentos":"despesas administrativas","desp admin":"despesas administrativas",
            "desp financ":"despesas financeiras líquidas","juros pagos":"despesas financeiras líquidas",
            "despesas financeiras":"despesas financeiras líquidas","juros bancários":"despesas financeiras líquidas",
            "tarifas bancárias":"despesas financeiras líquidas",
            "juros recebidos":"receitas não operacionais","receitas financeiras":"receitas não operacionais",
            "juros ativos":"receitas não operacionais",
            "deprecia":"despesas com depreciações e amortizações",
            "imposto de renda":"provisão para imposto de renda","irpj":"provisão para imposto de renda",
            "csll":"provisão para contribuição social",
            "distribuição de lucro":"pró-labore/distribuição de lucro",
        }
    elif tipo=="BALANCO":
        mapa={
            "caixa":"disponibilidades saldo","banco":"disponibilidades saldo",
            "disponib":"disponibilidades saldo","aplicaç":"disponibilidades saldo",
            "caixa e bancos":"disponibilidades saldo",
            "cliente":"contas a receber saldo","receber":"contas a receber saldo",
            "estoque":"estoque final do mês de mercadorias para revenda saldo",
            "impostos a recuperar":"Outros AC","tributos a recuperar":"Outros AC","outros ac":"Outros AC",
            "imobilizado":"Ativo NC","intangível":"Ativo NC","investimentos":"Ativo NC","ativo nc":"Ativo NC",
            "fornecedor":"contas a pagar de fornecedores saldo",
            "empréstimos cp":"Passivos Financeiros","empréstimo cp":"Passivos Financeiros",
            "financiamentos cp":"Passivos Financeiros","financiamento":"Passivos Financeiros",
            "salários e encargos":"Outros PC","impostos a recolher":"Outros PC","outros pc":"Outros PC",
            "empréstimos lp":"Passivo NC","empréstimo lp":"Passivo NC",
            "financiamentos lp":"Passivo NC","passivo nc":"Passivo NC",
            "capital social":"PL","reservas":"PL","lucros acumulados":"PL",
            "resultado exercício":"PL","resultado do exercício":"PL",
            "ativo total":"ativo total saldo","passivo total":"passivo total saldo",
        }
    elif tipo=="FLUXO":
        mapa={
            "total entradas":"Disponibilidades entradas",
            "receita serviç":"Centro de Custos Entradas 1","receita produt":"Centro de Custos Entradas 2",
            "receb":"Centro de Custos Entradas 3","outras receitas":"Centro de Custos Entradas 4",
            "total saídas":"Disponibilidades Saida","total saidas":"Disponibilidades Saida",
            "folha":"Centro de Custos Saidas 1","fornecedor":"Centro de Custos Saidas 2",
            "aluguel":"Centro de Custos Saidas 3","desp. operac":"Centro de Custos Saidas 4",
        }
    else:
        return []

    meses_pt={"jan":"jan","fev":"fev","mar":"mar","abr":"abr","mai":"mai","jun":"jun",
               "jul":"jul","ago":"ago","set":"set","out":"out","nov":"nov","dez":"dez",
               "janeiro":"jan","fevereiro":"fev","março":"mar","abril":"abr","maio":"mai",
               "junho":"jun","julho":"jul","agosto":"ago","setembro":"set",
               "outubro":"out","novembro":"nov","dezembro":"dez"}

    acumulado={}
    for c in celulas_brutas:
        conta_raw=str(c.get("conta","")).strip()
        conta=conta_raw.lower()
        mes_raw=str(c.get("mes","")).lower().strip()[:3]
        mes=meses_pt.get(mes_raw,mes_raw)
        ano=str(c.get("ano",""))
        if not conta or not mes or not ano: continue
        campo_dest=None
        for k,v in mapa.items():
            if k in conta: campo_dest=v; break
        if not campo_dest: continue
        try:
            v=abs(float(c.get("valor",0)))
            chave=(ano,mes,campo_dest)
            acumulado[chave]=acumulado.get(chave,0)+v
        except: pass

    return [{"ano":a,"mes":m,"campo":cp,"valor":v} for (a,m,cp),v in acumulado.items()]



    prompt=f"""Você é um extrator de dados financeiros. NÃO interprete, NÃO some, NÃO mapeie nada —
apenas extraia cada linha de dado bruto exatamente como está no arquivo.

TIPO DO DOCUMENTO: {tipo}

ARQUIVO:
{amostra}

INSTRUÇÕES:
1. Para cada valor numérico no arquivo, identifique: ANO, MÊS, nome da CONTA (exatamente como
   escrito no arquivo, sem traduzir ou simplificar) e o VALOR
2. Ignore linhas de total/subtotal SE houver linhas de detalhe que já somam ao mesmo resultado
   (evite duplicar). Se não tiver certeza, inclua a linha mesmo assim.
3. Ignore colunas de %, médias, ou texto sem valor numérico
4. Mantenha valores negativos como estão (não inverta sinais)
5. mês deve ser um dos: jan/fev/mar/abr/mai/jun/jul/ago/set/out/nov/dez (ou nome completo)

Retorne SOMENTE este JSON, sem texto adicional, sem markdown:
{{"celulas":[{{"ano":"2024","mes":"jan","conta":"Receita de Vendas","valor":552000}}]}}"""
    try:
        txt=""
        if not api_key.startswith("sk-ant-") and OPENAI_OK:
            client=OpenAI(api_key=api_key)
            r=client.chat.completions.create(model="gpt-4o",max_tokens=32000,
              messages=[{"role":"system","content":"Você é um extrator literal de dados. Retorne APENAS JSON válido sem markdown."},
                        {"role":"user","content":prompt}])
            txt=r.choices[0].message.content.strip()
        elif ANTHROPIC_OK:
            client=anthropic.Anthropic(api_key=api_key)
            r=client.messages.create(model="claude-sonnet-4-6",max_tokens=32000,
              system="Você é um extrator literal de dados. Retorne APENAS JSON válido sem markdown.",
              messages=[{"role":"user","content":prompt}])
            txt=r.content[0].text.strip()
        txt=re.sub(r'^```json\s*','',txt); txt=re.sub(r'\s*```$','',txt)
        return json.loads(txt).get("celulas",[])
    except Exception as e:
        return []

def _chamar_ia_extracao_bruta_lote(df_cli, api_key, tipo):
    amostra=df_cli.to_string(max_cols=df_cli.shape[1])
    prompt=f"""Você é um extrator de dados financeiros. NÃO interprete, NÃO some, NÃO mapeie nada —
apenas extraia cada linha de dado bruto exatamente como está no arquivo.

TIPO DO DOCUMENTO: {tipo}

ARQUIVO:
{amostra}

INSTRUÇÕES:
1. Para cada valor numérico no arquivo, identifique: ANO, MÊS, nome da CONTA (exatamente como
   escrito no arquivo, sem traduzir ou simplificar) e o VALOR
2. Ignore linhas de total/subtotal SE houver linhas de detalhe que já somam ao mesmo resultado
   (evite duplicar). Se não tiver certeza, inclua a linha mesmo assim.
3. Ignore colunas de %, médias, ou texto sem valor numérico
4. Mantenha valores negativos como estão (não inverta sinais)
5. mês deve ser um dos: jan/fev/mar/abr/mai/jun/jul/ago/set/out/nov/dez (ou nome completo)
6. Extraia TODOS os registros do arquivo, não pare antes do final

Retorne SOMENTE este JSON, sem texto adicional, sem markdown:
{{"celulas":[{{"ano":"2024","mes":"jan","conta":"Receita de Vendas","valor":552000}}]}}"""
    try:
        txt=""
        if not api_key.startswith("sk-ant-") and OPENAI_OK:
            client=OpenAI(api_key=api_key)
            r=client.chat.completions.create(model="gpt-4o",max_tokens=32000,
              messages=[{"role":"system","content":"Você é um extrator literal de dados. Retorne APENAS JSON válido sem markdown."},
                        {"role":"user","content":prompt}])
            txt=r.choices[0].message.content.strip()
        elif ANTHROPIC_OK:
            client=anthropic.Anthropic(api_key=api_key)
            r=client.messages.create(model="claude-sonnet-4-6",max_tokens=32000,
              system="Você é um extrator literal de dados. Retorne APENAS JSON válido sem markdown.",
              messages=[{"role":"user","content":prompt}])
            txt=r.content[0].text.strip()
        txt=re.sub(r'^```json\s*','',txt); txt=re.sub(r'\s*```$','',txt)
        return json.loads(txt).get("celulas",[])
    except Exception as e:
        import sys
        print(f"ERRO em _chamar_ia_extracao_bruta_lote: {e}", file=sys.stderr)
        return []

def _chamar_ia_extracao_bruta(df_cli, api_key, tipo):
    col_ano_bruta=None
    for col in df_cli.columns[:3]:
        if str(col).lower().strip() in ["ano","year"]:
            col_ano_bruta=col; break
    if col_ano_bruta is None or df_cli.shape[0]<=300:
        n_linhas=min(len(df_cli),300)
        return _chamar_ia_extracao_bruta_lote(df_cli.head(n_linhas),api_key,tipo)
    todas_celulas=[]
    anos_unicos=df_cli[col_ano_bruta].dropna().unique()
    for ano_lote in anos_unicos:
        df_lote=df_cli[df_cli[col_ano_bruta]==ano_lote]
        celulas_lote=_chamar_ia_extracao_bruta_lote(df_lote,api_key,tipo)
        todas_celulas.extend(celulas_lote)
    return todas_celulas

def _chamar_ia_extracao(df_cli, campos, api_key, tipo):
    amostra=df_cli.head(60).to_string(max_cols=df_cli.shape[1])
    prompt=f"""Você é um especialista em contabilidade brasileira analisando uma demonstração financeira.

TIPO PROVÁVEL: {tipo} (pode ser DRE, Balanço Patrimonial ou Fluxo de Caixa — confirme pelo conteúdo)

ARQUIVO (pode ter qualquer layout — vertical, horizontal, com subcontas, com totais em destaque):
{amostra}

CAMPOS PADRÃO DO SISTEMA (mapeie os termos do arquivo para estes, mesmo que os nomes sejam diferentes,
sinônimos, abreviações, plural/singular, ou em outro idioma):
{json.dumps(campos,ensure_ascii=False)}

REGRAS IMPORTANTES:
1. Identifique ANO e MÊS de cada valor (mês: jan/fev/mar/abr/mai/jun/jul/ago/set/out/nov/dez)
2. Se o arquivo tem SUBCONTAS detalhadas que pertencem a uma conta principal (ex: "Insumos",
   "Embalagem", "Frete" todos compondo o CMV), SOME essas subcontas no campo de destino correto
3. Se o arquivo já tem TOTAIS calculados (ex: "RECEITA BRUTA" em destaque/maiúsculo), use esse
   valor direto em vez de tentar somar subcontas, para não duplicar
4. Ignore colunas de %AV (análise vertical), %AH (análise horizontal), médias, ou variações percentuais
5. Valores devem ser sempre positivos (remova sinais negativos de despesas/deduções)
6. Se o arquivo for horizontal (meses como colunas), extraia corretamente cada coluna-mês
7. Se o arquivo for vertical (meses como linhas), extraia corretamente cada linha-mês
8. O ANO pode estar implícito no título/cabeçalho da planilha (ex: "DRE 2024") — use-o para todos os meses
9. Tente mapear o MÁXIMO de contas possível, mesmo termos não óbvios — use seu conhecimento contábil
10. Reconheça variações comuns: "Faturamento"="Receita Bruta", "Custo da Mercadoria Vendida"="CMV",
    "Despesas com Pessoal"="Despesas Administrativas", "Resultado Líquido"="Lucro Líquido", etc.

Retorne SOMENTE este JSON, sem texto adicional, sem markdown:
{{"celulas":[{{"ano":"2024","mes":"jan","campo":"receita bruta de vendas","valor":1234567.89}}]}}"""
    try:
        txt=""
        if not api_key.startswith("sk-ant-") and OPENAI_OK:
            client=OpenAI(api_key=api_key)
            r=client.chat.completions.create(model="gpt-4o",max_tokens=16000,
              messages=[{"role":"system","content":"Você é um especialista contábil. Retorne APENAS JSON válido sem markdown."},
                        {"role":"user","content":prompt}])
            txt=r.choices[0].message.content.strip()
        elif ANTHROPIC_OK:
            client=anthropic.Anthropic(api_key=api_key)
            r=client.messages.create(model="claude-sonnet-4-6",max_tokens=16000,
              system="Você é um especialista contábil. Retorne APENAS JSON válido sem markdown.",
              messages=[{"role":"user","content":prompt}])
            txt=r.content[0].text.strip()
        txt=re.sub(r'^```json\s*','',txt); txt=re.sub(r'\s*```$','',txt)
        return json.loads(txt).get("celulas",[])
    except Exception as e:
        return []

def ia_extrair(df_cli, campos, api_key):
    tipo=detectar_tipo(df_cli)
    celulas_local=[]
    if tipo=="DRE": celulas_local=parser_dre(df_cli)
    elif tipo=="BALANCO": celulas_local=parser_balanco(df_cli)
    elif tipo=="FLUXO": celulas_local=parser_fluxo(df_cli)
    if len(celulas_local)>=10:
        return celulas_local
    if not api_key: return celulas_local
    amostra=df_cli.head(50).to_string(max_cols=30)
    prompt=f"""Especialista contabilidade brasileira. Analise o arquivo e extraia dados financeiros.

TIPO DETECTADO: {tipo}

ARQUIVO:
{amostra}

CAMPOS NECESSÁRIOS:
{json.dumps(campos[:30],ensure_ascii=False)}

REGRAS:
1. ANO e MÊS de cada valor (mês: jan/fev/mar/abr/mai/jun/jul/ago/set/out/nov/dez)
2. Ignore totais acumulados, %AV, %AH, médias
3. Valores reais — não percentuais
4. Se horizontal (meses nas colunas), gire corretamente
5. ANO pode estar no nome da coluna

Retorne SOMENTE JSON:
{{"celulas":[{{"ano":"2024","mes":"jan","campo":"receita bruta de vendas","valor":1234567.89}}]}}"""
    try:
        txt=""
        if not api_key.startswith("sk-ant-") and OPENAI_OK:
            client=OpenAI(api_key=api_key)
            r=client.chat.completions.create(model="gpt-4o",max_tokens=16000,
              messages=[{"role":"system","content":"Retorne APENAS JSON válido sem markdown."},
                        {"role":"user","content":prompt}])
            txt=r.choices[0].message.content.strip()
        elif ANTHROPIC_OK:
            client=anthropic.Anthropic(api_key=api_key)
            r=client.messages.create(model="claude-sonnet-4-6",max_tokens=16000,
              system="Retorne APENAS JSON válido sem markdown.",
              messages=[{"role":"user","content":prompt}])
            txt=r.content[0].text.strip()
        txt=re.sub(r'^```json\s*','',txt); txt=re.sub(r'\s*```$','',txt)
        try:
            celulas_ia=json.loads(txt).get("celulas",[])
        except json.JSONDecodeError:
            # Tenta corrigir problemas comuns: aspas simples, vírgula decimal em número, texto extra antes/depois do JSON
            txt_corrigido=txt
            m=re.search(r'\{.*\}',txt_corrigido,re.DOTALL)
            if m: txt_corrigido=m.group(0)
            txt_corrigido=re.sub(r"(?<=[:\[,]\s*)'([^']*)'(?=\s*[,\]}])",r'"\1"',txt_corrigido)
            try:
                celulas_ia=json.loads(txt_corrigido).get("celulas",[])
            except json.JSONDecodeError:
                st.warning("⚠️ A IA retornou um formato inesperado para este arquivo — usando apenas os dados já capturados pelo motor de regras. Considere revisar manualmente ou ajustar o formato do arquivo.")
                return celulas_local
        chaves_local={(c["ano"],c["mes"],c["campo"]) for c in celulas_local}
        for c in celulas_ia:
            if (c.get("ano"),c.get("mes"),c.get("campo")) not in chaves_local:
                celulas_local.append(c)
        return celulas_local
    except Exception as e:
        st.error(f"Erro IA: {e}"); return celulas_local

def celulas_to_df(celulas):
    reg={}
    tem_filial=False
    for c in celulas:
        fil=c.get("filial")
        if fil: tem_filial=True
        k=(str(c.get("ano","")),str(c.get("mes","")).lower()[:3],fil)
        if k not in reg:
            reg[k]={"Ano":k[0],"mês":k[1]}
            if fil: reg[k]["Filial"]=fil
        try:
            reg[k][c["campo"]]=float(c["valor"])
        except:
            reg[k][c["campo"]]=c["valor"]
    if not reg: return pd.DataFrame()
    df=pd.DataFrame(list(reg.values()))
    if tem_filial and "Filial" not in df.columns:
        df["Filial"]="(Todas as filiais)"
    elif tem_filial:
        df["Filial"]=df["Filial"].fillna("(Todas as filiais)")
    # Garante TODOS os campos de TODAS as demonstrações, mesmo que não tenham vindo neste arquivo
    todos_campos=set()
    for r in reg.values(): todos_campos.update(r.keys())
    todos_campos.update(TODOS)  # garante cobertura total do padrão do sistema
    for campo in todos_campos:
        if campo in ["Ano","mês","Filial"]: continue
        if campo not in df.columns: df[campo]=0.
        else: df[campo]=df[campo].fillna(0.)
    try:
        df["Data"]=pd.to_datetime(
            df["Ano"].astype(str)+"-"+
            df["mês"].astype(str).str.lower().str[:3].map(MES_NUM).fillna("01")+"-01",
            errors="coerce")
        df=df.sort_values("Data").reset_index(drop=True)
    except: pass
    return df

def identificar_demonstracao(celulas):
    if not celulas: return None
    campos_presentes={c["campo"] for c in celulas}
    melhor_demo=None; melhor_score=0
    for nome_demo,campos_demo in DEMONSTRACOES_CAMPOS.items():
        score=len(campos_presentes & set(campos_demo))
        if score>melhor_score:
            melhor_score=score; melhor_demo=nome_demo
    return melhor_demo

def merge_banco_por_demonstracao(df_existente, celulas_novas, nome_demo):
    print(f"[MERGE DRE/BAL/FLUXO] Chamado pra demonstração: {nome_demo}")
    print(f"[MERGE DRE/BAL/FLUXO] df_existente é None? {df_existente is None}")
    if df_existente is not None:
        print(f"[MERGE DRE/BAL/FLUXO] df_existente tem {len(df_existente)} linhas, colunas: {list(df_existente.columns)[:10]}...")
    df_novo=celulas_to_df(celulas_novas)
    print(f"[MERGE DRE/BAL/FLUXO] df_novo (recém importado) tem {len(df_novo)} linhas")
    if df_novo.empty: return df_existente
    campos_desta_demo=DEMONSTRACOES_CAMPOS.get(nome_demo,[])
    if df_existente is None or df_existente.empty:
        print(f"[MERGE DRE/BAL/FLUXO] ⚠️ ENTROU NO CAMINHO 'df_existente vazio' — vai descartar tudo que já existia!")
        for outra_demo,campos in DEMONSTRACOES_CAMPOS.items():
            if outra_demo==nome_demo: continue
            for c in campos:
                if c not in df_novo.columns: df_novo[c]=0.
        return df_novo
    df_existente=df_existente.copy()
    if "Ano" not in df_existente.columns or "mês" not in df_existente.columns:
        return df_novo
    df_existente["_fil_tmp"]=df_existente["Filial"].astype(str) if "Filial" in df_existente.columns else ""
    df_novo["_fil_tmp"]=df_novo["Filial"].astype(str) if "Filial" in df_novo.columns else ""
    df_existente["_chave"]=df_existente["Ano"].astype(str)+"_"+df_existente["mês"].astype(str).str.lower().str[:3]+"_"+df_existente["_fil_tmp"]
    df_novo["_chave"]=df_novo["Ano"].astype(str)+"_"+df_novo["mês"].astype(str).str.lower().str[:3]+"_"+df_novo["_fil_tmp"]
    chaves_existentes=set(df_existente["_chave"])
    chaves_novas=set(df_novo["_chave"])
    todas_chaves=chaves_existentes | chaves_novas
    linhas_final=[]
    for chave in sorted(todas_chaves):
        linha={}
        row_exist=df_existente[df_existente["_chave"]==chave]
        if len(row_exist)>0:
            # Preserva TUDO da linha existente (de qualquer demonstração), inclusive
            # os campos desta demo, que só serão sobrescritos abaixo se vierem dados novos
            linha.update(row_exist.iloc[0].to_dict())
            linha.pop("_fil_tmp",None)
        else:
            partes=chave.split("_",2)
            ano,mes=partes[0],partes[1]
            fil_val=partes[2] if len(partes)>2 else ""
            linha={"Ano":ano,"mês":mes}
            if fil_val: linha["Filial"]=fil_val
        row_novo=df_novo[df_novo["_chave"]==chave]
        if len(row_novo)>0:
            # Só sobrescreve os campos desta demo quando esta chave específica
            # tem dados novos — preserva o valor antigo se não vier nada novo aqui
            for c in campos_desta_demo:
                if c in row_novo.columns:
                    linha[c]=float(row_novo.iloc[0].get(c,0) or 0)
        for c in campos_desta_demo:
            if c not in linha: linha[c]=0.
        linhas_final.append(linha)
    df_final=pd.DataFrame(linhas_final)
    df_final=df_final.drop(columns=["_chave","_fil_tmp"],errors="ignore")
    if "Filial" in df_final.columns:
        df_final["Filial"]=df_final["Filial"].fillna("(Todas as filiais)")
    for nome_demo2,campos2 in DEMONSTRACOES_CAMPOS.items():
        for c in campos2:
            if c not in df_final.columns: df_final[c]=0.
            else: df_final[c]=df_final[c].fillna(0.)
    try:
        df_final["Data"]=pd.to_datetime(
            df_final["Ano"].astype(str)+"-"+
            df_final["mês"].astype(str).str.lower().str[:3].map(MES_NUM).fillna("01")+"-01",
            errors="coerce")
        df_final=df_final.sort_values("Data").reset_index(drop=True)
    except: pass
    return df_final

def gerar_cobertura(df, demonstracoes_campos):
    if df.empty: return pd.DataFrame()
    cm=cm_(df); ca=ca_(df)
    linhas=[]
    for _,row in df.iterrows():
        linha={"Ano":row.get(ca,""),"Mês":row.get(cm,"")}
        for nome_demo,campos_demo in demonstracoes_campos.items():
            preenchidos=sum(1 for c in campos_demo if c in df.columns and float(row.get(c,0) or 0)!=0)
            total=len(campos_demo)
            linha[nome_demo]="✅" if preenchidos>0 else "⚪"
            linha[f"{nome_demo}_qtd"]=f"{preenchidos}/{total}"
        linhas.append(linha)
    return pd.DataFrame(linhas)

# ═══════════════════════════════════════════════════
# CÁLCULOS
# ═══════════════════════════════════════════════════
def calcular(df):
    d=df.copy()
    d["deduções"]         = cn(d,"impostos sobre vendas")+cn(d,"devoluções de vendas")
    d["receita líquida"]  = cn(d,"receita bruta de vendas")-d["deduções"]
    d["lucro bruto"]      = d["receita líquida"]-cn(d,"CMV (custo da mercadoria vendida)")
    d["margem contrib"]   = d["lucro bruto"]-cn(d,"despesas comerciais")
    d["desp op"]          = (cn(d,"despesas comerciais")+cn(d,"despesas administrativas")+
                             cn(d,"despesas financeiras líquidas")+cn(d,"despesas com depreciações e amortizações"))
    d["lucro operacional"]= d["lucro bruto"]-d["desp op"]
    d["resultado IR"]     = d["lucro operacional"]+cn(d,"receitas não operacionais")-cn(d,"despesas não operacionais")
    d["lucro líquido"]    = d["resultado IR"]-cn(d,"provisão para imposto de renda")-cn(d,"provisão para contribuição social")
    d["EBITDA"]           = (d["lucro líquido"]+cn(d,"provisão para imposto de renda")+
                             cn(d,"provisão para contribuição social")+
                             cn(d,"despesas financeiras líquidas")+
                             cn(d,"despesas com depreciações e amortizações"))
    rl=d["receita líquida"].replace(0,np.nan)
    d["margem bruta %"]   = d["lucro bruto"]/rl*100
    d["margem contrib %"] = d["margem contrib"]/rl*100
    d["margem op %"]      = d["lucro operacional"]/rl*100
    d["margem líquida %"] = d["lucro líquido"]/rl*100
    d["EBITDA %"]         = d["EBITDA"]/rl*100
    d["ativo circ"]       = (cn(d,"disponibilidades saldo")+cn(d,"contas a receber saldo")+
                             cn(d,"estoque final do mês de mercadorias para revenda saldo")+cn(d,"Outros AC"))
    d["ativo total"]      = d["ativo circ"]+cn(d,"Ativo NC")
    d["pass circ"]        = (cn(d,"contas a pagar de fornecedores saldo")+cn(d,"Passivos Financeiros")+cn(d,"Outros PC"))
    d["pass total"]       = d["pass circ"]+cn(d,"Passivo NC")
    d["PL"]               = d["ativo total"]-d["pass total"]
    pc=d["pass circ"].replace(0,np.nan); pt=d["pass total"].replace(0,np.nan)
    pl=d["PL"].replace(0,np.nan);       at=d["ativo total"].replace(0,np.nan)
    ac=d["ativo circ"]
    est=cn(d,"estoque final do mês de mercadorias para revenda saldo")
    ei =cn(d,"estoque inicial do mês de mercadorias para revenda saldo")
    cmv=cn(d,"CMV (custo da mercadoria vendida)")
    cr =cn(d,"contas a receber saldo"); cp=cn(d,"contas a pagar de fornecedores saldo")
    d["liquidez corrente"] = ac/pc
    d["liquidez imediata"] = cn(d,"disponibilidades saldo")/pc
    d["ROE"]    = d["lucro líquido"]/pl*100
    d["kanitz"] = (0.05*(d["lucro líquido"]/pl.fillna(1))+
                   1.65*(at.fillna(0)/pt.fillna(1))+
                   3.55*((ac-est)/pc.fillna(1))+
                   1.06*(ac/pc.fillna(1))+
                   0.33*(pt.fillna(0)/pl.fillna(1)))
    d["PMR"]  = (cr/(rl.fillna(1)*12))*365
    compras   = (est+cmv-ei)*12
    d["PMP"]  = (cp/compras.replace(0,np.nan).fillna(1))*365
    giro      = cmv/((ei+est)/2).replace(0,np.nan)
    d["PME"]  = 30/giro.replace(0,np.nan)
    d["ciclo de caixa"] = d["PMR"]+d["PME"]-d["PMP"]
    d["giro estoque"]   = giro
    d["ticket médio"]   = rl/cn(d,"numero de vendas").replace(0,np.nan)
    pf=cn(d,"Passivos Financeiros"); disp=cn(d,"disponibilidades saldo")
    d["ICD"]  = d["EBITDA"]/(pf-disp).replace(0,np.nan)*100
    ent=cn(d,"Disponibilidades entradas"); sai=cn(d,"Disponibilidades Saida")
    d["saldo período"]   = ent-sai
    d["saldo acumulado"] = d["saldo período"].cumsum()
    scores=[]
    for _,row in d.iterrows():
        try:
            k=float(row.get("kanitz",0) or 0)
            k_norm=max(0,min(1,(k-(-4))/(0-(-4))))
            li=float(row.get("liquidez imediata",0) or 0)*100
            li_norm=max(0,min(1,li/5))
            eb=float(row.get("EBITDA %",0) or 0)
            eb_norm=max(0,min(1,eb/5)) if eb>0 else 0
            cc=float(row.get("ciclo de caixa",0) or 0)
            cc_norm=max(0,min(1,(120-cc)/120))
            s=(k_norm*0.25+li_norm*0.25+eb_norm*0.25+cc_norm*0.25)*100
        except: s=0.
        scores.append(round(max(0,min(100,s)),2))
    d["score_risco"]=scores
    d["ativo circ"]  = (cn(d,"disponibilidades saldo")+cn(d,"contas a receber saldo")+
                        cn(d,"estoque final do mês de mercadorias para revenda saldo")+cn(d,"Outros AC"))
    d["ativo total"] = d["ativo circ"]+cn(d,"Ativo NC")
    d["pass circ"]   = (cn(d,"contas a pagar de fornecedores saldo")+cn(d,"Passivos Financeiros")+cn(d,"Outros PC"))
    d["pass total"]  = d["pass circ"]+cn(d,"Passivo NC")
    d["PL"]          = d["ativo total"]-d["pass total"]
    pc=d["pass circ"].replace(0,np.nan); pt=d["pass total"].replace(0,np.nan)
    pl=d["PL"].replace(0,np.nan);       at=d["ativo total"].replace(0,np.nan)
    ac=d["ativo circ"]
    est=cn(d,"estoque final do mês de mercadorias para revenda saldo")
    ei =cn(d,"estoque inicial do mês de mercadorias para revenda saldo")
    cmv=cn(d,"CMV (custo da mercadoria vendida)")
    cr =cn(d,"contas a receber saldo"); cp=cn(d,"contas a pagar de fornecedores saldo")
    d["liquidez corrente"] = ac/pc
    d["liquidez imediata"] = cn(d,"disponibilidades saldo")/pc
    d["ROE"]    = d["lucro líquido"]/pl*100
    d["kanitz"] = (0.05*(d["lucro líquido"]/pl.fillna(1))+1.65*(at.fillna(0)/pt.fillna(1))+
                   3.55*((ac-est)/pc.fillna(1))-1.06*(ac/pc.fillna(1))-0.33*(pt.fillna(0)/pl.fillna(1)))
    d["PMR"] = (cr/(rl.fillna(1)*12))*365
    compras  = (est+cmv-ei)*12
    d["PMP"] = (cp/compras.replace(0,np.nan).fillna(1))*365
    giro     = cmv/((ei+est)/2).replace(0,np.nan)
    d["PME"] = 30/giro.replace(0,np.nan)
    d["ciclo de caixa"] = d["PMR"]+d["PME"]-d["PMP"]
    d["giro estoque"]   = giro
    d["ticket médio"]   = rl/cn(d,"numero de vendas").replace(0,np.nan)
    pf=cn(d,"Passivos Financeiros"); disp=cn(d,"disponibilidades saldo")
    d["ICD"] = d["EBITDA"]/(pf-disp).replace(0,np.nan)*100
    ent=cn(d,"Disponibilidades entradas"); sai=cn(d,"Disponibilidades Saida")
    d["saldo período"]   = ent-sai
    d["saldo acumulado"] = d["saldo período"].cumsum()
    scores=[]
    for _,row in d.iterrows():
        try:
            k=float(row.get("kanitz",0) or 0)
            k_norm=max(0,min(1,(k-(-4))/(0-(-4))))
            li=float(row.get("liquidez imediata",0) or 0)*100
            li_norm=max(0,min(1,li/5))
            eb=float(row.get("EBITDA %",0) or 0)
            eb_norm=max(0,min(1,eb/5)) if eb>0 else 0
            cc=float(row.get("ciclo de caixa",0) or 0)
            cc_norm=max(0,min(1,(120-cc)/120))
            s=(k_norm*0.25+li_norm*0.25+eb_norm*0.25+cc_norm*0.25)*100
        except: s=0.
        scores.append(round(max(0,min(100,s)),2))
    d["score_risco"]=scores
    return d

def score_label(s):
    try:
        v=float(s)
        if v>=80: return "🟢 Baixo Risco","g","#059669"
        if v>=60: return "🟡 Risco Moderado","y","#D97706"
        if v>=40: return "🟠 Alto Risco","y","#EA580C"
        if v>=20: return "🔴 Risco Crítico","r","#DC2626"
        return "⚫ Insolvência Iminente","r","#6B7280"
    except: return "—","","#6B7280"

def detectar_anomalias(serie,janela=3,mult=1.3):
    s=pd.to_numeric(serie,errors="coerce").fillna(0)
    if len(s)<4: return pd.Series([False]*len(s),index=s.index)
    mm=s.rolling(janela,center=True,min_periods=1).mean()
    std=s.rolling(janela,center=True,min_periods=1).std().fillna(1)
    return (s-mm).abs()>mult*std

# ═══════════════════════════════════════════════════
# ML
# ═══════════════════════════════════════════════════
def treinar(serie,modelo,n=6):
    try:
        s=pd.to_numeric(serie,errors="coerce").dropna()
        if len(s)<6: return None
        if modelo=="ARIMA" and STATS_OK:
            return ARIMA(s,order=(1,1,1)).fit().forecast(n)
        if modelo=="SARIMAX" and STATS_OK and len(s)>=24:
            return SARIMAX(s,order=(1,1,1),seasonal_order=(1,1,1,12)).fit(disp=False).forecast(n)
        if modelo=="ExponentialSmoothing" and STATS_OK:
            kw={"trend":"add","damped_trend":True}
            if len(s)>=24: kw.update({"seasonal":"add","seasonal_periods":12})
            return ExponentialSmoothing(s,**kw).fit().forecast(n)
        if modelo=="Holt" and STATS_OK:
            return Holt(s,damped_trend=True).fit().forecast(n)
        if modelo=="Média Móvel":
            mm=s.rolling(3).mean().iloc[-1]; return pd.Series([mm]*n)
        if modelo=="Croston":
            return croston_tsb_forecast(s,n,variante="croston")
        if modelo=="TSB":
            return croston_tsb_forecast(s,n,variante="tsb")
        if modelo=="Ensemble":
            modelos_base=[m for m,ok in MODELOS_ML.items() if ok and m!="Ensemble"]
            proj_ens,_=ensemble_forecast(s,modelos_base,n)
            return proj_ens
        if modelo=="Prophet" and PROPHET_OK:
            idx=pd.date_range(end=pd.Timestamp.now(),periods=len(s),freq="MS")
            df_p=pd.DataFrame({"ds":idx,"y":s.values})
            m=Prophet(changepoint_prior_scale=0.01,yearly_seasonality=True); m.fit(df_p)
            fut=m.make_future_dataframe(periods=n,freq="MS"); fc=m.predict(fut)
            return pd.Series(fc["yhat"].tail(n).values)
    except: pass
    return None

def treinar_backtest(serie,modelo,timeout_s=10):
    """Treina escondendo os últimos 6 meses e devolve (mse, previsao, real) juntos —
    permite reaproveitar o resultado sem retreinar o mesmo modelo duas vezes.
    Tem limite de tempo: se o modelo não convergir em timeout_s segundos
    (ex: SARIMAX travado numa série ruim), desiste desse modelo e segue pro próximo,
    em vez de travar o processamento inteiro pra sempre."""
    try:
        s=pd.to_numeric(serie,errors="coerce").dropna()
        if len(s)<14: return float("inf"),None,None
        tr,te=s.iloc[:-6],s.iloc[-6:]
        _fut=_ML_EXECUTOR.submit(treinar,tr,modelo,6)
        try:
            pr=_fut.result(timeout=timeout_s)
        except concurrent.futures.TimeoutError:
            return float("inf"),None,None
        if pr is None: return float("inf"),None,None
        mse=float(mean_squared_error(te.values,pr.values[:6]))
        return mse,pr.values[:6],te.values
    except:
        return float("inf"),None,None

def treinar_backtest_exog(serie, meses_pico, meses_promo):
    """Mesma lógica do treinar_backtest, mas pro SARIMAX com variável exógena
    (sazonalidade/promoção) — esconde os últimos 6 meses e testa contra eles,
    pra ficar comparável com os outros modelos no ranking."""
    try:
        s=pd.to_numeric(serie,errors="coerce").dropna()
        if len(s)<14: return float("inf"),None,None
        tr,te=s.iloc[:-6],s.iloc[-6:]
        exog_tr=montar_exog_calendario(tr.index, meses_pico, meses_promo)
        exog_te=montar_exog_calendario(te.index, meses_pico, meses_promo)
        pr,_=treinar_com_exog(tr, exog_tr, exog_te, 6)
        if pr is None: return float("inf"),None,None
        pr_vals=pr.values[:6] if hasattr(pr,"values") else list(pr)[:6]
        mse=float(mean_squared_error(te.values,pr_vals))
        return mse,pr_vals,te.values
    except:
        return float("inf"),None,None

def mse_modelo(serie,modelo):
    mse,_,_=treinar_backtest(serie,modelo)
    return mse

def melhor_modelo(serie,modelos):
    modelos_base=[m for m in modelos if m!="Ensemble"]
    res={}
    previsoes_bt={}
    real_bt=None
    for m in modelos_base:
        mse,prev,real=treinar_backtest(serie,m)
        res[m]=mse
        if prev is not None:
            previsoes_bt[m]=prev
            real_bt=real
    validos={k:v for k,v in res.items() if v<float("inf")}

    # Só tenta o Ensemble se: (1) foi pedido, (2) tem pelo menos 2 modelos válidos,
    # (3) os melhores estão PRÓXIMOS entre si (senão misturar só piora o resultado)
    if "Ensemble" in modelos and len(validos)>=2 and real_bt is not None:
        top3=sorted(validos,key=validos.get)[:3]
        melhor_erro=validos[top3[0]]
        pior_erro=validos[top3[-1]]
        proximos=melhor_erro>0 and (pior_erro-melhor_erro)/melhor_erro<0.30
        if proximos:
            pesos={m:1.0/(validos[m]+1e-6) for m in top3 if m in previsoes_bt}
            soma_pesos=sum(pesos.values())
            if soma_pesos>0:
                comb=np.zeros(len(real_bt))
                for m,w in pesos.items():
                    comb+=(w/soma_pesos)*previsoes_bt[m]
                mse_ens=float(mean_squared_error(real_bt,comb))
                res["Ensemble"]=mse_ens
                validos["Ensemble"]=mse_ens

    if not validos: return "Média Móvel",res
    return min(validos,key=validos.get),res

def ensemble_forecast(serie, modelos, n, top_k=3):
    """Combina até `top_k` modelos com menor erro de backtest, com peso maior
    para o que errou menos. Se menos de top_k modelos conseguirem treinar
    com sucesso, usa só os que funcionaram (nunca menos de 1)."""
    try:
        s = pd.to_numeric(serie, errors="coerce").dropna()
        if len(s) < 14:
            return None, None
        candidatos = [m for m in modelos if m != "Ensemble"]
        erros = {m: mse_modelo(s, m) for m in candidatos}
        validos = {k: v for k, v in erros.items() if v < float("inf")}
        if not validos:
            return None, None

        melhores = sorted(validos, key=validos.get)[:top_k]
        previsoes = {}
        for m in melhores:
            p = treinar(s, m, n)
            if p is not None:
                previsoes[m] = np.asarray(p.values, dtype=float)
        if not previsoes:
            return None, None

        pesos = {m: 1.0 / (validos[m] + 1e-6) for m in previsoes}
        soma_pesos = sum(pesos.values())
        combinado = np.zeros(n)
        for m, p in previsoes.items():
            combinado += (pesos[m] / soma_pesos) * p

        nomes_usados = list(previsoes.keys())
        return pd.Series(combinado), f"Ensemble ({'+'.join(nomes_usados)})"
    except Exception:
        return None, None

def parse_valor_brl(serie):
    """Converte texto em formato brasileiro (ex: 'R$ 1.234,56') para número.
    Remove símbolos de moeda/espaços, remove pontos (milhar) e troca vírgula por ponto (decimal)."""
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce")
    s = serie.astype(str).str.strip()
    s = s.str.replace("R$", "", regex=False)
    s = s.str.replace("$", "", regex=False)
    s = s.str.replace(" ", "", regex=False)
    s = s.str.replace("\xa0", "", regex=False)
    tem_virgula = s.str.contains(",", regex=False)
    s_br = s.where(~(tem_virgula), s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    resultado = pd.to_numeric(s_br, errors="coerce")
    if resultado.isna().all():
        resultado = pd.to_numeric(s, errors="coerce")
    return resultado
def extrair_familia_produto(descricao):
    """Remove variações de tamanho/medida do nome do produto, pra agrupar
    itens que são a mesma família (ex: 'CAMISA POLO BRANCA TAM.G' e
    'CAMISA POLO BRANCA TAM.M' viram a mesma família 'CAMISA POLO BRANCA')."""
    import re
    if not isinstance(descricao, str):
        return str(descricao)
    txt = descricao.upper().strip()
    # Remove padrões de tamanho: TAM.G, TAM P, TAMANHO GG, tamanhos numéricos (P38, 3", 5,5")
    txt = re.sub(r'\bTAM\.?\s*[A-Z0-9]{1,3}\b', '', txt)
    txt = re.sub(r'\bTAMANHO\s*[A-Z0-9]{1,3}\b', '', txt)
    txt = re.sub(r'\b(PP|P|M|G|GG|XG|XGG|UNICO|ÚNICO)\b\s*$', '', txt)
    txt = re.sub(r'\d+[.,]?\d*\s*(MM|CM|M|KG|G|L|ML|")\b', '', txt)
    txt = re.sub(r'\s+', ' ', txt).strip()
    return txt if txt else descricao.upper().strip()

def montar_coluna_familia(df, col_descricao):
    """Adiciona a coluna '_FamiliaProduto' ao dataframe, agrupando variações
    de tamanho/medida do mesmo item base."""
    df = df.copy()
    df["_FamiliaProduto"] = df[col_descricao].apply(extrair_familia_produto)
    return df

def pareto_analysis(df, dimensao, metrica):
    """Agrupa por `dimensao`, soma `metrica`, ordena e calcula % acumulado + classe ABC."""
    dim_limpa = df[dimensao].astype(str).str.strip() if df[dimensao].dtype == object else df[dimensao]
    valores = parse_valor_brl(df[metrica])
    g = (df.assign(_dim_pareto=dim_limpa, _valor_pareto=valores)
           .groupby("_dim_pareto", dropna=False)["_valor_pareto"]
           .sum().sort_values(ascending=False).reset_index())
    g.columns = [dimensao, "valor"]
    total = g["valor"].sum()
    g["pct"] = g["valor"] / total * 100 if total else 0
    g["pct_acumulado"] = g["pct"].cumsum()
    def classe(p):
        if p <= 80: return "A"
        if p <= 95: return "B"
        return "C"
    g["classe_abc"] = g["pct_acumulado"].apply(classe)
    return g

def grafico_pareto_plotly(resultado, titulo, top_n=20):
    """Gráfico de Pareto com identidade visual mais limpa e profissional."""
    NAVY = "#14243B"
    BRASS = "#A9762F"
    dim_col = resultado.columns[0]
    r = resultado.head(top_n).copy()
    x = r[dim_col].astype(str).str.slice(0, 18)

    fig = go.Figure()
    fig.add_trace(go.Bar(x=x, y=r["valor"], name="Valor", marker_color="#0F6E56",
      marker_line_width=0, yaxis="y1",
      text=[fmt(v) for v in r["valor"]], textposition="outside",
      textfont=dict(size=10, color="#374151"),
      hovertemplate="<b>%{x}</b><br>%{text}<extra></extra>"))
    fig.add_trace(go.Scatter(x=x, y=r["pct_acumulado"], name="% Acumulado",
      mode="lines+markers", line=dict(color=BRASS, width=2.5),
      marker=dict(size=6, color=BRASS, line=dict(color="white", width=1)),
      yaxis="y2", hovertemplate="Acumulado: %{y:.1f}%<extra></extra>"))
    fig.add_hline(y=80, line_dash="dot", line_color="#9CA3AF", opacity=0.6, yref="y2",
      annotation_text="80%", annotation_font=dict(size=9, color="#9CA3AF"))

    n_ticks = min(len(x), 15)
    fig.update_layout(
      title=dict(text=titulo, font=dict(size=15, family="Georgia, serif", color=NAVY)),
      plot_bgcolor="white", paper_bgcolor="white",
      font=dict(color="#6B7280", size=11, family="Segoe UI, Arial"),
      margin=dict(l=10, r=10, t=50, b=110),
      xaxis=dict(tickangle=-40, tickfont=dict(size=9, color="#4B5563"),
        showgrid=False, nticks=n_ticks, linecolor="#E5E7EB"),
      yaxis=dict(title=None, gridcolor="#F3F4F6", showline=False, zeroline=False,
        tickfont=dict(size=9)),
      yaxis2=dict(overlaying="y", side="right", range=[0, 108],
        showgrid=False, ticksuffix="%", tickfont=dict(size=9)),
      legend=dict(bgcolor="rgba(0,0,0,0)", orientation="h", y=-0.28, x=0.5,
        xanchor="center", font=dict(size=10)),
      hovermode="x unified", height=420,
      hoverlabel=dict(bgcolor="white", bordercolor="#E5E7EB", font=dict(color=NAVY)))
    return fig

def serie_mensal_produto(df, produto_col, produto_valor, data_col, metrica):
    """Série histórica mensal (soma da métrica por mês) de um único produto."""
    sub = df[df[produto_col] == produto_valor].copy()
    sub["_data"] = pd.to_datetime(sub[data_col], errors="coerce", dayfirst=True)
    sub = sub.dropna(subset=["_data"])
    sub["_periodo"] = sub["_data"].dt.to_period("M")
    sub["_valor_serie"] = parse_valor_brl(sub[metrica])
    serie = sub.groupby("_periodo")["_valor_serie"].sum().sort_index()
    serie.index = serie.index.to_timestamp()
    return serie

def croston_tsb_forecast(serie, n, alpha=0.1, beta=0.1, variante="croston"):
    """Croston (variante='croston') ou TSB (variante='tsb') — feitos para demanda
    intermitente (muitos meses com valor zero, intercalados com picos).
    TSB corrige o viés do Croston original em séries com tendência de queda."""
    valores = np.asarray(serie.values, dtype=float)
    n_obs = len(valores)
    if n_obs < 4 or not (valores > 0).any():
        return None
    primeiro_idx = int(np.argmax(valores > 0))

    z = np.zeros(n_obs)
    p = np.zeros(n_obs)
    z[primeiro_idx] = valores[primeiro_idx]
    p[primeiro_idx] = 1.0

    if variante == "croston":
        q = 1
        for t in range(primeiro_idx + 1, n_obs):
            if valores[t] > 0:
                z[t] = alpha * valores[t] + (1 - alpha) * z[t - 1]
                p[t] = alpha * q + (1 - alpha) * p[t - 1]
                q = 1
            else:
                z[t] = z[t - 1]
                p[t] = p[t - 1]
                q += 1
        nivel_final = z[-1] / p[-1] if p[-1] > 0 else 0.0
    else:
        for t in range(primeiro_idx + 1, n_obs):
            ocorreu = 1.0 if valores[t] > 0 else 0.0
            if valores[t] > 0:
                z[t] = alpha * valores[t] + (1 - alpha) * z[t - 1]
            else:
                z[t] = z[t - 1]
            p[t] = beta * ocorreu + (1 - beta) * p[t - 1]
        nivel_final = z[-1] * p[-1]

    if nivel_final <= 0 or pd.isna(nivel_final):
        return None
    return pd.Series([nivel_final] * n)

def aplicar_correcao_precos(serie, reajustes):
    """Traz o histórico para a régua de preço 'de hoje': para cada ponto do histórico,
    aplica os reajustes que ainda iam acontecer depois daquele mês (efeito acumulado)."""
    if not reajustes:
        return serie
    s = serie.copy()
    for idx in s.index:
        periodo_idx = idx.to_period("M")
        fator = 1.0
        for rj in reajustes:
            try:
                rj_periodo = pd.Period(rj["data"])
                if periodo_idx < rj_periodo:
                    fator *= (1 + float(rj["pct"])/100)
            except Exception:
                pass
        s[idx] = s[idx] * fator
    return s

def montar_exog_calendario(indice_datas, meses_pico, meses_promo):
    """Monta as colunas exógenas (sazonal/promoção) alinhadas a um índice de datas mensais."""
    meses_nomes_exog = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    exog = pd.DataFrame(index=indice_datas)
    exog["sazonal_pico"] = [1 if meses_nomes_exog[d.month-1] in (meses_pico or []) else 0 for d in indice_datas]
    exog["promocao"] = [1 if meses_nomes_exog[d.month-1] in (meses_promo or []) else 0 for d in indice_datas]
    return exog

def treinar_com_exog(serie, exog_hist, exog_futuro, n):
    """Treina um SARIMAX considerando as variáveis exógenas (sazonalidade/promoção).
    Retorna (previsao, nome_modelo) ou (None, None) se falhar."""
    if not STATS_OK:
        return None, None
    try:
        ordem_sazonal = (1,1,1,12) if len(serie) >= 24 else (0,0,0,0)
        modelo = SARIMAX(serie, exog=exog_hist, order=(1,1,1), seasonal_order=ordem_sazonal,
                          enforce_stationarity=False, enforce_invertibility=False)
        fit = modelo.fit(disp=False)
        forecast = fit.forecast(n, exog=exog_futuro)
        return forecast, "SARIMAX (c/ sazonalidade+promoção)"
    except Exception:
        return None, None

def prever_top_produtos(df, produto_col, metrica, data_col, pct_top=0.20,
                         meses_previsao=3, min_periodos=6, modelos=None):
    """
    PRIMEIRO filtra só os produtos com histórico suficiente (>= min_periodos meses),
    DEPOIS classifica os que mais vendem dentro desse grupo e roda ML no Top `pct_top`.
    Assim nenhuma "vaga" do Top % é desperdiçada com produto sem dado suficiente.
    """
    df = df.copy()
    if df[produto_col].dtype == object:
        df[produto_col] = df[produto_col].astype(str).str.strip()
    modelos = modelos or [m for m,ok in MODELOS_ML.items() if ok]

    df["_data_ml"] = pd.to_datetime(df[data_col], errors="coerce", dayfirst=True)
    df["_periodo_ml"] = df["_data_ml"].dt.to_period("M")
    n_periodos_por_produto = df.dropna(subset=["_data_ml"]).groupby(produto_col)["_periodo_ml"].nunique()
    produtos_com_historico = n_periodos_por_produto[n_periodos_por_produto >= min_periodos].index

    df_elegiveis = df[df[produto_col].isin(produtos_com_historico)]
    ranking = pareto_analysis(df_elegiveis, produto_col, metrica)
    n_selecionar = max(1, int(np.ceil(len(ranking) * pct_top)))
    top_produtos = ranking.head(n_selecionar)[produto_col].tolist()

    linhas = []
    for prod in top_produtos:
        serie = serie_mensal_produto(df, produto_col, prod, data_col, metrica)
        melhor, _rank = melhor_modelo(serie, modelos)
        proj = treinar(serie, melhor, meses_previsao)
        ultimo_real = float(serie.iloc[-1])
        if proj is not None:
            prox = float(proj.iloc[0])
            var = safe(prox - ultimo_real, abs(ultimo_real)) * 100
            linhas.append({produto_col: prod, "n_periodos": len(serie),
                "modelo_escolhido": melhor, "ultimo_real": ultimo_real,
                "previsao": [round(v,2) for v in proj.tolist()],
                "var_pct_proximo_mes": round(var,1), "status": "ok",
                "rank": _rank})
        else:
            linhas.append({produto_col: prod, "n_periodos": len(serie),
                "modelo_escolhido": melhor, "ultimo_real": ultimo_real,
                "previsao": None, "var_pct_proximo_mes": None, "status": "falhou ao treinar",
                "rank": _rank})
    return pd.DataFrame(linhas)

def load_df(cid):
    p=os.path.join(PASTA,f"{gid(cid)}_dados.csv")
    if not os.path.exists(p): return None
    try: return pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
    except: return None

def path_vendas(cid): return os.path.join(PASTA,f"{gid(cid)}_vendas.csv")

def save_vendas_df(cid,df):
    df.to_csv(path_vendas(cid),sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_vendas_df(cid):
    p=path_vendas(cid)
    if not os.path.exists(p): return None
    try: return pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
    except: return None

def _limpar_texto_vendas(df):
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].astype(str).str.strip()
    return df

def get_vendas_df():
    if st.session_state.get("vendas_raw") is not None:
        return _limpar_texto_vendas(st.session_state.vendas_raw.copy())
    if st.session_state.cid:
        df=load_vendas_df(st.session_state.cid)
        if df is not None:
            df=_limpar_texto_vendas(df)
            st.session_state.vendas_raw=df
            return df
    return None
def path_estoque_compras(cid): return os.path.join(PASTA,f"{gid(cid)}_estoque_compras.csv")

def save_estoque_compras(cid,df):
    df.to_csv(path_estoque_compras(cid),sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_estoque_compras(cid):
    p=path_estoque_compras(cid)
    if not os.path.exists(p): return None
    try: return pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
    except: return None

def _sufixo_filial_compras(filial):
    if not filial or filial=="(Todas as filiais)": return "consolidado"
    return "".join(c if c.isalnum() else "_" for c in filial).strip("_").lower()

def path_resultado_compras(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_resultado_compras__{_sufixo_filial_compras(filial)}.csv")
def path_calendario_compras(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_calendario_compras__{_sufixo_filial_compras(filial)}.csv")
def path_morto_compras(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_morto_compras__{_sufixo_filial_compras(filial)}.csv")
def path_resultado_ml(cid): return os.path.join(PASTA,f"{gid(cid)}_resultado_ml.csv")

def _path_antigo_com_fallback(p_novo,p_antigo,filial):
    if os.path.exists(p_novo): return p_novo
    if (not filial or filial=="(Todas as filiais)") and os.path.exists(p_antigo): return p_antigo
    return p_novo

def save_resultado_compras(cid,df_res,df_cal,df_morto,filial=None):
    if df_res is not None: df_res.to_csv(path_resultado_compras(cid,filial),sep=";",decimal=",",index=False,encoding="utf-8-sig")
    if df_cal is not None: df_cal.to_csv(path_calendario_compras(cid,filial),sep=";",decimal=",",index=False,encoding="utf-8-sig")
    if df_morto is not None: df_morto.to_csv(path_morto_compras(cid,filial),sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_resultado_compras(cid,filial=None):
    res=cal=morto=None
    try:
        p=_path_antigo_com_fallback(path_resultado_compras(cid,filial),os.path.join(PASTA,f"{gid(cid)}_resultado_compras.csv"),filial)
        if os.path.exists(p): res=pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
    except: pass
    try:
        p=_path_antigo_com_fallback(path_calendario_compras(cid,filial),os.path.join(PASTA,f"{gid(cid)}_calendario_compras.csv"),filial)
        if os.path.exists(p): cal=pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
    except: pass
    try:
        p=_path_antigo_com_fallback(path_morto_compras(cid,filial),os.path.join(PASTA,f"{gid(cid)}_morto_compras.csv"),filial)
        if os.path.exists(p): morto=pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
    except: pass
    return res,cal,morto

def save_resultado_ml(cid,df):
    if df is not None:
        df_save=df.copy()
        for col in ["previsao","rank"]:
            if col in df_save.columns:
                df_save[col]=df_save[col].apply(lambda x: json.dumps(x) if x is not None else None)
        df_save.to_csv(path_resultado_ml(cid),sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_resultado_ml(cid):
    p=path_resultado_ml(cid)
    if not os.path.exists(p): return None
    try:
        df=pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
        for col in ["previsao","rank"]:
            if col in df.columns:
                df[col]=df[col].apply(lambda x: json.loads(x) if pd.notna(x) and x else None)
        return df
    except: return None

def get_estoque_compras():
    if st.session_state.get("compras_df_estoque") is not None:
        return st.session_state["compras_df_estoque"]
    if st.session_state.cid:
        df=load_estoque_compras(st.session_state.cid)
        if df is not None:
            st.session_state["compras_df_estoque"]=df
            return df
    return None
def path_scorecard_forn(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_scorecard_fornecedores__{_sufixo_filial(filial)}.csv")

def save_scorecard_forn(cid,df,filial=None):
    df.to_csv(path_scorecard_forn(cid,filial),sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_scorecard_forn(cid,filial=None):
    p=path_scorecard_forn(cid,filial)
    if not os.path.exists(p):
        _p_antigo=os.path.join(PASTA,f"{gid(cid)}_scorecard_fornecedores.csv")
        if (not filial or filial=="(Todas as filiais)") and os.path.exists(_p_antigo):
            p=_p_antigo
        else:
            return None
    try: return pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
    except: return None

    def path_config_ml(cid): return os.path.join(PASTA,f"{gid(cid)}_config_ml.json")

def path_config_ml(cid): return os.path.join(PASTA,f"{gid(cid)}_config_ml.json")

def save_config_ml(cid):
    campos=["cfgml_meses_pico","cfgml_meses_promo","cfgml_outliers","cfgml_reajustes",
            "cfgml_fatores_mercado","cfgml_meses_prev","cfgml_min_hist","cfgml_configs_ativas",
            "cfgml_escopo_tipo","cfgml_agrupar_familia","cfgml_nome_cenario","cfgml_limite_erro"]
    dados={c:st.session_state.get(c) for c in campos if c in st.session_state}
    _escopo_ml_atual=st.session_state.get("cfgml_escopo_tipo")
    if _escopo_ml_atual and "Categoria" in _escopo_ml_atual:
        dados["cfgml_categoria"]=st.session_state.get("cfgml_categoria")
    elif _escopo_ml_atual and "Produto específico" in _escopo_ml_atual:
        dados["cfgml_produto"]=st.session_state.get("cfgml_produto")
    with open(path_config_ml(cid),"w",encoding="utf-8") as f:
        json.dump(dados,f,ensure_ascii=False,default=str)

def load_config_ml(cid):
    p=path_config_ml(cid)
    if not os.path.exists(p): return None
    try:
        with open(p,"r",encoding="utf-8") as f:
            return json.load(f)
    except: return None

def path_config_compras(cid): return os.path.join(PASTA,f"{gid(cid)}_config_compras.json")

def save_config_compras(cid):
    _escopo_atual=st.session_state.get("compras_escopo_tipo")
    dados={"compras_escopo_tipo":_escopo_atual}
    if _escopo_atual and "Categoria" in _escopo_atual:
        dados["compras_categoria_sel"]=st.session_state.get("compras_categoria_sel")
    elif _escopo_atual and "Produto específico" in _escopo_atual:
        dados["compras_produtos_escopo"]=st.session_state.get("compras_produtos_escopo")
    with open(path_config_compras(cid),"w",encoding="utf-8") as f:
        json.dump(dados,f,ensure_ascii=False,default=str)

def load_config_compras(cid):
    p=path_config_compras(cid)
    if not os.path.exists(p): return None
    try:
        with open(p,"r",encoding="utf-8") as f:
            return json.load(f)
    except: return None    

def path_leadtime_compras(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_leadtime_compras__{_sufixo_filial_compras(filial)}.json")

def save_leadtime_compras(cid,filial=None):
    dados={
        "produto":[[k[0],k[1],v] for k,v in st.session_state.get("compras_leadtime_produto",{}).items()],
        "categoria":[[k[0],k[1],v] for k,v in st.session_state.get("compras_leadtime_categoria",{}).items()],
        "fornecedor":st.session_state.get("compras_leadtime_fornecedor",{}),
        "catalogo":st.session_state.get("compras_leadtime_catalogo",15),
        "ultimo_nivel":st.session_state.get("compras_nivel_lead"),
    }
    with open(path_leadtime_compras(cid,filial),"w",encoding="utf-8") as f:
        json.dump(dados,f,ensure_ascii=False,default=str)

def load_leadtime_compras(cid,filial=None):
    p=path_leadtime_compras(cid,filial)
    if not os.path.exists(p):
        _p_antigo=os.path.join(PASTA,f"{gid(cid)}_leadtime_compras.json")
        if (not filial or filial=="(Todas as filiais)") and os.path.exists(_p_antigo):
            p=_p_antigo
        else:
            return None
    try:
        with open(p,"r",encoding="utf-8") as f:
            dados=json.load(f)
        resultado={
            "compras_leadtime_produto":{(item[0],item[1]):item[2] for item in dados.get("produto",[])},
            "compras_leadtime_categoria":{(item[0],item[1]):item[2] for item in dados.get("categoria",[])},
            "compras_leadtime_fornecedor":dados.get("fornecedor",{}),
            "compras_leadtime_catalogo":dados.get("catalogo",15),
        }
        if dados.get("ultimo_nivel"):
            resultado["compras_nivel_lead"]=dados["ultimo_nivel"]
        return resultado
    except: return None    

def path_validacao_ml(cid): return os.path.join(PASTA,f"{gid(cid)}_validacao_ml.csv")

def save_validacao_ml(cid,df):
    if df is not None and not df.empty:
        df.to_csv(path_validacao_ml(cid),sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_validacao_ml(cid):
    p=path_validacao_ml(cid)
    if not os.path.exists(p): return None
    try: return pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
    except: return None

def path_ml_produtos_resultado(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_ml_produtos_resultado__{_sufixo_filial(filial)}.csv")

def save_ml_produtos_resultado(cid,df,filial=None):
    if df is not None and not df.empty:
        df_save=df.copy()
        for col in ["previsao","rank"]:
            if col in df_save.columns:
                df_save[col]=df_save[col].apply(lambda x: json.dumps(x) if x is not None else None)
        df_save.to_csv(path_ml_produtos_resultado(cid,filial),sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_ml_produtos_resultado(cid,filial=None):
    p=path_ml_produtos_resultado(cid,filial)
    if not os.path.exists(p): return None
    try:
        df=pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
        for col in ["previsao","rank"]:
            if col in df.columns:
                df[col]=df[col].apply(lambda x: json.loads(x) if pd.notna(x) and x else None)
        return df
    except: return None    

def path_config_mlp(cid): return os.path.join(PASTA,f"{gid(cid)}_config_mlp.json")

def save_config_mlp(cid):
    dados={"mlp_pct":st.session_state.get("mlp_pct"),
           "mlp_meses":st.session_state.get("mlp_meses"),
           "mlp_minper":st.session_state.get("mlp_minper")}
    with open(path_config_mlp(cid),"w",encoding="utf-8") as f:
        json.dump(dados,f,ensure_ascii=False,default=str)

def load_config_mlp(cid):
    p=path_config_mlp(cid)
    if not os.path.exists(p): return None
    try:
        with open(p,"r",encoding="utf-8") as f:
            return json.load(f)
    except: return None

def path_config_mlp(cid): return os.path.join(PASTA,f"{gid(cid)}_config_mlp.json")

def save_config_mlp(cid):
    dados={"mlp_pct":st.session_state.get("mlp_pct"),
           "mlp_meses":st.session_state.get("mlp_meses"),
           "mlp_minper":st.session_state.get("mlp_minper")}
    with open(path_config_mlp(cid),"w",encoding="utf-8") as f:
        json.dump(dados,f,ensure_ascii=False,default=str)

def load_config_mlp(cid):
    p=path_config_mlp(cid)
    if not os.path.exists(p): return None
    try:
        with open(p,"r",encoding="utf-8") as f:
            return json.load(f)
    except: return None        

def path_parecer_ia(cid,tipo="comercial"):
    if tipo=="comercial": return os.path.join(PASTA,f"{gid(cid)}_parecer_ia_historico.json")
    return os.path.join(PASTA,f"{gid(cid)}_parecer_ia_historico__{tipo}.json")

def save_parecer_ia(cid,parecer_dict,tipo="comercial"):
    hist=load_parecer_ia_historico(cid,tipo)
    hist.insert(0,parecer_dict)
    hist=hist[:24]
    with open(path_parecer_ia(cid,tipo),"w",encoding="utf-8") as f:
        json.dump(hist,f,ensure_ascii=False)

def load_parecer_ia_historico(cid,tipo="comercial"):
    p=path_parecer_ia(cid,tipo)
    if not os.path.exists(p): return []
    try:
        with open(p,encoding="utf-8") as f: return json.load(f)
    except: return []

def delete_parecer_ia(cid,indice,tipo="comercial"):
    hist=load_parecer_ia_historico(cid,tipo)
    if 0<=indice<len(hist):
        hist.pop(indice)
        with open(path_parecer_ia(cid,tipo),"w",encoding="utf-8") as f:
            json.dump(hist,f,ensure_ascii=False)

def update_parecer_ia(cid,indice,novo_texto,tipo="comercial"):
    hist=load_parecer_ia_historico(cid,tipo)
    if 0<=indice<len(hist):
        hist[indice]["texto"]=novo_texto
        with open(path_parecer_ia(cid,tipo),"w",encoding="utf-8") as f:
            json.dump(hist,f,ensure_ascii=False)                

def _md_simples_para_html(texto):
    """Converte o texto do parecer (markdown simples: **negrito**, listas numeradas/bullet) em HTML."""
    import re
    linhas=[l.strip() for l in texto.split("\n")]
    html_partes=[]; buffer_lista=[]; tipo_lista=[None]
    def fecha_lista():
        if buffer_lista:
            tag="ol" if tipo_lista[0]=="numerada" else "ul"
            html_partes.append(f"<{tag}>"+"".join(buffer_lista)+f"</{tag}>")
            buffer_lista.clear(); tipo_lista[0]=None
    def proxima_nao_vazia(i):
        j=i+1
        while j<len(linhas) and not linhas[j]: j+=1
        return linhas[j] if j<len(linhas) else ""
    for i,ls in enumerate(linhas):
        if not ls:
            prox=proxima_nao_vazia(i)
            continua_numerada=tipo_lista[0]=="numerada" and re.match(r'^\d+\.\s+',prox)
            continua_bullet=tipo_lista[0]=="bullet" and re.match(r'^-\s+',prox)
            if not (continua_numerada or continua_bullet):
                fecha_lista()
            continue
        m_header=re.match(r'^\*\*(.+?)\*\*$',ls)
        if m_header:
            fecha_lista(); html_partes.append(f'<h3>{m_header.group(1)}</h3>'); continue
        m_num=re.match(r'^\d+\.\s+(.*)$',ls)
        if m_num:
            if tipo_lista[0]!="numerada": fecha_lista(); tipo_lista[0]="numerada"
            item=re.sub(r'\*\*(.+?)\*\*',r'<strong>\1</strong>',m_num.group(1))
            buffer_lista.append(f"<li>{item}</li>"); continue
        m_bul=re.match(r'^-\s+(.*)$',ls)
        if m_bul:
            if tipo_lista[0]!="bullet": fecha_lista(); tipo_lista[0]="bullet"
            item=re.sub(r'\*\*(.+?)\*\*',r'<strong>\1</strong>',m_bul.group(1))
            buffer_lista.append(f"<li>{item}</li>"); continue
        fecha_lista()
        p=re.sub(r'\*\*(.+?)\*\*',r'<strong>\1</strong>',ls)
        html_partes.append(f"<p>{p}</p>")
    fecha_lista()
    return "\n".join(html_partes)

def gerar_html_parecer(titulo,subtitulo,texto,nome_empresa="",cor_faixa="#14243B",cor_clara="#C9A876"):
    """Gera um HTML pronto pra imprimir/salvar como PDF (Ctrl+P no navegador), com a identidade visual do app."""
    corpo_html=_md_simples_para_html(texto)
    data_geracao=datetime.now().strftime("%d/%m/%Y às %H:%M")
    return f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8">
<title>{titulo}</title>
<style>
  @page {{ margin: 2cm; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; color:#1F2937; max-width:860px; margin:0 auto; padding:0 20px 60px; background:#fff; }}
  .capa {{ background:linear-gradient(135deg,{cor_faixa} 0%,#0B1520 100%); color:#fff; padding:44px 40px;
           border-radius:14px; margin:30px 0 36px; box-shadow:0 4px 14px rgba(0,0,0,.15); }}
  .capa h1 {{ font-family:Georgia,serif; font-size:1.7rem; font-weight:600; margin:0 0 8px; }}
  .capa .sub {{ color:{cor_clara}; font-size:.95rem; margin:0 0 18px; }}
  .capa .meta {{ font-size:.8rem; color:#D1D5DB; border-top:1px solid rgba(255,255,255,.15); padding-top:14px; margin-top:14px; }}
  h3 {{ font-family:Georgia,serif; color:{cor_faixa}; font-size:1.15rem; margin:32px 0 12px; padding-bottom:8px;
        border-bottom:2px solid {cor_clara}; }}
  p {{ line-height:1.7; font-size:.94rem; margin:0 0 12px; text-align:justify; }}
  ol,ul {{ line-height:1.8; font-size:.94rem; margin:0 0 16px; padding-left:24px; }}
  li {{ margin-bottom:8px; }}
  strong {{ color:{cor_faixa}; }}
  .rodape {{ margin-top:50px; padding-top:16px; border-top:1px solid #E5E7EB; font-size:.72rem; color:#9CA3AF; text-align:center; }}
  h3 {{ page-break-after: avoid; page-break-inside: avoid; }}
  li, p {{ page-break-inside: avoid; }}
  @media print {{ .capa {{ -webkit-print-color-adjust:exact; print-color-adjust:exact; }} }}
</style></head>
<body>
  <div class="capa">
    <h1>{titulo}</h1>
    <div class="sub">{subtitulo}</div>
    <div class="meta">{nome_empresa+" · " if nome_empresa else ""}Gerado em {data_geracao}</div>
  </div>
  {corpo_html}
  <div class="rodape">NetExame Analytics BI — Parecer gerado com apoio de Inteligência Artificial, a partir de dados reais da operação.</div>
</body></html>"""

def path_snap(cid,nome,filial=None): return os.path.join(PASTA,f"{gid(cid)}_snap_{nome}__{_sufixo_filial(filial)}.json")

def save_snap(cid,nome,dados,filial=None):
    with open(path_snap(cid,nome,filial),"w",encoding="utf-8") as f:
        json.dump(dados,f,ensure_ascii=False,default=str)

def load_snap(cid,nome,filial=None):
    p=path_snap(cid,nome,filial)
    if not os.path.exists(p):
        _p_antigo=os.path.join(PASTA,f"{gid(cid)}_snap_{nome}.json")
        if (not filial or filial=="(Todas as filiais)") and os.path.exists(_p_antigo):
            p=_p_antigo
        else:
            return None
    try:
        with open(p,encoding="utf-8") as f: return json.load(f)
    except: return None

def path_pareto_snap(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_snap_pareto__{_sufixo_filial(filial)}.csv")

def save_pareto_snap(cid,df,filial=None):
    df.to_csv(path_pareto_snap(cid,filial),sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_pareto_snap(cid,filial=None):
    p=path_pareto_snap(cid,filial)
    if not os.path.exists(p):
        _p_antigo=os.path.join(PASTA,f"{gid(cid)}_snap_pareto.csv")
        if (not filial or filial=="(Todas as filiais)") and os.path.exists(_p_antigo):
            p=_p_antigo
        else:
            return None
    try: return pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
    except: return None

def _sufixo_filial(filial):
    if not filial or filial=="(Todas as filiais)": return "consolidado"
    return "".join(c if c.isalnum() else "_" for c in filial).strip("_").lower()

def path_cfgml_resultado(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_cfgml_resultado__{_sufixo_filial(filial)}.csv")

def save_cfgml_resultado(cid,df,filial=None):
    if df is not None and not df.empty:
        df_save=df.copy()
        for col in ["previsao","rank"]:
            if col in df_save.columns:
                df_save[col]=df_save[col].apply(lambda x: json.dumps(x) if x is not None else None)
        # Upsert por _ProdutoUnico: preserva previsões de OUTRAS categorias/escopos já salvos antes,
        # só substitui as linhas dos produtos que acabaram de ser recalculados agora.
        _p_cfgml=path_cfgml_resultado(cid,filial)
        if "_ProdutoUnico" in df_save.columns and os.path.exists(_p_cfgml):
            try:
                df_antigo_cfgml=pd.read_csv(_p_cfgml,sep=";",decimal=",",encoding="utf-8-sig")
                if "_ProdutoUnico" in df_antigo_cfgml.columns:
                    df_antigo_cfgml=df_antigo_cfgml[~df_antigo_cfgml["_ProdutoUnico"].isin(df_save["_ProdutoUnico"])]
                    df_save=pd.concat([df_antigo_cfgml,df_save],ignore_index=True)
            except: pass
        df_save.to_csv(_p_cfgml,sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_cfgml_resultado(cid,filial=None):
    p=path_cfgml_resultado(cid,filial)
    if not os.path.exists(p):
        _p_antigo=os.path.join(PASTA,f"{gid(cid)}_cfgml_resultado.csv")
        if (not filial or filial=="(Todas as filiais)") and os.path.exists(_p_antigo):
            p=_p_antigo
        else:
            return None
    try:
        df=pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
        for col in ["previsao","rank"]:
            if col in df.columns:
                df[col]=df[col].apply(lambda x: json.loads(x) if pd.notna(x) and x else None)
        return df
    except: return None

def path_cfgml_fora_previsao(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_cfgml_fora_previsao__{_sufixo_filial(filial)}.csv")

def save_cfgml_fora_previsao(cid,df,filial=None):
    if df is not None and not df.empty:
        df.to_csv(path_cfgml_fora_previsao(cid,filial),sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_cfgml_fora_previsao(cid,filial=None):
    p=path_cfgml_fora_previsao(cid,filial)
    if not os.path.exists(p):
        _p_antigo=os.path.join(PASTA,f"{gid(cid)}_cfgml_fora_previsao.csv")
        if (not filial or filial=="(Todas as filiais)") and os.path.exists(_p_antigo):
            p=_p_antigo
        else:
            return None
    try:
        return pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig")
    except: return None

def path_validacao_full(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_validacao_ml_full__{_sufixo_filial(filial)}.json")
def path_validacao_escopo(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_validacao_ml_escopo__{_sufixo_filial(filial)}.csv")

def save_validacao_full(cid,df_comp,df_escopo,produto_col,col_data,metrica,filial=None):
    dados={
        "df_comp": df_comp.to_dict(orient="records") if df_comp is not None else None,
        "produto_col": produto_col,
        "col_data": col_data,
        "metrica": metrica,
    }
    with open(path_validacao_full(cid,filial),"w",encoding="utf-8") as f:
        json.dump(dados,f,ensure_ascii=False,default=str)
    if df_escopo is not None:
        df_escopo_save=df_escopo.copy()
        for c in df_escopo_save.columns:
            if "period" in str(df_escopo_save[c].dtype).lower() or "datetime" in str(df_escopo_save[c].dtype).lower():
                df_escopo_save[c]=df_escopo_save[c].astype(str)
        df_escopo_save.to_csv(path_validacao_escopo(cid,filial),sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_validacao_full(cid,filial=None):
    p=path_validacao_full(cid,filial)
    if not os.path.exists(p):
        _p_antigo=os.path.join(PASTA,f"{gid(cid)}_validacao_ml_full.json")
        if (not filial or filial=="(Todas as filiais)") and os.path.exists(_p_antigo):
            p=_p_antigo
        else:
            return None
    try:
        with open(p,encoding="utf-8") as f: dados=json.load(f)
        df_comp=pd.DataFrame(dados["df_comp"]) if dados.get("df_comp") else None
        df_escopo=None
        pe=path_validacao_escopo(cid,filial)
        if not os.path.exists(pe):
            _pe_antigo=os.path.join(PASTA,f"{gid(cid)}_validacao_ml_escopo.csv")
            if (not filial or filial=="(Todas as filiais)") and os.path.exists(_pe_antigo):
                pe=_pe_antigo
        if os.path.exists(pe):
            try: df_escopo=pd.read_csv(pe,sep=";",decimal=",",encoding="utf-8-sig")
            except: df_escopo=None
        return {"df_comp":df_comp,"df_escopo":df_escopo,"produto_col":dados.get("produto_col"),
                "col_data":dados.get("col_data"),"metrica":dados.get("metrica")}
    except: return None

def path_fluxo_financeiro(cid,filial=None): return os.path.join(PASTA,f"{gid(cid)}_fluxo_financeiro__{_sufixo_filial(filial)}.json")

def save_fluxo_financeiro(cid,filial=None):
    dados={
        "ff_parcelas_pagar": st.session_state.get("ff_parcelas_pagar",[]),
        "ff_parcelas_receber": st.session_state.get("ff_parcelas_receber",[]),
        "ff_encargos_pct": st.session_state.get("ff_encargos_pct",0.0),
        "ff_saldo_inicial": st.session_state.get("ff_saldo_inicial",0.0),
        "ff_incluir_contas_pr": st.session_state.get("ff_incluir_contas_pr",False),
    }
    with open(path_fluxo_financeiro(cid,filial),"w",encoding="utf-8") as f:
        json.dump(dados,f,ensure_ascii=False)

def load_fluxo_financeiro(cid,filial=None):
    p=path_fluxo_financeiro(cid,filial)
    if not os.path.exists(p):
        _p_antigo=os.path.join(PASTA,f"{gid(cid)}_fluxo_financeiro.json")
        if (not filial or filial=="(Todas as filiais)") and os.path.exists(_p_antigo):
            p=_p_antigo
        else:
            return None
    try:
        with open(p,"r",encoding="utf-8") as f:
            return json.load(f)
    except: return None

def path_contas_pagar(cid): return os.path.join(PASTA,f"{gid(cid)}_ff_contas_pagar.csv")
def path_contas_receber(cid): return os.path.join(PASTA,f"{gid(cid)}_ff_contas_receber.csv")

def save_contas_pr(cid,tipo,df):
    p=path_contas_pagar(cid) if tipo=="pagar" else path_contas_receber(cid)
    df.to_csv(p,sep=";",decimal=",",index=False,encoding="utf-8-sig")

def load_contas_pr(cid,tipo):
    p=path_contas_pagar(cid) if tipo=="pagar" else path_contas_receber(cid)
    if not os.path.exists(p): return None
    try: return pd.read_csv(p,sep=";",decimal=",",encoding="utf-8-sig",parse_dates=["Vencimento"])
    except: return None


def path_cenarios(cid): return os.path.join(PASTA,f"{gid(cid)}_cenarios_ml.json")

def load_cenarios(cid):
    p=path_cenarios(cid)
    if not os.path.exists(p): return {}
    try:
        with open(p,encoding="utf-8") as f: return json.load(f)
    except: return {}

def save_cenario(cid, nome_cenario, config, resultado_resumo):
    cenarios=load_cenarios(cid)
    cenarios[nome_cenario]={
        "config": config,
        "resumo": resultado_resumo,
        "criado_em": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    with open(path_cenarios(cid),"w",encoding="utf-8") as f:
        json.dump(cenarios,f,ensure_ascii=False,indent=2)

def delete_cenario(cid, nome_cenario):
    cenarios=load_cenarios(cid)
    if nome_cenario in cenarios:
        del cenarios[nome_cenario]
        with open(path_cenarios(cid),"w",encoding="utf-8") as f:
            json.dump(cenarios,f,ensure_ascii=False,indent=2)

_cfg=load_cfg()
_defs={"pg":"boas_vindas","cid":None,
       "api_key":_cfg.get("anthropic_api_key",""),
       "df_raw":None,"projecoes":{},"saldo_ini":0.,
       "entradas_vista":0.,"freq_fluxo":"Mensal","log":[],
       "vendas_raw":None,"ml_produtos_resultado":None,
       "config_ml_atual":{}, "cenarios_comparar":[]}
for k,v in _defs.items():
    if k not in st.session_state: st.session_state[k]=v



# ═══════════════════════════════════════════════════
# PLOTLY
# ═══════════════════════════════════════════════════
TH=dict(plot_bgcolor="#0D1117",paper_bgcolor="#0D1117",
        font=dict(color="#6E7681",family="Inter"),
        xaxis=dict(gridcolor="#1C2333",linecolor="#1C2333",tickfont=dict(size=9)),
        yaxis=dict(gridcolor="#1C2333",linecolor="#1C2333",tickfont=dict(size=9)),
        margin=dict(l=8,r=8,t=32,b=8),
        legend=dict(bgcolor="#161B27",bordercolor="#21262D",font=dict(size=9)))

def gl(df,campos,titulo,cx=None):
    fig=go.Figure()
    for i,c in enumerate(campos):
        if c not in df.columns: continue
        x=df[cx] if cx and cx in df.columns else df.index
        fig.add_trace(go.Scatter(x=x,y=pd.to_numeric(df[c],errors="coerce"),name=c,
          mode="lines+markers",line=dict(color=CORES[i%len(CORES)],width=2),marker=dict(size=4)))
    fig.update_layout(title=dict(text=titulo,font=dict(size=12,color="#E6EDF3")),**TH)
    return fig

def gb(df,campo,titulo,cx=None):
    v=pd.to_numeric(df[campo],errors="coerce") if campo in df.columns else pd.Series()
    x=df[cx].astype(str) if cx and cx in df.columns else pd.Series(range(len(df))).astype(str)
    cs=["#00D4AA" if val>=0 else "#F85149" for val in v]
    fig=go.Figure(go.Bar(x=x,y=v,marker_color=cs,text=[fmt(val) for val in v],
      textposition="outside",textfont=dict(size=8)))
    fig.update_layout(title=dict(text=titulo,font=dict(size=12,color="#E6EDF3")),**TH)
    return fig

def gauge(val,titulo,mn=0,mx=100,c="#2176FF"):
    fig=go.Figure(go.Indicator(mode="gauge+number",value=val,
      title={"text":titulo,"font":{"color":"#C9D1D9","size":11}},
      gauge=dict(axis=dict(range=[mn,mx],tickcolor="#484F58"),bar=dict(color=c),
        bgcolor="#161B27",bordercolor="#21262D",
        steps=[dict(range=[mn,mn+(mx-mn)*.33],color="#F8514910"),
               dict(range=[mn+(mx-mn)*.33,mn+(mx-mn)*.66],color="#FFB62710"),
               dict(range=[mn+(mx-mn)*.66,mx],color="#00D4AA10")]),
      number=dict(font=dict(color="#E6EDF3",size=24))))
    fig.update_layout(paper_bgcolor="#161B27",font=dict(color="#6E7681"),
      margin=dict(l=16,r=16,t=44,b=8),height=190)
    return fig

# ═══════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════
_cfg=load_cfg()
_defs={"pg":"home","cid":None,
       "api_key":_cfg.get("anthropic_api_key",""),
       "df_raw":None,"projecoes":{},"saldo_ini":0.,
       "entradas_vista":0.,"freq_fluxo":"Mensal","log":[],
       "vendas_raw":None,"ml_produtos_resultado":None}
for k,v in _defs.items():
    if k not in st.session_state: st.session_state[k]=v

def ir(p): st.session_state.pg=p; st.rerun()
def limpar_sessao_cliente():
    """Remove dados de ML/Compras/Validação da sessão anterior, para não 'vazar'
    entre clientes diferentes quando o usuário troca de cliente sem reiniciar o app."""
    chaves_limpar = [
        "vendas_raw","ml_produtos_resultado","config_ml_atual",
        "cfgml_resultado_atual","cfgml_config_usada","cfgml_produto_col_usado",
        "cfgml_df_base_usado","cfgml_df_comp_bruto","cfgml_df_escopo_val",
        "cfgml_produto_col_val_usado","cfgml_col_data_val_usado","cfgml_metrica_val_usado",
        "compras_df_estoque","compras_resultado","compras_calendario","compras_morto",
        "compras_matriz","compras_leadtime_produto","compras_leadtime_categoria",
        "compras_leadtime_catalogo","compras_minmax_produto","compras_fam_map",
        "compras_leadtime_tabela","compras_leadtime_fornecedor",
        "pareto_resultado_atual","pareto_dim_atual","pareto_met_atual",
        "ff_parcelas_pagar","ff_parcelas_receber","ff_encargos_pct","ff_saldo_inicial","ff_config_carregada",
        "compras_escopo_tipo","compras_escopo_tipo_backup",
        "compras_categoria_sel","compras_categoria_sel_backup",
        "compras_produtos_escopo","compras_produtos_escopo_backup",
        "compras_nivel_lead","compras_nivel_lead_backup",
        "cfgml_escopo_tipo","cfgml_categoria","cfgml_produto","cfgml_escopo_tipo_backup",
        "cfgml_produtos_fora_previsao",
        "compras_filial_sel","compras_filial_sel_backup",
        "cfgml_filial_sel","cfgml_filial_sel_backup",
        "cfgml_filial_sel_pag","cfgml_filial_sel_pag_backup",
        "cfgml_df_comp_bruto","cfgml_df_escopo_val",
        "cfgml_produto_col_val_usado","cfgml_col_data_val_usado","cfgml_metrica_val_usado",
        "cfgml_meses_pico_backup","cfgml_meses_promo_backup",
        "cfgml_outliers_backup","cfgml_configs_ativas_backup",
        "cfgml_meses_pico","cfgml_meses_promo","cfgml_outliers","cfgml_configs_ativas",
        "mlp_filial_sel","mlp_filial_sel_backup",
        "mlp_pct","mlp_meses","mlp_minper","_mlp_config_sincronizada",
        "pareto_visao_backup",
        "ff_filial_sel","ff_filial_sel_backup","ff_expander_aberto","ff_checkbox_incluir_pr",
        "home_filial_sel","home_filial_sel_backup",
        "dre_filial_sel","dre_filial_sel_backup",
        "bal_filial_sel","bal_filial_sel_backup","dre_dd_expander_aberto","bal_dd_expander_aberto",
        "fx_filial_sel","fx_filial_sel_backup",
    ]
        
    
    for k in chaves_limpar:
        if k in st.session_state:
            del st.session_state[k]
def addlog(t,tp="ok"):
    i={"ok":"✅","w":"⚠️","e":"❌","i":"ℹ️"}.get(tp,"•")
    st.session_state.log.insert(0,f"{datetime.now().strftime('%H:%M')} {i} {t}")

def get_df():
    if st.session_state.df_raw is not None:
        return calcular(st.session_state.df_raw)
    if st.session_state.cid:
        df=load_df(st.session_state.cid)
        if df is not None:
            st.session_state.df_raw=df
            return calcular(df)
    return None

def montar_dre_categoria(filial_sel,categoria_sel):
    """Monta uma DRE 'como se fosse' só de uma categoria — SEM salvar nada em disco,
    sempre recalculada ao vivo a partir de 3 fontes:
    - DRE consolidada (rateada por participação de receita, pros custos sem detalhe por categoria)
    - Vendas (Receita REAL da categoria, soma direta)
    - Estoque (CMV REAL da categoria, Quantidade × Custo Unitário)
    Retorna (df_calculado, tem_cmv_real) ou (None, False) se faltar dado.
    """
    df_raw=get_df_raw_bruto()
    if df_raw is None or "Filial" not in df_raw.columns:
        return None,False
    if not filial_sel or filial_sel=="(Todas as filiais)":
        campos_soma=[c for c in df_raw.columns if c not in ("Ano","mês","Filial","Data")]
        d=df_raw.groupby(["Ano","mês"],as_index=False)[campos_soma].sum()
    else:
        d=df_raw[df_raw["Filial"].astype(str)==filial_sel].drop(columns=["Filial"]).copy()
    d=_ordenar_cronologico(d)
    cm=cm_(d); ca=ca_(d)
    if not cm or not ca: return None,False

    df_v=get_vendas_df()
    if df_v is None or df_v.empty: return None,False
    col_cat_v=next((c for c in df_v.columns if c.strip().lower()=="categoria"),None)
    col_prod_v=next((c for c in df_v.columns if c.strip().lower() in ["produto","codigo","código","sku"]),None)
    col_data_v=next((c for c in df_v.columns if c.strip().lower() in ["emissao","emissão","data"]),None)
    col_val_v=next((c for c in df_v.columns if c.strip().lower() in ["vlr.total","vlr total","valor total","valor"]),None)
    col_qtd_v=next((c for c in df_v.columns if c.strip().lower() in ["quantidade","qtd"]),None)
    col_fil_v=col_filial(df_v)
    if not all([col_cat_v,col_prod_v,col_data_v,col_val_v]): return None,False

    dv2=df_v.copy()
    dv2["_data"]=pd.to_datetime(dv2[col_data_v],errors="coerce",dayfirst=True)
    dv2["_ano"]=dv2["_data"].dt.year.astype(str)
    _meses_abv3={1:"jan",2:"fev",3:"mar",4:"abr",5:"mai",6:"jun",7:"jul",8:"ago",9:"set",10:"out",11:"nov",12:"dez"}
    dv2["_mes"]=dv2["_data"].dt.month.map(_meses_abv3)
    if col_fil_v and filial_sel and filial_sel!="(Todas as filiais)":
        dv2=dv2[dv2[col_fil_v].astype(str)==filial_sel]

    receita_total=dv2.groupby(["_ano","_mes"])[col_val_v].sum()
    dv_cat=dv2[dv2[col_cat_v].astype(str)==categoria_sel] if categoria_sel and categoria_sel!="(Nenhuma)" else dv2
    receita_cat=dv_cat.groupby(["_ano","_mes"])[col_val_v].sum()

    cmv_cat=None
    df_est=get_estoque_compras()
    if df_est is not None and not df_est.empty and "Produto" in df_est.columns and "CustoUnitario" in df_est.columns and col_qtd_v:
        custo_map=df_est.groupby("Produto")["CustoUnitario"].first().to_dict()
        dv_cat2=dv_cat.copy()
        dv_cat2["_custo_unit"]=dv_cat2[col_prod_v].map(custo_map)
        dv_cat2["_cmv"]=pd.to_numeric(dv_cat2[col_qtd_v],errors="coerce")*dv_cat2["_custo_unit"]
        cmv_cat=dv_cat2.groupby(["_ano","_mes"])["_cmv"].sum()

    campo_cmv=next((c for c in d.columns if "cmv" in c.lower()),None)
    for idx,row in d.iterrows():
        chave=(str(row[ca]),str(row[cm]).lower()[:3])
        rt=receita_total.get(chave)
        rc=receita_cat.get(chave,0)
        ratio=(rc/rt) if rt else 0
        for campo in d.columns:
            if campo in (ca,cm): continue
            if campo=="receita bruta de vendas":
                d.at[idx,campo]=rc
            elif campo_cmv and campo==campo_cmv and cmv_cat is not None:
                sinal=-1 if float(row[campo] or 0)<0 else 1
                d.at[idx,campo]=sinal*cmv_cat.get(chave,0)
            else:
                try: d.at[idx,campo]=float(row[campo] or 0)*ratio
                except: pass
    return calcular(d),(cmv_cat is not None)

def _aplicar_rateio_contas(df, cid):
    """Quebra linhas sem Filial preenchida em N linhas (uma por filial), usando o % configurado
    em Configurações > Rateio de Contas sem Filial. Contas sem rateio configurado ainda
    continuam sem filial (comportamento igual ao de antes, nada quebra)."""
    if df is None or "Filial" not in df.columns or not cid:
        return df
    rateio=load_rateio_contas(cid)
    if not rateio:
        return df
    mask_sem_filial=df["Filial"].isna() | (df["Filial"].astype(str).str.strip()=="")
    if not mask_sem_filial.any():
        return df
    campos_valor=[c for c in df.columns if c not in ("Ano","mês","Filial")]
    linhas_com_filial=df[~mask_sem_filial].copy()
    linhas_sem_filial=df[mask_sem_filial].copy()
    novas_linhas=[]
    linhas_restantes=[]
    for _,row in linhas_sem_filial.iterrows():
        linha_base=row.copy()
        for campo in campos_valor:
            try: valor=float(row.get(campo,0) or 0)
            except: valor=0
            if valor and campo in rateio:
                for fil,pct in rateio[campo].items():
                    nova=row.copy()
                    for c2 in campos_valor: nova[c2]=0
                    nova["Filial"]=fil
                    nova[campo]=valor*pct/100
                    novas_linhas.append(nova)
                linha_base[campo]=0
        if any((linha_base.get(c,0) or 0)!=0 for c in campos_valor):
            linhas_restantes.append(linha_base)
    partes=[linhas_com_filial]
    if novas_linhas: partes.append(pd.DataFrame(novas_linhas))
    if linhas_restantes: partes.append(pd.DataFrame(linhas_restantes))
    return pd.concat(partes,ignore_index=True)

def get_df_raw_bruto():
    """Como get_df(), mas devolve o banco ANTES de calcular() — usado pelo seletor de Filial."""
    if st.session_state.df_raw is not None:
        return _aplicar_rateio_contas(st.session_state.df_raw,st.session_state.cid)
    if st.session_state.cid:
        df=load_df(st.session_state.cid)
        if df is not None:
            st.session_state.df_raw=df
            return _aplicar_rateio_contas(df,st.session_state.cid)
    return None

def _ordenar_cronologico(df):
    """Garante ordem cronológica real (Ano+Mês), não alfabética (onde 'mar' vem depois de 'jul')."""
    if "Ano" not in df.columns or "mês" not in df.columns: return df
    df=df.copy()
    df["_ord"]=df["Ano"].astype(str)+df["mês"].astype(str).str.lower().map(MES_NUM).fillna("00")
    df=df.sort_values("_ord").drop(columns=["_ord"]).reset_index(drop=True)
    return df

def get_df_filial(filial=None):
    """Como get_df(), mas ciente de Filial. Sem Filial no banco (cliente antigo) ou filial=None,
    comportamento idêntico ao get_df() de sempre."""
    df_raw=get_df_raw_bruto()
    if df_raw is None: return None
    if "Filial" not in df_raw.columns or not filial:
        return calcular(_ordenar_cronologico(df_raw))
    if filial=="(Todas as filiais)":
        campos_soma=[c for c in df_raw.columns if c not in ("Ano","mês","Filial","Data")]
        df_agg=df_raw.groupby(["Ano","mês"],as_index=False)[campos_soma].sum()
        return calcular(_ordenar_cronologico(df_agg))
    df_filtrado=df_raw[df_raw["Filial"].astype(str)==filial].drop(columns=["Filial"])
    return calcular(_ordenar_cronologico(df_filtrado))

# ═══════════════════════════════════════════════════
# SIDEBAR — botões simples sem HTML wrapper
# ═══════════════════════════════════════════════════
with st.sidebar:

    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo_netexame.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode()
        st.markdown(f"""
        <div style="padding:10px 0 14px;text-align:center">
          <img src="data:image/png;base64,{logo_b64}"
            style="width:100%;max-width:190px;margin-bottom:4px;
            border:2px solid #A9762F;border-radius:10px;padding:4px"/>
          <div style="color:#484F58;font-size:.62rem;letter-spacing:.1em;text-transform:uppercase;margin-top:2px">
            Analytics BI
          </div>
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="padding:10px 0 14px;text-align:center">
          <div style="font-size:1rem;font-weight:800;color:#E6EDF3">📊 NetExame</div>
          <div style="color:#484F58;font-size:.65rem;margin-top:2px">Analytics BI</div>
        </div>""", unsafe_allow_html=True)    

    # API Key — protegida por senha master
    if st.session_state.api_key:
        tp="🟠 OpenAI" if not st.session_state.api_key.startswith("sk-ant-") else "🔵 Claude"
        st.caption(f"🟢 {tp} ativa")
    with st.expander("🔑 Alterar API Key"):
        senha_ak=st.text_input("Senha master",type="password",key="senha_ak")
        if senha_ak==SENHA_MASTER:
            ak=st.text_input("Nova API Key",value="",type="password",
              placeholder="sk-ant-... ou sk-proj-...",key="nova_ak")
            if st.button("💾 Salvar",key="btn_ak",use_container_width=True):
                if ak:
                    st.session_state.api_key=ak
                    c=load_cfg(); c["anthropic_api_key"]=ak; save_cfg(c)
                    st.success("✅ Salva!")
        elif senha_ak:
            st.markdown('<div class="al-d">❌ Senha incorreta</div>',unsafe_allow_html=True)

    # Cliente ativo
    perf=load_cli(st.session_state.cid) if st.session_state.cid else None
    if perf:
        st.markdown(f"""<div style="background:#085041;border:1px solid #0F6E56;border-radius:9px;
          padding:9px 12px;margin:6px 0 10px"><div style="color:#9FE1CB;font-size:.62rem;
          text-transform:uppercase;letter-spacing:.08em">Cliente</div>
          <div style="color:#F2EDE1;font-weight:600;font-size:.86rem">{perf['nome']}</div></div>""",
          unsafe_allow_html=True)

    st.divider()

    # GESTÃO
    st.markdown('<div style="background:#0F6E56;color:#9FE1CB;font-size:.68rem;font-weight:700;'
                'letter-spacing:.08em;text-transform:uppercase;padding:4px 10px;border-radius:5px;'
                'display:inline-block;margin-bottom:4px">GESTÃO</div>',unsafe_allow_html=True)
    if st.button("🏢 Clientes",        key="sb_clientes",   use_container_width=True): ir("clientes")
    if st.button("➕ Novo Cliente",     key="sb_novo",        use_container_width=True): ir("novo")
    if st.button("⚙️ Configurações",   key="sb_config",      use_container_width=True): ir("config")

    st.divider()
    st.markdown('<div style="background:#0F6E56;color:#9FE1CB;font-size:.68rem;font-weight:700;'
                'letter-spacing:.08em;text-transform:uppercase;padding:4px 10px;border-radius:5px;'
                'display:inline-block;margin-bottom:4px">IMPORTAR</div>',unsafe_allow_html=True)
    if st.button("📥 Importar Dados",  key="sb_importar",   use_container_width=True): ir("importar")
    if st.button("🔗 Integração ERP",  key="sb_erp",         use_container_width=True): ir("erp")
    if st.button("🧾 Importar Vendas (Pareto/ML)", key="sb_importar_vendas", use_container_width=True): ir("importar_vendas")

    st.divider()
    st.markdown('<div style="background:#0F6E56;color:#9FE1CB;font-size:.68rem;font-weight:700;'
                'letter-spacing:.08em;text-transform:uppercase;padding:4px 10px;border-radius:5px;'
                'display:inline-block;margin-bottom:4px">MÓDULO COMERCIAL</div>',unsafe_allow_html=True)
    if st.button("📊 Dashboard Executivo", key="sb_gestao_estoque", use_container_width=True): ir("gestao_estoque")
    if st.button("📦 Gestão Comercial de Compras", key="sb_compras", use_container_width=True): ir("compras")
    if st.button("💰 Fluxo de Caixa Comercial Projetado", key="sb_fluxo_compras", use_container_width=True): ir("fluxo_compras")
    if st.button("🧠 Motor de Previsão Estatística Avançada - ML", key="sb_config_ml", use_container_width=True): ir("config_ml")
    if st.button("📈 Motor de Previsão por Participação %", key="sb_ml_produtos", use_container_width=True): ir("ml_produtos")
    if st.button("📉 Curva de Pareto", key="sb_pareto", use_container_width=True): ir("pareto")

    st.divider()
    st.markdown('<div style="background:#0F6E56;color:#9FE1CB;font-size:.68rem;font-weight:700;'
                'letter-spacing:.08em;text-transform:uppercase;padding:4px 10px;border-radius:5px;'
                'display:inline-block;margin-bottom:4px">MÓDULO FINANCEIRO</div>',unsafe_allow_html=True)
    if st.button("📄 DRE",             key="sb_dre",         use_container_width=True): ir("dre")
    if st.button("🏦 Balanço",         key="sb_balanco",     use_container_width=True): ir("balanco")
    if st.button("💵 Fluxo de Caixa",  key="sb_fluxo",       use_container_width=True): ir("fluxo")
    if st.button("🎯 Indicadores",     key="sb_indicadores", use_container_width=True): ir("indicadores")
    if st.button("⚠️ Alertas",         key="sb_alertas",     use_container_width=True): ir("alertas")
    if st.button("🔭 Projeções ML - Financeiro", key="sb_ml", use_container_width=True): ir("ml")
    if st.button("🗺️ Cenários FP&A",   key="sb_cenarios",    use_container_width=True): ir("cenarios")
    if st.button("📤 Exportação",      key="sb_exportar",    use_container_width=True): ir("exportar")

    st.divider()
    st.markdown('<div style="background:#0F6E56;color:#9FE1CB;font-size:.68rem;font-weight:700;'
                'letter-spacing:.08em;text-transform:uppercase;padding:4px 10px;border-radius:5px;'
                'display:inline-block;margin-bottom:4px">ANÁLISE</div>',unsafe_allow_html=True)
    if st.button("🧭 Parecer Consolidado", key="sb_parecer_ia", use_container_width=True): ir("parecer_ia")
    

# ═══════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════
def hdr(titulo,sub=""):
    st.markdown(f'''<div style="background:linear-gradient(135deg,#0F6E56 0%,#085041 100%);
      border-radius:14px;padding:26px 32px;margin-bottom:20px;box-shadow:0 4px 14px rgba(0,0,0,.12)">
      <div style="font-family:Georgia,serif;font-size:1.5rem;font-weight:600;color:#fff">{titulo}</div>
      <div style="color:#9FE1CB;font-size:.85rem;margin-top:6px;letter-spacing:.02em">{sub}</div>
      </div>''',unsafe_allow_html=True)

def mc(col,lbl,val,cls="",sub=""):
    col.markdown(f'<div class="mc"><div class="mc-lbl">{lbl}</div>'
                 f'<div class="mc-val {cls}">{val}</div>'
                 f'<div class="mc-sub">{sub}</div></div>',unsafe_allow_html=True)

def sec(txt):
    st.markdown(f'''<div style="display:flex;align-items:center;gap:10px;
      background:linear-gradient(135deg,#F8F5EE 0%,#F1ECE0 100%);
      border-left:4px solid #A9762F;border-radius:0 8px 8px 0;
      padding:10px 18px;margin:22px 0 14px">
      <span style="font-family:Georgia,serif;font-size:1.05rem;font-weight:600;color:#14243B">{txt}</span>
      </div>''',unsafe_allow_html=True)

def no_data():
    st.markdown('<div class="al-w">⚠️ Nenhum dado carregado. Use <b>📥 Importar Dados</b> ou <b>🔌 Integração ERP</b> no menu.</div>',unsafe_allow_html=True)

def cls_pct(v,inv=False):
    if abs(v)<0.5: return "neu"
    if inv: return "neg" if v>0 else "pos"
    return "pos" if v>0 else "neg"

# ═══════════════════════════════════════════════════
# PÁGINAS
# ═══════════════════════════════════════════════════
pg=st.session_state.pg

st.session_state["_pg_mudou_agora"]=st.session_state.get("_ultima_pg_scroll")!=pg
if st.session_state["_pg_mudou_agora"]:
    st.session_state["_ultima_pg_scroll"]=pg
    components_v1.html("""
        <script>
        function rolarPraTopo() {
            try {
                var alvo = window.parent.document.querySelector('section.main');
                if (alvo) { alvo.scrollTo(0,0); }
                window.parent.scrollTo(0,0);
                window.parent.document.documentElement.scrollTop = 0;
                window.parent.document.body.scrollTop = 0;
            } catch(e) {}
        }
        rolarPraTopo();
        setTimeout(rolarPraTopo, 100);
        setTimeout(rolarPraTopo, 300);
        setTimeout(rolarPraTopo, 600);
        setTimeout(rolarPraTopo, 1000);
        </script>
    """,height=0)

if pg=="boas_vindas":
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo_netexame.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode()
        st.markdown(f'''<div style="text-align:center;margin-bottom:16px">
          <img src="data:image/png;base64,{logo_b64}" style="max-height:110px;
            border:4px solid #A9762F;border-radius:14px;padding:6px 10px;
            box-shadow:0 2px 10px rgba(169,118,47,.35)"/>
        </div>''',unsafe_allow_html=True)
    if st.session_state.get("cid"):
        _cli_bv=load_cli(st.session_state.cid)
        _nome_bv=_cli_bv.get("nome","sua empresa") if _cli_bv else "sua empresa"
        st.markdown(f"""<div class="wc-back">
          <div class="wc-back-emoji">👋</div>
          <h2>Bem-vindo ao ambiente da {_nome_bv}!</h2>
          <p>Escolha um dashboard no menu lateral para começar a análise.</p>
        </div>""",unsafe_allow_html=True)
        st.stop()

    st.markdown("""<div class="wc-hero">
      <h1>Bem-vindo ao NetExame Analytics BI</h1>
      <p>Inteligência comercial, financeira e preditiva, em uma só plataforma — testado e validado com rigor estatístico</p>
    </div>""",unsafe_allow_html=True)

    st.markdown("""<div class="wc-steps">
        <div class="wc-step">
          <div class="wc-step-n">1</div>
          <div class="wc-step-t">Cadastre um cliente</div>
          <div class="wc-step-s">Em Novo Cliente</div>
        </div>
        <div class="wc-step">
          <div class="wc-step-n">2</div>
          <div class="wc-step-t">Importe os dados</div>
          <div class="wc-step-s">Em Importar Dados</div>
        </div>
        <div class="wc-step">
          <div class="wc-step-n">3</div>
          <div class="wc-step-t">Acesse os dashboards</div>
          <div class="wc-step-s">No menu lateral</div>
        </div>
    </div>""",unsafe_allow_html=True)

    cf1,cf2,cf3=st.columns(3)
    cf1.markdown("""<details class="wc-card-exp">
      <summary>📊 Módulo Financeiro</summary>
      <div class="wc-card-exp-content">
        <ul>
          <li>📊 Demonstrações completas — DRE, Balanço e Fluxo de Caixa, sempre atualizados</li>
          <li>📈 Mais de 15 indicadores de saúde financeira, calculados automaticamente</li>
          <li>🚨 Alertas quando algo sai do esperado, sem precisar procurar o problema</li>
          <li>🎯 Simulação de cenários — o que acontece se as vendas caírem ou subirem</li>
        </ul>
      </div>
    </details>""",unsafe_allow_html=True)

    cf2.markdown("""<details class="wc-card-exp">
      <summary>📈 Módulo Comercial</summary>
      <div class="wc-card-exp-content">
        <ul>
          <li>📈 Previsão de vendas futuras, produto por produto, com base no histórico real</li>
          <li>📦 Sugestão automática de quanto comprar e quando, sem depender de achismo</li>
          <li>🚨 Aviso antecipado de produtos que podem faltar ou sobrar no estoque</li>
          <li>🎯 Ranking dos produtos que mais pesam no faturamento</li>
          <li>🚚 Avaliação de fornecedores por prazo, qualidade e entrega</li>
        </ul>
      </div>
    </details>""",unsafe_allow_html=True)

    cf3.markdown("""<details class="wc-card-exp">
      <summary>✨ Inteligência Artificial</summary>
      <div class="wc-card-exp-content">
        <ul>
          <li>🔮 O sistema testa várias formas de prever a demanda e escolhe a mais precisa pra cada produto</li>
          <li>✅ Toda previsão é testada contra o que realmente aconteceu, antes de confiar nela</li>
          <li>🧭 Relatórios executivos escritos automaticamente, cruzando Comercial e Financeiro</li>
          <li>📄 Pronto pra imprimir ou apresentar, sem trabalho manual</li>
        </ul>
      </div>
    </details>""",unsafe_allow_html=True)    



# ── CLIENTES ────────────────────────────────────────
elif pg=="clientes":
    hdr("👥 Clientes")
    if st.session_state.get("_limpar_busca_cliente"):
        st.session_state["busca_cliente_nome"]=""
        st.session_state["_limpar_busca_cliente"]=False
    cls=ls_cli()
    if not cls:
        st.markdown('<div class="al-i">Nenhum cliente ainda. Clique em <b>➕ Novo Cliente</b>.</div>',unsafe_allow_html=True)

    # PIN pendente
    if "pin_pendente" not in st.session_state: st.session_state.pin_pendente=None

    st.markdown('<div style="font-size:1.15rem;font-weight:600;color:#14243B;margin-bottom:6px">🏢 Digite o nome da empresa para acessar</div>',unsafe_allow_html=True)
    busca_cliente=st.text_input("Nome da empresa",key="busca_cliente_nome",
        label_visibility="collapsed")
    st.markdown('''<script>
    setTimeout(function(){
        var campos=window.parent.document.querySelectorAll('input[aria-label="Nome da empresa"]');
        campos.forEach(function(c){ c.setAttribute("autocomplete","off"); });
    },300);
    </script>''',unsafe_allow_html=True)

    if busca_cliente.strip():
        termo_busca=busca_cliente.strip().lower()
        encontrados=[c for c in cls if termo_busca in c["nome"].lower()]
        if encontrados:
            for c in encontrados:
                renderizar_card_cliente(c)
        else:
            st.markdown('<div class="al-i">Nenhum cliente encontrado com esse nome.</div>',unsafe_allow_html=True)
    else:
        st.markdown('<div class="al-i">Digite o nome de uma empresa acima para acessá-la diretamente.</div>',unsafe_allow_html=True)

    st.divider()
    with st.expander("🔐 Ver lista completa de clientes (acesso restrito)"):
        senha_lista=st.text_input("Senha master",type="password",key="senha_lista_clientes")
        if senha_lista==SENHA_MASTER:
            st.markdown('<div class="al-s">✅ Acesso autorizado</div>',unsafe_allow_html=True)
            for c in cls:
                renderizar_card_cliente(c)
        elif senha_lista:
            st.markdown('<div class="al-d">❌ Senha incorreta.</div>',unsafe_allow_html=True)

    # Caixa de PIN
    if st.session_state.pin_pendente:
        cid_p=st.session_state.pin_pendente
        p3=load_cli(cid_p)
        st.markdown(f'<div class="al-i">🔒 Digite o PIN para acessar <b>{p3["nome"] if p3 else "?"}</b></div>',unsafe_allow_html=True)
        pin_dig=st.text_input("PIN",type="password",key="pin_input",max_chars=10)
        col_ok,col_cancel=st.columns(2)
        if col_ok.button("✅ Confirmar",use_container_width=True):
            if verificar_pin(cid_p,pin_dig):
                limpar_sessao_cliente()
                p_pin=load_cli(cid_p)
                st.session_state.cid=cid_p
                st.session_state.df_raw=load_df(cid_p)
                st.session_state.projecoes={}
                st.session_state.saldo_ini=float(p_pin.get("saldo_ini",0)) if p_pin else 0.
                st.session_state.entradas_vista=float(p_pin.get("entradas_vista",0)) if p_pin else 0.
                st.session_state.freq_fluxo=p_pin.get("freq_fluxo","Mensal") if p_pin else "Mensal"
                st.session_state.pin_pendente=None
                pre_carregar_cliente(cid_p)
                addlog(f"Acesso autorizado"); ir("boas_vindas")
            else:
                st.markdown('<div class="al-d">❌ PIN incorreto.</div>',unsafe_allow_html=True)
        if col_cancel.button("❌ Cancelar",use_container_width=True):
            st.session_state.pin_pendente=None; st.rerun()

    # Gerenciar PINs (senha master)
    st.divider()
    with st.expander("🔐 Gerenciar PINs — Acesso Restrito"):
        senha=st.text_input("Senha master",type="password",key="senha_master")
        if senha==SENHA_MASTER:
            st.markdown('<div class="al-s">✅ Acesso autorizado</div>',unsafe_allow_html=True)
            cli_sel=st.selectbox("Cliente",["Selecione"]+[c["nome"] for c in cls])
            if cli_sel!="Selecione":
                cid_sel=next(c["id"] for c in cls if c["nome"]==cli_sel)
                p4=load_cli(cid_sel)
                pin_atual=p4.get("pin","") if p4 else ""
                st.markdown(f'PIN atual: **{"configurado" if pin_atual else "sem PIN"}**')
                novo_pin=st.text_input("Novo PIN (deixe vazio para remover)",
                                       type="password",key="novo_pin",max_chars=10)
                if st.button("💾 Salvar PIN",use_container_width=True):
                    if p4:
                        p4["pin"]=novo_pin
                        salvar_cli(cid_sel,p4)
                        if novo_pin:
                            st.markdown(f'<div class="al-s">✅ PIN configurado para {cli_sel}</div>',unsafe_allow_html=True)
                        else:
                            st.markdown(f'<div class="al-s">✅ PIN removido de {cli_sel}</div>',unsafe_allow_html=True)
            st.divider()
            st.markdown("**💾 Backup Completo de Todos os Clientes**")
            st.markdown('<div class="al-i">Gera um arquivo ZIP com todos os JSONs e CSVs de todos os clientes cadastrados — útil para versionar no Git ou guardar uma cópia de segurança fora da nuvem.</div>',unsafe_allow_html=True)
            if st.button("📦 Gerar Backup (ZIP)",use_container_width=True):
                import zipfile
                buf_zip=BytesIO()
                with zipfile.ZipFile(buf_zip,"w",zipfile.ZIP_DEFLATED) as zf:
                    for arquivo in os.listdir(PASTA):
                        caminho=os.path.join(PASTA,arquivo)
                        if os.path.isfile(caminho):
                            zf.write(caminho,arcname=arquivo)
                buf_zip.seek(0)
                st.download_button("⬇️ Baixar backup_clientes.zip",buf_zip.getvalue(),
                  file_name=f"backup_clientes_{datetime.now().strftime('%Y-%m-%d')}.zip",
                  mime="application/zip",use_container_width=True)

            st.divider()
            st.markdown("**📈 Log de Acessos de Visitantes**")
            path_acessos_view=os.path.join(PASTA,"acessos.log")
            if os.path.exists(path_acessos_view):
                with open(path_acessos_view,"r",encoding="utf-8") as f:
                    linhas_acesso=[l.strip() for l in f.readlines() if l.strip()]
                st.markdown(f'<div class="al-i">📊 Total de acessos registrados: <b>{len(linhas_acesso)}</b></div>',unsafe_allow_html=True)
                with st.expander("Ver todos os horários de acesso"):
                    for linha_ac in reversed(linhas_acesso[-100:]):
                        st.caption(linha_ac)
            else:
                st.markdown('<div class="al-i">Nenhum acesso de visitante registrado ainda.</div>',unsafe_allow_html=True)
        elif senha:
            st.markdown('<div class="al-d">❌ Senha incorreta</div>',unsafe_allow_html=True)

# ── NOVO CLIENTE ────────────────────────────────────
elif pg=="novo":
    hdr("➕ Novo Cliente")
    nome=st.text_input("Nome da empresa *",placeholder="Ex: Empresa ABC Ltda")
    senha_novo_cliente=st.text_input("Senha master *",type="password",key="senha_novo_cliente")
    if st.button("✅ Cadastrar",use_container_width=True):
        if not nome.strip(): st.error("Informe o nome.")
        elif senha_novo_cliente!=SENHA_MASTER: st.error("❌ Senha master incorreta.")
        else:
            cid=gid(nome); salvar_cli(cid,{"nome":nome.strip()})
            st.session_state.cid=cid
            st.session_state.df_raw=None
            st.session_state.projecoes={}
            st.session_state.saldo_ini=0.
            st.session_state.entradas_vista=0.
            st.session_state.freq_fluxo="Mensal"
            addlog(f"'{nome}' cadastrado"); st.success(f"✅ {nome} cadastrado!"); time.sleep(1); ir("config")

# ── CONFIGURAÇÕES ───────────────────────────────────
elif pg=="config":
    hdr("⚙️ Configurações","API Key, saldo inicial e preferências")
    p2=load_cli(st.session_state.cid) if st.session_state.cid else None
    if p2:
        st.markdown(f'<div class="al-i">👤 Cliente ativo: <b>{p2["nome"]}</b></div>',unsafe_allow_html=True)

    with st.expander("🔑 API Key — OpenAI ou Claude"):
        st.markdown('<div class="al-i">Cole aqui sua chave da OpenAI (sk-proj-...) ou Claude (sk-ant-...). A IA usa essa chave para ler os arquivos e extrair dados automaticamente.</div>',unsafe_allow_html=True)
        if st.session_state.api_key:
            tp2="🟠 OpenAI" if not st.session_state.api_key.startswith("sk-ant-") else "🔵 Claude"
            st.markdown(f'<div class="al-s">🟢 {tp2} ativa</div>',unsafe_allow_html=True)
        senha_cfg=st.text_input("Senha master para alterar",type="password",key="senha_cfg")
        if senha_cfg==SENHA_MASTER:
            ak2=st.text_input("Nova API Key",value="",type="password",
              placeholder="sk-proj-... ou sk-ant-...",key="ak2_cfg")
        elif senha_cfg:
            st.markdown('<div class="al-d">❌ Senha incorreta</div>',unsafe_allow_html=True)
            ak2=""
        else:
            ak2=""

    with st.expander("Saldo Inicial de Caixa"):
        senha_saldo_ini=st.text_input("Senha master para alterar",type="password",key="senha_saldo_ini")
        if senha_saldo_ini==SENHA_MASTER:
            si=st.number_input("Saldo inicial (R$)",value=float(st.session_state.saldo_ini),step=1000.,format="%.2f")
        elif senha_saldo_ini:
            st.markdown('<div class="al-d">❌ Senha incorreta</div>',unsafe_allow_html=True)
            si=st.session_state.saldo_ini
        else:
            st.markdown(f'<div class="al-i">Saldo atual: {fmt(st.session_state.saldo_ini)} — digite a senha master acima para alterar.</div>',unsafe_allow_html=True)
            si=st.session_state.saldo_ini

    with st.expander("Entradas à Vista"):
        st.markdown('<div class="al-i">Valor fixo de entradas à vista somado automaticamente ao Fluxo de Caixa.</div>',unsafe_allow_html=True)
        ev_atual=float(p2.get("entradas_vista",0)) if p2 else 0.
        freq_atual=p2.get("freq_fluxo","Mensal") if p2 else "Mensal"
        senha_entradas_vista=st.text_input("Senha master para alterar",type="password",key="senha_entradas_vista")
        if senha_entradas_vista==SENHA_MASTER:
            ev=st.number_input("Valor das Entradas à Vista (R$)",value=ev_atual,step=100.,format="%.2f")
            freq_sel=st.radio("Frequência",["Mensal","Diário"],
              index=0 if freq_atual=="Mensal" else 1,horizontal=True,
              help="Mensal: soma R$ X por mês | Diário: soma R$ X × 22 dias úteis")
        elif senha_entradas_vista:
            st.markdown('<div class="al-d">❌ Senha incorreta</div>',unsafe_allow_html=True)
            ev=ev_atual; freq_sel=freq_atual
        else:
            st.markdown(f'<div class="al-i">Valor atual: {fmt(ev_atual)} ({freq_atual}) — digite a senha master acima para alterar.</div>',unsafe_allow_html=True)
            ev=ev_atual; freq_sel=freq_atual

    if st.button("💾 Salvar Configurações",use_container_width=True):
        st.session_state.saldo_ini=si
        st.session_state.entradas_vista=ev
        st.session_state.freq_fluxo=freq_sel
        if ak2 and senha_cfg==SENHA_MASTER:
            st.session_state.api_key=ak2
            c=load_cfg(); c["anthropic_api_key"]=ak2; save_cfg(c)
        if p2:
            p2["saldo_ini"]=si
            p2["entradas_vista"]=ev
            p2["freq_fluxo"]=freq_sel
            salvar_cli(st.session_state.cid,p2)
        st.success("✅ Configurações salvas!"); time.sleep(1); ir("importar")

    with st.expander("Rateio de Contas sem Filial"):
        st.markdown('<div class="al-i">Contas importadas sem filial especificada (ex: impostos corporativos, despesas gerais) — configure aqui como distribuir cada uma entre as lojas. Contas sem rateio configurado continuam aparecendo só no consolidado, sem quebrar nada.</div>',unsafe_allow_html=True)
        _df_raw_rateio=load_df(st.session_state.cid) if st.session_state.cid else None
        if _df_raw_rateio is not None and "Filial" in _df_raw_rateio.columns:
            _mask_sf=_df_raw_rateio["Filial"].isna() | (_df_raw_rateio["Filial"].astype(str).str.strip()=="")
            _campos_valor_cfg=[c for c in _df_raw_rateio.columns if c not in ("Ano","mês","Filial")]
            _campos_sem_filial=[]
            for c in _campos_valor_cfg:
                try:
                    if _df_raw_rateio.loc[_mask_sf,c].fillna(0).astype(float).abs().sum()>0:
                        _campos_sem_filial.append(c)
                except: pass
            if not _campos_sem_filial:
                st.markdown('<div class="al-s">✅ Nenhuma conta sem filial encontrada nos dados importados.</div>',unsafe_allow_html=True)
            else:
                _filiais_cfg=sorted(_df_raw_rateio.loc[~_mask_sf,"Filial"].dropna().astype(str).unique().tolist())
                if not _filiais_cfg:
                    st.markdown('<div class="al-w">⚠️ Não há nenhuma linha com Filial preenchida pra basear a sugestão — importe dados com filial em pelo menos algumas contas primeiro.</div>',unsafe_allow_html=True)
                else:
                    _campo_receita_cfg=next((c for c in _campos_valor_cfg if "receita bruta" in c.lower()),None)
                    if _campo_receita_cfg:
                        _receita_fil=_df_raw_rateio[~_mask_sf].groupby("Filial")[_campo_receita_cfg].sum()
                        _participacao_cfg=(_receita_fil/_receita_fil.sum()*100).round(1) if _receita_fil.sum()>0 else None
                    else:
                        _participacao_cfg=None
                    _rateio_atual=load_rateio_contas(st.session_state.cid)
                    _linhas_editor=[]
                    for campo in _campos_sem_filial:
                        _linha={"Conta":campo}
                        _cfg_existente=_rateio_atual.get(campo,{})
                        for fil in _filiais_cfg:
                            if fil in _cfg_existente:
                                _linha[fil]=_cfg_existente[fil]
                            elif _participacao_cfg is not None and fil in _participacao_cfg.index:
                                _linha[fil]=float(_participacao_cfg[fil])
                            else:
                                _linha[fil]=round(100/len(_filiais_cfg),1)
                        _linhas_editor.append(_linha)
                    _df_editor_rateio=pd.DataFrame(_linhas_editor)
                    st.markdown('<div class="al-i">💡 Sugestão automática baseada na participação real de receita de cada loja — ajuste se quiser.</div>',unsafe_allow_html=True)
                    _df_editado_rateio=st.data_editor(_df_editor_rateio,use_container_width=True,hide_index=True,key="rateio_contas_editor",
                        column_config={fil:st.column_config.NumberColumn(f"{fil} (%)",min_value=0,max_value=100,step=0.1) for fil in _filiais_cfg})
                    _somas_rateio=_df_editado_rateio[_filiais_cfg].sum(axis=1)
                    _linhas_erradas=_df_editado_rateio[(_somas_rateio<99.5)|(_somas_rateio>100.5)]
                    if not _linhas_erradas.empty:
                        st.markdown(f'<div class="al-w">⚠️ Essas contas não somam 100%: {", ".join(_linhas_erradas["Conta"].tolist())}. Ajuste antes de salvar.</div>',unsafe_allow_html=True)
                    senha_rateio_contas=st.text_input("Senha master para salvar",type="password",key="senha_rateio_contas")
                    if st.button("💾 Salvar Rateio de Contas",key="btn_salvar_rateio_contas",use_container_width=True):
                        if senha_rateio_contas!=SENHA_MASTER:
                            st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi salvo.</div>',unsafe_allow_html=True)
                        elif not _linhas_erradas.empty:
                            st.markdown('<div class="al-d">❌ Corrija as contas que não somam 100% antes de salvar.</div>',unsafe_allow_html=True)
                        elif st.session_state.cid:
                            _novo_rateio={}
                            for _,r in _df_editado_rateio.iterrows():
                                _novo_rateio[r["Conta"]]={fil:float(r[fil]) for fil in _filiais_cfg}
                            save_rateio_contas(st.session_state.cid,_novo_rateio)
                            st.session_state.df_raw=None
                            st.markdown('<div class="al-s">✅ Rateio salvo — as telas por loja já vão refletir essa distribuição.</div>',unsafe_allow_html=True)

# ── IMPORTAR ────────────────────────────────────────
elif pg=="importar":
    hdr("📥 Importar Dados","A IA lê qualquer layout — DRE horizontal, vertical, PDF, CSV, Excel")
    if not st.session_state.cid:
        st.markdown('<div class="al-w">⚠️ Cadastre e selecione um cliente primeiro.</div>',unsafe_allow_html=True); st.stop()
    if not st.session_state.api_key:
        st.markdown('<div class="al-w">⚠️ Configure a API Key em <b>⚙️ Configurações</b> primeiro.</div>',unsafe_allow_html=True)
    p2=load_cli(st.session_state.cid)
    st.markdown(f'<div class="al-i">👤 <b>{p2["nome"] if p2 else "?"}</b></div>',unsafe_allow_html=True)
    arqs=st.file_uploader("Selecione o(s) arquivo(s)",type=["pdf","xlsx","xls","xlsm","csv"],accept_multiple_files=True)
    senha_import_dados=None
    if arqs:
        senha_import_dados=st.text_input("Senha master para confirmar a importação *",type="password",key="senha_import_dados")
    col_ia, col_dir = st.columns(2)
    btn_ia  = col_ia.button("🤖 Processar com IA",use_container_width=True)
    btn_dir = col_dir.button("⚡ Leitura Direta (CSV/Excel padrão)",use_container_width=True)
    if arqs and (btn_ia or btn_dir) and senha_import_dados!=SENHA_MASTER:
        st.error("❌ Senha master incorreta.")
    elif arqs and btn_dir:
        dfs_dir=[]
        for a in arqs:
            b=a.read(); n=a.name.lower()
            try:
                if n.endswith(".csv"):
                    for enc in ["utf-8-sig","utf-8","latin1","cp1252"]:
                        for sep in [";",","]:
                            try:
                                df_tmp=pd.read_csv(io.BytesIO(b),sep=sep,decimal=",",encoding=enc,on_bad_lines="skip")
                                # Corrige mojibake (UTF-8 lido como Latin-1/CP1252) nos nomes das colunas
                                if enc in ("latin1","cp1252"):
                                    novas_cols=[]
                                    for c in df_tmp.columns:
                                        try:
                                            c_fix=str(c).encode("latin1").decode("utf-8")
                                            novas_cols.append(c_fix)
                                        except: novas_cols.append(c)
                                    df_tmp.columns=novas_cols
                                if df_tmp.shape[1]>=5 and any(c.lower() in ["mês","mes"] for c in df_tmp.columns):
                                    dfs_dir.append(df_tmp); break
                            except: pass
                        else: continue
                        break
                else:
                    xls=pd.read_excel(io.BytesIO(b),sheet_name=None)
                    for s,df_tmp in xls.items():
                        if any(c.lower() in ["mês","mes"] for c in df_tmp.columns):
                            dfs_dir.append(df_tmp); break
            except Exception as e: st.markdown(f'<div class="al-d">❌ {a.name}: {e}</div>',unsafe_allow_html=True)
        if dfs_dir:
            df_final=pd.concat(dfs_dir,ignore_index=True)
            # Padroniza colunas
            rename_map={}
            for c in df_final.columns:
                cl=c.lower().strip()
                if cl=="receita líquida de vendas": rename_map[c]="receita líquida"
                if cl in ["mês","mes"]: rename_map[c]="mês"
                if cl in ["ano","year"]: rename_map[c]="Ano"
            df_final=df_final.rename(columns=rename_map)
            # Ordena por data
            try:
                mes_num={"jan":"01","fev":"02","mar":"03","abr":"04","mai":"05","jun":"06",
                         "jul":"07","ago":"08","set":"09","out":"10","nov":"11","dez":"12"}
                df_final["_data"]=pd.to_datetime(
                    df_final["Ano"].astype(str)+"-"+
                    df_final["mês"].astype(str).str.lower().str[:3].map(mes_num).fillna("01")+"-01",
                    errors="coerce")
                df_final=df_final.sort_values("_data").drop(columns=["_data"]).reset_index(drop=True)
            except: pass
            # Identifica colunas extras que não fazem parte do padrão do sistema (TODOS + Ano/mês/Data)
            # e soma automaticamente em Disponibilidades entradas/Saida, sem perder o dinheiro de vista
            campos_calculados_sistema={
                "deduções","receita líquida","lucro bruto","margem contrib","desp op",
                "lucro operacional","resultado IR","lucro líquido","EBITDA",
                "margem bruta %","margem contrib %","margem op %","margem líquida %","EBITDA %",
                "ativo circ","ativo total","pass circ","pass total","PL",
                "liquidez corrente","liquidez imediata","ROE","kanitz","PMR","PMP","PME",
                "ciclo de caixa","giro estoque","ticket médio","ICD",
                "saldo período","saldo acumulado","score_risco",
                # Variações de nome vindas do arquivo do cliente (sinônimos comuns)
                "deduções da receita bruta","receita líquida de vendas","despesas operacionais totais",
                "resultado antes da provisão para imposto de renda e contribuição social",
                "ativo total saldo","passivo total saldo","ativo circulante saldo","passivo circulante saldo",
                "patrimônio líquido","Lucratividade","Margem de Contribuição",
                "Prazo médio de pagamentos","Prazo médio de recebimentos","Prazo médio de estocagem",
                "giro do estoque","termômetro de kanitz","roe","ticket medio",
            }
            campos_padrao_conhecidos=set(TODOS)|campos_calculados_sistema|{"Ano","mês","Data","Item"}
            # Palavras que indicam que a coluna é um INDICADOR/MÉTRICA, não dinheiro real —
            # essas colunas NÃO devem ser somadas ao Fluxo de Caixa
            pistas_indicador=["inadimp","obsol","índice","indice","%","percentual","taxa de",
                              "ratio","score","nota","grau"]
            campos_extras=[c for c in df_final.columns if c not in campos_padrao_conhecidos
                          and pd.api.types.is_numeric_dtype(df_final[c])
                          and not any(p in c.lower() for p in pistas_indicador)]
            campos_ignorados=[c for c in df_final.columns if c not in campos_padrao_conhecidos
                             and pd.api.types.is_numeric_dtype(df_final[c])
                             and any(p in c.lower() for p in pistas_indicador)]

            pistas_entrada_dir=["receita","entrada","recebiment","venda","faturamento","comiss","repasse"]
            pistas_saida_dir=["despesa","saida","saída","pagamento","custo","gasto","manutenç","marketing","honorári","taxa"]

            detalhamento_dir=[]
            if campos_extras and "mês" in df_final.columns and "Ano" in df_final.columns:
                if "Disponibilidades entradas" not in df_final.columns:
                    df_final["Disponibilidades entradas"]=0.
                if "Disponibilidades Saida" not in df_final.columns:
                    df_final["Disponibilidades Saida"]=0.
                for campo_extra in campos_extras:
                    col_l=campo_extra.lower()
                    eh_saida=any(p in col_l for p in pistas_saida_dir)
                    campo_destino="Disponibilidades Saida" if eh_saida else "Disponibilidades entradas"
                    for idx,row in df_final.iterrows():
                        v=float(row.get(campo_extra,0) or 0)
                        if v==0: continue
                        df_final.at[idx,campo_destino]=float(df_final.at[idx,campo_destino] or 0)+v
                        detalhamento_dir.append({"ano":str(row.get("Ano","")),
                                                 "mes":str(row.get("mês","")).lower()[:3],
                                                 "campo_pai":campo_destino,
                                                 "subconta":campo_extra,"valor":v})

            if detalhamento_dir and st.session_state.cid:
                df_detalhe_dir=pd.DataFrame(detalhamento_dir)
                path_detalhe_dir=os.path.join(PASTA,f"{gid(st.session_state.cid)}_detalhamento.csv")
                if os.path.exists(path_detalhe_dir):
                    df_detalhe_antigo_dir=pd.read_csv(path_detalhe_dir,sep=";",decimal=",",encoding="utf-8-sig")
                    df_detalhe_antigo_dir["ano"]=df_detalhe_antigo_dir["ano"].astype(str)
                    df_detalhe_dir["ano"]=df_detalhe_dir["ano"].astype(str)
                    df_detalhe_dir=pd.concat([df_detalhe_antigo_dir,df_detalhe_dir],ignore_index=True).drop_duplicates(
                        subset=["ano","mes","campo_pai","subconta"],keep="last")
                df_detalhe_dir.to_csv(path_detalhe_dir,sep=";",decimal=",",index=False,encoding="utf-8-sig")

            st.session_state.df_raw=df_final
            save_df(st.session_state.cid,df_final)
            st.markdown(f'<div class="al-s">✅ Leitura direta: <b>{len(df_final)}</b> períodos × <b>{df_final.shape[1]}</b> campos</div>',unsafe_allow_html=True)
            if campos_extras:
                with st.expander(f"⚠️ {len(campos_extras)} coluna(s) extra(s) somadas em Entradas/Saídas (revisão recomendada)"):
                    st.markdown('<div class="al-w">Estas colunas não fazem parte do padrão do sistema, mas o valor foi preservado e somado ao total de Entradas ou Saídas do Fluxo de Caixa. Veja o detalhamento no drill-down da tela de Fluxo de Caixa.</div>',unsafe_allow_html=True)
                    st.write(campos_extras)
            if campos_ignorados:
                with st.expander(f"ℹ️ {len(campos_ignorados)} coluna(s) identificadas como indicador/métrica — não somadas ao Fluxo de Caixa"):
                    st.markdown('<div class="al-i">Estas colunas parecem ser indicadores (ex: % de inadimplência, obsolescência) e não valores monetários — por isso não foram somadas ao total de Entradas/Saídas. Os valores continuam disponíveis no banco de dados, caso queira usá-los em análises futuras.</div>',unsafe_allow_html=True)
                    st.write(campos_ignorados)
            addlog(f"Leitura direta: {len(df_final)} períodos")
        else:
            st.error("Não foi possível ler o arquivo. Tente o botão 🤖 Processar com IA.")
    elif arqs and btn_ia:
        if not st.session_state.api_key:
            st.error("Configure a API Key em ⚙️ Configurações.")
        else:
            for a in arqs:
                with st.spinner(f"Lendo {a.name}..."):
                    df_r,msg=ler(a.read(),a.name)
                if df_r is None:
                    st.markdown(f'<div class="al-d">❌ {a.name}: {msg}</div>',unsafe_allow_html=True); continue
                st.markdown(f'<div class="al-s">✅ {a.name}: {msg}</div>',unsafe_allow_html=True)

                # Caso especial: balancete em PDF já processado direto em células prontas
                # Pode conter DRE + Balanço + Fluxo de Caixa juntos — merge de cada um separadamente
                if isinstance(df_r,tuple) and len(df_r)==2 and df_r[0]=="_CELULAS_BALANCETE_":
                    cels=df_r[1]
                    detalhamento_arq=[]
                    cels_limpas=[]
                    for c in cels:
                        if "_detalhamento" in c:
                            detalhamento_arq.extend(c["_detalhamento"])
                        else:
                            cels_limpas.append(c)
                    cels=cels_limpas
                    if detalhamento_arq and st.session_state.cid:
                        df_detalhe=pd.DataFrame(detalhamento_arq)
                        path_detalhe=os.path.join(PASTA,f"{gid(st.session_state.cid)}_detalhamento.csv")
                        if os.path.exists(path_detalhe):
                            df_detalhe_antigo=pd.read_csv(path_detalhe,sep=";",decimal=",",encoding="utf-8-sig")
                            df_detalhe_antigo["ano"]=df_detalhe_antigo["ano"].astype(str)
                            df_detalhe["ano"]=df_detalhe["ano"].astype(str)
                            df_concat=pd.concat([df_detalhe_antigo,df_detalhe],ignore_index=True)
                            df_detalhe=df_concat.drop_duplicates(subset=["ano","mes","campo_pai","subconta"],keep="last")
                        df_detalhe.to_csv(path_detalhe,sep=";",decimal=",",index=False,encoding="utf-8-sig")

                    # Separa as células por demonstração real (DRE / Balanço / Fluxo)
                    cels_dre=[c for c in cels if c.get("campo") in CAMPOS_DRE]
                    cels_bal=[c for c in cels if c.get("campo") in CAMPOS_BAL]
                    cels_fluxo=[c for c in cels if c.get("campo") in CAMPOS_FLUXO]

                    df_atual=st.session_state.df_raw
                    resumo_demos=[]
                    for nome_demo,cels_grupo in [("DRE",cels_dre),("Balanço",cels_bal),("Fluxo",cels_fluxo)]:
                        if not cels_grupo: continue
                        df_atual=merge_banco_por_demonstracao(df_atual,cels_grupo,nome_demo)
                        resumo_demos.append(f"{nome_demo} ({len(cels_grupo)} células)")

                    if not resumo_demos:
                        st.markdown(f'<div class="al-w">⚠️ {a.name}: não foi possível identificar nenhuma demonstração.</div>',unsafe_allow_html=True)
                        continue

                    st.session_state.df_raw=df_atual
                    save_df(st.session_state.cid,df_atual)
                    st.markdown(f'<div class="al-s">✅ <b>{a.name}</b> processado — '
                               f'{" + ".join(resumo_demos)}, banco atualizado com <b>{len(df_atual)}</b> períodos totais.</div>',
                               unsafe_allow_html=True)
                    addlog(f"{a.name}: balancete — {' + '.join(resumo_demos)}")
                    
                    continue

                tipo_det=detectar_tipo(df_r,a.name)

                # Se o arquivo está no formato longo (vertical), usa o parser determinístico dedicado
                if eh_formato_longo(df_r) and tipo_det!="DESCONHECIDO":
                    cels=parser_formato_longo(df_r,tipo_det)
                    st.markdown(f'<div class="al-i">ℹ️ Formato vertical detectado — processado com parser determinístico (sem IA).</div>',unsafe_allow_html=True)
                else:
                    with st.spinner(f"🤖 IA extraindo dados..."):
                        cels=ia_extrair(df_r,TODOS,st.session_state.api_key)

                # Verifica se campos CRÍTICOS vieram zerados/ausentes
                campos_criticos={"DRE":["receita bruta de vendas","CMV (custo da mercadoria vendida)"],
                                "BALANCO":["ativo total saldo","passivo total saldo"],
                                "FLUXO":["Disponibilidades entradas","Disponibilidades Saida"]}.get(tipo_det,[])
                campos_presentes={c.get("campo") for c in cels if "_detalhamento" not in c}
                faltando_critico=any(cc not in campos_presentes for cc in campos_criticos)

                # Verifica se a quantidade de PERÍODOS distintos é suspeita (poucos meses cobertos)
                periodos_cobertos={(c.get("ano"),c.get("mes")) for c in cels if "_detalhamento" not in c}
                poucos_periodos=len(periodos_cobertos)<6

                # Se faltou campo crítico OU poucos períodos, tenta extração bruta + regras determinísticas
                if (len(cels)<10 or faltando_critico or poucos_periodos) and st.session_state.api_key:
                    with st.spinner("🤖 Tentando extração bruta + regras determinísticas..."):
                        celulas_brutas=_chamar_ia_extracao_bruta(df_r,st.session_state.api_key,tipo_det)
                        st.caption(f"🔧 Debug: IA retornou {len(celulas_brutas)} linhas brutas")
                        cels_via_regras=mapear_celulas_brutas(celulas_brutas,tipo_det) if celulas_brutas else []
                        st.caption(f"🔧 Debug: após mapeamento por regras = {len(cels_via_regras)} campos")
                        if len(cels_via_regras)>len(cels):
                            cels=cels_via_regras
                            st.markdown(f'<div class="al-i">ℹ️ Layout não padrão — usada extração bruta por IA + soma por regras determinísticas ({len(celulas_brutas)} linhas → {len(cels_via_regras)} campos).</div>',unsafe_allow_html=True)

                if not cels:
                    st.markdown(f'<div class="al-w">⚠️ {a.name}: nenhum dado extraído.</div>',unsafe_allow_html=True)
                    continue
                # Separa o detalhamento (drill-down) e as colunas não reconhecidas das células normais
                detalhamento_arq=[]
                nao_reconhecidas_arq=[]
                cels_limpas=[]
                for c in cels:
                    if "_detalhamento" in c:
                        detalhamento_arq.extend(c["_detalhamento"])
                    elif "_nao_reconhecidas" in c:
                        nao_reconhecidas_arq.extend(c["_nao_reconhecidas"])
                    else:
                        cels_limpas.append(c)
                cels=cels_limpas

                if nao_reconhecidas_arq:
                    colunas_unicas=sorted(set(item["coluna"] for item in nao_reconhecidas_arq))
                    total_valor_nr=sum(item["valor"] for item in nao_reconhecidas_arq)
                    with st.expander(f"⚠️ {len(colunas_unicas)} coluna(s) não reconhecida(s) automaticamente — classificadas como 'Outras' (revisão recomendada)"):
                        st.markdown(f'<div class="al-w">O dinheiro foi mantido no total (R$ {total_valor_nr:,.2f} somados em "Outras"), mas o sistema não conseguiu identificar uma categoria específica para estas colunas. Para classificar corretamente, renomeie a coluna no arquivo original usando um termo mais comum (ex: "Receita de Serviços", "Fornecedores") e reimporte.</div>',unsafe_allow_html=True)
                        df_nr=pd.DataFrame(nao_reconhecidas_arq)
                        df_nr_resumo=df_nr.groupby(["coluna","destino"])["valor"].agg(["sum","count"]).reset_index()
                        df_nr_resumo.columns=["Coluna do Arquivo","Classificada como","Total (R$)","Nº de períodos"]
                        df_nr_resumo["Total (R$)"]=df_nr_resumo["Total (R$)"].apply(lambda v: fmt(v))
                        st.dataframe(df_nr_resumo,use_container_width=True,hide_index=True)
                if detalhamento_arq and st.session_state.cid:
                    df_detalhe=pd.DataFrame(detalhamento_arq)
                    path_detalhe=os.path.join(PASTA,f"{gid(st.session_state.cid)}_detalhamento.csv")
                    if os.path.exists(path_detalhe):
                        df_detalhe_antigo=pd.read_csv(path_detalhe,sep=";",decimal=",",encoding="utf-8-sig")
                        df_detalhe_antigo["ano"]=df_detalhe_antigo["ano"].astype(str)
                        df_detalhe["ano"]=df_detalhe["ano"].astype(str)
                        df_concat=pd.concat([df_detalhe_antigo,df_detalhe],ignore_index=True)
                        df_detalhe=df_concat.drop_duplicates(subset=["ano","mes","campo_pai","subconta"],keep="last")
                    df_detalhe.to_csv(path_detalhe,sep=";",decimal=",",index=False,encoding="utf-8-sig")
                demo_detectada=identificar_demonstracao(cels)
                if not demo_detectada:
                    st.markdown(f'<div class="al-w">⚠️ {a.name}: não foi possível identificar a demonstração.</div>',unsafe_allow_html=True)
                    continue
                
                # Se a memória da sessão estiver vazia (ex: servidor reiniciado desde o último
                # import), recarrega do disco antes de juntar — sem isso, um import novo apagava
                # tudo que já estava salvo em disco de sessões/reinícios anteriores.
                df_atual=st.session_state.df_raw
                if df_atual is None and st.session_state.cid:
                    df_atual=load_df(st.session_state.cid)
                df_merged=merge_banco_por_demonstracao(df_atual,cels,demo_detectada)
                st.session_state.df_raw=df_merged
                save_df(st.session_state.cid,df_merged)
                st.markdown(f'<div class="al-s">✅ <b>{a.name}</b> identificado como <b>{demo_detectada}</b> — '
                           f'{len(cels)} células processadas, banco atualizado com <b>{len(df_merged)}</b> períodos totais.</div>',
                           unsafe_allow_html=True)
                addlog(f"{a.name}: {demo_detectada} — {len(cels)} células")

                # AUDITORIA NÃO-BLOQUEANTE: IA roda em paralelo só pra avisar discrepâncias
                if st.session_state.api_key and tipo_det in ("DRE","BALANCO","FLUXO"):
                    with st.spinner("🔍 Auditoria automática com IA (não bloqueia)..."):
                        try:
                            cels_ia_aud=_chamar_ia_extracao(df_r,TODOS,st.session_state.api_key,tipo_det)
                            st.caption(f"🔧 Debug temporário: IA retornou {len(cels_ia_aud)} células")
                            idx_parser={(c.get("ano"),c.get("mes"),c.get("campo")):c.get("valor") for c in cels}
                            idx_ia_aud={(c.get("ano"),c.get("mes"),c.get("campo")):c.get("valor") for c in cels_ia_aud}
                            alertas_aud=[]
                            campos_comparados=0
                            for chave,v_p in idx_parser.items():
                                v_i=idx_ia_aud.get(chave)
                                if v_i is not None:
                                    campos_comparados+=1
                                    try:
                                        diff=abs(float(v_p)-float(v_i))/max(abs(float(v_p)),1)*100
                                        if diff>10:
                                            alertas_aud.append((chave,v_p,v_i,diff))
                                    except: pass
                            if alertas_aud:
                                with st.expander(f"🔍 Auditoria IA — {len(alertas_aud)} possível(is) divergência(s) encontrada(s) (revisão opcional)"):
                                    st.markdown('<div class="al-i">Estes valores foram conferidos por IA e mostraram diferença relevante em relação ao parser. O banco já foi salvo com os valores do parser — revise se necessário.</div>',unsafe_allow_html=True)
                                    for (ano,mes,campo),v_p,v_i,diff in sorted(alertas_aud,key=lambda x:-x[3])[:15]:
                                        st.markdown(f'<div class="al-w">⚠️ <b>{campo}</b> — {mes}/{ano}: Parser={fmt(v_p)} vs IA={fmt(v_i)} ({diff:.0f}% diferença)</div>',unsafe_allow_html=True)
                            else:
                                st.markdown(f'<div class="al-s">🔍 Auditoria IA concluída: {campos_comparados} valores comparados, '
                                           f'nenhuma divergência relevante (>10%) encontrada. Os dados do parser foram confirmados pela IA.</div>',unsafe_allow_html=True)
                        except Exception as e:
                            st.markdown(f'<div class="al-w">⚠️ Auditoria IA não pôde ser concluída ({e}). O banco já foi salvo com os valores do parser local, que são confiáveis independentemente da auditoria.</div>',unsafe_allow_html=True)
    if st.session_state.df_raw is not None and not st.session_state.df_raw.empty:
        st.divider()
        df_ex=st.session_state.df_raw; cm2=cm_(df_ex); ca2=ca_(df_ex)
        st.markdown(f'<div class="al-s">📊 Banco de dados: <b>{len(df_ex)}</b> períodos × <b>{df_ex.shape[1]}</b> campos</div>',unsafe_allow_html=True)
        if ca2: st.markdown(f'<div class="al-i">Anos: {", ".join(sorted(str(a) for a in df_ex[ca2].dropna().unique()))}</div>',unsafe_allow_html=True)

        sec("📋 Tela de Revisão — Cobertura por Demonstração")
        cobertura=gerar_cobertura(df_ex,DEMONSTRACOES_CAMPOS)
        if not cobertura.empty:
            cols_show=["Ano","Mês"]+[d for d in DEMONSTRACOES_CAMPOS.keys() if d in cobertura.columns]
            cols_qtd=[f"{d}_qtd" for d in DEMONSTRACOES_CAMPOS.keys() if f"{d}_qtd" in cobertura.columns]
            st.dataframe(cobertura[cols_show],use_container_width=True,hide_index=True,height=min(400,40+len(cobertura)*36))
            with st.expander("📊 Ver detalhamento de campos preenchidos por período"):
                st.dataframe(cobertura[["Ano","Mês"]+cols_qtd],use_container_width=True,hide_index=True)

        c1,c2=st.columns(2)
        with c1:
            if st.button("👁️ Ver Dados Completos",use_container_width=True):
                st.session_state.ver_dados_completos=not st.session_state.get("ver_dados_completos",False)
        with c2:
            senha_limpar_dados=st.text_input("Senha master para limpar",type="password",key="senha_limpar_dados")
            if st.button("🗑 Limpar e Reimportar Tudo",use_container_width=True):
                if senha_limpar_dados!=SENHA_MASTER:
                    st.error("❌ Senha master incorreta — nada foi apagado.")
                else:
                    st.session_state.df_raw=None; st.session_state.projecoes={}
                    if st.session_state.cid:
                        p_path=os.path.join(PASTA,f"{gid(st.session_state.cid)}_dados.csv")
                        if os.path.exists(p_path): os.remove(p_path)
                    st.rerun()

        if st.session_state.get("ver_dados_completos",False):
            st.dataframe(df_ex,use_container_width=True,height=400)

# ── ERP ─────────────────────────────────────────────
elif pg=="erp":
    hdr("🔌 Integração ERP","Omie e Conta Azul")
    if not st.session_state.cid:
        st.markdown('<div class="al-w">⚠️ Selecione um cliente primeiro.</div>',unsafe_allow_html=True); st.stop()
    erp=st.radio("ERP:",["🟠 Omie","🔵 Conta Azul"],horizontal=True)
    if "Omie" in erp:
        st.markdown('<div class="al-i">📌 Omie → Configurações → API → Criar Aplicação → copie App Key e App Secret</div>',unsafe_allow_html=True)
        c1,c2=st.columns(2)
        ak_o=c1.text_input("App Key",type="password",key="omie_ak")
        as_o=c2.text_input("App Secret",type="password",key="omie_as")
        c3,c4,c5=st.columns(3)
        ano_o=c3.selectbox("Ano",list(range(2020,2031)),index=4)
        mes_i=c4.selectbox("Início",list(range(1,13)),format_func=lambda n:MESES[n-1])
        mes_f2=c5.selectbox("Fim",list(range(1,13)),format_func=lambda n:MESES[n-1],index=11)
        senha_omie=st.text_input("Senha master *",type="password",key="senha_import_omie")
        if st.button("🔍 Importar do Omie",use_container_width=True):
            if senha_omie!=SENHA_MASTER:
                st.error("❌ Senha master incorreta.")
            else:
                dfs2=[]; errs2=[]
                pb=st.progress(0); rng=list(range(mes_i,mes_f2+1))
                with st.spinner("Buscando..."):
                    for idx_m,m in enumerate(rng):
                        try:
                            ms=f"{m:02d}/{ano_o}"
                            r=requests.post("https://app.omie.com.br/api/v1/financas/dre/",
                              json={"call":"ObterRelDRE","app_key":ak_o,"app_secret":as_o,
                                    "param":[{"dDtInicio":f"01/{ms}","dDtFim":f"28/{ms}"}]},timeout=30)
                            dre2=r.json(); linha={"Ano":str(ano_o),"mês":MESES[m-1]}
                            def bk(obj,k):
                                if isinstance(obj,dict):
                                    if k in obj: return obj[k]
                                    for v2 in obj.values():
                                        r2=bk(v2,k)
                                        if r2 is not None: return r2
                                elif isinstance(obj,list):
                                    for i2 in obj:
                                        r2=bk(i2,k)
                                        if r2 is not None: return r2
                            for k2,c2_ in {"nReceitaBruta":"receita bruta de vendas",
                              "nCMV":"CMV (custo da mercadoria vendida)",
                              "nDespesasComerciais":"despesas comerciais",
                              "nDespesasAdministrativas":"despesas administrativas",
                              "nDespesasFinanceiras":"despesas financeiras líquidas"}.items():
                                v2=bk(dre2,k2)
                                if v2 is not None:
                                    try: linha[c2_]=float(str(v2).replace(",","."))
                                    except: pass
                            dfs2.append(linha)
                        except Exception as e: errs2.append(str(e))
                        pb.progress((idx_m+1)/len(rng))
                df_om=pd.DataFrame(dfs2) if dfs2 else pd.DataFrame()
                _tem_dado_financeiro_om=any(c not in ("Ano","mês") for c in df_om.columns)
                if dfs2 and _tem_dado_financeiro_om:
                    st.session_state.df_raw=df_om
                    save_df(st.session_state.cid,df_om)
                    st.markdown(f'<div class="al-s">✅ {len(df_om)} meses do Omie</div>',unsafe_allow_html=True)
                    addlog(f"Omie: {len(df_om)} meses")
                elif dfs2:
                    st.markdown('<div class="al-w">⚠️ A busca não retornou nenhum dado financeiro (só Ano/mês) — nada foi salvo. Confira a App Key e App Secret.</div>',unsafe_allow_html=True)
                if errs2: st.warning("; ".join(errs2[:3]))
    else:
        st.markdown('<div class="al-i">📌 Conta Azul → Integrações → API → Gerar Token</div>',unsafe_allow_html=True)
        tok=st.text_input("Access Token",type="password")
        c1,c2=st.columns(2)
        ano_ca=c1.selectbox("Ano",list(range(2020,2031)),index=4)
        mes_ca=c2.selectbox("Mês",list(range(1,13)),format_func=lambda n:MESES[n-1])
        senha_conta_azul=st.text_input("Senha master *",type="password",key="senha_import_conta_azul")
        if st.button("🔍 Importar Conta Azul",use_container_width=True):
            if senha_conta_azul!=SENHA_MASTER:
                st.error("❌ Senha master incorreta.")
            else:
                h={"Authorization":f"Bearer {tok}"}
                ini=f"{ano_ca}-{mes_ca:02d}-01"; fim=f"{ano_ca}-{mes_ca:02d}-28"
                linha_ca={"Ano":str(ano_ca),"mês":MESES[mes_ca-1]}
                try:
                    r=requests.get("https://api.contaazul.com/v1/sales",headers=h,
                      params={"emission_start":ini,"emission_end":fim,"size":500},timeout=20)
                    if r.status_code==200:
                        linha_ca["receita bruta de vendas"]=sum(float(v.get("total",0)) for v in r.json() if isinstance(v,dict))
                    if "receita bruta de vendas" in linha_ca:
                        df_ca=pd.DataFrame([linha_ca]); st.session_state.df_raw=df_ca
                        save_df(st.session_state.cid,df_ca)
                        st.markdown('<div class="al-s">✅ Conta Azul importado</div>',unsafe_allow_html=True)
                    else:
                        st.markdown('<div class="al-w">⚠️ A busca não retornou nenhum dado financeiro — nada foi salvo. Confira o Access Token.</div>',unsafe_allow_html=True)
                except Exception as e: st.error(f"Erro: {e}")

# ── DRE ─────────────────────────────────────────────
elif pg=="dre":
    hdr("📊 DRE","Demonstração do Resultado — Real | AV% | AH%")

    _df_raw_dre=get_df_raw_bruto()
    _col_fil_dre=col_filial(_df_raw_dre) if _df_raw_dre is not None else None
    filial_sel_dre=None
    if _col_fil_dre:
        _filiais_disp_dre=sorted(v for v in _df_raw_dre[_col_fil_dre].dropna().astype(str).unique().tolist() if v!="(Todas as filiais)")
        _opcoes_filial_dre=["(Todas as filiais)"]+_filiais_disp_dre

        def _on_change_filial_dre():
            st.session_state["dre_filial_sel_backup"]=st.session_state["dre_filial_sel"]

        if "dre_filial_sel_backup" not in st.session_state:
            st.session_state["dre_filial_sel_backup"]=_opcoes_filial_dre[0]
        if st.session_state["dre_filial_sel_backup"] not in _opcoes_filial_dre:
            st.session_state["dre_filial_sel_backup"]=_opcoes_filial_dre[0]
        st.session_state["dre_filial_sel"]=st.session_state["dre_filial_sel_backup"]

        filial_sel_dre=st.selectbox("🏬 Filial",_opcoes_filial_dre,
          key="dre_filial_sel",on_change=_on_change_filial_dre)

    # Seletor de Categoria — só aparece se houver Vendas importada com coluna Categoria.
    # Cruza com a Filial escolhida acima (categoria dentro de uma loja específica).
    categoria_sel_dre="(Nenhuma)"
    _df_v_dre=get_vendas_df()
    _col_cat_dre=next((c for c in _df_v_dre.columns if c.strip().lower()=="categoria"),None) if _df_v_dre is not None else None
    if _col_cat_dre:
        _categorias_disp_dre=sorted(_df_v_dre[_col_cat_dre].dropna().astype(str).unique().tolist())
        categoria_sel_dre=st.selectbox("📦 Categoria (opcional)",["(Nenhuma)"]+_categorias_disp_dre,key="dre_categoria_sel")

    if categoria_sel_dre!="(Nenhuma)":
        df,_tem_cmv_real_dre=montar_dre_categoria(filial_sel_dre,categoria_sel_dre)
        if df is None:
            st.markdown('<div class="al-w">⚠️ Não consegui montar a DRE por categoria — confira se a base de Vendas está importada corretamente.</div>',unsafe_allow_html=True)
            no_data(); st.stop()
        _msg_cmv_dre="com CMV real (custo × quantidade vendida)" if _tem_cmv_real_dre else "⚠️ sem arquivo de Estoque/Custo importado — CMV também rateado, não é real"
        with st.expander(f"📦 Visão da categoria {categoria_sel_dre} — como isso foi calculado"):
            st.markdown(f'<div class="al-i">📦 Visão da categoria <b>{categoria_sel_dre}</b>{" — "+filial_sel_dre if filial_sel_dre and filial_sel_dre!="(Todas as filiais)" else " — todas as filiais"}. '
                        f'<b>Receita</b> é real (soma direta da base de Vendas), <b>CMV</b> {_msg_cmv_dre}. '
                        f'As demais contas (Aluguel, Salários, etc.) são <b>estimadas por rateio</b>, proporcionalmente à participação de receita dessa categoria — não são valores reais, porque o Financeiro não registra essas contas por categoria.</div>',unsafe_allow_html=True)
    else:
        df=get_df_filial(filial_sel_dre)
    if df is None: no_data(); st.stop()
    cm=cm_(df); ca=ca_(df)
    anos=sorted({str(a) for a in df[ca].dropna().unique()}) if ca else []
    ano_f=st.selectbox("Filtrar por ano",["Todos"]+anos) if anos else "Todos"
    df_v=df[df[ca].astype(str)==ano_f] if ano_f!="Todos" else df
    ul=df_v.iloc[-1]
    k=st.columns(5)
    for col_k,lbl,campo,t in [(k[0],"Rec. Bruta","receita bruta de vendas","brl"),
      (k[1],"Rec. Líquida","receita líquida","brl"),(k[2],"Lucro Bruto","lucro bruto","brl"),
      (k[3],"Lucro Líquido","lucro líquido","brl"),(k[4],"EBITDA %","EBITDA %","pct")]:
        try:
            v=float(ul.get(campo,0))
            col_k.markdown(f'<div class="mc"><div class="mc-lbl">{lbl}</div>'
                          f'<div class="mc-val {cor(v)}">{fmt(v,t)}</div></div>',unsafe_allow_html=True)
        except: pass
    g1,g2=st.columns(2)
    TH_LIGHT=dict(plot_bgcolor="white",paper_bgcolor="white",
        font=dict(color="#6B7280",size=10,family="Inter"),
        xaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=False,tickangle=-35),
        yaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=True),
        margin=dict(l=8,r=8,t=36,b=8),
        legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.3,font=dict(size=9)),
        hovermode="x unified",
        hoverlabel=dict(bgcolor="white",bordercolor="#E8ECF0",font=dict(color="#111827",size=11)))
    with g1:
        fig_rb=go.Figure()
        v=pd.to_numeric(df_v["receita bruta de vendas"],errors="coerce") if "receita bruta de vendas" in df_v.columns else pd.Series()
        x=df_v[cm].astype(str) if cm else pd.Series(range(len(df_v))).astype(str)
        cs=["#2563EB" if val>=0 else "#DC2626" for val in v]
        fig_rb.add_trace(go.Bar(x=x,y=v,marker_color=cs,
          text=[fmt(val) for val in v],textposition="outside",textfont=dict(size=8,color="#6B7280")))
        fig_rb.update_layout(title=dict(text="💰 Receita Bruta",font=dict(size=12,color="#111827")),**TH_LIGHT)
        st.plotly_chart(fig_rb,use_container_width=True)
    with g2:
        fig_mg=go.Figure()
        x=df_v[cm].astype(str) if cm else pd.Series(range(len(df_v))).astype(str)
        if ca and cm:
            x=df_v[cm].astype(str)+"/"+df_v[ca].astype(str).str[-2:]
        for c,cor_c,nm in [
          ("margem bruta %","#14243B","Margem Bruta"),
          ("margem contrib %","#5B7B9A","Margem Contrib."),
          ("margem líquida %","#A9762F","Margem Líquida"),
          ("EBITDA %","#8B5E34","EBITDA")]:
            if c not in df_v.columns: continue
            fig_mg.add_trace(go.Scatter(x=x,y=pd.to_numeric(df_v[c],errors="coerce"),
              name=nm,mode="lines+markers",
              line=dict(color=cor_c,width=2.4),
              marker=dict(size=6,color=cor_c,line=dict(color="white",width=1.5)),
              hovertemplate=f"<b>{nm}</b><br>%{{x}}<br>%{{y:.1f}}%<extra></extra>"))
        fig_mg.update_layout(
          title=dict(text="Margens (%)",font=dict(size=15,family="Georgia, serif",color="#14243B")),
          plot_bgcolor="white",paper_bgcolor="white",
          font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
          margin=dict(l=10,r=10,t=50,b=60),
          xaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=False,
            tickangle=-40,tickfont=dict(size=9,color="#4B5563")),
          yaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=True,zeroline=False,tickfont=dict(size=9)),
          legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.34,x=0.5,xanchor="center",font=dict(size=10)),
          hovermode="x unified",height=420,
          hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B",size=11)))
        st.plotly_chart(fig_mg,use_container_width=True)
    df_12=df_v
    n=len(df_12)
    meses_cols=list(df_12[cm].astype(str)) if cm else [str(i) for i in range(n)]
    def dv(c,i):
        try: return float(df_12.iloc[i].get(c,0))
        except: return 0.
    def av(c,i):
        # AV% em relação à Receita Bruta (base = 100%)
        try: return dv(c,i)/(float(df_12.iloc[i].get("receita bruta de vendas",1)) or 1)*100
        except: return 0.
    def av_pct(c,i):
        # Para linhas que já são percentual — retorna o valor direto
        try: return float(df_12.iloc[i].get(c,0))
        except: return 0.
    def ah(c,i):
        try:
            v=dv(c,i)
            if i==0: return 0.
            v0=float(df_12.iloc[i-1].get(c,1)) or 1
            return (v-v0)/abs(v0)*100
        except:
            return 0.
    def ah_pct(c,i):
        try:
            v=float(df_12.iloc[i].get(c,0))
            if i==0: return 0.
            v0=float(df_12.iloc[i-1].get(c,0))
            return v-v0
        except:
            return 0.
    linhas=[("cat","(+) RECEITA BRUTA","receita bruta de vendas",False),
            ("sub","  Impostos s/ Vendas","impostos sobre vendas",True),
            ("sub","  Devoluções","devoluções de vendas",True),
            ("tot","= RECEITA LÍQUIDA","receita líquida",False),
            ("cat","(-) CMV","CMV (custo da mercadoria vendida)",True),
            ("tot","= LUCRO BRUTO","lucro bruto",False),
            ("pct","  Margem Bruta %","margem bruta %",False),
            ("sub","  (-) Desp. Comerciais","despesas comerciais",True),
            ("tot","= MARGEM CONTRIB.","margem contrib",False),
            ("pct","  Margem Contrib. %","margem contrib %",False),
            ("sub","  (-) Desp. Adm.","despesas administrativas",True),
            ("sub","  (-) Desp. Fin.","despesas financeiras líquidas",True),
            ("sub","  (-) Depreciação","despesas com depreciações e amortizações",True),
            ("tot","= LUCRO OPERACIONAL","lucro operacional",False),
            ("pct","  Margem Op. %","margem op %",False),
            ("sub","  (+/-) Não Operac.","receitas não operacionais",False),
            ("sub","  (-) IR/CSLL","provisão para imposto de renda",True),
            ("tot","= LUCRO LÍQUIDO","lucro líquido",False),
            ("pct","  Margem Líquida %","margem líquida %",False),
            ("tot","  EBITDA","EBITDA",False),
            ("pct","  EBITDA %","EBITDA %",False)]
    header="<tr><th>Descrição</th>"+"".join(f"<th>{m}</th><th>AV%</th><th>AH%</th>" for m in meses_cols)+"</tr>"
    rows=""
    for tipo,desc,campo,inv in linhas:
        cls_tr={"cat":"cat","tot":"tot","pct":"pct","sub":"sub"}.get(tipo,"")
        row=f'<tr class="{cls_tr}"><td>{desc}</td>'
        for i in range(n):
            if tipo=="pct":
                v=av_pct(campo,i)
                delta=ah_pct(campo,i)
                cls_d="pos" if delta>0 else ("neg" if delta<0 else "neu")
                row+=f'<td class="pct">{fmt(v,"pct")}</td>'
                row+=f'<td class="{cls_d}" style="font-size:.7rem">{"▲" if delta>0 else "▼"}{abs(delta):.1f}pp</td>'
                row+=f'<td></td>'
            else:
                v=dv(campo,i); a_v=av(campo,i); a_h=ah(campo,i)
                cls_v=""
                row+=f'<td class="{cls_v}">{fmt(v)}</td>'
                row+=f'<td class="{cls_pct(a_v,inv)}">{fmt(a_v,"pct")}</td>'
                row+=f'<td class="{cls_pct(a_h,inv)}">{fmt(a_h,"pct")}</td>'
        rows+=row+"</tr>"
    st.markdown(f'<div class="dre-wrap"><table class="dre">{header}{rows}</table></div>',unsafe_allow_html=True)

    # Botão de exportação formatada em Excel
    def gerar_excel_dre():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb=Workbook()
        ws=wb.active
        ws.title="DRE"

        cor_azul=PatternFill("solid",start_color="2563EB",end_color="2563EB")
        cor_cinza_clara=PatternFill("solid",start_color="F3F4F6",end_color="F3F4F6")
        cor_branco_bold=Font(bold=True,color="FFFFFF",size=11)
        cor_preto=Font(color="111827",size=10)
        cor_verde=Font(color="059669",size=10)
        cor_vermelho=Font(color="DC2626",size=10)
        cor_cinza_texto=Font(color="9CA3AF",size=10)
        borda_fina=Border(bottom=Side(style="thin",color="E8ECF0"))

        # Linha de contexto: qual filtro estava ativo na exportação
        filtro_txt=f"Filtro aplicado: Ano = {ano_f}" if ano_f!="Todos" else "Filtro aplicado: Todos os períodos"
        ws.cell(row=1,column=1,value=filtro_txt).font=Font(italic=True,color="6B7280",size=9)
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=4)

        # Cabeçalho (linha 2) — mês/ano combinados para evitar ambiguidade entre anos diferentes
        meses_cols_completo=list(df_12[ca].astype(str)+"/"+df_12[cm].astype(str)) if ca and cm else meses_cols
        ws.cell(row=2,column=1,value="Descrição").font=cor_branco_bold
        ws.cell(row=2,column=1).fill=cor_azul
        col=2
        for m in meses_cols_completo:
            for sub in ["Real","AV%","AH%"]:
                c=ws.cell(row=2,column=col,value=f"{m} - {sub}")
                c.font=cor_branco_bold; c.fill=cor_azul
                c.alignment=Alignment(horizontal="center")
                col+=1

        linha_excel=3
        for tipo,desc,campo,inv in linhas:
            desc_limpo=desc.strip()
            if desc_limpo.startswith("="):
                desc_limpo="'"+desc_limpo
            ws.cell(row=linha_excel,column=1,value=desc_limpo)
            if tipo=="tot":
                ws.cell(row=linha_excel,column=1).font=Font(bold=True,color="2563EB",size=10)
                ws.cell(row=linha_excel,column=1).fill=cor_cinza_clara
            elif tipo=="cat":
                ws.cell(row=linha_excel,column=1).font=Font(bold=True,color="111827",size=10)
            else:
                ws.cell(row=linha_excel,column=1).font=cor_preto

            col=2
            for i in range(n):
                if tipo=="pct":
                    v=av_pct(campo,i); delta=ah_pct(campo,i)
                    c1=ws.cell(row=linha_excel,column=col,value=v/100); c1.number_format="0.0%"
                    c2=ws.cell(row=linha_excel,column=col+1,value=delta/100); c2.number_format="+0.0%;-0.0%"
                    c2.font=cor_verde if delta>0 else (cor_vermelho if delta<0 else cor_cinza_texto)
                    col+=3
                else:
                    v=dv(campo,i); a_v=av(campo,i); a_h=ah(campo,i)
                    c1=ws.cell(row=linha_excel,column=col,value=v)
                    c1.number_format='R$ #,##0;[RED](R$ #,##0)'
                    c1.font=cor_vermelho if (inv and v!=0) else (cor_verde if v>0 else cor_preto)
                    c2=ws.cell(row=linha_excel,column=col+1,value=a_v/100); c2.number_format="0.0%"
                    c3=ws.cell(row=linha_excel,column=col+2,value=a_h/100); c3.number_format="+0.0%;-0.0%"
                    col+=3
            for cc in range(1,col):
                ws.cell(row=linha_excel,column=cc).border=borda_fina
            linha_excel+=1

        ws.column_dimensions["A"].width=28
        for cc in range(2,col):
            ws.column_dimensions[get_column_letter(cc)].width=14
        ws.freeze_panes="B2"

        # Aba de Drill-down (subcontas), se existir
        if st.session_state.cid:
            path_detalhe_xls=os.path.join(PASTA,f"{gid(st.session_state.cid)}_detalhamento.csv")
            if os.path.exists(path_detalhe_xls):
                df_det_xls=pd.read_csv(path_detalhe_xls,sep=";",decimal=",",encoding="utf-8-sig")
                campos_dre_xls=DEMONSTRACOES_CAMPOS.get("DRE",[])
                df_det_dre_xls=df_det_xls[df_det_xls["campo_pai"].isin(campos_dre_xls)]
                if ano_f!="Todos":
                    df_det_dre_xls=df_det_dre_xls[df_det_dre_xls["ano"].astype(str)==ano_f]
                if not df_det_dre_xls.empty:
                    ws2=wb.create_sheet("Drill-down")
                    ws2.cell(row=1,column=1,value="Conta Principal").font=cor_branco_bold
                    ws2.cell(row=1,column=2,value="Subconta").font=cor_branco_bold
                    ws2.cell(row=1,column=3,value="Ano").font=cor_branco_bold
                    ws2.cell(row=1,column=4,value="Mês").font=cor_branco_bold
                    ws2.cell(row=1,column=5,value="Valor (R$)").font=cor_branco_bold
                    for cc in range(1,6):
                        ws2.cell(row=1,column=cc).fill=cor_azul
                    linha2=2
                    for _,r in df_det_dre_xls.sort_values(["campo_pai","subconta","ano","mes"]).iterrows():
                        ws2.cell(row=linha2,column=1,value=r["campo_pai"])
                        ws2.cell(row=linha2,column=2,value=r["subconta"])
                        ws2.cell(row=linha2,column=3,value=str(r["ano"]))
                        ws2.cell(row=linha2,column=4,value=r["mes"])
                        c_val=ws2.cell(row=linha2,column=5,value=float(r["valor"]))
                        c_val.number_format='R$ #,##0'
                        linha2+=1
                    for cc,w in zip("ABCDE",[26,28,8,8,16]):
                        ws2.column_dimensions[cc].width=w
                    ws2.freeze_panes="A2"

        buf=BytesIO(); wb.save(buf); buf.seek(0)
        return buf.getvalue()

    st.download_button("📥 Exportar esta DRE (Excel formatado)",gerar_excel_dre(),
      file_name=f"DRE_{ano_f if ano_f!='Todos' else 'completa'}.xlsx",
      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
      use_container_width=True)

    # Drill-down de subcontas
    if st.session_state.cid:
        path_detalhe=os.path.join(PASTA,f"{gid(st.session_state.cid)}_detalhamento.csv")
        if os.path.exists(path_detalhe):
            df_det=pd.read_csv(path_detalhe,sep=";",decimal=",",encoding="utf-8-sig")
            campos_dre_dd=DEMONSTRACOES_CAMPOS.get("DRE",[])
            df_det_dre=df_det[df_det["campo_pai"].isin(campos_dre_dd)]
            if not df_det_dre.empty:
              if "dre_dd_expander_aberto" not in st.session_state:
                  st.session_state["dre_dd_expander_aberto"]=False
              with st.expander("🔍 Ver detalhamento por subconta (drill-down)",expanded=st.session_state["dre_dd_expander_aberto"]):
                st.session_state["dre_dd_expander_aberto"]=True
                campos_pai=sorted(df_det_dre["campo_pai"].unique().tolist())
                campo_sel_dd=st.selectbox("Selecione a conta",campos_pai,key="dd_campo")
                df_det_f=df_det_dre[df_det_dre["campo_pai"]==campo_sel_dd]
                if ano_f!="Todos":
                    df_det_f=df_det_f[df_det_f["ano"].astype(str)==ano_f]
                if not df_det_f.empty:
                    df_det_f=df_det_f.copy()
                    df_det_f["periodo"]=df_det_f["mes"].astype(str)+"/"+df_det_f["ano"].astype(str).str[-2:]
                    pivot=df_det_f.pivot_table(index="subconta",columns="periodo",values="valor",
                                                aggfunc="sum",fill_value=0)
                    ordem_meses=["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]
                    cols_ordenadas=sorted(pivot.columns,
                      key=lambda x: (x.split("/")[1], ordem_meses.index(x.split("/")[0]) if x.split("/")[0] in ordem_meses else 99))
                    pivot=pivot[cols_ordenadas]
                    pivot["Total"]=pivot.sum(axis=1)
                    pivot=pivot.sort_values("Total",ascending=False)
                    pivot_fmt=pivot.copy()
                    for c in pivot_fmt.columns:
                        pivot_fmt[c]=pivot_fmt[c].apply(lambda v: fmt(v))
                    st.dataframe(pivot_fmt,use_container_width=True)
                else:
                    st.info("Sem detalhamento disponível para esta conta no período selecionado.")

    sec("🌊 Waterfall — Último Período")
    wf_c=["receita bruta de vendas","deduções","CMV (custo da mercadoria vendida)",
          "despesas comerciais","despesas administrativas","despesas financeiras líquidas","lucro líquido"]
    wf_l=["Rec. Bruta","(-) Deduções","(-) CMV","(-) D.Com","(-) D.Adm","(-) D.Fin","= Luc. Líq."]
    wf_v=[float(ul.get(c,0)) for c in wf_c]
    if any(v!=0 for v in wf_v):
        fig_wf=go.Figure(go.Waterfall(orientation="v",measure=["relative"]*len(wf_l),x=wf_l,y=wf_v,
          connector=dict(line=dict(color="#E8ECF0",width=1)),
          increasing=dict(marker=dict(color="#059669")),
          decreasing=dict(marker=dict(color="#DC2626")),
          totals=dict(marker=dict(color="#2563EB")),
          text=[fmt(v) for v in wf_v],textposition="outside",
          textfont=dict(size=9,color="#6B7280")))
        fig_wf.update_layout(
          title=dict(text="🌊 Composição do Resultado — Último Período",
                     font=dict(size=12,color="#111827")),
          plot_bgcolor="white",paper_bgcolor="white",
          font=dict(color="#6B7280",size=10,family="Inter"),
          margin=dict(l=8,r=8,t=44,b=8),
          xaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=False),
          yaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=True),
          showlegend=False)
        st.plotly_chart(fig_wf,use_container_width=True)

# ── FLUXO ───────────────────────────────────────────
elif pg=="fluxo":
    hdr("💰 Fluxo de Caixa","Entradas, saídas, saldo e alertas automáticos")

    _df_raw_fx=get_df_raw_bruto()
    _col_fil_fx=col_filial(_df_raw_fx) if _df_raw_fx is not None else None
    filial_sel_fx=None
    if _col_fil_fx:
        _filiais_disp_fx=sorted(v for v in _df_raw_fx[_col_fil_fx].dropna().astype(str).unique().tolist() if v!="(Todas as filiais)")
        _opcoes_filial_fx=["(Todas as filiais)"]+_filiais_disp_fx

        def _on_change_filial_fx():
            st.session_state["fx_filial_sel_backup"]=st.session_state["fx_filial_sel"]

        if "fx_filial_sel_backup" not in st.session_state:
            st.session_state["fx_filial_sel_backup"]=_opcoes_filial_fx[0]
        if st.session_state["fx_filial_sel_backup"] not in _opcoes_filial_fx:
            st.session_state["fx_filial_sel_backup"]=_opcoes_filial_fx[0]
        st.session_state["fx_filial_sel"]=st.session_state["fx_filial_sel_backup"]

        filial_sel_fx=st.selectbox("🏬 Filial",_opcoes_filial_fx,
          key="fx_filial_sel",on_change=_on_change_filial_fx)

    df=get_df_filial(filial_sel_fx)
    if df is None: no_data(); st.stop()
    cm=cm_(df); ca=ca_(df); si=st.session_state.saldo_ini

    # Filtros
    col_f1,col_f2=st.columns(2)
    anos_disp=["Todos"]+sorted({str(a) for a in df[ca].dropna().unique().tolist()}) if ca else ["Todos"]
    ano_sel=col_f1.selectbox("Ano",anos_disp,key="fluxo_ano")
    df=df[df[ca].astype(str)==ano_sel].reset_index(drop=True) if ano_sel!="Todos" else df.reset_index(drop=True)
    meses_disp=["Todos"]+df[cm].dropna().unique().tolist() if cm else ["Todos"]
    mes_sel=col_f2.selectbox("Mês",meses_disp,key="fluxo_mes")
    df=df[df[cm]==mes_sel].reset_index(drop=True) if mes_sel!="Todos" else df.reset_index(drop=True)
    if df.empty: st.warning("Sem dados para este período."); st.stop()
    df_f=df.copy()
    ent=pd.to_numeric(df_f.get("Disponibilidades entradas",pd.Series(0,index=df_f.index)),errors="coerce").fillna(0)
    sai=pd.to_numeric(df_f.get("Disponibilidades Saida",pd.Series(0,index=df_f.index)),errors="coerce").fillna(0)
    ev=float(st.session_state.get("entradas_vista",0))
    freq=st.session_state.get("freq_fluxo","Mensal")
    if ev>0:
        ent_vista=pd.Series([ev*22 if freq=="Diário" else ev]*len(df_f),index=df_f.index)
        df_f["entradas à vista"]=ent_vista
        ent=ent+ent_vista
    df_f["tot entradas"]=ent; df_f["tot saidas"]=sai
    df_f["saldo período"]=ent-sai; df_f["saldo acumulado"]=si+(ent-sai).cumsum()
    ul=df_f.iloc[-1]
    k=st.columns(4)
    mc(k[0],"Saldo Inicial",fmt(si),"b")
    mc(k[1],"Total Entradas",fmt(ent.sum()),"g",f"Média: {fmt(ent.mean())}/mês")
    mc(k[2],"Total Saídas",fmt(sai.sum()),"r",f"Média: {fmt(sai.mean())}/mês")
    sf=float(ul.get("saldo acumulado",si)); mc(k[3],"Saldo Final",fmt(sf),cor(sf))
    sec("🚨 Alertas de Caixa")
    neg=df_f[df_f["saldo acumulado"]<0]
    if len(neg)>0:
        mn=neg[cm].tolist() if cm else list(neg.index)
        st.markdown(f'<div class="al-d">🔴 Caixa NEGATIVO em <b>{len(neg)}</b> período(s): {", ".join(str(m) for m in mn)}<br>Ação: negociar prazos com fornecedores ou antecipar recebíveis.</div>',unsafe_allow_html=True)
    elif len(df_f[df_f["saldo acumulado"]<df_f["saldo acumulado"].mean()*.3])>0:
        st.markdown('<div class="al-w">⚠️ Caixa abaixo de 30% da média em alguns períodos.</div>',unsafe_allow_html=True)
    else:
        st.markdown('<div class="al-s">✅ Caixa positivo em todos os períodos.</div>',unsafe_allow_html=True)

    
    sec("📈 Evolução")
    g1,g2=st.columns(2)
    x=df_f[cm].astype(str) if cm else pd.Series(range(len(df_f))).astype(str)
    if ca and cm:
        x=df_f[cm].astype(str)+"/"+df_f[ca].astype(str).str[-2:]
    TH_FC=dict(plot_bgcolor="white",paper_bgcolor="white",
        font=dict(color="#6B7280",size=10,family="Inter"),
        xaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=False,tickangle=-35),
        yaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=True),
        margin=dict(l=8,r=8,t=40,b=40),
        legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.3,font=dict(size=9)),
        hovermode="x unified",
        hoverlabel=dict(bgcolor="white",bordercolor="#E8ECF0",font=dict(color="#111827",size=11)))
    with g1:
        fig_fc=go.Figure()
        fig_fc.add_trace(go.Bar(x=x,y=df_f["tot entradas"],name="Entradas",
          marker=dict(color="#059669",line=dict(color="white",width=0.6)),opacity=.9,
          hovertemplate="<b>Entradas</b><br>%{x}<br>R$ %{y:,.0f}<extra></extra>"))
        fig_fc.add_trace(go.Bar(x=x,y=-df_f["tot saidas"],name="Saídas",
          marker=dict(color="#DC2626",line=dict(color="white",width=0.6)),opacity=.9,
          hovertemplate="<b>Saídas</b><br>%{x}<br>R$ %{customdata:,.0f}<extra></extra>",
          customdata=df_f["tot saidas"]))
        fig_fc.add_hline(y=0,line_dash="dot",line_color="#9CA3AF",opacity=0.5)
        fig_fc.update_layout(title=dict(text="Entradas vs Saídas",
          font=dict(size=13,family="Georgia, serif",color="#14243B")),
          barmode="relative",bargap=0.28,**TH_FC)
        st.plotly_chart(fig_fc,use_container_width=True)
    with g2:
        saldo=df_f["saldo acumulado"]
        cor_saldo="#059669" if float(saldo.iloc[-1])>=0 else "#DC2626"
        fig_sal=go.Figure()
        fig_sal.add_trace(go.Scatter(x=x,y=saldo,fill="tozeroy",mode="lines+markers",
          line=dict(color=cor_saldo,width=2.8),
          fillcolor="rgba(5,150,105,.08)" if cor_saldo=="#059669" else "rgba(220,38,38,.08)",
          marker=dict(size=6,color=cor_saldo,line=dict(color="white",width=2)),
          hovertemplate="<b>Saldo Acumulado</b><br>%{x}<br>R$ %{y:,.0f}<extra></extra>"))
        fig_sal.add_hline(y=0,line_dash="dot",line_color="#9CA3AF",opacity=0.5)
        fig_sal.update_layout(title=dict(text="Saldo Acumulado",
          font=dict(size=13,family="Georgia, serif",color="#14243B")),**TH_FC,showlegend=False)
        st.plotly_chart(fig_sal,use_container_width=True)
    sec("📊 Detalhamento")
    t1,t2=st.tabs(["Entradas por tipo","Saídas por Centro"])
    CORES_LIGHT=["#0F6E56","#A9762F","#059669","#2563EB","#7C3AED","#0891B2","#6b7280","#14B8A6"]

    # Lê do mesmo arquivo de detalhamento que alimenta o drill-down — garante que o gráfico
    # mostre exatamente as mesmas subcontas reais, com os mesmos nomes, em qualquer formato de origem
    # Só lê o detalhamento se o Fluxo de Caixa realmente foi importado nesse cliente (totais reais
    # na tabela, não zerados) — evita mostrar gráfico "fantasma" de uma importação antiga
    df_det_graf=None
    fluxo_tem_dado_real=float(df_f["tot entradas"].sum() or 0)!=0 or float(df_f["tot saidas"].sum() or 0)!=0
    if st.session_state.cid and fluxo_tem_dado_real:
        path_det_graf=os.path.join(PASTA,f"{gid(st.session_state.cid)}_detalhamento.csv")
        if os.path.exists(path_det_graf):
            df_det_graf=pd.read_csv(path_det_graf,sep=";",decimal=",",encoding="utf-8-sig")

    def _totais_por_subconta(campo_pai_alvo,top_n=7):
        """Agrupa o detalhamento por subconta, somando todos os períodos visíveis (df_f).
        Retorna as top_n maiores + 'Outras' com o restante, para manter o gráfico legível."""
        if df_det_graf is None: return {}
        dfd=df_det_graf[df_det_graf["campo_pai"]==campo_pai_alvo]
        if dfd.empty: return {}
        # Restringe aos períodos que estão sendo exibidos no momento (respeita filtro de ano/mês)
        anos_visiveis=set(df_f[ca].astype(str)) if ca and ca in df_f.columns else None
        if anos_visiveis:
            dfd=dfd[dfd["ano"].astype(str).isin(anos_visiveis)]
        soma_sub=dfd.groupby("subconta")["valor"].sum().sort_values(ascending=False)
        if len(soma_sub)<=top_n:
            return soma_sub.to_dict()
        top=soma_sub.head(top_n)
        resto=soma_sub.iloc[top_n:].sum()
        resultado=top.to_dict()
        if resto>0: resultado["Outras"]=resto
        return resultado

    with t1:
        totais_ent=_totais_por_subconta("Disponibilidades entradas")
        # Entradas à Vista (estimativa configurada manualmente, não vem do detalhamento)
        ev=float(st.session_state.get("entradas_vista",0))
        freq=st.session_state.get("freq_fluxo","Mensal")
        if ev>0:
            vals_vista_total=(ev*22 if freq=="Diário" else ev)*len(df_f)
            totais_ent["À Vista"]=vals_vista_total
        if totais_ent:
            fig_ent=go.Figure(go.Bar(x=list(totais_ent.keys()),y=list(totais_ent.values()),
              marker=dict(color=CORES_LIGHT[:len(totais_ent)],line=dict(color="white",width=0.8)),
              opacity=.92,text=[f"R$ {v/1e6:.1f}M" for v in totais_ent.values()],
              textposition="outside",textfont=dict(size=9,color="#374151")))
            fig_ent.update_layout(title=dict(text="Entradas por Subconta",
              font=dict(size=13,family="Georgia, serif",color="#14243B")),
              bargap=0.35,**TH_FC,showlegend=False)
            st.plotly_chart(fig_ent,use_container_width=True)
            fig_pie=go.Figure(go.Pie(
              labels=list(totais_ent.keys()),
              values=list(totais_ent.values()),
              marker=dict(colors=CORES_LIGHT[:len(totais_ent)],line=dict(color="white",width=2)),
              hole=.5,textfont=dict(size=10,color="#FFFFFF"),
              textinfo="percent",hovertemplate="<b>%{label}</b><br>R$ %{value:,.0f}<br>%{percent}<extra></extra>"))
            fig_pie.update_layout(
              title=dict(text="Distribuição das Entradas",font=dict(size=13,family="Georgia, serif",color="#14243B")),
              plot_bgcolor="white",paper_bgcolor="white",
              font=dict(color="#6B7280",size=10),
              margin=dict(l=8,r=8,t=40,b=8),height=300,
              legend=dict(bgcolor="rgba(0,0,0,0)",font=dict(size=9),orientation="v",x=1.02))
            st.plotly_chart(fig_pie,use_container_width=True)
        else:
            st.markdown('<div class="al-i">Dados de entradas não disponíveis neste arquivo.</div>',unsafe_allow_html=True)
    with t2:
        totais_sai=_totais_por_subconta("Disponibilidades Saida")
        if totais_sai:
            fig_sai=go.Figure(go.Bar(x=list(totais_sai.keys()),y=list(totais_sai.values()),
              marker=dict(color=CORES_LIGHT[:len(totais_sai)],line=dict(color="white",width=0.8)),
              opacity=.92,text=[f"R$ {v/1e6:.1f}M" for v in totais_sai.values()],
              textposition="outside",textfont=dict(size=9,color="#374151")))
            fig_sai.update_layout(title=dict(text="Saídas por Subconta",
              font=dict(size=13,family="Georgia, serif",color="#14243B")),
              bargap=0.35,**TH_FC,showlegend=False)
            st.plotly_chart(fig_sai,use_container_width=True)
            fig_pie_sai=go.Figure(go.Pie(
              labels=list(totais_sai.keys()),
              values=list(totais_sai.values()),
              marker=dict(colors=CORES_LIGHT[:len(totais_sai)],line=dict(color="white",width=2)),
              hole=.5,textfont=dict(size=10,color="#111827"),
              textinfo="percent",hovertemplate="<b>%{label}</b><br>R$ %{value:,.0f}<br>%{percent}<extra></extra>"))
            fig_pie_sai.update_layout(
              title=dict(text="Distribuição das Saídas",font=dict(size=13,family="Georgia, serif",color="#14243B")),
              plot_bgcolor="white",paper_bgcolor="white",
              font=dict(color="#6B7280",size=10),
              margin=dict(l=8,r=8,t=40,b=8),height=300,
              legend=dict(bgcolor="rgba(0,0,0,0)",font=dict(size=9),orientation="v",x=1.02))
            st.plotly_chart(fig_pie_sai,use_container_width=True)
        else:
            st.markdown('<div class="al-i">Dados de centro de custo não disponíveis.</div>',unsafe_allow_html=True)

    if st.session_state.cid and fluxo_tem_dado_real:
        path_detalhe_fc=os.path.join(PASTA,f"{gid(st.session_state.cid)}_detalhamento.csv")
        if os.path.exists(path_detalhe_fc):
            df_det_fc=pd.read_csv(path_detalhe_fc,sep=";",decimal=",",encoding="utf-8-sig")
            df_det_fc_fluxo=df_det_fc[df_det_fc["campo_pai"].isin(["Disponibilidades entradas","Disponibilidades Saida"])]
            if not df_det_fc_fluxo.empty:
                with st.expander("🔍 Ver detalhamento por categoria (drill-down)"):
                    campos_pai_fc=sorted(df_det_fc_fluxo["campo_pai"].unique().tolist())
                    campo_sel_fc=st.selectbox("Selecione",campos_pai_fc,key="dd_fluxo_campo",
                      format_func=lambda x: "Entradas" if x=="Disponibilidades entradas" else "Saídas")
                    df_det_f_fc=df_det_fc_fluxo[df_det_fc_fluxo["campo_pai"]==campo_sel_fc]
                    if ano_sel!="Todos":
                        df_det_f_fc=df_det_f_fc[df_det_f_fc["ano"].astype(str)==ano_sel]
                    if mes_sel!="Todos":
                        mes_sel_3=str(mes_sel).lower()[:3]
                        df_det_f_fc=df_det_f_fc[df_det_f_fc["mes"].astype(str).str.lower()==mes_sel_3]
                    if not df_det_f_fc.empty:
                        df_det_f_fc=df_det_f_fc.copy()
                        df_det_f_fc["periodo"]=df_det_f_fc["mes"].astype(str)+"/"+df_det_f_fc["ano"].astype(str).str[-2:]
                        pivot_fc=df_det_f_fc.pivot_table(index="subconta",columns="periodo",values="valor",
                                                        aggfunc="sum",fill_value=0)
                        ordem_meses_fc=["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]
                        cols_ordenadas_fc=sorted(pivot_fc.columns,
                          key=lambda x: (x.split("/")[1], ordem_meses_fc.index(x.split("/")[0]) if x.split("/")[0] in ordem_meses_fc else 99))
                        pivot_fc=pivot_fc[cols_ordenadas_fc]
                        pivot_fc["Total"]=pivot_fc.sum(axis=1)
                        pivot_fc=pivot_fc.sort_values("Total",ascending=False)
                        pivot_fc_fmt=pivot_fc.copy()
                        for c in pivot_fc_fmt.columns:
                            pivot_fc_fmt[c]=pivot_fc_fmt[c].apply(lambda v: fmt(v))
                        st.dataframe(pivot_fc_fmt,use_container_width=True)
                    else:
                        st.info("Sem detalhamento disponível para este período.")

    def gerar_excel_fluxo():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb=Workbook()
        ws=wb.active
        ws.title="Fluxo de Caixa"

        cor_azul=PatternFill("solid",start_color="2563EB",end_color="2563EB")
        cor_cinza_clara=PatternFill("solid",start_color="F3F4F6",end_color="F3F4F6")
        cor_branco_bold=Font(bold=True,color="FFFFFF",size=11)
        cor_preto=Font(color="111827",size=10)
        cor_verde=Font(color="059669",size=10)
        cor_vermelho=Font(color="DC2626",size=10)
        cor_cinza_texto=Font(color="9CA3AF",size=10)
        borda_fina=Border(bottom=Side(style="thin",color="E8ECF0"))

        filtro_txt=f"Filtro aplicado: Ano = {ano_sel}" if ano_sel!="Todos" else "Filtro aplicado: Todos os períodos"
        if mes_sel!="Todos": filtro_txt+=f" | Mês = {mes_sel}"
        ws.cell(row=1,column=1,value=filtro_txt).font=Font(italic=True,color="6B7280",size=9)
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=3)

        meses_cols_fc=list(df_f[ca].astype(str)+"/"+df_f[cm].astype(str)) if ca and cm and ca in df_f.columns else list(df_f[cm].astype(str)) if cm else []
        ws.cell(row=2,column=1,value="Descrição").font=cor_branco_bold
        ws.cell(row=2,column=1).fill=cor_azul
        col=2
        for m in meses_cols_fc:
            c=ws.cell(row=2,column=col,value=m)
            c.font=cor_branco_bold; c.fill=cor_azul
            c.alignment=Alignment(horizontal="center")
            col+=1

        linhas_fc_xls=[
            ("tot","Total Entradas","Disponibilidades entradas",False),
            ("sub","  Receita Serviços","Centro de Custos Entradas 1",False),
            ("sub","  Receita Produtos","Centro de Custos Entradas 2",False),
            ("sub","  Recebimento Clientes","Centro de Custos Entradas 3",False),
            ("sub","  Receita Financeira/Outras","Centro de Custos Entradas 4",False),
            ("tot","Total Saídas","Disponibilidades Saida",True),
            ("sub","  Folha","Centro de Custos Saidas 1",True),
            ("sub","  Fornecedores","Centro de Custos Saidas 2",True),
            ("sub","  Impostos/Operacionais","Centro de Custos Saidas 3",True),
            ("sub","  Investimentos","Centro de Custos Saidas 4",True),
            ("tot","Saldo Acumulado","saldo acumulado",False),
        ]

        linha_excel=3
        for tipo,desc,campo,inv in linhas_fc_xls:
            if campo not in df_f.columns: continue
            ws.cell(row=linha_excel,column=1,value=desc.strip())
            if tipo=="tot":
                ws.cell(row=linha_excel,column=1).font=Font(bold=True,color="2563EB",size=10)
                ws.cell(row=linha_excel,column=1).fill=cor_cinza_clara
            else:
                ws.cell(row=linha_excel,column=1).font=cor_preto

            col=2
            for i in range(len(df_f)):
                v=float(df_f.iloc[i].get(campo,0) or 0)
                c1=ws.cell(row=linha_excel,column=col,value=v)
                c1.number_format='R$ #,##0;[RED](R$ #,##0)'
                c1.font=cor_vermelho if (inv and v!=0) else (cor_verde if v>0 else cor_preto)
                ws.cell(row=linha_excel,column=col).border=borda_fina
                col+=1
            linha_excel+=1

        ws.column_dimensions["A"].width=26
        for cc in range(2,col):
            ws.column_dimensions[get_column_letter(cc)].width=13
        ws.freeze_panes="B3"

        if st.session_state.cid:
            path_detalhe_xls=os.path.join(PASTA,f"{gid(st.session_state.cid)}_detalhamento.csv")
            if os.path.exists(path_detalhe_xls):
                df_det_xls=pd.read_csv(path_detalhe_xls,sep=";",decimal=",",encoding="utf-8-sig")
                df_det_fc_xls=df_det_xls[df_det_xls["campo_pai"].isin(["Disponibilidades entradas","Disponibilidades Saida"])]
                if ano_sel!="Todos":
                    df_det_fc_xls=df_det_fc_xls[df_det_fc_xls["ano"].astype(str)==ano_sel]
                if not df_det_fc_xls.empty:
                    ws2=wb.create_sheet("Drill-down")
                    ws2.cell(row=1,column=1,value="Conta Principal").font=cor_branco_bold
                    ws2.cell(row=1,column=2,value="Subconta").font=cor_branco_bold
                    ws2.cell(row=1,column=3,value="Ano").font=cor_branco_bold
                    ws2.cell(row=1,column=4,value="Mês").font=cor_branco_bold
                    ws2.cell(row=1,column=5,value="Valor (R$)").font=cor_branco_bold
                    for cc in range(1,6):
                        ws2.cell(row=1,column=cc).fill=cor_azul
                    linha2=2
                    for _,r in df_det_fc_xls.sort_values(["campo_pai","subconta","ano","mes"]).iterrows():
                        ws2.cell(row=linha2,column=1,value=r["campo_pai"])
                        ws2.cell(row=linha2,column=2,value=r["subconta"])
                        ws2.cell(row=linha2,column=3,value=str(r["ano"]))
                        ws2.cell(row=linha2,column=4,value=r["mes"])
                        c_val=ws2.cell(row=linha2,column=5,value=float(r["valor"]))
                        c_val.number_format='R$ #,##0'
                        linha2+=1
                    for cc,w in zip("ABCDE",[26,28,8,8,16]):
                        ws2.column_dimensions[cc].width=w
                    ws2.freeze_panes="A2"

        buf=BytesIO(); wb.save(buf); buf.seek(0)
        return buf.getvalue()

    st.download_button("📥 Exportar este Fluxo de Caixa (Excel formatado)",gerar_excel_fluxo(),
      file_name=f"FluxoCaixa_{ano_sel if ano_sel!='Todos' else 'completo'}.xlsx",
      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
      use_container_width=True)

    cols_show=[c for c in [cm,"tot entradas","tot saidas","saldo período","saldo acumulado"] if c and c in df_f.columns]
    with st.expander(f"📋 Tabela Completa ({len(df_f)} períodos)"):
        st.dataframe(df_f[cols_show] if cols_show else df_f,use_container_width=True)

# ── BALANÇO ─────────────────────────────────────────
elif pg=="balanco":
    hdr("🏦 Balanço Patrimonial","Posição financeira com AV% e AH%")

    _df_raw_bal=get_df_raw_bruto()
    _col_fil_bal=col_filial(_df_raw_bal) if _df_raw_bal is not None else None
    filial_sel_bal=None
    if _col_fil_bal:
        _filiais_disp_bal=sorted(v for v in _df_raw_bal[_col_fil_bal].dropna().astype(str).unique().tolist() if v!="(Todas as filiais)")
        _opcoes_filial_bal=["(Todas as filiais)"]+_filiais_disp_bal

        def _on_change_filial_bal():
            st.session_state["bal_filial_sel_backup"]=st.session_state["bal_filial_sel"]

        if "bal_filial_sel_backup" not in st.session_state:
            st.session_state["bal_filial_sel_backup"]=_opcoes_filial_bal[0]
        if st.session_state["bal_filial_sel_backup"] not in _opcoes_filial_bal:
            st.session_state["bal_filial_sel_backup"]=_opcoes_filial_bal[0]
        st.session_state["bal_filial_sel"]=st.session_state["bal_filial_sel_backup"]

        filial_sel_bal=st.selectbox("🏬 Filial",_opcoes_filial_bal,
          key="bal_filial_sel",on_change=_on_change_filial_bal)

    df=get_df_filial(filial_sel_bal)
    if df is None: no_data(); st.stop()
    cm=cm_(df); ca=ca_(df)
    anos=sorted({str(a) for a in df[ca].dropna().unique()}) if ca else []
    col_b1,col_b2=st.columns(2)
    ano_b=col_b1.selectbox("Ano",["Todos"]+anos,key="bal_ano")
    df_b=df[df[ca].astype(str)==ano_b].reset_index(drop=True) if ano_b!="Todos" else df.reset_index(drop=True)
    meses_b=["Todos"]+df_b[cm].dropna().unique().tolist() if cm else ["Todos"]
    mes_b=col_b2.selectbox("Mês",meses_b,key="bal_mes")
    df_b=df_b[df_b[cm]==mes_b].reset_index(drop=True) if mes_b!="Todos" else df_b
    if df_b.empty: st.warning("Sem dados para este período."); st.stop()
    df_b4=df_b  # todos os períodos filtrados
    ul=df_b4.iloc[-1]
    k=st.columns(4)
    mc(k[0],"Ativo Total",fmt(ul.get("ativo total",0)),"b")
    mc(k[1],"Passivo Total",fmt(ul.get("pass total",0)),"r")
    mc(k[2],"Patrimônio Líquido",fmt(ul.get("PL",0)),cor(ul.get("PL",0)))
    li=float(ul.get("liquidez imediata",0) or 0)
    mc(k[3],"Liq. Imediata",fmt(li,"x"),"g" if li>=0.5 else ("y" if li>=0.3 else "r"),
       "✅ Saudável" if li>=0.5 else ("⚠️ Atenção" if li>=0.3 else "🔴 Risco"))
    x_b=list(df_b4[cm].astype(str)) if cm else [str(i) for i in range(len(df_b4))]
    sec("📋 Balanço — Real | AV% | AH%")
    def bal_row(desc,campo,tipo="sub",inv=False):
        row=f'<tr class="{tipo}"><td>{desc}</td>'
        for i in range(len(df_b4)):
            v=float(df_b4.iloc[i].get(campo,0))
            # AV% usa Ativo Total do mesmo período como base
            at_i=float(df_b4.iloc[i].get("ativo total",1) or 1)
            a_v=v/at_i*100
            a_h=safe(v-float(df_b4.iloc[i-1].get(campo,1)),
                     abs(float(df_b4.iloc[i-1].get(campo,1)) or 1))*100 if i>0 else 0.
            row+=f'<td class="{cor(v,inv)}">{fmt(v)}</td>'
            row+=f'<td class="{cls_pct(a_v,inv)}">{fmt(a_v,"pct")}</td>'
            row+=f'<td class="{cls_pct(a_h,inv)}">{fmt(a_h,"pct")}</td>'
        return row+"</tr>"
    header_b="<tr><th>Descrição</th>"+"".join(f"<th>{m}</th><th>AV%</th><th>AH%</th>" for m in x_b)+"</tr>"
    def tot_row(desc,campo,cls="b"):
        row=f'<tr class="tot"><td>{desc}</td>'
        for i in range(len(df_b4)):
            v=float(df_b4.iloc[i].get(campo,0))
            at_i=float(df_b4.iloc[i].get("ativo total",1) or 1)
            a_v=v/at_i*100
            a_h=safe(v-float(df_b4.iloc[i-1].get(campo,1)),
                     abs(float(df_b4.iloc[i-1].get(campo,1)) or 1))*100 if i>0 else 0.
            row+=f'<td class="{cls}">{fmt(v)}</td>'
            row+=f'<td>{fmt(a_v,"pct")}</td>'
            row+=f'<td class="{cls_pct(a_h)}">{fmt(a_h,"pct")}</td>'
        return row+"</tr>"
    rows_b=('<tr class="cat"><td>ATIVO</td>'+"".join("<td></td><td></td><td></td>" for _ in x_b)+"</tr>"+
        bal_row("  Disponibilidades","disponibilidades saldo")+
        bal_row("  Contas a Receber","contas a receber saldo")+
        bal_row("  Estoques","estoque final do mês de mercadorias para revenda saldo")+
        bal_row("  Outros AC","Outros AC")+
        tot_row("= ATIVO CIRCULANTE","ativo circ")+
        bal_row("  Ativo NC","Ativo NC")+
        tot_row("= ATIVO TOTAL","ativo total")+
        '<tr class="cat"><td>PASSIVO</td>'+"".join("<td></td><td></td><td></td>" for _ in x_b)+"</tr>"+
        bal_row("  Fornecedores","contas a pagar de fornecedores saldo",inv=True)+
        bal_row("  Pass. Financeiros","Passivos Financeiros",inv=True)+
        bal_row("  Outros PC","Outros PC",inv=True)+
        tot_row("= PASSIVO CIRCULANTE","pass circ","r")+
        bal_row("  Passivo NC","Passivo NC",inv=True)+
        tot_row("= PASSIVO TOTAL","pass total","r")+
        tot_row("= PATRIMÔNIO LÍQUIDO","PL","g"))
    st.markdown(f'<div class="dre-wrap" style="overflow-x:auto;max-width:100%"><table class="dre" style="min-width:1200px">{header_b}{rows_b}</table></div>',unsafe_allow_html=True)
    def gerar_excel_balanco():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb=Workbook()
        ws=wb.active
        ws.title="Balanço"

        cor_azul=PatternFill("solid",start_color="2563EB",end_color="2563EB")
        cor_cinza_clara=PatternFill("solid",start_color="F3F4F6",end_color="F3F4F6")
        cor_branco_bold=Font(bold=True,color="FFFFFF",size=11)
        cor_preto=Font(color="111827",size=10)
        cor_verde=Font(color="059669",size=10)
        cor_vermelho=Font(color="DC2626",size=10)
        cor_cinza_texto=Font(color="9CA3AF",size=10)
        borda_fina=Border(bottom=Side(style="thin",color="E8ECF0"))

        filtro_txt=f"Filtro aplicado: Ano = {ano_b}" if ano_b!="Todos" else "Filtro aplicado: Todos os períodos"
        if mes_b!="Todos": filtro_txt+=f" | Mês = {mes_b}"
        ws.cell(row=1,column=1,value=filtro_txt).font=Font(italic=True,color="6B7280",size=9)
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=4)

        meses_cols_bal=list(df_b4[ca].astype(str)+"/"+df_b4[cm].astype(str)) if ca and cm else x_b
        ws.cell(row=2,column=1,value="Descrição").font=cor_branco_bold
        ws.cell(row=2,column=1).fill=cor_azul
        col=2
        for m in meses_cols_bal:
            for sub in ["Real","AV%","AH%"]:
                c=ws.cell(row=2,column=col,value=f"{m} - {sub}")
                c.font=cor_branco_bold; c.fill=cor_azul
                c.alignment=Alignment(horizontal="center")
                col+=1

        linhas_bal=[
            ("cat","ATIVO",None,False),
            ("sub","  Disponibilidades","disponibilidades saldo",False),
            ("sub","  Contas a Receber","contas a receber saldo",False),
            ("sub","  Estoques","estoque final do mês de mercadorias para revenda saldo",False),
            ("sub","  Outros AC","Outros AC",False),
            ("tot","ATIVO CIRCULANTE","ativo circ",False),
            ("sub","  Ativo NC","Ativo NC",False),
            ("tot","ATIVO TOTAL","ativo total",False),
            ("cat","PASSIVO",None,False),
            ("sub","  Fornecedores","contas a pagar de fornecedores saldo",True),
            ("sub","  Pass. Financeiros","Passivos Financeiros",True),
            ("sub","  Outros PC","Outros PC",True),
            ("tot","PASSIVO CIRCULANTE","pass circ",True),
            ("sub","  Passivo NC","Passivo NC",True),
            ("tot","PASSIVO TOTAL","pass total",True),
            ("tot","PATRIMÔNIO LÍQUIDO","PL",False),
        ]

        linha_excel=3
        for tipo,desc,campo,inv in linhas_bal:
            ws.cell(row=linha_excel,column=1,value=desc.strip())
            if tipo=="tot":
                ws.cell(row=linha_excel,column=1).font=Font(bold=True,color="2563EB",size=10)
                ws.cell(row=linha_excel,column=1).fill=cor_cinza_clara
            elif tipo=="cat":
                ws.cell(row=linha_excel,column=1).font=Font(bold=True,color="111827",size=10)
            else:
                ws.cell(row=linha_excel,column=1).font=cor_preto

            col=2
            if tipo=="cat":
                for i in range(len(df_b4)):
                    col+=3
                for cc in range(1,col):
                    ws.cell(row=linha_excel,column=cc).border=borda_fina
                linha_excel+=1
                continue

            for i in range(len(df_b4)):
                v=float(df_b4.iloc[i].get(campo,0))
                at_i=float(df_b4.iloc[i].get("ativo total",1) or 1)
                a_v=v/at_i*100
                a_h=safe(v-float(df_b4.iloc[i-1].get(campo,1)),
                         abs(float(df_b4.iloc[i-1].get(campo,1)) or 1))*100 if i>0 else 0.
                c1=ws.cell(row=linha_excel,column=col,value=v)
                c1.number_format='R$ #,##0;[RED](R$ #,##0)'
                c1.font=cor_vermelho if (inv and v!=0) else (cor_verde if v>0 else cor_preto)
                c2=ws.cell(row=linha_excel,column=col+1,value=a_v/100); c2.number_format="0.0%"
                c3=ws.cell(row=linha_excel,column=col+2,value=a_h/100); c3.number_format="+0.0%;-0.0%"
                c3.font=cor_verde if a_h>0 else (cor_vermelho if a_h<0 else cor_cinza_texto)
                col+=3
            for cc in range(1,col):
                ws.cell(row=linha_excel,column=cc).border=borda_fina
            linha_excel+=1

        ws.column_dimensions["A"].width=28
        for cc in range(2,col):
            ws.column_dimensions[get_column_letter(cc)].width=14
        ws.freeze_panes="B3"

        if st.session_state.cid:
            path_detalhe_xls=os.path.join(PASTA,f"{gid(st.session_state.cid)}_detalhamento.csv")
            if os.path.exists(path_detalhe_xls):
                df_det_xls=pd.read_csv(path_detalhe_xls,sep=";",decimal=",",encoding="utf-8-sig")
                campos_bal_xls=["disponibilidades saldo","contas a receber saldo",
                    "estoque final do mês de mercadorias para revenda saldo","Outros AC","Ativo NC",
                    "contas a pagar de fornecedores saldo","Passivos Financeiros","Outros PC","Passivo NC","PL"]
                df_det_bal_xls=df_det_xls[df_det_xls["campo_pai"].isin(campos_bal_xls)]
                if ano_b!="Todos":
                    df_det_bal_xls=df_det_bal_xls[df_det_bal_xls["ano"].astype(str)==ano_b]
                if not df_det_bal_xls.empty:
                    ws2=wb.create_sheet("Drill-down")
                    ws2.cell(row=1,column=1,value="Conta Principal").font=cor_branco_bold
                    ws2.cell(row=1,column=2,value="Subconta").font=cor_branco_bold
                    ws2.cell(row=1,column=3,value="Ano").font=cor_branco_bold
                    ws2.cell(row=1,column=4,value="Mês").font=cor_branco_bold
                    ws2.cell(row=1,column=5,value="Valor (R$)").font=cor_branco_bold
                    for cc in range(1,6):
                        ws2.cell(row=1,column=cc).fill=cor_azul
                    linha2=2
                    for _,r in df_det_bal_xls.sort_values(["campo_pai","subconta","ano","mes"]).iterrows():
                        ws2.cell(row=linha2,column=1,value=r["campo_pai"])
                        ws2.cell(row=linha2,column=2,value=r["subconta"])
                        ws2.cell(row=linha2,column=3,value=str(r["ano"]))
                        ws2.cell(row=linha2,column=4,value=r["mes"])
                        c_val=ws2.cell(row=linha2,column=5,value=float(r["valor"]))
                        c_val.number_format='R$ #,##0'
                        linha2+=1
                    for cc,w in zip("ABCDE",[26,28,8,8,16]):
                        ws2.column_dimensions[cc].width=w
                    ws2.freeze_panes="A2"

        buf=BytesIO(); wb.save(buf); buf.seek(0)
        return buf.getvalue()

    st.download_button("📥 Exportar este Balanço (Excel formatado)",gerar_excel_balanco(),
      file_name=f"Balanco_{ano_b if ano_b!='Todos' else 'completo'}.xlsx",
      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
      use_container_width=True)
    # Só mostra o drill-down se o Balanço realmente foi importado nesse cliente (cards com valor real) —
    # evita mostrar detalhamento "fantasma" de uma importação antiga quando o banco atual está zerado
    ativo_total_atual=float(ul.get("ativo total",0) or 0)
    if st.session_state.cid and ativo_total_atual!=0:
        path_detalhe_bal=os.path.join(PASTA,f"{gid(st.session_state.cid)}_detalhamento.csv")
        if os.path.exists(path_detalhe_bal):
            df_det_bal=pd.read_csv(path_detalhe_bal,sep=";",decimal=",",encoding="utf-8-sig")
            campos_balanco_dd=["disponibilidades saldo","contas a receber saldo",
                "estoque final do mês de mercadorias para revenda saldo","Outros AC","Ativo NC",
                "contas a pagar de fornecedores saldo","Passivos Financeiros","Outros PC","Passivo NC","PL"]
            df_det_bal_f=df_det_bal[df_det_bal["campo_pai"].isin(campos_balanco_dd)]
            if not df_det_bal_f.empty:
                if "bal_dd_expander_aberto" not in st.session_state:
                    st.session_state["bal_dd_expander_aberto"]=False
                with st.expander("🔍 Ver detalhamento por subconta (drill-down)",expanded=st.session_state["bal_dd_expander_aberto"]):
                    st.session_state["bal_dd_expander_aberto"]=True
                    campos_pai_bal=sorted(df_det_bal_f["campo_pai"].unique().tolist())
                    campo_sel_bal=st.selectbox("Selecione a conta",campos_pai_bal,key="dd_bal_campo")
                    df_det_bal_sel=df_det_bal_f[df_det_bal_f["campo_pai"]==campo_sel_bal]
                    if ano_b!="Todos":
                        df_det_bal_sel=df_det_bal_sel[df_det_bal_sel["ano"].astype(str)==ano_b]
                    if mes_b!="Todos":
                        mes_b_3=str(mes_b).lower()[:3]
                        df_det_bal_sel=df_det_bal_sel[df_det_bal_sel["mes"].astype(str).str.lower()==mes_b_3]
                    if not df_det_bal_sel.empty:
                        df_det_bal_sel=df_det_bal_sel.copy()
                        df_det_bal_sel["periodo"]=df_det_bal_sel["mes"].astype(str)+"/"+df_det_bal_sel["ano"].astype(str).str[-2:]
                        pivot_bal=df_det_bal_sel.pivot_table(index="subconta",columns="periodo",values="valor",
                                                            aggfunc="sum",fill_value=0)
                        ordem_meses_bal=["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]
                        cols_ordenadas_bal=sorted(pivot_bal.columns,
                          key=lambda x: (x.split("/")[1], ordem_meses_bal.index(x.split("/")[0]) if x.split("/")[0] in ordem_meses_bal else 99))
                        pivot_bal=pivot_bal[cols_ordenadas_bal]
                        pivot_bal_fmt=pivot_bal.copy()
                        for c in pivot_bal_fmt.columns:
                            pivot_bal_fmt[c]=pivot_bal_fmt[c].apply(lambda v: fmt(v))
                        st.dataframe(pivot_bal_fmt,use_container_width=True)
                    else:
                        st.info("Sem detalhamento disponível para esta conta no período selecionado.")

    g1,g2=st.columns(2)
    TH_BAL=dict(plot_bgcolor="white",paper_bgcolor="white",
        font=dict(color="#6B7280",size=10,family="Inter"),
        xaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=False,tickangle=-35),
        yaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=True),
        margin=dict(l=8,r=8,t=40,b=40),
        legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.3,font=dict(size=9)),
        hovermode="x unified",
        hoverlabel=dict(bgcolor="white",bordercolor="#E8ECF0",font=dict(color="#111827",size=11)))
    x_bal=df[cm].astype(str) if cm else pd.Series(range(len(df))).astype(str)
    if ca and cm:
        x_bal=df[cm].astype(str)+"/"+df[ca].astype(str).str[-2:]
    TH_NOVO=dict(plot_bgcolor="white",paper_bgcolor="white",
      font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
      margin=dict(l=10,r=10,t=50,b=60),
      xaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=False,
        tickangle=-40,tickfont=dict(size=9,color="#4B5563")),
      yaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=True,zeroline=False,tickfont=dict(size=9)),
      legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.34,x=0.5,xanchor="center",font=dict(size=10)),
      hovermode="x unified",height=420,
      hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B",size=11)))

    with g1:
        fig_pat=go.Figure()
        for c,cor_c,nm in [("ativo total","#14243B","Ativo Total"),
                            ("pass total","#A9762F","Passivo Total"),
                            ("PL","#5B7B9A","Patrimônio Líquido")]:
            if c not in df.columns: continue
            fig_pat.add_trace(go.Scatter(x=x_bal,y=pd.to_numeric(df[c],errors="coerce"),
              name=nm,mode="lines+markers",line=dict(color=cor_c,width=2.4),
              marker=dict(size=6,color=cor_c,line=dict(color="white",width=1.5)),
              hovertemplate=f"<b>{nm}</b><br>%{{x}}<br>R$ %{{y:,.0f}}<extra></extra>"))
        fig_pat.update_layout(title=dict(text="Estrutura Patrimonial",
          font=dict(size=15,family="Georgia, serif",color="#14243B")),**TH_NOVO)
        st.plotly_chart(fig_pat,use_container_width=True)
    with g2:
        fig_liq=go.Figure()
        for c,cor_c,nm,ref in [("liquidez corrente","#14243B","Liq. Corrente",1.5),
                                ("liquidez imediata","#A9762F","Liq. Imediata",0.5)]:
            if c not in df.columns: continue
            fig_liq.add_trace(go.Scatter(x=x_bal,y=pd.to_numeric(df[c],errors="coerce"),
              name=nm,mode="lines+markers",line=dict(color=cor_c,width=2.4),
              marker=dict(size=6,color=cor_c,line=dict(color="white",width=1.5)),
              hovertemplate=f"<b>{nm}</b><br>%{{x}}<br>%{{y:.2f}}x<extra></extra>"))
            fig_liq.add_hline(y=ref,line_dash="dot",line_color=cor_c,opacity=0.4,
              annotation_text=f"Referência {nm}: {ref}x",
              annotation_font=dict(size=8.5,color=cor_c),
              annotation_position="top left")
        fig_liq.update_layout(title=dict(text="Índices de Liquidez",
          font=dict(size=15,family="Georgia, serif",color="#14243B")),**TH_NOVO)
        st.plotly_chart(fig_liq,use_container_width=True)

    

# ── INDICADORES ─────────────────────────────────────
elif pg=="indicadores":
    hdr("📈 Indicadores","KPIs com gauge, histórico, máx. e mín.")
    _df_raw_ind=get_df_raw_bruto()
    _col_fil_ind=col_filial(_df_raw_ind) if _df_raw_ind is not None else None
    filial_sel_ind=None
    if _col_fil_ind:
        _filiais_disp_ind=sorted(v for v in _df_raw_ind[_col_fil_ind].dropna().astype(str).unique().tolist() if v!="(Todas as filiais)")
        _opcoes_filial_ind=["(Todas as filiais)"]+_filiais_disp_ind

        def _on_change_filial_ind():
            st.session_state["ind_filial_sel_backup"]=st.session_state["ind_filial_sel"]

        if "ind_filial_sel_backup" not in st.session_state:
            st.session_state["ind_filial_sel_backup"]=_opcoes_filial_ind[0]
        if st.session_state["ind_filial_sel_backup"] not in _opcoes_filial_ind:
            st.session_state["ind_filial_sel_backup"]=_opcoes_filial_ind[0]
        st.session_state["ind_filial_sel"]=st.session_state["ind_filial_sel_backup"]

        filial_sel_ind=st.selectbox("🏬 Filial",_opcoes_filial_ind,
          key="ind_filial_sel",on_change=_on_change_filial_ind)

    df=get_df_filial(filial_sel_ind)
    if df is None: no_data(); st.stop()
    cm=cm_(df); ca=ca_(df)

    # Filtros
    col_i1,col_i2=st.columns(2)
    anos_i=["Todos"]+[str(a) for a in sorted(df[ca].dropna().unique().tolist())] if ca else ["Todos"]
    ano_i=col_i1.selectbox("Ano",anos_i,key="ind_ano")
    df_i=df[df[ca].astype(str)==ano_i].reset_index(drop=True) if ano_i!="Todos" else df.reset_index(drop=True)
    meses_i=["Todos"]+df_i[cm].dropna().unique().tolist() if cm else ["Todos"]
    mes_i=col_i2.selectbox("Mês",meses_i,key="ind_mes")
    df_i=df_i[df_i[cm]==mes_i].reset_index(drop=True) if mes_i!="Todos" else df_i
    if df_i.empty: st.warning("Sem dados para este período."); st.stop()
    ul=df_i.iloc[-1]

    # Snapshot financeiro pro Parecer Financeiro (mesma classificação de alertas usada em Alertas e Diagnóstico)
    if st.session_state.cid:
        _alertas_ind=[]
        def _chk_ind(tipo,msg,detalhe=""): _alertas_ind.append([tipo,msg,detalhe])
        try:
            _ml_ind=float(ul.get("margem líquida %",0))
            if _ml_ind<0: _chk_ind("d","💸 Prejuízo operacional",f"Margem Líquida: {_ml_ind:.1f}% — empresa gastando mais do que fatura")
            elif _ml_ind<5: _chk_ind("w","⚠️ Margem Líquida muito baixa",f"{_ml_ind:.1f}% — mínimo recomendado é 5%")
            else: _chk_ind("s","✅ Margem Líquida saudável",f"{_ml_ind:.1f}%")
        except: pass
        try:
            _mb_ind=float(ul.get("margem bruta %",0))
            if _mb_ind<20: _chk_ind("w","⚠️ Margem Bruta baixa",f"{_mb_ind:.1f}% — revisar CMV e precificação")
            else: _chk_ind("s","✅ Margem Bruta adequada",f"{_mb_ind:.1f}%")
        except: pass
        try:
            _lc_ind=float(ul.get("liquidez corrente",0) or 0)
            if _lc_ind<1: _chk_ind("d","🔴 Liquidez Corrente crítica",f"{_lc_ind:.2f}x — dívidas de curto prazo maiores que ativos circulantes")
            elif _lc_ind<1.5: _chk_ind("w","⚠️ Liquidez Corrente abaixo do ideal",f"{_lc_ind:.2f}x — recomendado ≥ 1,5x")
            else: _chk_ind("s","✅ Liquidez Corrente saudável",f"{_lc_ind:.2f}x")
        except: pass
        try:
            _li_ind=float(ul.get("liquidez imediata",0) or 0)
            if _li_ind<0.3: _chk_ind("d","🔴 Liquidez Imediata crítica",f"{_li_ind:.2f}x — caixa insuficiente para cobrir dívidas imediatas")
            elif _li_ind<0.5: _chk_ind("w","⚠️ Liquidez Imediata baixa",f"{_li_ind:.2f}x — recomendado ≥ 0,5x")
            else: _chk_ind("s","✅ Liquidez Imediata saudável",f"{_li_ind:.2f}x")
        except: pass
        try:
            _kz_ind=float(ul.get("kanitz",0))
            if _kz_ind<-3: _chk_ind("d","🔴 Kanitz — Zona Insolvente",f"{_kz_ind:.2f} — alto risco de insolvência")
            elif _kz_ind<0: _chk_ind("w","⚠️ Kanitz — Zona de Penumbra",f"{_kz_ind:.2f} — situação financeira incerta")
            else: _chk_ind("s","✅ Kanitz — Zona Solvente",f"{_kz_ind:.2f} — empresa financeiramente saudável")
        except: pass
        try:
            _cc_ind=float(ul.get("ciclo de caixa",0) or 0)
            if _cc_ind>90: _chk_ind("d","🔴 Ciclo de Caixa muito elevado",f"{_cc_ind:.0f} dias — capital de giro muito comprometido")
            elif _cc_ind>60: _chk_ind("w","⚠️ Ciclo de Caixa elevado",f"{_cc_ind:.0f} dias — monitorar capital de giro")
            elif _cc_ind<0: _chk_ind("s","✅ Ciclo de Caixa negativo (favorável)",f"{_cc_ind:.0f} dias — empresa recebe antes de pagar")
            else: _chk_ind("s","✅ Ciclo de Caixa adequado",f"{_cc_ind:.0f} dias")
        except: pass
        try:
            _roe_ind=float(ul.get("ROE",0))
            if _roe_ind<0: _chk_ind("d","💸 ROE negativo",f"{_roe_ind:.1f}% — retorno negativo sobre patrimônio")
            elif _roe_ind<6: _chk_ind("w","⚠️ ROE baixo",f"{_roe_ind:.1f}% — abaixo da taxa mínima de atratividade")
            else: _chk_ind("s","✅ ROE adequado",f"{_roe_ind:.1f}%")
        except: pass

        _periodo_snap_ind=f"{mes_i}/{ano_i}" if mes_i!="Todos" else (f"Ano {ano_i}" if ano_i!="Todos" else f"{ul.get(cm,'')}/{ul.get(ca,'')}")
        _financeiro_snapshot={
            "periodo":_periodo_snap_ind,
            "receita_bruta":float(ul.get("receita bruta de vendas",0) or 0),
            "receita_liquida":float(ul.get("receita líquida",0) or 0),
            "lucro_bruto":float(ul.get("lucro bruto",0) or 0),
            "lucro_liquido":float(ul.get("lucro líquido",0) or 0),
            "ebitda_rs":float(ul.get("EBITDA",0) or 0),
            "margem_bruta_pct":float(ul.get("margem bruta %",0) or 0),
            "margem_liquida_pct":float(ul.get("margem líquida %",0) or 0),
            "margem_contrib_pct":float(ul.get("margem contrib %",0) or 0),
            "ebitda_pct":float(ul.get("EBITDA %",0) or 0),
            "pmr":float(ul.get("PMR",0) or 0),"pmp":float(ul.get("PMP",0) or 0),
            "pme":float(ul.get("PME",0) or 0),"ciclo_caixa":float(ul.get("ciclo de caixa",0) or 0),
            "giro_estoque":float(ul.get("giro estoque",0) or 0),
            "liquidez_corrente":float(ul.get("liquidez corrente",0) or 0),
            "liquidez_imediata":float(ul.get("liquidez imediata",0) or 0),
            "kanitz":float(ul.get("kanitz",0) or 0),
            "roe":float(ul.get("ROE",0) or 0),
            "icd":float(ul.get("ICD",0) or 0),
            "ativo_total":float(ul.get("ativo total",0) or 0),
            "passivo_total":float(ul.get("pass total",0) or 0),
            "patrimonio_liquido":float(ul.get("PL",0) or 0),
            "score_saude":float(df_i["score_risco"].mean()) if "score_risco" in df_i.columns else 0.0,
            "alertas":_alertas_ind,
            "gerado_em":datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
        save_snap(st.session_state.cid,"financeiro",_financeiro_snapshot,filial=filial_sel_ind)

    # Score de Saúde (0-100) — vinha da antiga página Home
    if "score_risco" in df_i.columns:
        _periodo_ind=f"{mes_i}/{ano_i}" if mes_i!="Todos" else (f"Ano {ano_i}" if ano_i!="Todos" else f"Último: {ul.get(cm,'')}/{ul.get(ca,'')}")
        _sc_ind=float(df_i["score_risco"].mean())
        _lbl_ind,_cls_ind,_cor_s_ind=score_label(_sc_ind)
        _cor_num_ind="#00D4AA" if _sc_ind>=65 else ("#FFB627" if _sc_ind>=45 else "#F85149")
        sec(f"🎯 Score de Saúde — {_periodo_ind}")
        col_score_i,col_graf_i=st.columns([1,3])
        with col_score_i:
            st.markdown(f"""<div style="background:white;border:2px solid {_cor_num_ind};border-radius:16px;
              padding:28px 16px;text-align:center;margin-bottom:8px;
              box-shadow:0 4px 12px rgba(0,0,0,.08)">
              <div style="font-size:3.2rem;font-weight:800;color:{_cor_num_ind};line-height:1">{_sc_ind:.0f}</div>
              <div style="color:#9CA3AF;font-size:.72rem;margin:4px 0 10px">/ 100</div>
              <div style="color:{_cor_num_ind};font-size:.88rem;font-weight:700">{_lbl_ind}</div>
            </div>""",unsafe_allow_html=True)
            with st.expander("ℹ️ Como é calculado"):
                st.markdown("""**Score de Saúde (0–100)**

4 indicadores com peso igual de **25% cada**:

🔵 **Kanitz** — normalizado entre -4 e 0
🔵 **Liquidez Imediata** — normalizado entre 0% e 5%
🔵 **EBITDA %** — normalizado entre 0% e 5%
🔵 **Ciclo de Caixa** — normalizado entre 0 e 120 dias (invertido)

**Faixas:**
- 80–100 → 🟢 Baixo Risco
- 60–79 → 🟡 Risco Moderado
- 40–59 → 🟠 Alto Risco
- 20–39 → 🔴 Risco Crítico
- 0–19 → ⚫ Insolvência Iminente""")
        with col_graf_i:
            if len(df_i)>1:
                x_sc_i=df_i[cm].astype(str) if cm else pd.Series(range(len(df_i))).astype(str)
                fig_sc_i=go.Figure()
                fig_sc_i.add_trace(go.Scatter(x=x_sc_i,y=df_i["score_risco"],fill="tozeroy",mode="lines+markers",
                  line=dict(color=_cor_num_ind,width=2.5),fillcolor="rgba(33,212,170,.06)",
                  marker=dict(size=5,color=_cor_num_ind,line=dict(color="#0D1117",width=1.5))))
                fig_sc_i.add_hline(y=65,line_dash="dash",line_color="#00D4AA",opacity=0.25)
                fig_sc_i.add_hline(y=45,line_dash="dash",line_color="#FFB627",opacity=0.25)
                fig_sc_i.add_hline(y=30,line_dash="dash",line_color="#F85149",opacity=0.25)
                fig_sc_i.update_layout(plot_bgcolor="white",paper_bgcolor="white",
                  font=dict(color="#6B7280",size=10),margin=dict(l=8,r=8,t=12,b=8),
                  xaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=False),
                  yaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",range=[0,100]),
                  height=200,showlegend=False)
                st.plotly_chart(fig_sc_i,use_container_width=True)

    def _ultimo_valor_real(campo):
        """Pega o último valor não-nulo desse indicador específico (não da última linha do
        banco como um todo) — evita 'nan' quando a demonstração que alimenta esse indicador
        ainda não foi importada para o período mais recente."""
        if campo not in df_i.columns: return 0.0
        s=pd.to_numeric(df_i[campo],errors="coerce").dropna()
        return float(s.iloc[-1]) if not s.empty else 0.0

    sec("🌡️ Termômetro de Kanitz")
    try:
        k_v=_ultimo_valor_real("kanitz")
        if k_v>0: st.markdown(f'<div class="kz-s"><b style="color:#00D4AA">✅ SOLVENTE — {k_v:.2f}</b> &nbsp; Acima de 0: boa saúde financeira.</div>',unsafe_allow_html=True)
        elif k_v>=-3: st.markdown(f'<div class="kz-w"><b style="color:#FFB627">⚠️ PENUMBRA — {k_v:.2f}</b> &nbsp; Entre -3 e 0: atenção necessária.</div>',unsafe_allow_html=True)
        else: st.markdown(f'<div class="kz-d"><b style="color:#F85149">🔴 INSOLVENTE — {k_v:.2f}</b> &nbsp; Abaixo de -3: alto risco.</div>',unsafe_allow_html=True)
    except: pass
    sec("📊 Gauges")
    g1,g2,g3,g4=st.columns(4)
    try: g1.plotly_chart(gauge(_ultimo_valor_real("margem líquida %"),"Margem Líquida %",-20,50,"#2176FF"),use_container_width=True)
    except: pass
    try: g2.plotly_chart(gauge(_ultimo_valor_real("EBITDA %"),"EBITDA %",-10,60,"#00D4AA"),use_container_width=True)
    except: pass
    try:
        lc2=_ultimo_valor_real("liquidez corrente")
        g3.plotly_chart(gauge(min(lc2,3),"Liquidez Corrente",0,3,"#00D4AA" if lc2>=1.5 else ("#FFB627" if lc2>=1 else "#F85149")),use_container_width=True)
    except: pass
    try:
        k_v2=_ultimo_valor_real("kanitz")
        g4.plotly_chart(gauge(max(min(k_v2,15),-7),"Kanitz",-7,15,"#00D4AA" if k_v2>0 else ("#FFB627" if k_v2>=-3 else "#F85149")),use_container_width=True)
    except: pass
    sec("📋 Todos os Indicadores")
    inds=[("Receita Bruta","receita bruta de vendas","brl",False),("Receita Líquida","receita líquida","brl",False),
          ("Lucro Bruto","lucro bruto","brl",False),
          ("Lucratividade","margem líquida %","pct",False),("Margem EBITDA","EBITDA %","pct",False),
          ("Margem Bruta","margem bruta %","pct",False),("Margem Contrib.","margem contrib %","pct",False),
          ("ROE","ROE","pct",False),("ICD","ICD","pct",False),
          ("Liq. Corrente","liquidez corrente","x",False),("Liq. Imediata","liquidez imediata","x",False),
          ("PMR (dias)","PMR","d",True),("PMP (dias)","PMP","d",False),
          ("PME (dias)","PME","d",True),("Ciclo Caixa","ciclo de caixa","d",True),
          ("Giro Estoque","giro estoque","x",False),("Ticket Médio","ticket médio","brl",False),
          ("Lucro Líquido","lucro líquido","brl",False),("EBITDA R$","EBITDA","brl",False)]
    cols_ind=st.columns(4)
    for i,(lbl,campo,t,inv) in enumerate(inds):
        try:
            # Usa o último valor REAL (não-nulo) desse indicador específico, em vez da última
            # linha do banco como um todo — evita "nan%" quando a demonstração que alimenta
            # esse indicador ainda não foi importada para o período mais recente
            if campo in df_i.columns:
                _serie_campo=pd.to_numeric(df_i[campo],errors="coerce").dropna()
                v=float(_serie_campo.iloc[-1]) if not _serie_campo.empty else 0.0
                _var_ind=0.0
                if len(_serie_campo)>=2:
                    _v_ant=float(_serie_campo.iloc[-2])
                    _var_ind=safe(v-_v_ant,abs(_v_ant))*100 if _v_ant!=0 else 0.0
            else:
                v=0.0; _var_ind=0.0
            mx_v=float(df_i[campo].max() or 0) if campo in df_i.columns else 0
            mn_v=float(df_i[campo].min() or 0) if campo in df_i.columns else 0
            _cls_ind=cor(_var_ind,inv)
            cols_ind[i%4].markdown(
                f'<div class="mc"><div class="mc-lbl">{lbl}</div>'
                f'<div class="mc-val {_cls_ind}">{fmt(v,t)}</div>'
                f'<div class="mc-sub">Max: {fmt(mx_v,t)} | Min: {fmt(mn_v,t)}</div></div>',unsafe_allow_html=True)
        except: pass
    if mes_i=="Todos" and len(df_i)>=6:
        sec("🔮 Projeção — Próximo Mês")
        _mds_prox=[m for m,ok in MODELOS_ML.items() if ok]
        _cols_proj_i=st.columns(3)
        for _cp_i,_campo_p_i,_lbl_p_i in zip(_cols_proj_i,["receita bruta de vendas","lucro líquido","EBITDA"],["Receita Bruta","Lucro Líquido","EBITDA"]):
            if _campo_p_i not in df_i.columns: continue
            try:
                _melhor_i,_=melhor_modelo(df_i[_campo_p_i],_mds_prox[:3])
                _proj_i=treinar(df_i[_campo_p_i],_melhor_i,1)
                if _proj_i is not None:
                    _v_at_i=float(df_i[_campo_p_i].iloc[-1]); _v_pr_i=float(_proj_i.iloc[0])
                    _var_i=safe(_v_pr_i-_v_at_i,abs(_v_at_i))*100
                    _mds_testados_i=", ".join([m for m,ok in MODELOS_ML.items() if ok])
                    _cp_i.markdown(f'<div class="mc"><div class="mc-lbl">📮 {_lbl_p_i}</div>'
                                f'<div class="mc-val {cor(_var_i)}">{fmt(_v_pr_i)}</div>'
                                f'<div class="mc-sub">{_var_i:+.1f}% vs último mês<br>'
                                f'<span style="font-size:.65rem;color:#9CA3AF">'
                                f'Escolhido por menor MSE: <b>{_melhor_i}</b><br>'
                                f'Modelos testados: {_mds_testados_i}</span></div></div>',unsafe_allow_html=True)
            except: pass
    sec("📈 Evolução")
    _x_evol_i=df_i[cm].astype(str) if cm else pd.Series(range(len(df_i))).astype(str)
    if ca and cm:
        _x_evol_i=df_i[cm].astype(str)+"/"+df_i[ca].astype(str).str[-2:]
    _ge1,_ge2=st.columns(2)
    def _layout_graf_i(titulo,ytitle):
        return dict(
            title=dict(text=titulo,font=dict(size=13,color="#111827",family="Inter"),x=0),
            plot_bgcolor="white",paper_bgcolor="white",
            font=dict(color="#6B7280",size=10,family="Inter"),
            margin=dict(l=12,r=12,t=44,b=40),
            xaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",tickfont=dict(size=9),
                       tickangle=-35,showgrid=False),
            yaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",
                       tickfont=dict(size=9),
                       title=dict(text=ytitle,font=dict(size=9,color="#9CA3AF")),showgrid=True),
            legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",
                        y=-0.35,font=dict(size=9),itemsizing="constant"),
            hovermode="x unified",
            hoverlabel=dict(bgcolor="white",bordercolor="#E8ECF0",
                            font=dict(color="#111827",size=11)),
            shapes=[dict(type="rect",xref="paper",yref="paper",
                        x0=0,y0=0,x1=1,y1=1,
                        line=dict(color="#E8ECF0",width=1))])
    with _ge1:
        _fig_rf_i=go.Figure()
        for c,cor_c,nm,eixo,tracado in [
          ("receita bruta de vendas","#0EA5E9","Receita Bruta","y","solid"),
          ("receita líquida","#16A34A","Receita Líquida","y","dash"),
          ("lucro bruto","#A9762F","Lucro Bruto","y2","solid"),
          ("lucro líquido","#DC2626","Lucro Líquido","y2","solid")]:
            if c not in df_i.columns: continue
            y=pd.to_numeric(df_i[c],errors="coerce")
            _fig_rf_i.add_trace(go.Scatter(x=_x_evol_i,y=y,name=nm,mode="lines+markers",
              line=dict(color=cor_c,width=2.5,dash=tracado),
              marker=dict(size=6,color=cor_c,line=dict(color="#161B27",width=1.5)),
              yaxis=eixo,
              hovertemplate=f"<b>{nm}</b><br>%{{x}}<br>R$ %{{y:,.0f}}<extra></extra>"))
        _layout_ge1_i=_layout_graf_i("💰 Resultado Financeiro (R$)","Receita (R$)")
        _layout_ge1_i["yaxis"]["title"]["font"]["color"]="#0EA5E9"
        _layout_ge1_i["yaxis2"]=dict(overlaying="y",side="right",showgrid=False,
          tickfont=dict(size=9),title=dict(text="Lucro (R$)",font=dict(size=9,color="#A9762F")))
        _fig_rf_i.update_layout(**_layout_ge1_i)
        st.plotly_chart(_fig_rf_i,use_container_width=True)
    with _ge2:
        _fig_mg_i=go.Figure()
        for c,cor_c,nm in [
          ("margem bruta %","#0EA5E9","Margem Bruta"),
          ("margem contrib %","#16A34A","Margem Contribuição"),
          ("margem líquida %","#A9762F","Margem Líquida"),
          ("EBITDA %","#DC2626","EBITDA")]:
            if c not in df_i.columns: continue
            y=pd.to_numeric(df_i[c],errors="coerce")
            _fig_mg_i.add_trace(go.Scatter(x=_x_evol_i,y=y,name=nm,mode="lines+markers",
              line=dict(color=cor_c,width=2.5),
              marker=dict(size=6,color=cor_c,line=dict(color="#161B27",width=1.5)),
              hovertemplate=f"<b>{nm}</b><br>%{{x}}<br>%{{y:.1f}}%<extra></extra>"))
        _fig_mg_i.update_layout(**_layout_graf_i("📊 Margens (%)","% sobre Receita Líquida"))
        st.plotly_chart(_fig_mg_i,use_container_width=True)

    t3,t4=st.tabs(["Prazos","ROE / ICD"])
    TH_IND=dict(plot_bgcolor="white",paper_bgcolor="white",
        font=dict(color="#6B7280",size=10,family="Inter"),
        xaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=False,tickangle=-35),
        yaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=True),
        margin=dict(l=8,r=8,t=40,b=40),
        legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.3,font=dict(size=9)),
        hovermode="x unified",
        hoverlabel=dict(bgcolor="white",bordercolor="#E8ECF0",font=dict(color="#111827",size=11)))

    def gl_light(df,campos,titulo,cx=None):
        fig=go.Figure()
        cores_l=["#14243B","#A9762F","#5B7B9A","#8B5E34","#345678"]
        x=df[cx] if cx and cx in df.columns else df.index
        if ca and cm and cx==cm:
            x=df[cm].astype(str)+"/"+df[ca].astype(str).str[-2:]
        for i,c in enumerate(campos):
            if c not in df.columns: continue
            fig.add_trace(go.Scatter(x=x,y=pd.to_numeric(df[c],errors="coerce"),
              name=c,mode="lines+markers",
              line=dict(color=cores_l[i%len(cores_l)],width=2.4),
              marker=dict(size=6,color=cores_l[i%len(cores_l)],
                          line=dict(color="white",width=1.5)),
              hovertemplate=f"<b>{c}</b><br>%{{x}}<br>%{{y:.2f}}<extra></extra>"))
        fig.update_layout(
          title=dict(text=titulo,font=dict(size=15,family="Georgia, serif",color="#14243B")),
          plot_bgcolor="white",paper_bgcolor="white",
          font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
          margin=dict(l=10,r=10,t=50,b=60),
          xaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=False,
            tickangle=-40,tickfont=dict(size=9,color="#4B5563")),
          yaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=True,zeroline=False,tickfont=dict(size=9)),
          legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.34,x=0.5,xanchor="center",font=dict(size=9)),
          hovermode="x unified",height=420,
          hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B",size=11)))
        return fig

    with t3: st.plotly_chart(gl_light(df_i,["PMR","PMP","PME","ciclo de caixa"],"Prazos (dias)",cm),use_container_width=True)
    with t4: st.plotly_chart(gl_light(df_i,["ROE","ICD"],"ROE e ICD %",cm),use_container_width=True)

# ── ALERTAS ─────────────────────────────────────────
elif pg=="alertas":
    hdr("🚨 Alertas e Diagnóstico","Análise automática da saúde financeira por período")
    _df_raw_al=get_df_raw_bruto()
    _col_fil_al=col_filial(_df_raw_al) if _df_raw_al is not None else None
    filial_sel_al=None
    if _col_fil_al:
        _filiais_disp_al=sorted(v for v in _df_raw_al[_col_fil_al].dropna().astype(str).unique().tolist() if v!="(Todas as filiais)")
        _opcoes_filial_al=["(Todas as filiais)"]+_filiais_disp_al

        def _on_change_filial_al():
            st.session_state["al_filial_sel_backup"]=st.session_state["al_filial_sel"]

        if "al_filial_sel_backup" not in st.session_state:
            st.session_state["al_filial_sel_backup"]=_opcoes_filial_al[0]
        if st.session_state["al_filial_sel_backup"] not in _opcoes_filial_al:
            st.session_state["al_filial_sel_backup"]=_opcoes_filial_al[0]
        st.session_state["al_filial_sel"]=st.session_state["al_filial_sel_backup"]

        filial_sel_al=st.selectbox("🏬 Filial",_opcoes_filial_al,
          key="al_filial_sel",on_change=_on_change_filial_al)

    df=get_df_filial(filial_sel_al)
    if df is None: no_data(); st.stop()
    cm=cm_(df); ca=ca_(df)

    # Filtros
    col_a1,col_a2=st.columns(2)
    anos_a=["Todos"]+[str(a) for a in sorted(df[ca].dropna().unique().tolist())] if ca else ["Todos"]
    ano_a=col_a1.selectbox("Ano",anos_a,key="al_ano")
    df_a=df[df[ca].astype(str)==ano_a].reset_index(drop=True) if ano_a!="Todos" else df.reset_index(drop=True)
    meses_a=["Todos"]+df_a[cm].dropna().unique().tolist() if cm else ["Todos"]
    mes_a=col_a2.selectbox("Mês",meses_a,key="al_mes")
    df_a=df_a[df_a[cm]==mes_a].reset_index(drop=True) if mes_a!="Todos" else df_a
    if df_a.empty: st.warning("Sem dados para este período."); st.stop()
    ul=df_a.iloc[-1]
    periodo_a=f"{ul.get(cm,'')} {ul.get(ca,'')}".strip()

    # Score
    sc=float(ul.get("score_risco",50)); lbl,cls_s,cor_s=score_label(sc)
    cor_num="#059669" if sc>=80 else ("#D97706" if sc>=60 else ("#EA580C" if sc>=40 else "#DC2626"))

    sec(f"🎯 Score de Saúde — {periodo_a}")
    c1,c2,c3,c4=st.columns(4)
    mc(c1,"Score",f"{sc:.0f}/100","g" if sc>=80 else ("y" if sc>=60 else "r"))

    # Alertas automáticos
    alertas=[]
    def chk(tipo,msg,detalhe=""): alertas.append((tipo,msg,detalhe))

    try:
        ml=float(ul.get("margem líquida %",0))
        if ml<0: chk("d","💸 Prejuízo operacional",f"Margem Líquida: {ml:.1f}% — empresa gastando mais do que fatura")
        elif ml<5: chk("w","⚠️ Margem Líquida muito baixa",f"{ml:.1f}% — mínimo recomendado é 5%")
        else: chk("s","✅ Margem Líquida saudável",f"{ml:.1f}%")
    except: pass
    try:
        mb=float(ul.get("margem bruta %",0))
        if mb<20: chk("w","⚠️ Margem Bruta baixa",f"{mb:.1f}% — revisar CMV e precificação")
        else: chk("s","✅ Margem Bruta adequada",f"{mb:.1f}%")
    except: pass
    try:
        lc=float(ul.get("liquidez corrente",0) or 0)
        if lc<1: chk("d","🔴 Liquidez Corrente crítica",f"{lc:.2f}x — dívidas de curto prazo maiores que ativos circulantes")
        elif lc<1.5: chk("w","⚠️ Liquidez Corrente abaixo do ideal",f"{lc:.2f}x — recomendado ≥ 1,5x")
        else: chk("s","✅ Liquidez Corrente saudável",f"{lc:.2f}x")
    except: pass
    try:
        li=float(ul.get("liquidez imediata",0) or 0)
        if li<0.3: chk("d","🔴 Liquidez Imediata crítica",f"{li:.2f}x — caixa insuficiente para cobrir dívidas imediatas")
        elif li<0.5: chk("w","⚠️ Liquidez Imediata baixa",f"{li:.2f}x — recomendado ≥ 0,5x")
        else: chk("s","✅ Liquidez Imediata saudável",f"{li:.2f}x")
    except: pass
    try:
        kz=float(ul.get("kanitz",0))
        if kz<-3: chk("d","🔴 Kanitz — Zona Insolvente",f"{kz:.2f} — alto risco de insolvência")
        elif kz<0: chk("w","⚠️ Kanitz — Zona de Penumbra",f"{kz:.2f} — situação financeira incerta")
        else: chk("s","✅ Kanitz — Zona Solvente",f"{kz:.2f} — empresa financeiramente saudável")
    except: pass
    try:
        cc=float(ul.get("ciclo de caixa",0) or 0)
        if cc>90: chk("d","🔴 Ciclo de Caixa muito elevado",f"{cc:.0f} dias — capital de giro muito comprometido")
        elif cc>60: chk("w","⚠️ Ciclo de Caixa elevado",f"{cc:.0f} dias — monitorar capital de giro")
        elif cc<0: chk("s","✅ Ciclo de Caixa negativo (favorável)",f"{cc:.0f} dias — empresa recebe antes de pagar")
        else: chk("s","✅ Ciclo de Caixa adequado",f"{cc:.0f} dias")
    except: pass
    try:
        roe=float(ul.get("ROE",0))
        if roe<0: chk("d","💸 ROE negativo",f"{roe:.1f}% — retorno negativo sobre patrimônio")
        elif roe<6: chk("w","⚠️ ROE baixo",f"{roe:.1f}% — abaixo da taxa mínima de atratividade")
        else: chk("s","✅ ROE adequado",f"{roe:.1f}%")
    except: pass

    # Tendência receita
    if "receita bruta de vendas" in df_a.columns and len(df_a)>=3:
        u3=pd.to_numeric(df_a["receita bruta de vendas"],errors="coerce").tail(3)
        if len(u3)==3:
            if u3.iloc[-1]<u3.iloc[-2]<u3.iloc[-3]:
                chk("d","📉 Receita em QUEDA","3 períodos consecutivos de queda — atenção urgente")
            elif u3.iloc[-1]>u3.iloc[-2]>u3.iloc[-3]:
                chk("s","📈 Receita em CRESCIMENTO","3 períodos consecutivos de crescimento")

    pd_=sum(1 for a in alertas if a[0]=="d")
    pw_=sum(1 for a in alertas if a[0]=="w")
    ps_=sum(1 for a in alertas if a[0]=="s")
    mc(c2,"🔴 Críticos",str(pd_),"r")
    mc(c3,"⚠️ Atenção",str(pw_),"y")
    mc(c4,"✅ Saudável",str(ps_),"g")

    # Diagnóstico
    sec("📋 Diagnóstico Detalhado")
    for tipo,msg,detalhe in sorted(alertas,key=lambda x:{"d":0,"w":1,"s":2}[x[0]]):
        st.markdown(f'<div class="al-{tipo}"><b>{msg}</b>'
                    f'{"<br><span style=\'font-size:.8rem;opacity:.85\'>"+detalhe+"</span>" if detalhe else ""}'
                    f'</div>',unsafe_allow_html=True)

    # Anomalias com gráfico
    sec("🔍 Detecção de Anomalias")
    st.markdown('<div class="al-i">Pontos fora do padrão histórico — variação acima de 1,3x o desvio padrão móvel.</div>',unsafe_allow_html=True)
    campos_anom=[
        ("receita bruta de vendas","#2563EB","Receita Bruta"),
        ("CMV (custo da mercadoria vendida)","#DC2626","CMV"),
        ("despesas administrativas","#D97706","Desp. Administrativas"),
        ("lucro líquido","#059669","Lucro Líquido"),
        ("EBITDA","#7C3AED","EBITDA")]
    achou_anom=False
    for campo_a2,cor_a,nm_a in campos_anom:
        if campo_a2 not in df_a.columns: continue
        anom=detectar_anomalias(df_a[campo_a2]); n_anom=int(anom.sum())
        if n_anom>0:
            achou_anom=True
            ma=df_a[cm][anom].tolist() if cm else list(df_a.index[anom])
            st.markdown(f'<div class="al-w">⚠️ <b>{nm_a}</b>: {n_anom} ponto(s) atípico(s) — '
                       f'períodos: {", ".join(str(m) for m in ma[:6])}</div>',unsafe_allow_html=True)
    if not achou_anom:
        st.markdown('<div class="al-s">✅ Nenhuma anomalia detectada nos campos analisados.</div>',unsafe_allow_html=True)

    # Tabela resumo por período
    sec("📊 Resumo de Saúde por Período")
    st.markdown('<div class="al-i">✅ = indicador positivo/saudável | ⚠️ = atenção | 🔴 = crítico</div>',unsafe_allow_html=True)
    if ca and cm:
        res=[]
        for _,row in df_a.iterrows():
            def status(campo,inv=False,limites=None):
                try:
                    v=float(row.get(campo,0) or 0)
                    if limites:
                        if v>=limites[1]: return "✅"
                        if v>=limites[0]: return "⚠️"
                        return "🔴"
                    if inv: return "⚠️" if v>60 else "✅"
                    return "✅" if v>0 else "🔴"
                except: return "—"
            sc_r=float(row.get("score_risco",0))
            sc_lbl="🟢" if sc_r>=80 else ("🟡" if sc_r>=60 else ("🟠" if sc_r>=40 else "🔴"))
            res.append({
                "Ano":row.get(ca,""),"Mês":row.get(cm,""),
                "Score":f"{sc_lbl} {sc_r:.0f}",
                "Receita":status("receita bruta de vendas"),
                "Lucro":status("lucro líquido"),
                "Mg Líquida":status("margem líquida %",limites=(0,5)),
                "Liquidez":status("liquidez corrente",limites=(1,1.5)),
                "Kanitz":status("kanitz",limites=(0,0)),
                "Ciclo Cx":status("ciclo de caixa",inv=True)})
        st.dataframe(pd.DataFrame(res),use_container_width=True,hide_index=True)

# ── ML ───────────────────────────────────────────────
elif pg=="ml":
    hdr("🔮 Projeções ML","Machine Learning — o sistema testa todos os modelos e escolhe o mais preciso")
    if not STATS_OK:
        st.markdown('<div class="al-d">❌ pip install statsmodels scikit-learn</div>',unsafe_allow_html=True)
    _df_raw_mlp2=get_df_raw_bruto()
    _col_fil_mlp2=col_filial(_df_raw_mlp2) if _df_raw_mlp2 is not None else None
    filial_sel_mlp2=None
    if _col_fil_mlp2:
        _filiais_disp_mlp2=sorted(v for v in _df_raw_mlp2[_col_fil_mlp2].dropna().astype(str).unique().tolist() if v!="(Todas as filiais)")
        _opcoes_filial_mlp2=["(Todas as filiais)"]+_filiais_disp_mlp2

        def _on_change_filial_mlp2():
            st.session_state["mlp2_filial_sel_backup"]=st.session_state["mlp2_filial_sel"]

        if "mlp2_filial_sel_backup" not in st.session_state:
            st.session_state["mlp2_filial_sel_backup"]=_opcoes_filial_mlp2[0]
        if st.session_state["mlp2_filial_sel_backup"] not in _opcoes_filial_mlp2:
            st.session_state["mlp2_filial_sel_backup"]=_opcoes_filial_mlp2[0]
        st.session_state["mlp2_filial_sel"]=st.session_state["mlp2_filial_sel_backup"]

        filial_sel_mlp2=st.selectbox("🏬 Filial",_opcoes_filial_mlp2,
          key="mlp2_filial_sel",on_change=_on_change_filial_mlp2)

    df=get_df_filial(filial_sel_mlp2)
    if df is None: no_data(); st.stop()
    cm=cm_(df); ca=ca_(df)

    if st.session_state.get("projecoes_filial_atual")!=filial_sel_mlp2:
        st.session_state.projecoes={}
        st.session_state["projecoes_filial_atual"]=filial_sel_mlp2
        if st.session_state.cid:
            _proj_disco_ml=load_projecoes_ml(st.session_state.cid,filial_sel_mlp2)
            if _proj_disco_ml:
                st.session_state.projecoes=_proj_disco_ml

    

    # Filtro período + meses a projetar
    sec("⚙️ Configurações")
    col1,col2,col3=st.columns(3)
    anos_ml=["Todos"]+[str(a) for a in sorted(df[ca].dropna().unique().tolist())] if ca else ["Todos"]
    ano_ml=col1.selectbox("Ano base",anos_ml,key="ml_ano")
    df_ml=df[df[ca].astype(str)==ano_ml].reset_index(drop=True) if ano_ml!="Todos" else df.reset_index(drop=True)
    n_m=col2.slider("Meses a projetar",1,24,6)
    var_pct=col3.slider("Variação cenários (%)",5,30,15,step=5)

    if len(df_ml)<6:
        st.markdown('<div class="al-w">⚠️ Mínimo 6 períodos de dados para projetar.</div>',unsafe_allow_html=True)
        st.stop()

    mds_disp=[m for m,ok in MODELOS_ML.items() if ok]

    # Campos fixos principais
    campos_fixos=[
        # DRE
        ("receita bruta de vendas","Receita Bruta","#2563EB"),
        ("receita líquida","Receita Líquida","#2563EB"),
        ("lucro bruto","Lucro Bruto","#059669"),
        ("lucro líquido","Lucro Líquido","#059669"),
        ("EBITDA","EBITDA","#D97706"),
        ("margem bruta %","Margem Bruta %","#7C3AED"),
        ("margem contrib %","Margem Contribuição %","#7C3AED"),
        ("margem líquida %","Margem Líquida %","#DC2626"),
        ("EBITDA %","EBITDA %","#D97706"),
        ("despesas comerciais","Desp. Comerciais","#DC2626"),
        ("despesas administrativas","Desp. Administrativas","#DC2626"),
        # Balanço
        ("disponibilidades saldo","Disponibilidades","#0891B2"),
        ("contas a receber saldo","Contas a Receber","#0891B2"),
        ("estoque final do mês de mercadorias para revenda saldo","Estoques","#D97706"),
        ("ativo total","Ativo Total","#2563EB"),
        ("pass total","Passivo Total","#DC2626"),
        ("PL","Patrimônio Líquido","#059669"),
        # Indicadores
        ("liquidez corrente","Liquidez Corrente","#0891B2"),
        ("liquidez imediata","Liquidez Imediata","#0891B2"),
        ("kanitz","Kanitz","#7C3AED"),
        ("ROE","ROE %","#059669"),
        ("PMR","PMR (dias)","#D97706"),
        ("PMP","PMP (dias)","#059669"),
        ("ciclo de caixa","Ciclo de Caixa","#DC2626"),
        # Fluxo
        ("Disponibilidades entradas","Entradas Caixa","#059669"),
        ("Disponibilidades Saida","Saídas Caixa","#DC2626"),
    ]
    campos_proj=[(c,l,cor_c) for c,l,cor_c in campos_fixos if c in df_ml.columns]

    senha_gerar_projecoes=st.text_input("Senha master para gerar",type="password",key="senha_gerar_projecoes")
    if st.button("🚀 Gerar Projeções",use_container_width=True):
        if senha_gerar_projecoes!=SENHA_MASTER:
            st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi executado.</div>',unsafe_allow_html=True)
            st.stop()
        projs={}; pb=st.progress(0)
        with st.spinner("🤖 Testando modelos e gerando projeções..."):
            for i,(campo,lbl,cor_c) in enumerate(campos_proj):
                try:
                    melhor,rank=melhor_modelo(df_ml[campo],mds_disp)
                    proj=treinar(df_ml[campo],melhor,n_m)
                    if proj is not None:
                        projs[campo]={"modelo":melhor,"valores":proj.tolist(),"rank":rank,"lbl":lbl,"cor":cor_c}
                except: pass
                pb.progress((i+1)/len(campos_proj))
        pb.empty()
        st.session_state.projecoes=projs
        st.session_state["projecoes_filial_atual"]=filial_sel_mlp2
        if st.session_state.cid:
            save_projecoes_ml(st.session_state.cid,projs,filial_sel_mlp2)
        st.markdown(f'<div class="al-s">✅ {len(projs)} campos projetados para {n_m} meses.</div>',unsafe_allow_html=True)
        addlog(f"ML: {len(projs)} campos × {n_m} meses")

    if st.session_state.projecoes:
        projs=st.session_state.projecoes

        TH_ML=dict(plot_bgcolor="white",paper_bgcolor="white",
            font=dict(color="#6B7280",size=10,family="Inter"),
            xaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=False,tickangle=-35),
            yaxis=dict(gridcolor="#F3F4F6",linecolor="#E8ECF0",showgrid=True),
            margin=dict(l=8,r=8,t=44,b=40),
            legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.3,font=dict(size=9)),
            hovermode="x unified",
            hoverlabel=dict(bgcolor="white",bordercolor="#E8ECF0",font=dict(color="#111827",size=11)))

        x_h=df_ml[cm].astype(str) if cm else pd.Series(range(len(df_ml))).astype(str)
        if ca and cm:
            x_h=df_ml[cm].astype(str)+"/"+df_ml[ca].astype(str).str[-2:]
        _meses_abv3_ml={1:"jan",2:"fev",3:"mar",4:"abr",5:"mai",6:"jun",7:"jul",8:"ago",9:"set",10:"out",11:"nov",12:"dez"}
        _meses_inv_ml={v:k for k,v in _meses_abv3_ml.items()}
        if ca and cm and cm in df_ml.columns and ca in df_ml.columns and len(df_ml)>0:
            _ult_mes_ml=str(df_ml[cm].iloc[-1]).strip().lower()[:3]
            _ult_ano_ml=int(df_ml[ca].iloc[-1])
            _mn_ml=_meses_inv_ml.get(_ult_mes_ml)
            if _mn_ml:
                x_p=[]
                for _i_ml in range(n_m):
                    _mn_ml+=1
                    if _mn_ml>12: _mn_ml=1; _ult_ano_ml+=1
                    x_p.append(f"{_meses_abv3_ml[_mn_ml]}/{str(_ult_ano_ml)[-2:]}")
            else:
                x_p=[f"M+{i+1}" for i in range(n_m)]
        else:
            x_p=[f"M+{i+1}" for i in range(n_m)]

        # Gráficos 2 por linha
        campos_list=list(projs.items())
        _max_graf_ml=len(campos_list)
        _n_graf_ml=st.slider("Quantidade de gráficos exibidos",1,_max_graf_ml,min(10,_max_graf_ml),key="ml_qtd_graficos")
        campos_list=campos_list[:_n_graf_ml]
        for idx in range(0,len(campos_list),2):
            cols=st.columns(2)
            for j,col in enumerate(cols):
                if idx+j>=len(campos_list): break
                campo,(p)=campos_list[idx+j]
                lbl=p.get("lbl",campo)
                cor_c=p.get("cor","#2563EB")
                v_hist=pd.to_numeric(df_ml[campo],errors="coerce")
                v_proj=p["valores"]
                # Sempre pega o último mês real do dataset completo — não do filtrado
                v_at=float(pd.to_numeric(df[campo],errors="coerce").dropna().iloc[-1]) if campo in df.columns else 0
                v_pr=v_proj[-1] if v_proj else 0
                var=safe(v_pr-v_at,abs(v_at))*100

                # Ranking modelos
                rank_s=sorted(p.get("rank",{}).items(),key=lambda x:x[1])
                melhor_nm=p["modelo"]

                fig=go.Figure()
                # Histórico
                fig.add_trace(go.Scatter(x=x_h,y=v_hist,name="Histórico",
                  mode="lines+markers",line=dict(color=cor_c,width=2.2),
                  marker=dict(size=4,color=cor_c,line=dict(color="white",width=1.5)),
                  hovertemplate=f"<b>Histórico</b><br>%{{x}}<br>%{{y:,.1f}}<extra></extra>"))
                # Projeção base
                fig.add_trace(go.Scatter(x=x_p,y=v_proj,name=f"Projeção ({melhor_nm})",
                  mode="lines+markers",line=dict(color=cor_c,width=2,dash="dash"),
                  marker=dict(size=6,color=cor_c,symbol="diamond",line=dict(color="white",width=1.5)),
                  hovertemplate=f"<b>Projeção</b><br>%{{x}}<br>%{{y:,.1f}}<extra></extra>"))
                # Banda incerteza
                y_up=[v*(1+var_pct/100) for v in v_proj]
                y_dn=[v*(1-var_pct/100) for v in v_proj]
                fig.add_trace(go.Scatter(x=x_p+x_p[::-1],y=y_up+y_dn[::-1],
                  fill="toself",fillcolor="rgba(107,114,128,.08)",
                  line=dict(color="rgba(0,0,0,0)"),name=f"±{var_pct}%",showlegend=True))

                fig.update_layout(
                  title=dict(text=f"{lbl} — {var:+.1f}% em {n_m} meses ({melhor_nm})",
                    font=dict(size=11,color="#111827")),**TH_ML)
                col.plotly_chart(fig,use_container_width=True)

        # Cenários e ranking
        sec("🎯 Cenários e Ranking de Modelos")
        for campo,(p) in projs.items():
            lbl=p.get("lbl",campo)
            v_hist=pd.to_numeric(df_ml[campo],errors="coerce").dropna()
            v_at=float(df[campo].dropna().iloc[-1]) if campo in df.columns and len(df[campo].dropna())>0 else 0
            v_pr=p["valores"][-1] if p["valores"] else 0
            var=safe(v_pr-v_at,abs(v_at))*100
            rank_s=sorted(p.get("rank",{}).items(),key=lambda x:x[1])

            def fmt_titulo(campo,v):
                pct_c=["margem bruta %","margem contrib %","margem líquida %",
                        "EBITDA %","liquidez corrente","liquidez imediata","ROE"]
                dias_c=["PMR","PMP","ciclo de caixa"]
                if campo in pct_c: return fmt(v,"pct")
                if campo in dias_c: return fmt(v,"d")
                if campo=="kanitz": return f"{v:.2f}"
                return fmt(v)
            with st.expander(f"{lbl} — Melhor modelo: {p['modelo']} | Projeção: {fmt_titulo(campo,v_pr)} ({var:+.1f}%)"):
                # Cenários
                c1,c2,c3=st.columns(3)
                def fmt_c(v):
                    pct_c=["margem bruta %","margem contrib %","margem líquida %",
                            "EBITDA %","liquidez corrente","liquidez imediata","ROE"]
                    dias_c=["PMR","PMP","ciclo de caixa"]
                    if campo in pct_c: return fmt(v,"pct")
                    if campo in dias_c: return fmt(v,"d")
                    if campo=="kanitz": return f"{v:.2f}"
                    return fmt(v)

                mc(c1,f"🐻 Pessimista (-{var_pct}%)",fmt_c(v_pr*(1-var_pct/100)),"r",f"vs atual: {fmt_c(v_at)}")
                mc(c2,"📊 Base (projeção)",fmt_c(v_pr),"y",f"{var:+.1f}% vs atual")
                mc(c3,f"🐂 Otimista (+{var_pct}%)",fmt_c(v_pr*(1+var_pct/100)),"g",f"vs atual: {fmt_c(v_at)}")
                # Ranking modelos
                v_hist=pd.to_numeric(df_ml[campo],errors="coerce").dropna()
                media_serie=float(v_hist.mean()) if len(v_hist)>0 else 1
                # Formata média histórica corretamente por tipo de campo
                pct_campos=["margem bruta %","margem contrib %","margem líquida %",
                            "EBITDA %","liquidez corrente","liquidez imediata","ROE"]
                dias_campos=["PMR","PMP","ciclo de caixa"]
                if campo in pct_campos:
                    media_fmt=fmt(media_serie,"pct")
                    unidade="pontos percentuais"
                elif campo in dias_campos:
                    media_fmt=fmt(media_serie,"d")
                    unidade="dias"
                elif campo=="kanitz":
                    media_fmt=f"{media_serie:.2f}"
                    unidade="pontos no termômetro de Kanitz"
                else:
                    media_fmt=fmt(media_serie)
                    unidade="reais"

                st.markdown(f"""**Como chegamos a essa projeção:**

O sistema usou os últimos **{len(v_hist)}** períodos históricos para treinar cada modelo.
Depois testou a precisão de cada um nos últimos **6 meses reais** — comparando o que o modelo teria previsto com o que realmente aconteceu.
O vencedor foi **{p['modelo']}** com menor erro relativo.
O erro % abaixo mostra o desvio médio da previsão em relação à média histórica de **{media_fmt}** ({unidade}).
Quanto menor o %, mais preciso o modelo foi nos dados reais.

**Ranking de modelos — erro relativo (% sobre a média histórica):**""")
                cols_r=st.columns(max(min(len(rank_s),5),1))
                for i,(mod,mse) in enumerate(rank_s[:5]):
                    ico=["🥇","🥈","🥉","4️⃣","5️⃣"][i]
                    desc={"ARIMA":"Captura tendências e autocorrelações",
                          "ExponentialSmoothing":"Pesa mais os dados recentes",
                          "SARIMAX":"Captura sazonalidade anual",
                          "Holt":"Tendência com amortecimento",
                          "Prophet":"IA do Meta para séries temporais"}.get(mod,"")
                    # Converte MSE para RMSE% (erro relativo à média da série)
                    media_serie=float(v_hist.mean()) if len(v_hist)>0 else 1
                    rmse=float(mse**0.5) if mse<float("inf") else 0
                    erro_pct=safe(rmse,abs(media_serie))*100
                    cls_err="g" if erro_pct<5 else ("y" if erro_pct<15 else "r")
                    cols_r[i].markdown(f'<div class="mc"><div class="mc-lbl">{ico} {mod}</div>'
                      f'<div class="mc-val {cls_err}" style="font-size:1rem">{erro_pct:.1f}% erro</div>'
                      f'<div class="mc-sub">{desc}</div></div>',unsafe_allow_html=True)

        # Tabela resumo
        sec("📋 Resumo de Todas as Projeções")
        def fmt_campo(campo_f, v):
            pct_campos=["margem bruta %","margem contrib %","margem líquida %",
                        "EBITDA %","liquidez corrente","liquidez imediata","ROE"]
            dias_campos=["PMR","PMP","ciclo de caixa"]
            if campo_f in pct_campos: return fmt(v,"pct")
            if campo_f in dias_campos: return fmt(v,"d")
            if campo_f=="kanitz": return f"{v:.2f}"
            return fmt(v)
        res=[]
        for campo_r,p in projs.items():
            lbl=p.get("lbl",campo_r)
            v_a=float(df[campo_r].dropna().iloc[-1]) if campo_r in df.columns and len(df[campo_r].dropna())>0 else 0
            v_p2=p["valores"][-1] if p["valores"] else 0
            var2=safe(v_p2-v_a,abs(v_a))*100
            res.append({
                "Campo":lbl,"Modelo":p["modelo"],
                "Atual":fmt_campo(campo_r,v_a),
                "Projeção":fmt_campo(campo_r,v_p2),
                f"Pessimista (-{var_pct}%)":fmt_campo(campo_r,v_p2*(1-var_pct/100)),
                f"Otimista (+{var_pct}%)":fmt_campo(campo_r,v_p2*(1+var_pct/100)),
                "Var%":f"{var2:+.1f}%"})
        if res:
            with st.expander(f"📋 Ver tabela completa ({len(res)} campos)"):
                st.dataframe(pd.DataFrame(res),use_container_width=True,hide_index=True)

# ── CENÁRIOS FP&A ───────────────────────────────────
elif pg=="cenarios":
    hdr("📊 Cenários FP&A","DRE, Balanço e Fluxo de Caixa projetados com ML")
    _df_raw_cen=get_df_raw_bruto()
    _col_fil_cen=col_filial(_df_raw_cen) if _df_raw_cen is not None else None
    filial_sel_cen=None
    if _col_fil_cen:
        _filiais_disp_cen=sorted(v for v in _df_raw_cen[_col_fil_cen].dropna().astype(str).unique().tolist() if v!="(Todas as filiais)")
        _opcoes_filial_cen=["(Todas as filiais)"]+_filiais_disp_cen

        def _on_change_filial_cen():
            st.session_state["cen_filial_sel_backup"]=st.session_state["cen_filial_sel"]

        if "cen_filial_sel_backup" not in st.session_state:
            st.session_state["cen_filial_sel_backup"]=_opcoes_filial_cen[0]
        if st.session_state["cen_filial_sel_backup"] not in _opcoes_filial_cen:
            st.session_state["cen_filial_sel_backup"]=_opcoes_filial_cen[0]
        st.session_state["cen_filial_sel"]=st.session_state["cen_filial_sel_backup"]

        filial_sel_cen=st.selectbox("🏬 Filial",_opcoes_filial_cen,
          key="cen_filial_sel",on_change=_on_change_filial_cen)

    df=get_df_filial(filial_sel_cen)
    if df is None: no_data(); st.stop()

    # Garante que a projeção em memória é da MESMA filial selecionada aqui — recarrega do disco se não bater
    if st.session_state.get("projecoes_filial_atual")!=filial_sel_cen:
        st.session_state.projecoes={}
        st.session_state["projecoes_filial_atual"]=filial_sel_cen
        if st.session_state.cid:
            _proj_disco_cen=load_projecoes_ml(st.session_state.cid,filial_sel_cen)
            if _proj_disco_cen:
                st.session_state.projecoes=_proj_disco_cen

    if not st.session_state.projecoes:
        st.markdown(f'<div class="al-w">⚠️ Gere as projeções primeiro em <b>🔮 Projeções ML</b>, para a filial <b>{filial_sel_cen or "(Todas as filiais)"}</b>.</div>',unsafe_allow_html=True)
        st.stop()

    projs=st.session_state.projecoes
    n_m=len(list(projs.values())[0]["valores"]) if projs else 6
    cm=cm_(df); ca=ca_(df)

    sec("⚙️ Configurações")
    c1,c2=st.columns(2)
    cenario=c1.radio("Cenário",["Base","Pessimista","Otimista"],horizontal=True)
    demo=c2.radio("Demonstração",["DRE","Balanço","Fluxo de Caixa"],horizontal=True)
    var_pct=st.session_state.get("var_pct_ml",15)

    # Fator do cenário
    fator=1.0
    if "Pessimista" in cenario: fator=1-var_pct/100
    elif "Otimista" in cenario: fator=1+var_pct/100

    # Último valor real de cada campo
    ul_real=df.iloc[-1]

    # Session state para edições
    if "fp_edicoes" not in st.session_state: st.session_state.fp_edicoes={}
    chave_cen=f"{cenario}_{demo}"
    if chave_cen not in st.session_state.fp_edicoes:
        st.session_state.fp_edicoes[chave_cen]={}

    def get_proj(campo):
        base=[]
        if campo in projs:
            base=[v*fator for v in projs[campo]["valores"]]
        else:
            v=float(ul_real.get(campo,0) or 0)
            base=[v]*n_m
        edicoes=st.session_state.fp_edicoes[chave_cen].get(campo,{})
        for i,v_edit in edicoes.items():
            if 0<=int(i)<len(base): base[int(i)]=v_edit
        return base

    def painel_edicao(campos_editaveis):
        st.divider()
        with st.expander("✏️ Ajustar Premissas — edite valores e recalcule"):
            tem_edicao=any(st.session_state.fp_edicoes[chave_cen].get(c) for c,_ in campos_editaveis)
            if tem_edicao:
                c_rst=st.columns([3,1])
                c_rst[0].markdown('<div class="al-w">✏️ Há valores editados — demonstração recalculada.</div>',unsafe_allow_html=True)
                if c_rst[1].button("🔄 Resetar tudo",key=f"rst_all_{chave_cen}",use_container_width=True):
                    st.session_state.fp_edicoes[chave_cen]={}; st.rerun()
            for campo,lbl in campos_editaveis:
                if campo in projs:
                    base_orig=[v*fator for v in projs[campo]["valores"]]
                else:
                    v0=float(ul_real.get(campo,0) or 0)
                    base_orig=[v0]*n_m
                edicoes=st.session_state.fp_edicoes[chave_cen].get(campo,{})
                st.markdown(f"**{lbl}** — Base ML: {fmt(base_orig[-1])}")
                modo=st.radio("Modo",["% variação","R$ valor"],horizontal=True,
                  key=f"modo_{campo}_{chave_cen}",label_visibility="collapsed")
                aplicar=st.radio("Aplicar a",["Todos os meses","Mês a mês"],horizontal=True,
                  key=f"aplic_{campo}_{chave_cen}",label_visibility="collapsed")
                novos={}; mudou=False
                if aplicar=="Todos os meses":
                    if modo=="% variação":
                        pct_g=st.slider(f"Variação %",min_value=-50,max_value=100,value=0,step=1,
                          key=f"pct_g_{campo}_{chave_cen}")
                        if pct_g!=0:
                            for i in range(n_m):
                                novos[str(i)]=round(base_orig[i]*(1+pct_g/100),2)
                            mudou=novos!=edicoes
                        elif edicoes:
                            mudou=True
                    else:
                        v_ref=round(float(base_orig[0]),0)
                        v_g=st.number_input(f"Novo valor R$ (todos os meses)",
                          value=v_ref,step=max(abs(v_ref)*0.05,1000.),format="%.0f",
                          key=f"val_g_{campo}_{chave_cen}")
                        if v_g!=v_ref:
                            for i in range(n_m): novos[str(i)]=float(v_g)
                            mudou=novos!=edicoes
                        elif edicoes: mudou=True
                else:
                    n_cols=min(n_m,6)
                    for bloco in range(0,n_m,n_cols):
                        cols_m=st.columns(n_cols)
                        for j in range(n_cols):
                            i=bloco+j
                            if i>=n_m: break
                            m=meses_proj[i]
                            v_base=round(float(base_orig[i]),0)
                            v_edit=float(edicoes.get(str(i),v_base))
                            if modo=="% variação":
                                pct_atual=round((v_edit/v_base-1)*100 if v_base!=0 else 0,1)
                                pct_m=cols_m[j].number_input(m,value=pct_atual,
                                  step=1.,format="%.1f",
                                  key=f"pct_m_{campo}_{i}_{chave_cen}")
                                novo_v=round(v_base*(1+pct_m/100),2)
                                if abs(novo_v-v_base)>0.01: novos[str(i)]=novo_v
                            else:
                                novo_v=cols_m[j].number_input(m,value=v_edit,
                                  step=max(abs(v_base)*0.05,1000.),format="%.0f",
                                  key=f"val_m_{campo}_{i}_{chave_cen}")
                                if abs(novo_v-v_base)>0.01: novos[str(i)]=round(float(novo_v),2)
                    mudou=novos!=edicoes
                if mudou:
                    if novos: st.session_state.fp_edicoes[chave_cen][campo]=novos
                    else: st.session_state.fp_edicoes[chave_cen].pop(campo,None)
                    st.rerun()
                if edicoes:
                    if st.button(f"↩️ Resetar {lbl}",key=f"rst_{campo}_{chave_cen}"):
                        st.session_state.fp_edicoes[chave_cen].pop(campo,None); st.rerun()
                st.divider()

    # Colunas dos meses — calcula meses reais após o último período
    MESES_NOMES=["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]
    ul_real=df.iloc[-1]
    try:
        ult_mes=str(ul_real.get(cm,"jan")).lower()[:3]
        ult_ano=int(ul_real.get(ca,2024))
        idx_mes=MESES_NOMES.index(ult_mes) if ult_mes in MESES_NOMES else 11
        meses_proj=[]
        for i in range(n_m):
            idx=(idx_mes+1+i)%12
            ano=ult_ano+((idx_mes+1+i)//12)
            meses_proj.append(f"{MESES_NOMES[idx]}/{ano}")
    except:
        meses_proj=[f"M+{i+1}" for i in range(n_m)]

    col_info,col_rst=st.columns([4,1])
    col_info.markdown(f'<div class="al-i">📌 Cenário: <b>{cenario}</b> | {n_m} meses projetados | Fator: {"×"+str(round(fator,2)) if fator!=1 else "Base"}</div>',unsafe_allow_html=True)
    if col_rst.button("🔄 Resetar tudo",key="rst_all",use_container_width=True):
        st.session_state.fp_edicoes[chave_cen]={}
        st.rerun()
    # Aviso se há edições ativas
    if st.session_state.fp_edicoes.get(chave_cen):
        campos_edit=list(st.session_state.fp_edicoes[chave_cen].keys())
        st.markdown(f'<div class="al-w">✏️ Valores editados manualmente em: <b>{", ".join(campos_edit)}</b> — clique em Resetar para voltar ao ML.</div>',unsafe_allow_html=True)

    # ── DRE PROJETADA
    if "DRE" in demo:
        sec("📋 DRE Projetada")

        rb  = get_proj("receita bruta de vendas")
        imp = get_proj("impostos sobre vendas")
        dev = get_proj("devoluções de vendas")
        imp = get_proj("impostos sobre vendas")
        dev = get_proj("devoluções de vendas")
        rl  = [rb[i]-imp[i]-dev[i] for i in range(n_m)]
        cmv = get_proj("CMV (custo da mercadoria vendida)")
        lb  = [rl[i]-cmv[i] for i in range(n_m)]
        dc  = get_proj("despesas comerciais")
        mc_ = [lb[i]-dc[i] for i in range(n_m)]
        da  = get_proj("despesas administrativas")
        df_ = get_proj("despesas financeiras líquidas")
        dep = get_proj("despesas com depreciações e amortizações")
        lo  = [lb[i]-dc[i]-da[i]-df_[i]-dep[i] for i in range(n_m)]
        rno = get_proj("receitas não operacionais")
        dno = get_proj("despesas não operacionais")
        ir  = get_proj("provisão para imposto de renda")
        cs  = get_proj("provisão para contribuição social")
        ll  = [lo[i]+rno[i]-dno[i]-ir[i]-cs[i] for i in range(n_m)]
        ebt = [ll[i]+ir[i]+cs[i]+df_[i]+dep[i] for i in range(n_m)]

        # Margens
        def pct(num,den,i): return safe(num[i],den[i])*100 if den[i]!=0 else 0
        def ah(serie,i): return safe(serie[i]-serie[i-1],abs(serie[i-1]))*100 if i>0 else 0

        linhas_dre=[
            ("cat","(+) RECEITA BRUTA",rb,False),
            ("sub","  (-) Impostos",imp,True),
            ("sub","  (-) Devoluções",dev,True),
            ("tot","= RECEITA LÍQUIDA",rl,False),
            ("sub","  (-) CMV",cmv,True),
            ("tot","= LUCRO BRUTO",lb,False),
            ("pct_rb","  Margem Bruta %",lb,False),
            ("sub","  (-) Desp. Comerciais",dc,True),
            ("tot","= MARGEM CONTRIB.",mc_,False),
            ("pct_rb","  Margem Contrib. %",mc_,False),
            ("sub","  (-) Desp. Adm.",da,True),
            ("sub","  (-) Desp. Fin.",df_,True),
            ("sub","  (-) Depreciação",dep,True),
            ("tot","= LUCRO OPERACIONAL",lo,False),
            ("pct_rb","  Margem Op. %",lo,False),
            ("sub","  (+/-) Não Operac.",rno,False),
            ("sub","  (-) IR/CSLL",[ir[i]+cs[i] for i in range(n_m)],True),
            ("tot","= LUCRO LÍQUIDO",ll,False),
            ("pct_rb","  Margem Líquida %",ll,False),
            ("tot","  EBITDA",ebt,False),
            ("pct_rb","  EBITDA %",ebt,False),
        ]

        header_d="<tr><th>Descrição</th>"+"".join(f"<th>{m}</th><th>AV%</th><th>AH%</th>" for m in meses_proj)+"</tr>"
        rows_d=""
        for tipo,desc,serie,inv in linhas_dre:
            cls_tr={"cat":"cat","tot":"tot","pct_rb":"pct","sub":"sub"}.get(tipo,"")
            row=f'<tr class="{cls_tr}"><td>{desc}</td>'
            for i in range(n_m):
                if tipo=="pct_rb":
                    v=pct(serie,rb,i)
                    dlt=v-pct(serie,rb,i-1) if i>0 else 0
                    row+=f'<td class="pct">{fmt(v,"pct")}</td>'
                    row+=f'<td class="{"pos" if dlt>0 else "neg" if dlt<0 else "neu"}">{dlt:+.1f}pp</td>'
                    row+=f'<td></td>'
                else:
                    v=float(serie[i])
                    a_v=safe(v,rb[i])*100
                    a_h=ah(serie,i)
                    cls_v="neg" if (inv and v!=0) else cor(v)
                    row+=f'<td class="{cls_v}">{fmt(v)}</td>'
                    row+=f'<td class="{cls_pct(a_v,inv)}">{fmt(a_v,"pct")}</td>'
                    row+=f'<td class="{cls_pct(a_h,inv)}">{fmt(a_h,"pct")}</td>'
            rows_d+=row+"</tr>"
        st.markdown(f'<div class="dre-wrap" style="overflow-x:auto"><table class="dre" style="min-width:900px">{header_d}{rows_d}</table></div>',unsafe_allow_html=True)
        def gerar_excel_dre_3_cenarios():
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb=Workbook()
            wb.remove(wb.active)

            cor_azul=PatternFill("solid",start_color="2563EB",end_color="2563EB")
            cor_cinza_clara=PatternFill("solid",start_color="F3F4F6",end_color="F3F4F6")
            cor_branco_bold=Font(bold=True,color="FFFFFF",size=11)
            cor_preto=Font(color="111827",size=10)
            cor_verde=Font(color="059669",size=10)
            cor_vermelho=Font(color="DC2626",size=10)
            borda_fina=Border(bottom=Side(style="thin",color="E8ECF0"))

            campos_dre_proj=["receita bruta de vendas","impostos sobre vendas","devoluções de vendas",
                "CMV (custo da mercadoria vendida)","despesas comerciais","despesas administrativas",
                "despesas financeiras líquidas","despesas com depreciações e amortizações",
                "receitas não operacionais","despesas não operacionais",
                "provisão para imposto de renda","provisão para contribuição social"]

            for nome_cenario,fator_c in [("Pessimista",1-var_pct/100),("Base",1.0),("Otimista",1+var_pct/100)]:
                vals={}
                for campo in campos_dre_proj:
                    if campo in projs:
                        vals[campo]=[v*fator_c for v in projs[campo]["valores"]]
                    else:
                        v=float(ul_real.get(campo,0) or 0)
                        vals[campo]=[v]*n_m

                rb_c=vals["receita bruta de vendas"]; imp_c=vals["impostos sobre vendas"]; dev_c=vals["devoluções de vendas"]
                rl_c=[rb_c[i]-imp_c[i]-dev_c[i] for i in range(n_m)]
                cmv_c=vals["CMV (custo da mercadoria vendida)"]
                lb_c=[rl_c[i]-cmv_c[i] for i in range(n_m)]
                dc_c=vals["despesas comerciais"]
                mc_c=[lb_c[i]-dc_c[i] for i in range(n_m)]
                da_c=vals["despesas administrativas"]; dff_c=vals["despesas financeiras líquidas"]; dep_c=vals["despesas com depreciações e amortizações"]
                lo_c=[mc_c[i]-da_c[i]-dff_c[i]-dep_c[i] for i in range(n_m)]
                rno_c=vals["receitas não operacionais"]; dno_c=vals["despesas não operacionais"]
                ir_c=vals["provisão para imposto de renda"]; cs_c=vals["provisão para contribuição social"]
                ll_c=[lo_c[i]+rno_c[i]-dno_c[i]-ir_c[i]-cs_c[i] for i in range(n_m)]
                ebt_c=[ll_c[i]+ir_c[i]+cs_c[i]+dff_c[i]+dep_c[i] for i in range(n_m)]

                linhas_c=[
                    ("cat","(+) RECEITA BRUTA",rb_c,False),("sub","  (-) Impostos",imp_c,True),
                    ("sub","  (-) Devoluções",dev_c,True),("tot","RECEITA LÍQUIDA",rl_c,False),
                    ("sub","  (-) CMV",cmv_c,True),("tot","LUCRO BRUTO",lb_c,False),
                    ("sub","  (-) Desp. Comerciais",dc_c,True),("tot","MARGEM CONTRIB.",mc_c,False),
                    ("sub","  (-) Desp. Adm.",da_c,True),("sub","  (-) Desp. Fin.",dff_c,True),
                    ("sub","  (-) Depreciação",dep_c,True),("tot","LUCRO OPERACIONAL",lo_c,False),
                    ("sub","  (+/-) Não Operac.",rno_c,False),("sub","  (-) IR/CSLL",[ir_c[i]+cs_c[i] for i in range(n_m)],True),
                    ("tot","LUCRO LÍQUIDO",ll_c,False),("tot","EBITDA",ebt_c,False),
                ]

                ws=wb.create_sheet(nome_cenario)
                ws.cell(row=1,column=1,value=f"Cenário: {nome_cenario} | Variação aplicada: {var_pct}%").font=Font(italic=True,color="6B7280",size=9)
                ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=4)
                ws.cell(row=2,column=1,value="Descrição").font=cor_branco_bold
                ws.cell(row=2,column=1).fill=cor_azul
                col=2
                for m in meses_proj:
                    c=ws.cell(row=2,column=col,value=m)
                    c.font=cor_branco_bold; c.fill=cor_azul; c.alignment=Alignment(horizontal="center")
                    col+=1

                linha_excel=3
                for tipo,desc,serie,inv in linhas_c:
                    ws.cell(row=linha_excel,column=1,value=desc.strip())
                    if tipo=="tot":
                        ws.cell(row=linha_excel,column=1).font=Font(bold=True,color="2563EB",size=10)
                        ws.cell(row=linha_excel,column=1).fill=cor_cinza_clara
                    else:
                        ws.cell(row=linha_excel,column=1).font=cor_preto
                    col=2
                    for i in range(n_m):
                        v=float(serie[i])
                        c1=ws.cell(row=linha_excel,column=col,value=v)
                        c1.number_format='R$ #,##0;[RED](R$ #,##0)'
                        c1.font=cor_vermelho if (inv and v!=0) else (cor_verde if v>0 else cor_preto)
                        ws.cell(row=linha_excel,column=col).border=borda_fina
                        col+=1
                    linha_excel+=1

                ws.column_dimensions["A"].width=26
                for cc in range(2,col):
                    ws.column_dimensions[get_column_letter(cc)].width=13
                ws.freeze_panes="B3"

            buf=BytesIO(); wb.save(buf); buf.seek(0)
            return buf.getvalue()

        st.download_button("📥 Exportar DRE — Todos os 3 Cenários (Excel)",gerar_excel_dre_3_cenarios(),
          file_name="DRE_Projetada_3_Cenarios.xlsx",
          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
          use_container_width=True,key="dl_dre_3cen")
        painel_edicao([
            ("receita bruta de vendas","💰 Receita Bruta"),
            ("impostos sobre vendas","🏛️ Impostos s/ Vendas"),
            ("devoluções de vendas","↩️ Devoluções"),
            ("CMV (custo da mercadoria vendida)","📦 CMV"),
            ("despesas comerciais","🛒 Desp. Comerciais"),
            ("despesas administrativas","🏢 Desp. Administrativas"),
            ("despesas financeiras líquidas","💳 Desp. Financeiras"),
            ("despesas com depreciações e amortizações","📉 Depreciação"),
        ])

    # ── BALANÇO PROJETADO
    elif "Balanço" in demo:
        sec("🏦 Balanço Projetado")
        disp = get_proj("disponibilidades saldo")
        cr   = get_proj("contas a receber saldo")
        est  = get_proj("estoque final do mês de mercadorias para revenda saldo")
        oac  = get_proj("Outros AC")
        anc  = get_proj("Ativo NC")
        ac   = [disp[i]+cr[i]+est[i]+oac[i] for i in range(n_m)]
        at   = [ac[i]+anc[i] for i in range(n_m)]
        forn = get_proj("contas a pagar de fornecedores saldo")
        pf   = get_proj("Passivos Financeiros")
        opc  = get_proj("Outros PC")
        pnc  = get_proj("Passivo NC")
        pc   = [forn[i]+pf[i]+opc[i] for i in range(n_m)]
        pt   = [pc[i]+pnc[i] for i in range(n_m)]
        pl   = [at[i]-pt[i] for i in range(n_m)]

        def ah(serie,i): return safe(serie[i]-serie[i-1],abs(serie[i-1]))*100 if i>0 else 0

        linhas_bal=[
            ("cat","ATIVO",None),
            ("sub","  Disponibilidades",disp,False),
            ("sub","  Contas a Receber",cr,False),
            ("sub","  Estoques",est,False),
            ("sub","  Outros AC",oac,False),
            ("tot","= ATIVO CIRCULANTE",ac,False),
            ("sub","  Ativo NC",anc,False),
            ("tot","= ATIVO TOTAL",at,False),
            ("cat","PASSIVO",None),
            ("sub","  Fornecedores",forn,True),
            ("sub","  Pass. Financeiros",pf,True),
            ("sub","  Outros PC",opc,True),
            ("tot","= PASSIVO CIRCULANTE",pc,True),
            ("sub","  Passivo NC",pnc,True),
            ("tot","= PASSIVO TOTAL",pt,True),
            ("tot","= PATRIMÔNIO LÍQUIDO",pl,False),
        ]

        header_b="<tr><th>Descrição</th>"+"".join(f"<th>{m}</th><th>AV%</th><th>AH%</th>" for m in meses_proj)+"</tr>"
        rows_b=""
        for item in linhas_bal:
            if item[2] is None:
                rows_b+=f'<tr class="cat"><td>{item[1]}</td>'+"".join("<td></td><td></td><td></td>" for _ in meses_proj)+"</tr>"
                continue
            tipo,desc,serie,inv=item
            cls_tr={"tot":"tot","sub":"sub"}.get(tipo,"")
            row=f'<tr class="{cls_tr}"><td>{desc}</td>'
            for i in range(n_m):
                v=float(serie[i])
                at_i=float(at[i]) if at[i] else 1
                a_v=safe(v,at_i)*100
                a_h=ah(serie,i)
                cls_v="neg" if (inv and v!=0) else cor(v)
                row+=f'<td class="{cls_v}">{fmt(v)}</td>'
                row+=f'<td class="{cls_pct(a_v,inv)}">{fmt(a_v,"pct")}</td>'
                row+=f'<td class="{cls_pct(a_h,inv)}">{fmt(a_h,"pct")}</td>'
            rows_b+=row+"</tr>"
        st.markdown(f'<div class="dre-wrap" style="overflow-x:auto"><table class="dre" style="min-width:900px">{header_b}{rows_b}</table></div>',unsafe_allow_html=True)
        def gerar_excel_balanco_3_cenarios():
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb=Workbook()
            wb.remove(wb.active)

            cor_azul=PatternFill("solid",start_color="2563EB",end_color="2563EB")
            cor_cinza_clara=PatternFill("solid",start_color="F3F4F6",end_color="F3F4F6")
            cor_branco_bold=Font(bold=True,color="FFFFFF",size=11)
            cor_preto=Font(color="111827",size=10)
            cor_verde=Font(color="059669",size=10)
            cor_vermelho=Font(color="DC2626",size=10)
            borda_fina=Border(bottom=Side(style="thin",color="E8ECF0"))

            campos_bal_proj=["disponibilidades saldo","contas a receber saldo",
                "estoque final do mês de mercadorias para revenda saldo","Outros AC","Ativo NC",
                "contas a pagar de fornecedores saldo","Passivos Financeiros","Outros PC","Passivo NC"]

            for nome_cenario,fator_c in [("Pessimista",1-var_pct/100),("Base",1.0),("Otimista",1+var_pct/100)]:
                vals={}
                for campo in campos_bal_proj:
                    if campo in projs:
                        vals[campo]=[v*fator_c for v in projs[campo]["valores"]]
                    else:
                        v=float(ul_real.get(campo,0) or 0)
                        vals[campo]=[v]*n_m

                disp_c=vals["disponibilidades saldo"]; cr_c=vals["contas a receber saldo"]
                est_c=vals["estoque final do mês de mercadorias para revenda saldo"]; oac_c=vals["Outros AC"]
                anc_c=vals["Ativo NC"]
                ac_c=[disp_c[i]+cr_c[i]+est_c[i]+oac_c[i] for i in range(n_m)]
                at_c=[ac_c[i]+anc_c[i] for i in range(n_m)]
                forn_c=vals["contas a pagar de fornecedores saldo"]; pf_c=vals["Passivos Financeiros"]
                opc_c=vals["Outros PC"]; pnc_c=vals["Passivo NC"]
                pc_c=[forn_c[i]+pf_c[i]+opc_c[i] for i in range(n_m)]
                pt_c=[pc_c[i]+pnc_c[i] for i in range(n_m)]
                pl_c=[at_c[i]-pt_c[i] for i in range(n_m)]

                linhas_c=[
                    ("cat","ATIVO",None,False),
                    ("sub","  Disponibilidades",disp_c,False),
                    ("sub","  Contas a Receber",cr_c,False),
                    ("sub","  Estoques",est_c,False),
                    ("sub","  Outros AC",oac_c,False),
                    ("tot","ATIVO CIRCULANTE",ac_c,False),
                    ("sub","  Ativo NC",anc_c,False),
                    ("tot","ATIVO TOTAL",at_c,False),
                    ("cat","PASSIVO",None,False),
                    ("sub","  Fornecedores",forn_c,True),
                    ("sub","  Pass. Financeiros",pf_c,True),
                    ("sub","  Outros PC",opc_c,True),
                    ("tot","PASSIVO CIRCULANTE",pc_c,True),
                    ("sub","  Passivo NC",pnc_c,True),
                    ("tot","PASSIVO TOTAL",pt_c,True),
                    ("tot","PATRIMÔNIO LÍQUIDO",pl_c,False),
                ]

                ws=wb.create_sheet(nome_cenario)
                ws.cell(row=1,column=1,value=f"Cenário: {nome_cenario} | Variação aplicada: {var_pct}%").font=Font(italic=True,color="6B7280",size=9)
                ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=4)
                ws.cell(row=2,column=1,value="Descrição").font=cor_branco_bold
                ws.cell(row=2,column=1).fill=cor_azul
                col=2
                for m in meses_proj:
                    c=ws.cell(row=2,column=col,value=m)
                    c.font=cor_branco_bold; c.fill=cor_azul; c.alignment=Alignment(horizontal="center")
                    col+=1

                linha_excel=3
                for tipo,desc,serie,inv in linhas_c:
                    ws.cell(row=linha_excel,column=1,value=desc.strip())
                    if tipo=="tot":
                        ws.cell(row=linha_excel,column=1).font=Font(bold=True,color="2563EB",size=10)
                        ws.cell(row=linha_excel,column=1).fill=cor_cinza_clara
                    elif tipo=="cat":
                        ws.cell(row=linha_excel,column=1).font=Font(bold=True,color="111827",size=10)
                    else:
                        ws.cell(row=linha_excel,column=1).font=cor_preto

                    if tipo=="cat":
                        for cc in range(1,col):
                            ws.cell(row=linha_excel,column=cc).border=borda_fina
                        linha_excel+=1
                        continue

                    col=2
                    for i in range(n_m):
                        v=float(serie[i])
                        c1=ws.cell(row=linha_excel,column=col,value=v)
                        c1.number_format='R$ #,##0;[RED](R$ #,##0)'
                        c1.font=cor_vermelho if (inv and v!=0) else (cor_verde if v>0 else cor_preto)
                        ws.cell(row=linha_excel,column=col).border=borda_fina
                        col+=1
                    linha_excel+=1

                ws.column_dimensions["A"].width=26
                for cc in range(2,col):
                    ws.column_dimensions[get_column_letter(cc)].width=13
                ws.freeze_panes="B3"

            buf=BytesIO(); wb.save(buf); buf.seek(0)
            return buf.getvalue()

        st.download_button("📥 Exportar Balanço — Todos os 3 Cenários (Excel)",gerar_excel_balanco_3_cenarios(),
          file_name="Balanco_Projetado_3_Cenarios.xlsx",
          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
          use_container_width=True,key="dl_bal_3cen")
        painel_edicao([
            ("disponibilidades saldo","🏦 Disponibilidades"),
            ("contas a receber saldo","📥 Contas a Receber"),
            ("estoque final do mês de mercadorias para revenda saldo","📦 Estoques"),
            ("Outros AC","📋 Outros AC"),
            ("Ativo NC","🏛️ Ativo NC"),
            ("contas a pagar de fornecedores saldo","🏭 Fornecedores"),
            ("Passivos Financeiros","💳 Pass. Financeiros"),
            ("Outros PC","📋 Outros PC"),
            ("Passivo NC","🏛️ Passivo NC"),
        ])

    # ── FLUXO PROJETADO
    elif "Fluxo" in demo:
        sec("💰 Fluxo de Caixa Projetado")
        ent  = get_proj("Disponibilidades entradas")
        sai  = get_proj("Disponibilidades Saida")
        ev=float(st.session_state.get("entradas_vista",0))
        freq=st.session_state.get("freq_fluxo","Mensal")
        ev_m=ev*22 if freq=="Diário" else ev
        ent=[ent[i]+ev_m for i in range(n_m)]
        sp   = [ent[i]-sai[i] for i in range(n_m)]
        si   = float(st.session_state.get("saldo_ini",0))
        sa   = []
        for i in range(n_m):
            sa.append((sa[i-1] if i>0 else si)+sp[i])

        def ah(serie,i): return safe(serie[i]-serie[i-1],abs(serie[i-1]))*100 if i>0 else 0

        linhas_fc=[
            ("tot","Total Entradas",ent,False),
            ("sub","  (-) Total Saídas",sai,True),
            ("tot","= Saldo do Período",sp,False),
            ("tot","= Saldo Acumulado",sa,False),
        ]

        header_f="<tr><th>Descrição</th>"+"".join(f"<th>{m}</th><th>AH%</th>" for m in meses_proj)+"</tr>"
        rows_f=""
        for tipo,desc,serie,inv in linhas_fc:
            cls_tr={"tot":"tot","sub":"sub"}.get(tipo,"")
            row=f'<tr class="{cls_tr}"><td>{desc}</td>'
            for i in range(n_m):
                v=float(serie[i])
                a_h=ah(serie,i)
                cls_v="neg" if (inv and v!=0) else cor(v)
                row+=f'<td class="{cls_v}">{fmt(v)}</td>'
                row+=f'<td class="{cls_pct(a_h,inv)}">{fmt(a_h,"pct")}</td>'
            rows_f+=row+"</tr>"
        st.markdown(f'<div class="dre-wrap" style="overflow-x:auto"><table class="dre" style="min-width:700px">{header_f}{rows_f}</table></div>',unsafe_allow_html=True)
        def gerar_excel_fluxo_3_cenarios():
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb=Workbook()
            wb.remove(wb.active)

            cor_azul=PatternFill("solid",start_color="2563EB",end_color="2563EB")
            cor_cinza_clara=PatternFill("solid",start_color="F3F4F6",end_color="F3F4F6")
            cor_branco_bold=Font(bold=True,color="FFFFFF",size=11)
            cor_preto=Font(color="111827",size=10)
            cor_verde=Font(color="059669",size=10)
            cor_vermelho=Font(color="DC2626",size=10)
            borda_fina=Border(bottom=Side(style="thin",color="E8ECF0"))

            ev_fc=float(st.session_state.get("entradas_vista",0))
            freq_fc=st.session_state.get("freq_fluxo","Mensal")
            ev_m_fc=ev_fc*22 if freq_fc=="Diário" else ev_fc
            si_fc=float(st.session_state.get("saldo_ini",0))

            for nome_cenario,fator_c in [("Pessimista",1-var_pct/100),("Base",1.0),("Otimista",1+var_pct/100)]:
                if "Disponibilidades entradas" in projs:
                    ent_c=[v*fator_c for v in projs["Disponibilidades entradas"]["valores"]]
                else:
                    v=float(ul_real.get("Disponibilidades entradas",0) or 0)
                    ent_c=[v]*n_m
                ent_c=[ent_c[i]+ev_m_fc for i in range(n_m)]

                if "Disponibilidades Saida" in projs:
                    sai_c=[v*fator_c for v in projs["Disponibilidades Saida"]["valores"]]
                else:
                    v=float(ul_real.get("Disponibilidades Saida",0) or 0)
                    sai_c=[v]*n_m

                sp_c=[ent_c[i]-sai_c[i] for i in range(n_m)]
                sa_c=[]
                for i in range(n_m):
                    sa_c.append((sa_c[i-1] if i>0 else si_fc)+sp_c[i])

                linhas_c=[
                    ("tot","Total Entradas",ent_c,False),
                    ("sub","  (-) Total Saídas",sai_c,True),
                    ("tot","Saldo do Período",sp_c,False),
                    ("tot","Saldo Acumulado",sa_c,False),
                ]

                ws=wb.create_sheet(nome_cenario)
                ws.cell(row=1,column=1,value=f"Cenário: {nome_cenario} | Variação aplicada: {var_pct}%").font=Font(italic=True,color="6B7280",size=9)
                ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=4)
                ws.cell(row=2,column=1,value="Descrição").font=cor_branco_bold
                ws.cell(row=2,column=1).fill=cor_azul
                col=2
                for m in meses_proj:
                    c=ws.cell(row=2,column=col,value=m)
                    c.font=cor_branco_bold; c.fill=cor_azul; c.alignment=Alignment(horizontal="center")
                    col+=1

                linha_excel=3
                for tipo,desc,serie,inv in linhas_c:
                    ws.cell(row=linha_excel,column=1,value=desc.strip())
                    if tipo=="tot":
                        ws.cell(row=linha_excel,column=1).font=Font(bold=True,color="2563EB",size=10)
                        ws.cell(row=linha_excel,column=1).fill=cor_cinza_clara
                    else:
                        ws.cell(row=linha_excel,column=1).font=cor_preto
                    col=2
                    for i in range(n_m):
                        v=float(serie[i])
                        c1=ws.cell(row=linha_excel,column=col,value=v)
                        c1.number_format='R$ #,##0;[RED](R$ #,##0)'
                        c1.font=cor_vermelho if (inv and v!=0) else (cor_verde if v>0 else cor_preto)
                        ws.cell(row=linha_excel,column=col).border=borda_fina
                        col+=1
                    linha_excel+=1

                ws.column_dimensions["A"].width=26
                for cc in range(2,col):
                    ws.column_dimensions[get_column_letter(cc)].width=13
                ws.freeze_panes="B3"

            buf=BytesIO(); wb.save(buf); buf.seek(0)
            return buf.getvalue()

        painel_edicao([
            ("Disponibilidades entradas","💵 Entradas"),
            ("Disponibilidades Saida","💸 Saídas"),
        ])

elif pg=="importar_vendas":
    hdr("📥 Importar Vendas","Base de vendas para Curva de Pareto e ML por Produto (ex: exportação do ERP)")
    if not st.session_state.cid:
        st.markdown('<div class="al-w">⚠️ Cadastre e selecione um cliente primeiro.</div>',unsafe_allow_html=True); st.stop()
    arq_v=st.file_uploader("Selecione o arquivo (CSV ou Excel)",type=["csv","xlsx","xls"],key="up_vendas")
    if arq_v is not None:
        senha_import_vendas=st.text_input("Senha master para confirmar a importação *",type="password",key="senha_import_vendas")
        if st.button("📤 Processar arquivo importado",use_container_width=True):
            if senha_import_vendas!=SENHA_MASTER:
                st.error("❌ Senha master incorreta.")
            else:
                b=arq_v.read()
                df_v,msg_v=ler(b,arq_v.name)
                if df_v is None or isinstance(df_v,tuple):
                    st.markdown(f'<div class="al-d">❌ Não foi possível ler: {msg_v}</div>',unsafe_allow_html=True)
                else:
                    df_v.columns=[str(c).strip() for c in df_v.columns]
                    st.markdown(f'<div class="al-s">✅ {msg_v} — {len(df_v)} linhas, {df_v.shape[1]} colunas</div>',unsafe_allow_html=True)
                    st.session_state.vendas_raw=df_v
                    save_vendas_df(st.session_state.cid,df_v)
                    addlog(f"Vendas importadas: {len(df_v)} linhas")
                    sec("📊 Prévia dos dados importados")
                    st.dataframe(df_v.head(400),use_container_width=True,height=320)
    else:
        df_v_atual=get_vendas_df()
        if df_v_atual is not None and not df_v_atual.empty:
            st.markdown(f'<div class="al-i">ℹ️ Já existe uma base salva para este cliente ({len(df_v_atual)} linhas). Suba um novo arquivo acima para substituí-la.</div>',unsafe_allow_html=True)
            senha_limpar_vendas=st.text_input("Senha master para confirmar a limpeza *",type="password",key="senha_limpar_vendas")
            if st.button("🗑 Limpar base de vendas",use_container_width=True):
                if senha_limpar_vendas!=SENHA_MASTER:
                    st.error("❌ Senha master incorreta.")
                else:
                    st.session_state.vendas_raw=None
                    p_v=path_vendas(st.session_state.cid)
                    if os.path.exists(p_v): os.remove(p_v)
                    st.rerun()

elif pg=="pareto":
    hdr("📐 Curva de Pareto","Concentração de vendas por Cliente ou por Produto")
    df_v=get_vendas_df()
    if df_v is None or df_v.empty:
        st.markdown('<div class="al-w">⚠️ Importe a base de vendas primeiro em <b>📥 Importar Vendas</b>.</div>',unsafe_allow_html=True); st.stop()

    _col_fil_ml=col_filial(df_v)
    if _col_fil_ml:
        _filiais_disp_ml=sorted(df_v[_col_fil_ml].dropna().astype(str).unique().tolist())
        _opcoes_filial_ml=["(Todas as filiais)"]+_filiais_disp_ml

        def _on_change_filial_ml():
            st.session_state["cfgml_filial_sel_backup"]=st.session_state["cfgml_filial_sel"]
            for _k_limpar_pareto in ["pareto_resultado_atual","pareto_dim_atual","pareto_met_atual"]:
                if _k_limpar_pareto in st.session_state:
                    del st.session_state[_k_limpar_pareto]

        if "cfgml_filial_sel_backup" not in st.session_state:
            st.session_state["cfgml_filial_sel_backup"]=_opcoes_filial_ml[0]
        if st.session_state["cfgml_filial_sel_backup"] not in _opcoes_filial_ml:
            st.session_state["cfgml_filial_sel_backup"]=_opcoes_filial_ml[0]
        st.session_state["cfgml_filial_sel"]=st.session_state["cfgml_filial_sel_backup"]

        filial_sel_ml=st.selectbox("🏬 Filial",_opcoes_filial_ml,
          key="cfgml_filial_sel",on_change=_on_change_filial_ml)

        if filial_sel_ml!="(Todas as filiais)":
            df_v=df_v[df_v[_col_fil_ml].astype(str)==filial_sel_ml].copy()

    cols_v=list(df_v.columns)

    col_cli=next((c for c in cols_v if c.strip().lower() in ["nome","cliente","cliente/loja"]),None)
    col_prod=next((c for c in cols_v if c.strip().lower() in ["descricao","descrição","produto"]),None)
    col_valor=next((c for c in cols_v if c.strip().lower() in ["vlr.total","vlr total","valor total","valor"]),None)

    if not col_valor:
        st.markdown('<div class="al-d">❌ Não encontrei a coluna de valor total (ex: "Vlr.Total") nesta base.</div>',unsafe_allow_html=True); st.stop()

    sec("⚙️ Ver curva por")
    _opcoes_visao_pareto=["👥 Cliente (Vendas por Cliente)","📦 Produto (Vendas por Produto)"]

    def _on_change_visao_pareto():
        st.session_state["pareto_visao_backup"]=st.session_state["pareto_visao"]

    if "pareto_visao_backup" not in st.session_state:
        st.session_state["pareto_visao_backup"]=st.session_state.get("pareto_visao",_opcoes_visao_pareto[0])
    if st.session_state["pareto_visao_backup"] not in _opcoes_visao_pareto:
        st.session_state["pareto_visao_backup"]=_opcoes_visao_pareto[0]
    st.session_state["pareto_visao"]=st.session_state["pareto_visao_backup"]

    visao=st.radio("Escolha",_opcoes_visao_pareto,horizontal=True,
      key="pareto_visao",on_change=_on_change_visao_pareto)
    dim_sel=col_cli if "Cliente" in visao else col_prod
    if not dim_sel:
        st.markdown('<div class="al-d">❌ Não encontrei a coluna correspondente nesta base.</div>',unsafe_allow_html=True); st.stop()

    # Recarrega o resultado salvo em disco ANTES de desenhar o slider, pra que o padrão
    # de "20% da Classe A" já apareça certo mesmo numa sessão nova (não só na 2ª interação).
    if "pareto_resultado_atual" not in st.session_state and st.session_state.cid:
        _pareto_disco=load_pareto_snap(st.session_state.cid,st.session_state.get("cfgml_filial_sel"))
        if _pareto_disco is not None:
            st.session_state["pareto_resultado_atual"]=_pareto_disco
            st.session_state["pareto_dim_atual"]=dim_sel

    n_disponivel=df_v[dim_sel].nunique() if dim_sel else 20
    top_n_max=max(5,min(300,n_disponivel))
    if n_disponivel<=5:
        top_n=n_disponivel
        st.caption(f"ℹ️ Mostrando todos os {n_disponivel} itens disponíveis (menos que 5 no total)")
    else:
        # Padrão inteligente: sugere 20% da Classe A (os itens que, juntos, somam 80% do
        # valor total — a "fatia vital" do Pareto). Esse número muda a cada vez que você
        # gera a curva, porque depende de quantos itens formam a Classe A NESSA análise.
        # O slider vai até 300 pra você conseguir explorar além da sugestão, se quiser —
        # não é um limite do Pareto em si, é só o teto de itens que cabem legíveis no gráfico.
        _resultado_prev_pareto=st.session_state.get("pareto_resultado_atual")
        if _resultado_prev_pareto is not None and "classe_abc" in _resultado_prev_pareto.columns:
            _n_a_prev=int((_resultado_prev_pareto["classe_abc"]=="A").sum())
            _default_topn=max(5,min(top_n_max,round(_n_a_prev*0.2)))
            _fonte_sugestao=f"{_n_a_prev} itens na Classe A da última rodada"
        else:
            _default_topn=min(20,top_n_max)
            _fonte_sugestao="ainda sem rodada anterior pra calcular — gere a curva 1x pra ativar a sugestão"
        top_n=st.slider("Itens no gráfico",5,top_n_max,_default_topn,key="pareto_topn")
        st.caption(f"💡 Sugestão atual: {_default_topn} itens (20% de {_fonte_sugestao}). Ajuste o slider se quiser ver mais ou menos.")
    senha_gerar_pareto=st.text_input("Senha master para gerar",type="password",key="senha_gerar_pareto")
    if st.button("🚀 Gerar Curva de Pareto",use_container_width=True):
        if senha_gerar_pareto!=SENHA_MASTER:
            st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi executado.</div>',unsafe_allow_html=True)
            st.stop()
        df_v_calc=df_v.copy()
        resultado=pareto_analysis(df_v_calc,dim_sel,col_valor)
        st.session_state["pareto_resultado_atual"]=resultado
        st.session_state["pareto_dim_atual"]=dim_sel
        st.session_state["pareto_met_atual"]=col_valor
        if st.session_state.cid:
            save_pareto_snap(st.session_state.cid,resultado,st.session_state.get("cfgml_filial_sel"))
            save_snap(st.session_state.cid,"pareto_visao",{"visao":visao})
    resultado=st.session_state.get("pareto_resultado_atual")
    dim_salva=st.session_state.get("pareto_dim_atual")
    if resultado is not None and dim_salva!=dim_sel:
        st.markdown(f'<div class="al-w">⚠️ Você trocou a visão — clique em "🚀 Gerar Curva de Pareto" de novo para atualizar.</div>',unsafe_allow_html=True)
        resultado=None
    if resultado is not None:
        n_a=int((resultado["classe_abc"]=="A").sum())
        n_total=len(resultado)
        mc_cols=st.columns(3)
        mc(mc_cols[0],"Itens Classe A (80%)",f"{n_a} de {n_total}","g",f"{n_a/n_total*100:.1f}% dos itens")
        mc(mc_cols[1],"Itens Classe B (80-95%)",str(int((resultado["classe_abc"]=="B").sum())),"y")
        mc(mc_cols[2],"Itens Classe C (95-100%)",str(int((resultado["classe_abc"]=="C").sum())),"r")
        dim_atual=st.session_state.get("pareto_dim_atual",dim_sel)
        st.plotly_chart(grafico_pareto_plotly(resultado,f"Vlr.Total por {dim_atual}",top_n),use_container_width=True)
        with st.expander(f"📋 Tabela Completa ({len(resultado)} produtos)"):
            st.dataframe(resultado,use_container_width=True,height=400)
        csv_pareto=resultado.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
        st.download_button("📥 Exportar tabela (CSV)",csv_pareto,file_name=f"pareto_{gid(dim_atual)}.csv",use_container_width=True)

elif pg=="ml_produtos":
    hdr("🧮 ML por Produto (Top N%)","Escolhe o melhor modelo de ML por produto, restrito aos que mais vendem")
    if not STATS_OK:
        st.markdown('<div class="al-d">❌ pip install statsmodels scikit-learn</div>',unsafe_allow_html=True)
    df_v=get_vendas_df()
    if df_v is None or df_v.empty:
        st.markdown('<div class="al-w">⚠️ Importe a base de vendas primeiro em <b>📥 Importar Vendas</b>.</div>',unsafe_allow_html=True); st.stop()

    _col_fil_mlp=col_filial(df_v)
    if _col_fil_mlp:
        _filiais_disp_mlp=sorted(df_v[_col_fil_mlp].dropna().astype(str).unique().tolist())
        _opcoes_filial_mlp=["(Todas as filiais)"]+_filiais_disp_mlp

        def _on_change_filial_mlp():
            st.session_state["mlp_filial_sel_backup"]=st.session_state["mlp_filial_sel"]
            for _k_limpar_mlp in ["ml_produtos_resultado","mlp_produto_col_atual",
                                   "mlp_metrica_col_atual","mlp_data_col_atual"]:
                if _k_limpar_mlp in st.session_state:
                    del st.session_state[_k_limpar_mlp]

        if "mlp_filial_sel_backup" not in st.session_state:
            st.session_state["mlp_filial_sel_backup"]=_opcoes_filial_mlp[0]
        if st.session_state["mlp_filial_sel_backup"] not in _opcoes_filial_mlp:
            st.session_state["mlp_filial_sel_backup"]=_opcoes_filial_mlp[0]
        st.session_state["mlp_filial_sel"]=st.session_state["mlp_filial_sel_backup"]

        filial_sel_mlp=st.selectbox("🏬 Filial",_opcoes_filial_mlp,
          key="mlp_filial_sel",on_change=_on_change_filial_mlp)

        if filial_sel_mlp!="(Todas as filiais)":
            df_v=df_v[df_v[_col_fil_mlp].astype(str)==filial_sel_mlp].copy()

    cols_v=list(df_v.columns)

    cod_col=next((c for c in cols_v if c.strip().lower() in ["produto","codigo","código","sku"]),None)
    desc_col=next((c for c in cols_v if c.strip().lower() in ["descricao","descrição"]),None)
    metrica_col=next((c for c in cols_v if c.strip().lower() in ["vlr.total","vlr total","valor total","valor"]),None)
    data_col=next((c for c in cols_v if c.strip().lower() in ["emissao","emissão","data"]),None)

    if not metrica_col or not data_col:
        st.markdown('<div class="al-d">❌ Não encontrei as colunas Vlr.Total / Emissao nesta base.</div>',unsafe_allow_html=True); st.stop()
    if not cod_col and not desc_col:
        st.markdown('<div class="al-d">❌ Não encontrei nenhuma coluna de Produto (Código ou Descrição) nesta base.</div>',unsafe_allow_html=True); st.stop()

    # Chave única do produto: SÓ o código (é o que não muda). A descrição, quando existe,
    # vira apenas rótulo de exibição, usando a versão mais frequente pra cada código —
    # evita que pequenas variações de texto no cadastro (ex: espaços, "C/2" vs
    # "C / 2") dividam o mesmo produto em previsões separadas e contraditórias.
    # Se a base não tiver coluna de Descrição, usa só o código como rótulo.
    df_v=df_v.copy()
    if desc_col:
        df_v[desc_col]=df_v[desc_col].astype(str).str.strip()
    if cod_col:
        df_v[cod_col]=df_v[cod_col].astype(str).str.strip()
        if desc_col:
            desc_mais_comum=df_v.groupby(cod_col)[desc_col].agg(lambda s: s.value_counts().idxmax())
            df_v["_ProdutoUnico"]=df_v[cod_col].map(
                lambda c: f"{c} - {desc_mais_comum.get(c,'')}")
        else:
            df_v["_ProdutoUnico"]=df_v[cod_col]
        produto_col_agrupar=cod_col
    else:
        df_v["_ProdutoUnico"]=df_v[desc_col]
        produto_col_agrupar=desc_col
    produto_col="_ProdutoUnico"

    sec("⚙️ Configurações")
    if "_mlp_config_sincronizada" not in st.session_state:
        st.session_state["_mlp_config_sincronizada"]=True
        if st.session_state.cid:
            _cfg_mlp_atual=load_config_mlp(st.session_state.cid)
            if _cfg_mlp_atual:
                if _cfg_mlp_atual.get("mlp_pct") is not None:
                    st.session_state["mlp_pct"]=_cfg_mlp_atual["mlp_pct"]
                if _cfg_mlp_atual.get("mlp_meses") is not None:
                    st.session_state["mlp_meses"]=_cfg_mlp_atual["mlp_meses"]
                if _cfg_mlp_atual.get("mlp_minper") is not None:
                    st.session_state["mlp_minper"]=_cfg_mlp_atual["mlp_minper"]
    if "mlp_pct" not in st.session_state: st.session_state["mlp_pct"]=0.20
    if "mlp_meses" not in st.session_state: st.session_state["mlp_meses"]=3
    if "mlp_minper" not in st.session_state: st.session_state["mlp_minper"]=12
    pct_top=st.select_slider("Top % de produtos (por venda)",
      options=[0.05,0.10,0.20,0.30,0.50,1.00],
      format_func=lambda v:f"Top {int(v*100)}%",key="mlp_pct")
    c5,c6=st.columns(2)
    meses_previsao=c5.slider("Meses a prever",1,12,key="mlp_meses")
    min_periodos=c6.slider("Mínimo de meses de histórico",4,24,key="mlp_minper")

    senha_rodar_mlp=st.text_input("Senha master para rodar",type="password",key="senha_rodar_mlp")
    if st.button("🚀 Rodar ML nos Produtos",use_container_width=True):
        if senha_rodar_mlp!=SENHA_MASTER:
            st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi executado.</div>',unsafe_allow_html=True)
            st.stop()
        if st.session_state.cid:
            save_config_mlp(st.session_state.cid)
        df_v_calc=df_v.copy()
        df_v_calc[metrica_col]=pd.to_numeric(df_v_calc[metrica_col],errors="coerce")

        ranking_mlp=pareto_analysis(df_v_calc,"_ProdutoUnico",metrica_col)
        n_selecionar_mlp=max(1,int(np.ceil(len(ranking_mlp)*pct_top)))
        df_v_calc["_data_ml_mlp"]=pd.to_datetime(df_v_calc[data_col],errors="coerce",dayfirst=True)
        df_v_calc["_periodo_ml_mlp"]=df_v_calc["_data_ml_mlp"].dt.to_period("M")
        n_periodos_mlp=df_v_calc.dropna(subset=["_data_ml_mlp"]).groupby("_ProdutoUnico")["_periodo_ml_mlp"].nunique()
        produtos_com_hist_mlp=n_periodos_mlp[n_periodos_mlp>=min_periodos].index
        ranking_elegivel_mlp=ranking_mlp[ranking_mlp["_ProdutoUnico"].isin(produtos_com_hist_mlp)]
        top_produtos_mlp=ranking_elegivel_mlp.head(n_selecionar_mlp)["_ProdutoUnico"].tolist()

        st.markdown(f'<div class="al-i">🔎 {len(top_produtos_mlp)} produtos elegíveis (Top {int(pct_top*100)}% com histórico mínimo).</div>',unsafe_allow_html=True)
        pb_mlp=st.progress(0)
        texto_pb_mlp=st.empty()
        modelos_mlp=[m for m,ok in MODELOS_ML.items() if ok]
        linhas_mlp=[]
        for idx_mlp,prod_mlp in enumerate(top_produtos_mlp):
            texto_pb_mlp.caption(f"Processando {idx_mlp+1} de {len(top_produtos_mlp)}: {str(prod_mlp)[:50]}")
            serie_mlp=serie_mensal_produto(df_v_calc,"_ProdutoUnico",prod_mlp,data_col,metrica_col)
            melhor_mlp,rank_mlp=melhor_modelo(serie_mlp,modelos_mlp)
            proj_mlp=treinar(serie_mlp,melhor_mlp,meses_previsao)
            ultimo_mlp=float(serie_mlp.iloc[-1])
            if proj_mlp is not None:
                prox_mlp=float(proj_mlp.iloc[0])
                var_mlp=safe(prox_mlp-ultimo_mlp,abs(ultimo_mlp))*100
                linhas_mlp.append({"_ProdutoUnico":prod_mlp,"n_periodos":len(serie_mlp),
                    "modelo_escolhido":melhor_mlp,"ultimo_real":ultimo_mlp,
                    "previsao":[round(v,2) for v in proj_mlp.tolist()],
                    "var_pct_proximo_mes":round(var_mlp,1),"status":"ok","rank":rank_mlp})
            else:
                linhas_mlp.append({"_ProdutoUnico":prod_mlp,"n_periodos":len(serie_mlp),
                    "modelo_escolhido":melhor_mlp,"ultimo_real":ultimo_mlp,
                    "previsao":None,"var_pct_proximo_mes":None,"status":"falhou ao treinar","rank":rank_mlp})
            if len(top_produtos_mlp)>0:
                pb_mlp.progress((idx_mlp+1)/len(top_produtos_mlp))
        pb_mlp.empty(); texto_pb_mlp.empty()
        resultado=pd.DataFrame(linhas_mlp)

        st.session_state.ml_produtos_resultado=resultado
        st.session_state.vendas_raw_com_chave=df_v
        st.session_state["mlp_produto_col_atual"]=produto_col
        st.session_state["mlp_metrica_col_atual"]=metrica_col
        st.session_state["mlp_data_col_atual"]=data_col
        if st.session_state.cid:
            save_ml_produtos_resultado(st.session_state.cid,resultado,st.session_state.get("mlp_filial_sel"))
        addlog(f"ML por Produto: {len(resultado)} produtos (top {int(pct_top*100)}%)")

    resultado=st.session_state.get("ml_produtos_resultado")
    if resultado is None and st.session_state.cid:
        resultado=load_ml_produtos_resultado(st.session_state.cid,st.session_state.get("mlp_filial_sel"))
        if resultado is not None:
            st.session_state["ml_produtos_resultado"]=resultado
    if resultado is not None and not resultado.empty:
        produto_col_r=st.session_state.get("mlp_produto_col_atual",produto_col)
        metrica_col_r=st.session_state.get("mlp_metrica_col_atual",metrica_col)
        data_col_r=st.session_state.get("mlp_data_col_atual",data_col)
        df_v_r=st.session_state.get("vendas_raw_com_chave",df_v)
        ok_mask=resultado["status"]=="ok"
        n_ok=int(ok_mask.sum())
        mc_cols=st.columns(3)
        mc(mc_cols[0],"Produtos analisados",str(len(resultado)),"b")
        mc(mc_cols[1],"Previsões geradas",str(n_ok),"g")
        mc(mc_cols[2],"Sem dados suficientes",str(len(resultado)-n_ok),"y")
        sec("📋 Resultado — Melhor Modelo por Produto")
        def confiab_label(n):
            if n>=18: return "🟢 Alta"
            if n>=12: return "🟡 Média"
            return "🔴 Baixa"
        tabela_show=resultado.copy()
        tabela_show["confiabilidade"]=tabela_show["n_periodos"].apply(confiab_label)
        tabela_show["ultimo_real"]=tabela_show["ultimo_real"].apply(lambda v: fmt(v) if pd.notna(v) else "—")
        tabela_show["previsao_proximo_mes"]=tabela_show["previsao"].apply(
          lambda p: fmt(p[0]) if isinstance(p,list) and len(p)>0 else "—")
        with st.expander(f"📋 Ver tabela completa ({len(tabela_show)} produtos)"):
            st.dataframe(tabela_show[[produto_col_r,"n_periodos","confiabilidade","modelo_escolhido","ultimo_real",
              "previsao_proximo_mes","var_pct_proximo_mes","status"]],
              use_container_width=True,height=420)
        csv_ml=resultado.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
        c_exp1,c_exp2=st.columns(2)
        c_exp1.download_button("📥 Exportar resultado (CSV)",csv_ml,file_name="ml_produtos.csv",use_container_width=True)

        # CSV "editável": uma linha por mês (real + previsão), com coluna em branco
        # pra quem for usar preencher um ajuste manual por cima do número do modelo
        linhas_editavel=[]
        for _,linha_ex in resultado[ok_mask].iterrows():
            prod_ex=linha_ex[produto_col_r]
            serie_ex=serie_mensal_produto(df_v_r if 'df_v_r' in dir() else df_v,produto_col_r,prod_ex,data_col_r,metrica_col_r)
            for periodo_ex,valor_ex in serie_ex.items():
                linhas_editavel.append({"Produto":prod_ex,"Tipo":"Real","Mes":str(periodo_ex),
                    "Valor":round(float(valor_ex),2),"Modelo":"","Ajustado (preencha se quiser)":""})
            proj_ex=linha_ex["previsao"] or []
            modelo_ex=linha_ex["modelo_escolhido"]
            ultimo_periodo=serie_ex.index[-1]
            for i_ex,v_ex in enumerate(proj_ex):
                mes_futuro=ultimo_periodo+pd.DateOffset(months=i_ex+1)
                linhas_editavel.append({"Produto":prod_ex,"Tipo":"Previsão","Mes":mes_futuro.strftime("%Y-%m"),
                    "Valor":round(float(v_ex),2),"Modelo":modelo_ex,"Ajustado (preencha se quiser)":""})
        df_editavel=pd.DataFrame(linhas_editavel)
        csv_editavel=df_editavel.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
        c_exp2.download_button("📝 Exportar para edição (1 linha por mês)",csv_editavel,
          file_name="ml_produtos_editavel.csv",use_container_width=True)

        sec("📈 Evolução dos Produtos — Histórico + Previsão")
        n_graficos=st.slider("Quantos produtos mostrar",2,min(20,n_ok) if n_ok>=2 else 2,
          min(8,n_ok) if n_ok>=2 else n_ok,key="mlp_n_graficos")
        resultado_ok=resultado[ok_mask].head(n_graficos).reset_index(drop=True)
        cores_prod=["#14243B","#A9762F","#3D5A80","#8B5E34","#5B7B9A","#C99A54","#2C3E50","#B08D57",
                    "#4A6482","#996B3D","#1F3552","#D4A76A","#345678","#8C6239","#6889A8","#BF9B5F",
                    "#2E4761","#A0763B","#527092","#9C7443"]

        fig_lin=go.Figure()
        cards_var=[]
        for i,(_,linha_g) in enumerate(resultado_ok.iterrows()):
            prod_g=linha_g[produto_col_r]
            proj_g=linha_g["previsao"]
            melhor_g=linha_g["modelo_escolhido"]
            serie_g=serie_mensal_produto(df_v_r,produto_col_r,prod_g,data_col_r,metrica_col_r)
            x_h=[str(p) for p in serie_g.index]
            x_p=[f"M+{i2+1}" for i2 in range(len(proj_g))]
            v_at=float(serie_g.iloc[-1])
            v_pr=proj_g[-1]
            var_g=safe(v_pr-v_at,abs(v_at))*100
            cor_g=cores_prod[i%len(cores_prod)]
            nome_curto=str(prod_g)[:30]

            fig_lin.add_trace(go.Scatter(x=x_h,y=serie_g.values,name=nome_curto,
              legendgroup=f"g{i}",mode="lines",line=dict(color=cor_g,width=2)))
            fig_lin.add_trace(go.Scatter(x=[x_h[-1]]+x_p,y=[serie_g.values[-1]]+proj_g,
              name=nome_curto,legendgroup=f"g{i}",showlegend=False,
              mode="lines+markers",line=dict(color=cor_g,width=2,dash="dash"),
              marker=dict(size=5,color=cor_g,symbol="diamond")))

            n_per_g=len(serie_g)
            confiab_g="🟢 Alta" if n_per_g>=18 else ("🟡 Média" if n_per_g>=12 else "🔴 Baixa")
            cards_var.append({"produto":nome_curto,"cor":cor_g,"v_at":v_at,"v_pr":v_pr,
                              "var":var_g,"modelo":melhor_g,"confiab":confiab_g,"n_per":n_per_g})

        if len(resultado_ok)>0:
            primeira_serie=serie_mensal_produto(df_v_r,produto_col_r,
              resultado_ok.iloc[0][produto_col_r],data_col_r,metrica_col_r)
            ultimo_mes_real=str(primeira_serie.index[-1])
            fig_lin.add_vline(x=ultimo_mes_real,line_dash="dash",line_color="#9CA3AF",opacity=0.6)
            fig_lin.add_annotation(x=ultimo_mes_real,y=1,yref="paper",yanchor="bottom",
              text="Início da previsão",showarrow=False,font=dict(size=9,color="#9CA3AF"))
        fig_lin.update_layout(
          title=dict(text=f"Histórico e Previsão — Top {len(resultado_ok)} Produtos",
            font=dict(size=15,family="Georgia, serif",color="#14243B")),
          plot_bgcolor="white",paper_bgcolor="white",
          font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
          margin=dict(l=10,r=10,t=50,b=70),
          xaxis=dict(type="category",gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=False,
            tickangle=-40,tickfont=dict(size=9,color="#4B5563")),
          yaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=True,zeroline=False,tickfont=dict(size=9)),
          legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.38,x=0.5,xanchor="center",font=dict(size=9)),
          hovermode="x unified",height=500,
          hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B")))
        st.plotly_chart(fig_lin,use_container_width=True)

        sec("🎯 Variação Prevista por Produto")
        for i in range(0,len(cards_var),3):
            cols_c=st.columns(3)
            for j,col_c in enumerate(cols_c):
                if i+j>=len(cards_var): break
                cv=cards_var[i+j]
                cls_v="g" if cv["var"]>0 else ("r" if cv["var"]<0 else "")
                col_c.markdown(f'''<div class="mc" style="border-left:3px solid {cv["cor"]}">
                  <div class="mc-lbl">{cv["produto"]}</div>
                  <div class="mc-val {cls_v}">{cv["var"]:+.1f}%</div>
                  <div class="mc-sub">{fmt(cv["v_at"])} → {fmt(cv["v_pr"])} · {cv["modelo"]}</div>
                  <div class="mc-sub">{cv["confiab"]} · {cv["n_per"]} meses de histórico</div>
                  </div>''',unsafe_allow_html=True)

        sec("📈 Ver previsão de um produto específico")
        var_pct_prod=st.slider("Variação cenários (%)",5,30,15,step=5,key="mlp_var_cenario")
        produtos_ok=resultado[ok_mask][produto_col_r].tolist()
        if produtos_ok:
            prod_sel=st.selectbox("Produto",produtos_ok,key="mlp_prod_detalhe")
            serie=serie_mensal_produto(df_v_r,produto_col_r,prod_sel,data_col_r,metrica_col_r)
            linha=resultado[resultado[produto_col_r]==prod_sel].iloc[0]
            proj=linha["previsao"]
            melhor_nm=linha["modelo_escolhido"]
            rank_prod=linha.get("rank",{}) or {}
            if proj:
                x_h=[str(p) for p in serie.index]
                x_p=[f"M+{i+1}" for i in range(len(proj))]
                v_at=float(serie.iloc[-1])
                v_pr=proj[-1]
                var_tot=safe(v_pr-v_at,abs(v_at))*100

                fig=go.Figure()
                fig.add_trace(go.Scatter(x=x_h,y=serie.values,name="Histórico",
                  mode="lines+markers",line=dict(color="#14243B",width=2.3),
                  marker=dict(size=5,color="#14243B",line=dict(color="white",width=1.5))))
                fig.add_trace(go.Scatter(x=x_p,y=proj,name=f"Previsão ({linha['modelo_escolhido']})",
                  mode="lines+markers",line=dict(color="#A9762F",width=2.3,dash="dash"),
                  marker=dict(size=7,color="#A9762F",symbol="diamond",line=dict(color="white",width=1))))
                y_up=[v*(1+var_pct_prod/100) for v in proj]
                y_dn=[v*(1-var_pct_prod/100) for v in proj]
                fig.add_trace(go.Scatter(x=x_p+x_p[::-1],y=y_up+y_dn[::-1],
                  fill="toself",fillcolor="rgba(107,114,128,.08)",
                  line=dict(color="rgba(0,0,0,0)"),name=f"±{var_pct_prod}%"))
                fig.add_vline(x=x_h[-1],line_dash="dash",line_color="#9CA3AF",opacity=0.6)
                fig.add_annotation(x=x_h[-1],y=1,yref="paper",yanchor="bottom",
                  text="Início da previsão",showarrow=False,font=dict(size=9,color="#9CA3AF"))
                fig.update_layout(
                  title=dict(text=f"{prod_sel} — {var_tot:+.1f}% em {len(proj)} meses ({melhor_nm})",
                    font=dict(size=14,family="Georgia, serif",color="#14243B")),
                  plot_bgcolor="white",paper_bgcolor="white",font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
                  margin=dict(l=10,r=10,t=48,b=60),
                  xaxis=dict(type="category",tickangle=-40,tickfont=dict(size=9,color="#4B5563"),
                    linecolor="#E5E7EB",showgrid=False),
                  yaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=True,zeroline=False,tickfont=dict(size=9)),
                  legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.32,x=0.5,xanchor="center",font=dict(size=10)),
                  hovermode="x unified",height=400,
                  hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B")))
                st.plotly_chart(fig,use_container_width=True)

                c1,c2,c3=st.columns(3)
                mc(c1,f"🐻 Pessimista (-{var_pct_prod}%)",fmt(v_pr*(1-var_pct_prod/100)),"r",f"vs atual: {fmt(v_at)}")
                mc(c2,"📊 Base (projeção)",fmt(v_pr),"y",f"{var_tot:+.1f}% vs atual")
                mc(c3,f"🐂 Otimista (+{var_pct_prod}%)",fmt(v_pr*(1+var_pct_prod/100)),"g",f"vs atual: {fmt(v_at)}")

                media_serie=float(serie.mean()) if len(serie)>0 else 1
                st.markdown(f"""**Como chegamos a essa projeção:**

O sistema usou os últimos **{len(serie)}** períodos históricos para treinar cada modelo.
Depois testou a precisão de cada um nos últimos meses reais — comparando o que o modelo teria previsto com o que realmente aconteceu.
O vencedor foi **{melhor_nm}** com menor erro relativo.
O erro % abaixo mostra o desvio médio da previsão em relação à média histórica de **{fmt(media_serie)}**.
Quanto menor o %, mais preciso o modelo foi nos dados reais.

**Ranking de modelos — erro relativo (% sobre a média histórica):**""")
                rank_s=sorted(rank_prod.items(),key=lambda x:x[1])
                if rank_s:
                    desc_modelos={"ARIMA":"Captura tendências e autocorrelações",
                      "ExponentialSmoothing":"Pesa mais os dados recentes",
                      "SARIMAX":"Captura sazonalidade anual",
                      "Holt":"Tendência com amortecimento",
                      "Média Móvel":"Média dos últimos períodos",
                      "Prophet":"IA do Meta para séries temporais"}
                    icones_rank=["🥇","🥈","🥉","4️⃣","5️⃣","6️⃣","7️⃣","8️⃣","9️⃣","🔟"]
                    cols_r=st.columns(max(min(len(rank_s),4),1))
                    for i,(mod,mse) in enumerate(rank_s):
                        ico=icones_rank[i] if i<len(icones_rank) else f"{i+1}º"
                        col_atual=cols_r[i%len(cols_r)]
                        rmse=float(mse**0.5) if mse<float("inf") else 0
                        erro_pct=safe(rmse,abs(media_serie))*100
                        cls_err="g" if erro_pct<5 else ("y" if erro_pct<15 else "r")
                        col_atual.markdown(f'<div class="mc"><div class="mc-lbl">{ico} {mod}</div>'
                          f'<div class="mc-val {cls_err}" style="font-size:1rem">{erro_pct:.1f}% erro</div>'
                          f'<div class="mc-sub">{desc_modelos.get(mod,"")}</div></div>',unsafe_allow_html=True)
elif pg=="config_ml":
    hdr("🎛️ Motor de Previsão — Configuração Avançada","Parametrize sazonalidade, promoções e reajustes — e valide antes de aplicar")
    if "cfgml_config_carregada" not in st.session_state:
        _cfg_ml_salva=load_config_ml(st.session_state.cid) if st.session_state.cid else None
        if _cfg_ml_salva:
            for _k_cfg,_v_cfg in _cfg_ml_salva.items():
                st.session_state[_k_cfg]=_v_cfg
        st.session_state["cfgml_config_carregada"]=True
    _df_v_top_cfgml=get_vendas_df()
    if _df_v_top_cfgml is None or _df_v_top_cfgml.empty:
        st.markdown('<div class="al-w">⚠️ Importe a base de vendas primeiro em <b>📥 Importar Vendas</b>.</div>',unsafe_allow_html=True); st.stop()

    # Seletor de Filial — fica visível acima das abas, disponível tanto em "Resultado" quanto em "Configurar"
    _col_fil_top_cfgml=col_filial(_df_v_top_cfgml)
    if _col_fil_top_cfgml:
        _filiais_disp_top_cfgml=sorted(_df_v_top_cfgml[_col_fil_top_cfgml].dropna().astype(str).unique().tolist())
        _opcoes_filial_top_cfgml=["(Todas as filiais)"]+_filiais_disp_top_cfgml

        def _on_change_filial_cfgml():
            st.session_state["cfgml_filial_sel_pag_backup"]=st.session_state["cfgml_filial_sel_pag"]
            for _k_limpar in ["cfgml_resultado_atual","cfgml_df_comp_bruto","cfgml_df_escopo_val",
                            "cfgml_produto_col_val_usado","cfgml_col_data_val_usado","cfgml_metrica_val_usado",
                            "cfgml_produtos_fora_previsao"]:
                if _k_limpar in st.session_state:
                    del st.session_state[_k_limpar]

        if "cfgml_filial_sel_pag_backup" not in st.session_state:
            st.session_state["cfgml_filial_sel_pag_backup"]=_opcoes_filial_top_cfgml[0]
        if st.session_state["cfgml_filial_sel_pag_backup"] not in _opcoes_filial_top_cfgml:
            st.session_state["cfgml_filial_sel_pag_backup"]=_opcoes_filial_top_cfgml[0]
        st.session_state["cfgml_filial_sel_pag"]=st.session_state["cfgml_filial_sel_pag_backup"]

        st.selectbox("🏬 Filial",_opcoes_filial_top_cfgml,
          key="cfgml_filial_sel_pag",on_change=_on_change_filial_cfgml)

    tab_resultado,tab_config=st.tabs(["📊 Resultado","⚙️ Configurar"])
    with tab_config:
        df_v=_df_v_top_cfgml
        filial_sel_cfgml=st.session_state.get("cfgml_filial_sel_pag","(Todas as filiais)")
        if filial_sel_cfgml!="(Todas as filiais)" and _col_fil_top_cfgml:
            df_v=df_v[df_v[_col_fil_top_cfgml].astype(str)==filial_sel_cfgml].copy()

        cols_v=list(df_v.columns)

        col_prod=next((c for c in cols_v if c.strip().lower() in ["produto","codigo","código","sku"]),None)
        col_cat=next((c for c in cols_v if c.strip().lower() in ["categoria","segmento","grupo"]),None)
        col_desc=next((c for c in cols_v if c.strip().lower() in ["descricao","descrição"]),None)
        col_data_cfg=next((c for c in cols_v if c.strip().lower() in ["emissao","emissão","data"]),None)
        meses_disponiveis_cfg=[]
        if col_data_cfg:
            datas_cfg=pd.to_datetime(df_v[col_data_cfg],errors="coerce",dayfirst=True).dropna()
            meses_disponiveis_cfg=sorted(datas_cfg.dt.to_period("M").astype(str).unique().tolist())

        sec("1️⃣ Escopo")
        _opcoes_escopo_ml=["🏷️ Categoria/Grupo inteiro","📦 Produto específico","🌐 Catálogo inteiro"]

        def _on_change_escopo_ml():
            st.session_state["cfgml_escopo_tipo_backup"]=st.session_state["cfgml_escopo_tipo"]

        if "cfgml_escopo_tipo_backup" not in st.session_state:
            st.session_state["cfgml_escopo_tipo_backup"]=st.session_state.get("cfgml_escopo_tipo",_opcoes_escopo_ml[0])
        st.session_state["cfgml_escopo_tipo"]=st.session_state["cfgml_escopo_tipo_backup"]

        escopo_tipo=st.radio("Aplicar configuração a",_opcoes_escopo_ml,
        horizontal=True,key="cfgml_escopo_tipo",on_change=_on_change_escopo_ml)

        if "cfgml_agrupar_familia" not in st.session_state: st.session_state["cfgml_agrupar_familia"]=False
        agrupar_familia=st.checkbox("👕 Agrupar variações de tamanho/medida na mesma família (ex: Camisa P+M+G vira 1 item)",
        key="cfgml_agrupar_familia",
        help="Útil para produtos vendidos em vários tamanhos (roupas, EPI) — reduz a fragmentação de SKU e aumenta o volume/estabilidade de cada série")

        escopo_valor=None
        if "Categoria" in escopo_tipo:
            if not col_cat:
                st.markdown('<div class="al-w">⚠️ Esta base não tem coluna de Categoria/Segmento/Grupo — use "Produto específico".</div>',unsafe_allow_html=True)
            else:
                categorias_disp=sorted(df_v[col_cat].dropna().astype(str).unique().tolist())
                if st.session_state.get("cfgml_categoria") not in categorias_disp:
                    st.session_state.pop("cfgml_categoria",None)
                escopo_valor=st.selectbox("Categoria",categorias_disp,key="cfgml_categoria")
        elif "Catálogo" in escopo_tipo:
            if not col_prod:
                st.markdown('<div class="al-d">❌ Não encontrei coluna de Produto/Código nesta base.</div>',unsafe_allow_html=True)
            else:
                escopo_valor=sorted(df_v[col_prod].dropna().astype(str).unique().tolist())
                st.caption(f"ℹ️ Todo o catálogo será processado: {len(escopo_valor)} produtos. Isso pode levar bastante tempo em bases grandes.")
        else:
            if col_prod:
                produtos_disp=sorted(df_v[col_prod].dropna().astype(str).unique().tolist())
                col_busca,col_limpar=st.columns([4,1])
                _busca_key=f"cfgml_busca_prod_{st.session_state.get('cfgml_busca_versao',0)}"
                busca=col_busca.text_input("Buscar SKU/código (busca parcial, separe por vírgula para múltiplos)",key=_busca_key,placeholder="Ex: PC000001, PC000002")
                if col_limpar.button("🔄 Limpar tudo",key="cfgml_limpar_busca",use_container_width=True):
                    v=st.session_state.get("cfgml_busca_versao",0)+1
                    st.session_state["cfgml_busca_versao"]=v
                    if "cfgml_produto" in st.session_state:
                        del st.session_state["cfgml_produto"]
                    st.rerun()
                ja_selecionados=st.session_state.get("cfgml_produto",[])
                if busca:
                    termos=[t.strip().upper() for t in busca.split(",") if t.strip()]
                    produtos_filtrados=[p for p in produtos_disp if any(t in p.upper() for t in termos)]
                else:
                    produtos_filtrados=produtos_disp
                # Garante que os já selecionados sempre aparecem na lista
                lista_final=sorted(set(ja_selecionados+produtos_filtrados[:500]))
                produtos_selecionados=st.multiselect(f"Produto(s) ({len(produtos_filtrados)} encontrados) — escolha 1 ou mais",
                lista_final,key="cfgml_produto",placeholder="Digite ou selecione os produtos")
                if len(produtos_disp)>500:
                    st.caption("ℹ️ Mostrando os 500 primeiros da busca — refine o texto para achar outros.")
                escopo_valor=produtos_selecionados if produtos_selecionados else None
            else:
                st.markdown('<div class="al-d">❌ Não encontrei coluna de Produto/Código nesta base.</div>',unsafe_allow_html=True)

        if escopo_valor:
            if "Categoria" in escopo_tipo and col_prod:
                n_itens_escopo=df_v[df_v[col_cat].astype(str)==str(escopo_valor)][col_prod].nunique()
            else:
                n_itens_escopo=len(escopo_valor) if isinstance(escopo_valor,list) else 1
            if "Catálogo" in escopo_tipo:
                escopo_valor_texto="Catálogo inteiro"
            elif isinstance(escopo_valor,list):
                escopo_valor_texto=", ".join(escopo_valor)
            else:
                escopo_valor_texto=str(escopo_valor)
            st.markdown(f'<div class="al-i">📌 Configurando: <b>{escopo_tipo}</b> = <b>{escopo_valor_texto}</b> ({n_itens_escopo} item(ns))</div>',unsafe_allow_html=True)
            if st.button("💾 Salvar este Escopo (sem rodar o Cenário)",key="btn_salvar_escopo_leve"):
                if st.session_state.cid:
                    save_config_ml(st.session_state.cid)
                    st.markdown('<div class="al-s">✅ Escopo salvo — vai persistir mesmo após reiniciar.</div>',unsafe_allow_html=True)
            MESES_NOMES_CFG=["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
            col_data_cfg=next((c for c in cols_v if c.strip().lower() in ["emissao","emissão","data"]),None)

            if "Catálogo" in escopo_tipo:
                col_escopo_filtro=col_prod if col_prod else col_cat
                df_escopo=df_v
            else:
                col_escopo_filtro=col_cat if "Categoria" in escopo_tipo else col_prod
                if col_escopo_filtro and isinstance(escopo_valor,list):
                    df_escopo=df_v[df_v[col_escopo_filtro].astype(str).isin(escopo_valor)]
                elif col_escopo_filtro:
                    df_escopo=df_v[df_v[col_escopo_filtro].astype(str)==str(escopo_valor)]
                else:
                    df_escopo=df_v

            if agrupar_familia and col_desc:
                df_escopo=montar_coluna_familia(df_escopo,col_desc)
                n_antes_familia=df_escopo[col_prod].nunique() if col_prod else 0
                n_depois_familia=df_escopo["_FamiliaProduto"].nunique()
                st.markdown(f'<div class="al-i">👕 Agrupamento por família ativo: {n_antes_familia} SKUs → {n_depois_familia} famílias.</div>',unsafe_allow_html=True)
            meses_com_dado_escopo=MESES_NOMES_CFG
            if col_data_cfg:
                datas_escopo=pd.to_datetime(df_escopo[col_data_cfg],errors="coerce",dayfirst=True).dropna()
                idx_meses_presentes=sorted(datas_escopo.dt.month.unique().tolist())
                meses_com_dado_escopo=[MESES_NOMES_CFG[m-1] for m in idx_meses_presentes]

            sec("2️⃣ Sazonalidade")
            st.markdown('<div class="al-i">Marque os meses de pico, se este grupo/produto tiver um padrão sazonal conhecido. Deixe em branco se não houver sazonalidade conhecida.</div>',unsafe_allow_html=True)

            def _on_change_meses_pico():
                st.session_state["cfgml_meses_pico_backup"]=st.session_state["cfgml_meses_pico"]

            if "cfgml_meses_pico_backup" not in st.session_state:
                st.session_state["cfgml_meses_pico_backup"]=st.session_state.get("cfgml_meses_pico",[])
            st.session_state["cfgml_meses_pico_backup"]=[m for m in st.session_state["cfgml_meses_pico_backup"] if m in meses_com_dado_escopo]
            st.session_state["cfgml_meses_pico"]=st.session_state["cfgml_meses_pico_backup"]

            meses_pico=st.multiselect("Meses de pico de demanda",meses_com_dado_escopo,key="cfgml_meses_pico",
            on_change=_on_change_meses_pico,placeholder="Selecione os meses (opcional)")
            tem_sazonalidade=len(meses_pico)>0

            sec("3️⃣ Promoções / Eventos")
            st.markdown('<div class="al-i">Meses historicamente promocionais entram como variável exógena — o modelo aprende a separar esse efeito da sazonalidade normal.</div>',unsafe_allow_html=True)

            def _on_change_meses_promo():
                st.session_state["cfgml_meses_promo_backup"]=st.session_state["cfgml_meses_promo"]

            if "cfgml_meses_promo_backup" not in st.session_state:
                st.session_state["cfgml_meses_promo_backup"]=st.session_state.get("cfgml_meses_promo",[])
            st.session_state["cfgml_meses_promo_backup"]=[m for m in st.session_state["cfgml_meses_promo_backup"] if m in meses_com_dado_escopo]
            st.session_state["cfgml_meses_promo"]=st.session_state["cfgml_meses_promo_backup"]

            meses_promo_cfg=st.multiselect("Meses com promoção/campanha recorrente",meses_com_dado_escopo,key="cfgml_meses_promo",
            on_change=_on_change_meses_promo,placeholder="Selecione os meses (opcional)")

            meses_disponiveis_cfg=[]
            if col_data_cfg:
                datas_cfg=pd.to_datetime(df_v[col_data_cfg],errors="coerce",dayfirst=True).dropna()
                meses_disponiveis_cfg=sorted(datas_cfg.dt.to_period("M").astype(str).unique().tolist())

            sec("4️⃣ Exclusão de Outliers")
            st.markdown('<div class="al-i">Marque meses que tiveram evento atípico (greve, erro de lançamento, pico isolado) e não devem entrar no treino do modelo.</div>',unsafe_allow_html=True)
            if meses_disponiveis_cfg:
                def _on_change_outliers():
                    st.session_state["cfgml_outliers_backup"]=st.session_state["cfgml_outliers"]

                if "cfgml_outliers_backup" not in st.session_state:
                    st.session_state["cfgml_outliers_backup"]=st.session_state.get("cfgml_outliers",[])
                st.session_state["cfgml_outliers_backup"]=[m for m in st.session_state["cfgml_outliers_backup"] if m in meses_disponiveis_cfg]
                st.session_state["cfgml_outliers"]=st.session_state["cfgml_outliers_backup"]

                meses_excluir=st.multiselect("Meses a excluir do treino",meses_disponiveis_cfg,key="cfgml_outliers",
                on_change=_on_change_outliers,placeholder="Selecione os meses a excluir (opcional)")
            else:
                st.markdown('<div class="al-w">⚠️ Não encontrei coluna de data válida nesta base para listar os meses.</div>',unsafe_allow_html=True)
                meses_excluir=[]

            sec("5️⃣ Correção de Preço (Reajustes)")
            st.markdown('<div class="al-i">Cadastre datas de reajuste de preço/custo para "trazer" o histórico para a régua de preço atual, sem distorcer a tendência real de demanda.</div>',unsafe_allow_html=True)
            if "cfgml_reajustes" not in st.session_state:
                st.session_state.cfgml_reajustes=[]
            c_rj1,c_rj2,c_rj3=st.columns([2,1,1])
            if meses_disponiveis_cfg:
                data_reajuste=c_rj1.selectbox("Mês do reajuste",meses_disponiveis_cfg,key="cfgml_data_reajuste")
            else:
                data_reajuste=None
                c_rj1.markdown('<div class="al-w">Sem datas disponíveis.</div>',unsafe_allow_html=True)
            pct_reajuste=c_rj2.number_input("% Aplicado",value=0.0,step=0.5,format="%.1f",key="cfgml_pct_reajuste")
            if c_rj3.button("➕ Adicionar",key="cfgml_add_reajuste",use_container_width=True):
                if data_reajuste and pct_reajuste!=0:
                    st.session_state.cfgml_reajustes.append({"data":data_reajuste,"pct":pct_reajuste})
                    st.rerun()
            if st.session_state.cfgml_reajustes:
                for idx_rj,rj in enumerate(st.session_state.cfgml_reajustes):
                    c_show1,c_show2=st.columns([5,1])
                    c_show1.markdown(f"📅 **{rj['data']}** — {rj['pct']:+.1f}%")
                    if c_show2.button("🗑",key=f"cfgml_rm_reajuste_{idx_rj}"):
                        st.session_state.cfgml_reajustes.pop(idx_rj)
                        st.rerun()
            else:
                st.caption("Nenhum reajuste cadastrado ainda.")

            sec("6️⃣ Fatores Condicionantes de Mercado")
            st.markdown('<div class="al-i">Cadastre eventos externos que impactam a demanda — crises de abastecimento, eventos favoráveis, políticas cambiais, etc. O sistema aplica o fator como multiplicador de ajuste sobre as previsões geradas pelo ML.</div>',unsafe_allow_html=True)

            if "cfgml_fatores_mercado" not in st.session_state:
                st.session_state["cfgml_fatores_mercado"]=[]

            with st.expander("➕ Adicionar Fator de Mercado"):
                c_fm1,c_fm2=st.columns(2)
                nome_fator=c_fm1.text_input("Nome do evento",placeholder="Ex: Crise de abastecimento EUA",key="cfgml_fator_nome")
                tipo_fator=c_fm2.selectbox("Tipo de impacto",["📉 Redução de demanda","📈 Aumento de demanda"],key="cfgml_fator_tipo")
                c_fm3,c_fm4,c_fm5=st.columns(3)
                data_ini_fator=c_fm3.text_input("Mês início (AAAA-MM)",placeholder="2026-01",key="cfgml_fator_ini")
                data_fim_fator=c_fm4.text_input("Mês fim (AAAA-MM)",placeholder="2026-06",key="cfgml_fator_fim")
                pct_fator=c_fm5.number_input("Intensidade (%)",min_value=1,max_value=200,value=20,step=5,key="cfgml_fator_pct",
                    help="Ex: 30 = redução ou aumento de 30% na demanda prevista no período")
                if st.button("💾 Adicionar fator",key="cfgml_add_fator"):
                    if nome_fator and data_ini_fator and data_fim_fator:
                        sinal = -1 if "Redução" in tipo_fator else 1
                        st.session_state["cfgml_fatores_mercado"].append({
                            "nome": nome_fator,
                            "tipo": tipo_fator,
                            "ini": data_ini_fator,
                            "fim": data_fim_fator,
                            "pct": pct_fator,
                            "sinal": sinal,
                        })
                        st.rerun()
                    else:
                        st.markdown('<div class="al-w">⚠️ Preencha nome e período do evento.</div>',unsafe_allow_html=True)

            if st.session_state["cfgml_fatores_mercado"]:
                for idx_fm,fm in enumerate(st.session_state["cfgml_fatores_mercado"]):
                    ico = "📉" if fm["sinal"]==-1 else "📈"
                    c_f1,c_f2=st.columns([5,1])
                    c_f1.markdown(f'<div class="al-w" style="margin:2px 0">{ico} <b>{fm["nome"]}</b> — {fm["ini"]} a {fm["fim"]} — {fm["pct"]}% de {"redução" if fm["sinal"]==-1 else "aumento"}</div>',unsafe_allow_html=True)
                    if c_f2.button("🗑",key=f"cfgml_rm_fator_{idx_fm}"):
                        st.session_state["cfgml_fatores_mercado"].pop(idx_fm)
                        st.rerun()
            else:
                st.caption("Nenhum fator de mercado cadastrado ainda.")

            sec("7️⃣ Horizonte e Histórico")
            c_h1,c_h2=st.columns(2)
            if "cfgml_meses_prev" not in st.session_state: st.session_state["cfgml_meses_prev"]=3
            if "cfgml_min_hist" not in st.session_state: st.session_state["cfgml_min_hist"]=12
            meses_prev_cfg=c_h1.slider("Meses a prever",1,12,key="cfgml_meses_prev")
            min_hist_cfg=c_h2.slider("Mínimo de meses de histórico",4,24,key="cfgml_min_hist")

            sec("8️⃣ Cenário")
            nome_cenario_padrao=(f"{', '.join(escopo_valor[:2])}{'...' if len(escopo_valor)>2 else ''}"
            if isinstance(escopo_valor,list) else str(escopo_valor or ""))
            if "cfgml_nome_cenario" not in st.session_state:
                st.session_state["cfgml_nome_cenario"]=f"{nome_cenario_padrao} - {datetime.now().strftime('%d/%m %H:%M')}" if escopo_valor else ""
            nome_cenario_cfg=st.text_input("Nome deste cenário",key="cfgml_nome_cenario")

            st.markdown('''<div style="background:#EFF6FF;border:1px solid #BFDBFE;border-radius:9px;
            padding:12px 16px;margin:8px 0"><div style="color:#1E40AF;font-weight:700;font-size:.9rem;
            margin-bottom:8px">🎛️ Quais configurações usar nesta rodada?</div>
            <div style="color:#1E3A8A;font-size:.8rem">Ligue/desligue cada uma para isolar o efeito individual, ou combine várias para testar juntas.</div>
            </div>''',unsafe_allow_html=True)

            _opcoes_configs_ativas=["📅 Sazonalidade","🏷️ Promoção","💰 Reajuste de Preço","🚫 Excluir Outliers","🌐 Fatores de Mercado"]

            def _on_change_configs_ativas():
                st.session_state["cfgml_configs_ativas_backup"]=st.session_state["cfgml_configs_ativas"]

            if "cfgml_configs_ativas_backup" not in st.session_state:
                st.session_state["cfgml_configs_ativas_backup"]=st.session_state.get("cfgml_configs_ativas",["💰 Reajuste de Preço","🚫 Excluir Outliers"])
            st.session_state["cfgml_configs_ativas"]=st.session_state["cfgml_configs_ativas_backup"]

            configs_ativas=st.multiselect("Configurações ativas nesta rodada",
            _opcoes_configs_ativas,
            key="cfgml_configs_ativas",on_change=_on_change_configs_ativas,
            placeholder="Selecione as configurações")
            usar_sazonalidade="📅 Sazonalidade" in configs_ativas
            usar_promocao="🏷️ Promoção" in configs_ativas
            usar_reajuste="💰 Reajuste de Preço" in configs_ativas
            usar_outliers="🚫 Excluir Outliers" in configs_ativas
            usar_fatores_mercado="🌐 Fatores de Mercado" in configs_ativas

            partes_ativas=[]
            if usar_sazonalidade: partes_ativas.append("Sazonalidade")
            if usar_promocao: partes_ativas.append("Promoção")
            if usar_reajuste: partes_ativas.append("Reajuste de Preço")
            if usar_outliers: partes_ativas.append("Exclusão de Outliers")
            efeito_testar=" + ".join(partes_ativas) if partes_ativas else "Nenhuma (linha de base)"

            senha_rodar_cenario=st.text_input("Senha master para rodar",type="password",key="senha_rodar_cenario")
            c_run1,c_run2=st.columns(2)
            rodar_cenario=c_run1.button("🚀 Rodar Cenário",use_container_width=True,key="cfgml_btn_rodar")
            salvar_apos_rodar=c_run2.checkbox("💾 Salvar este cenário ao rodar",value=True,key="cfgml_salvar_check")
            if rodar_cenario and senha_rodar_cenario!=SENHA_MASTER:
                st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi executado.</div>',unsafe_allow_html=True)
                rodar_cenario=False

            meses_pico_usar=meses_pico if usar_sazonalidade else []
            meses_promo_usar=meses_promo_cfg if usar_promocao else []
            reajustes_usar=st.session_state.cfgml_reajustes if usar_reajuste else []
            meses_excluir_usar=meses_excluir if usar_outliers else []

            if rodar_cenario and escopo_valor:
                config_atual={
                    "escopo_tipo": escopo_tipo, "escopo_valor": escopo_valor,
                    "meses_pico": meses_pico, "meses_promo": meses_promo_cfg,
                    "meses_excluir": meses_excluir,
                    "reajustes": st.session_state.cfgml_reajustes,
                    "meses_previsao": meses_prev_cfg, "min_periodos": min_hist_cfg,
                }
                st.session_state["mlconfig_snapshot"]={
                    "escopo_tipo":escopo_tipo,"escopo_valor":escopo_valor,
                    "configs_ativas":partes_ativas,
                    "usar_fatores_mercado":usar_fatores_mercado,
                    "n_fatores_mercado":len(st.session_state.get("cfgml_fatores_mercado",[])) if usar_fatores_mercado else 0,
                    "meses_pico":meses_pico if usar_sazonalidade else [],
                    "n_reajustes":len(st.session_state.cfgml_reajustes) if usar_reajuste else 0,
                    "meses_previsao":meses_prev_cfg,"min_periodos":min_hist_cfg,
                    "nome_cenario":nome_cenario_cfg,
                    "gerado_em":datetime.now().strftime("%Y-%m-%d %H:%M"),
                }
                if st.session_state.cid: save_snap(st.session_state.cid,"mlconfig",st.session_state["mlconfig_snapshot"],filial=filial_sel_cfgml)
                df_rodar=df_escopo.copy()
                metrica_cfg=next((c for c in cols_v if c.strip().lower() in ["vlr.total","vlr total","valor total","valor"]),None)
                if agrupar_familia and "_FamiliaProduto" in df_escopo.columns:
                    produto_col_cfg="_FamiliaProduto"
                else:
                    produto_col_cfg=col_prod if col_prod else col_cat
                if metrica_cfg and col_data_cfg and produto_col_cfg:
                    df_rodar[metrica_cfg]=parse_valor_brl(df_rodar[metrica_cfg])
                    # Barra de progresso real: mostra quantos produtos já foram processados,
                    # em vez de um spinner cego sem noção de quanto falta
                    ranking_prog=pareto_analysis(df_rodar,produto_col_cfg,metrica_cfg)
                    total_prog=len(ranking_prog)
                    st.markdown(f'<div class="al-i">🔎 {total_prog} produtos elegíveis nesta categoria/escopo.</div>',unsafe_allow_html=True)
                    pb_cfg=st.progress(0)
                    texto_pb=st.empty()
                    linhas_cfg=[]
                    modelos_cfg=[m for m,ok in MODELOS_ML.items() if ok]
                    df_rodar["_data_ml"]=pd.to_datetime(df_rodar[col_data_cfg],errors="coerce",dayfirst=True)
                    df_rodar["_periodo_ml"]=df_rodar["_data_ml"].dt.to_period("M")
                    n_periodos_prod=df_rodar.dropna(subset=["_data_ml"]).groupby(produto_col_cfg)["_periodo_ml"].nunique()
                    produtos_com_hist=n_periodos_prod[n_periodos_prod>=min_hist_cfg].index
                    ranking_elegivel=ranking_prog[ranking_prog[produto_col_cfg].isin(produtos_com_hist)]
                    top_produtos_cfg=ranking_elegivel[produto_col_cfg].tolist()
                    usa_exog_cfg = bool(meses_pico_usar) or bool(meses_promo_usar)
                    if meses_excluir_usar:
                        st.markdown(f'<div class="al-i">🚫 {len(meses_excluir_usar)} mês(es) excluído(s) do treino: {", ".join(meses_excluir_usar)}</div>',unsafe_allow_html=True)
                    if reajustes_usar:
                        st.markdown(f'<div class="al-i">💰 {len(reajustes_usar)} reajuste(s) de preço aplicado(s) ao histórico antes de treinar.</div>',unsafe_allow_html=True)

                    if meses_excluir_usar and col_data_cfg:
                        df_rodar["_periodo_str_ex"]=pd.to_datetime(df_rodar[col_data_cfg],errors="coerce",dayfirst=True).dt.to_period("M").astype(str)
                        df_rodar=df_rodar[~df_rodar["_periodo_str_ex"].isin(meses_excluir_usar)]

                    for idx_p,prod_p in enumerate(top_produtos_cfg):
                        texto_pb.caption(f"Processando {idx_p+1} de {len(top_produtos_cfg)}: {str(prod_p)[:50]}")
                        serie_p=serie_mensal_produto(df_rodar,produto_col_cfg,prod_p,col_data_cfg,metrica_cfg)
                        serie_p=aplicar_correcao_precos(serie_p, reajustes_usar)
                        ultimo_p=float(serie_p.iloc[-1])

                        # Croston/TSB só valem a pena testar em produto de venda intermitente
                        # (muitos meses com zero venda). Pular eles quando o produto vende
                        # regularmente economiza 2 treinos por produto, sem tirar nenhum
                        # modelo que teria chance real de vencer.
                        _pct_zeros_p=float((pd.to_numeric(serie_p,errors="coerce").fillna(0)==0).mean()) if len(serie_p)>0 else 0.0
                        if _pct_zeros_p<0.30:
                            modelos_cfg_p=[m for m in modelos_cfg if m not in ("Croston","TSB")]
                        else:
                            modelos_cfg_p=modelos_cfg

                        proj_p=None; melhor_p=None
                        melhor_comparativo_p,_rank_p=melhor_modelo(serie_p,modelos_cfg_p)
                        exog_venceu_p=False
                        if usa_exog_cfg and len(serie_p)>=12:
                            mse_exog_p,_,_=treinar_backtest_exog(serie_p, meses_pico_usar, meses_promo_usar)
                            if mse_exog_p<float("inf"):
                                _rank_p["SARIMAX (c/ sazonalidade+promoção)"]=mse_exog_p
                                mse_comparativo_p=_rank_p.get(melhor_comparativo_p,float("inf"))
                                if mse_exog_p<mse_comparativo_p:
                                    exog_venceu_p=True
                        if exog_venceu_p:
                            exog_hist_p=montar_exog_calendario(serie_p.index, meses_pico_usar, meses_promo_usar)
                            datas_futuras_p=pd.date_range(serie_p.index[-1],periods=meses_prev_cfg+1,freq="MS")[1:]
                            exog_fut_p=montar_exog_calendario(datas_futuras_p, meses_pico_usar, meses_promo_usar)
                            proj_p,nome_exog_p=treinar_com_exog(serie_p,exog_hist_p,exog_fut_p,meses_prev_cfg)
                            if proj_p is not None:
                                melhor_p=nome_exog_p
                        if proj_p is None:
                            melhor_p=melhor_comparativo_p
                            proj_p=treinar(serie_p,melhor_p,meses_prev_cfg)

                        if proj_p is not None:
                            proj_lista_p=proj_p.tolist() if hasattr(proj_p,"tolist") else list(proj_p)
                            proj_lista_p=[max(0.0,v) for v in proj_lista_p]  # venda nunca pode ser negativa
                            # Aplicar fatores condicionantes de mercado
                            fatores_mercado=st.session_state.get("cfgml_fatores_mercado",[]) if usar_fatores_mercado else []
                            if fatores_mercado:
                                ultima_data=serie_p.index[-1] if hasattr(serie_p,"index") else pd.Timestamp.now()
                                for i_proj in range(len(proj_lista_p)):
                                    data_proj=ultima_data+pd.DateOffset(months=i_proj+1)
                                    fator_acum=1.0
                                    for fm in fatores_mercado:
                                        try:
                                            ini=pd.Timestamp(fm["ini"]+"-01")
                                            fim=pd.Timestamp(fm["fim"]+"-01")+pd.DateOffset(months=1)-pd.Timedelta(days=1)
                                            if ini<=data_proj<=fim:
                                                fator_acum*=(1+fm["sinal"]*fm["pct"]/100)
                                        except: pass
                                    proj_lista_p[i_proj]=max(0.0,proj_lista_p[i_proj]*fator_acum)
                            prox_p=proj_lista_p[0]
                            var_p=safe(prox_p-ultimo_p,abs(ultimo_p))*100
                            linhas_cfg.append({produto_col_cfg:prod_p,"n_periodos":len(serie_p),
                                "modelo_escolhido":melhor_p,"ultimo_real":ultimo_p,
                                "previsao":[round(v,2) for v in proj_lista_p],
                                "var_pct_proximo_mes":round(var_p,1),"status":"ok","rank":_rank_p})
                        else:
                            linhas_cfg.append({produto_col_cfg:prod_p,"n_periodos":len(serie_p),
                                "modelo_escolhido":melhor_p,"ultimo_real":ultimo_p,
                                "previsao":None,"var_pct_proximo_mes":None,"status":"falhou ao treinar","rank":_rank_p})
                        if len(top_produtos_cfg)>0:
                            pb_cfg.progress((idx_p+1)/len(top_produtos_cfg))
                    pb_cfg.empty(); texto_pb.empty()
                    resultado_cenario=pd.DataFrame(linhas_cfg)
                    st.session_state["cfgml_resultado_atual"]=resultado_cenario

                    # Produtos fora da previsão: histórico insuficiente (nunca entraram no loop) + falha ao treinar (entraram mas não geraram previsão)
                    excluidos_hist_p=ranking_prog[~ranking_prog[produto_col_cfg].isin(produtos_com_hist)].copy()
                    excluidos_hist_p["Motivo"]="Histórico insuficiente"
                    excluidos_hist_p["MesesHistorico"]=excluidos_hist_p[produto_col_cfg].map(n_periodos_prod).fillna(0).astype(int)
                    excluidos_hist_p=excluidos_hist_p.rename(columns={produto_col_cfg:"Produto"})[["Produto","Motivo","MesesHistorico"]]

                    if not resultado_cenario.empty and "status" in resultado_cenario.columns:
                        falhou_treino_p=resultado_cenario[resultado_cenario["status"]=="falhou ao treinar"].copy()
                    else:
                        falhou_treino_p=pd.DataFrame()
                    if not falhou_treino_p.empty:
                        falhou_treino_p["Motivo"]="Falha ao treinar"
                        falhou_treino_p["MesesHistorico"]=falhou_treino_p["n_periodos"]
                        falhou_treino_p=falhou_treino_p.rename(columns={produto_col_cfg:"Produto"})[["Produto","Motivo","MesesHistorico"]]
                    else:
                        falhou_treino_p=pd.DataFrame(columns=["Produto","Motivo","MesesHistorico"])

                    produtos_fora_previsao=pd.concat([excluidos_hist_p,falhou_treino_p],ignore_index=True)
                    produtos_fora_previsao["MinimoExigido"]=min_hist_cfg
                    st.session_state["cfgml_produtos_fora_previsao"]=produtos_fora_previsao
                    if st.session_state.cid:
                        save_cfgml_fora_previsao(st.session_state.cid,produtos_fora_previsao,st.session_state.get("cfgml_filial_sel_pag"))

                    if st.session_state.cid:
                        _resultado_cenario_save=resultado_cenario.copy()
                        if "_ProdutoUnico" not in _resultado_cenario_save.columns:
                            _resultado_cenario_save["_ProdutoUnico"]=_resultado_cenario_save[produto_col_cfg]
                        save_cfgml_resultado(st.session_state.cid,_resultado_cenario_save,st.session_state.get("cfgml_filial_sel_pag"))
                    st.session_state["cfgml_config_usada"]=config_atual
                    st.session_state["cfgml_produto_col_usado"]=produto_col_cfg
                    st.session_state["cfgml_df_base_usado"]=df_rodar
                    if st.session_state.cid:
                        save_config_ml(st.session_state.cid)
                    if salvar_apos_rodar and nome_cenario_cfg:
                        ok_mask_cfg=resultado_cenario["status"]=="ok"
                        resumo_cfg={
                            "n_produtos": int(len(resultado_cenario)),
                            "n_ok": int(ok_mask_cfg.sum()),
                            "erro_medio_pct": None,
                        }
                        cenarios_existentes_cfg=load_cenarios(st.session_state.cid)
                        nome_final_cfg=nome_cenario_cfg
                        if nome_final_cfg in cenarios_existentes_cfg:
                            nome_final_cfg=f"{nome_cenario_cfg} ({datetime.now().strftime('%H:%M:%S')})"
                            st.markdown(f'<div class="al-w">⚠️ Já existia um cenário com esse nome — salvo como "{nome_final_cfg}" para não sobrescrever.</div>',unsafe_allow_html=True)
                        save_cenario(st.session_state.cid, nome_final_cfg, config_atual, resumo_cfg)
                        st.markdown(f'<div class="al-s">✅ Cenário "{nome_final_cfg}" salvo e rodado.</div>',unsafe_allow_html=True)
                    else:
                        st.markdown('<div class="al-s">✅ Cenário rodado (não salvo).</div>',unsafe_allow_html=True)
                else:
                    st.markdown('<div class="al-d">❌ Não foi possível identificar as colunas necessárias (Valor, Data, Produto/Categoria) nesta base.</div>',unsafe_allow_html=True)
    with tab_resultado:            

        resultado_cfg=st.session_state.get("cfgml_resultado_atual")
        if resultado_cfg is None and st.session_state.cid:
            resultado_cfg=load_cfgml_resultado(st.session_state.cid,st.session_state.get("cfgml_filial_sel_pag"))
            if resultado_cfg is not None: st.session_state["cfgml_resultado_atual"]=resultado_cfg
        if "cfgml_produtos_fora_previsao" not in st.session_state and st.session_state.cid:
            _fora_prev_disco=load_cfgml_fora_previsao(st.session_state.cid,st.session_state.get("cfgml_filial_sel_pag"))
            if _fora_prev_disco is not None:
                st.session_state["cfgml_produtos_fora_previsao"]=_fora_prev_disco
        if resultado_cfg is not None and not resultado_cfg.empty:
            config_usada_cfg=st.session_state.get("cfgml_config_usada",{})
            produto_col_cfg_r=st.session_state.get("cfgml_produto_col_usado", col_prod if col_prod else col_cat)
            metrica_cfg_r=next((c for c in cols_v if c.strip().lower() in ["vlr.total","vlr total","valor total","valor"]),None)
            data_col_cfg_r=col_data_cfg
            df_v=st.session_state.get("cfgml_df_base_usado", df_v)

            ok_mask_show=resultado_cfg["status"]=="ok"
            n_ok_cfg=int(ok_mask_show.sum())
            mc_cols_cfg=st.columns(3)
            mc(mc_cols_cfg[0],"Itens no cenário",str(len(resultado_cfg)),"b")
            mc(mc_cols_cfg[1],"Previsões geradas",str(n_ok_cfg),"g")
            mc(mc_cols_cfg[2],"Sem dados suficientes",str(len(resultado_cfg)-n_ok_cfg),"y")

            tabela_cfg_show=resultado_cfg.copy()
            tabela_cfg_show["ultimo_real"]=tabela_cfg_show["ultimo_real"].apply(lambda v: fmt(v) if pd.notna(v) else "—")
            tabela_cfg_show["previsao_proximo_mes"]=tabela_cfg_show["previsao"].apply(
            lambda p: fmt(p[0]) if isinstance(p,list) and len(p)>0 else "—")
            with st.expander(f"📋 Ver tabela completa ({len(tabela_cfg_show)} produtos)"):
                st.dataframe(tabela_cfg_show[[produto_col_cfg_r,"n_periodos","modelo_escolhido","ultimo_real",
                "previsao_proximo_mes","var_pct_proximo_mes","status"]],use_container_width=True,height=320)

            csv_cfg=resultado_cfg.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
            df_fora_previsao=st.session_state.get("cfgml_produtos_fora_previsao")
            n_fora_previsao=len(df_fora_previsao) if df_fora_previsao is not None else None
            c_exp_cfg1,c_exp_cfg2,c_exp_cfg3=st.columns(3)
            c_exp_cfg1.download_button("📥 Exportar resultado (CSV)",csv_cfg,file_name="cenario_ml.csv",use_container_width=True)
            if n_fora_previsao is None:
                c_exp_cfg3.caption("ℹ️ Rode o cenário nesta sessão para poder exportar os que ficaram de fora")
            elif n_fora_previsao==0:
                c_exp_cfg3.caption("✅ Nenhum produto ficou de fora da previsão")
            else:
                csv_fora_previsao=df_fora_previsao.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
                c_exp_cfg3.download_button(f"🚫 Exportar fora da previsão ({n_fora_previsao})",csv_fora_previsao,
                file_name="produtos_fora_previsao.csv",use_container_width=True)

            # CSV editável: uma linha por mês (real + previsão), com coluna em branco pra ajuste manual
            linhas_edit_cfg=[]
            for _,linha_e in resultado_cfg[ok_mask_show].iterrows():
                prod_e=linha_e[produto_col_cfg_r]
                serie_e=serie_mensal_produto(df_v,produto_col_cfg_r,prod_e,data_col_cfg_r,metrica_cfg_r)
                if len(serie_e)==0:
                    continue  # produto sem venda na base/filial atual — pula esse, sem quebrar a exportação inteira
                for periodo_e,valor_e in serie_e.items():
                    linhas_edit_cfg.append({"Produto":prod_e,"Tipo":"Real","Mes":str(periodo_e),
                        "Valor":round(float(valor_e),2),"Modelo":"","Ajustado (preencha se quiser)":""})
                proj_e=linha_e["previsao"] or []
                modelo_e=linha_e["modelo_escolhido"]
                ultimo_periodo_e=serie_e.index[-1]
                for i_e,v_e in enumerate(proj_e):
                    mes_fut_e=ultimo_periodo_e+pd.DateOffset(months=i_e+1)
                    linhas_edit_cfg.append({"Produto":prod_e,"Tipo":"Previsão","Mes":mes_fut_e.strftime("%Y-%m"),
                        "Valor":round(float(v_e),2),"Modelo":modelo_e,"Ajustado (preencha se quiser)":""})
            if linhas_edit_cfg:
                df_edit_cfg=pd.DataFrame(linhas_edit_cfg)
                csv_edit_cfg=df_edit_cfg.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
                c_exp_cfg2.download_button("📝 Exportar para edição (1 linha por mês)",csv_edit_cfg,
                file_name="cenario_ml_editavel.csv",use_container_width=True)

            # Um gráfico separado por produto (mais fácil de examinar individualmente)
            sec("📈 Evolução — Histórico + Previsão por Produto")
            resultado_ok_cfg=resultado_cfg[ok_mask_show].reset_index(drop=True)
            with st.expander("⚙️ Configurar exibição do gráfico"):
                mostrar_media_movel=st.checkbox("📉 Mostrar linha de média móvel (3 meses) por cima do histórico",
                value=True,key="cfgml_media_movel")
                fator_escala_cfg=st.slider("📏 Alongar escala do gráfico (mais = linha mais suave/achatada)",
                0.5,3.0,3.0,step=0.25,key="cfgml_fator_escala")
                max_mostrar_cfg=min(300,len(resultado_ok_cfg)) if len(resultado_ok_cfg)>0 else 1
                if max_mostrar_cfg<=1:
                    n_mostrar_cfg=max_mostrar_cfg
                    st.caption(f"Mostrando o único produto disponível.")
                else:
                    n_mostrar_cfg=st.slider("Quantos produtos mostrar",1,max_mostrar_cfg,
                    min(6,max_mostrar_cfg),key="cfgml_n_mostrar")

            resultado_ok_cfg=resultado_ok_cfg.sort_values(produto_col_cfg_r).reset_index(drop=True)
            for i_c,(_,linha_c) in enumerate(resultado_ok_cfg.head(n_mostrar_cfg).iterrows()):
                prod_c=linha_c[produto_col_cfg_r]
                proj_c=linha_c["previsao"]
                modelo_c=linha_c["modelo_escolhido"]
                serie_c=serie_mensal_produto(df_v,produto_col_cfg_r,prod_c,data_col_cfg_r,metrica_cfg_r)
                x_h_c=[str(p) for p in serie_c.index]
                x_p_c=[f"M+{i2+1}" for i2 in range(len(proj_c))]

                fig_c=go.Figure()
                fig_c.add_trace(go.Scatter(x=x_h_c,y=serie_c.values,name="Histórico",
                mode="lines+markers",line=dict(color="#14243B",width=2.3),
                marker=dict(size=5,color="#14243B",line=dict(color="white",width=1.5))))
                if mostrar_media_movel and len(serie_c)>=3:
                    media_movel_c=serie_c.rolling(window=3,min_periods=1).mean()
                    serie_completa_c=list(serie_c.values)+list(proj_c)
                    media_movel_prev_c=[]
                    for k in range(len(serie_c),len(serie_completa_c)):
                        janela=serie_completa_c[max(0,k-2):k+1]
                        media_movel_prev_c.append(sum(janela)/len(janela))
                    fig_c.add_trace(go.Scatter(x=x_h_c,y=media_movel_c.values,name="Média 3m",
                    mode="lines",line=dict(color="#374151",width=1.5,dash="dot")))
                    fig_c.add_trace(go.Scatter(x=[x_h_c[-1]]+x_p_c,
                    y=[media_movel_c.values[-1]]+media_movel_prev_c,name="Média 3m (prevista)",
                    mode="lines",line=dict(color="#374151",width=1.5,dash="dot"),opacity=0.6,showlegend=False))
                fig_c.add_trace(go.Scatter(x=[x_h_c[-1]]+x_p_c,y=[serie_c.values[-1]]+proj_c,
                name=f"Previsão ({modelo_c})",mode="lines+markers",
                line=dict(color="#A9762F",width=2.3,dash="dash"),
                marker=dict(size=7,color="#A9762F",symbol="diamond",line=dict(color="white",width=1))))
                fig_c.add_vline(x=x_h_c[-1],line_dash="dash",line_color="#9CA3AF",opacity=0.6)
                fig_c.add_annotation(x=x_h_c[-1],y=1,yref="paper",yanchor="bottom",
                text="Início da previsão",showarrow=False,font=dict(size=9,color="#9CA3AF"))
                todos_valores_c=list(serie_c.values)+list(proj_c)
                v_min_c,v_max_c=min(todos_valores_c),max(todos_valores_c)
                media_c_esc=sum(todos_valores_c)/len(todos_valores_c)
                folga_c=max((v_max_c-v_min_c)*fator_escala_cfg, media_c_esc*0.3)
                n_pontos_c=len(x_h_c)+len(x_p_c)
                fig_c.update_layout(
                title=dict(text=str(prod_c)[:60],font=dict(size=14,family="Georgia, serif",color="#14243B")),
                plot_bgcolor="white",paper_bgcolor="white",font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
                margin=dict(l=10,r=10,t=44,b=60),
                xaxis=dict(type="category",gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=False,
                    tickangle=-40,tickfont=dict(size=9,color="#4B5563"),nticks=min(n_pontos_c,12)),
                yaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=True,zeroline=False,tickfont=dict(size=9)),
                legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.32,x=0.5,xanchor="center",font=dict(size=10)),
                hovermode="x unified",height=380,
                hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B")))
                st.plotly_chart(fig_c,use_container_width=True)

            if len(resultado_ok_cfg)>n_mostrar_cfg:
                st.caption(f"Mostrando {n_mostrar_cfg} de {len(resultado_ok_cfg)} produtos com previsão — ajuste o slider acima ou use o CSV para ver todos.")

            # Ranking de erro dos modelos, por produto selecionado
            sec("🎯 Ranking de Modelos por Produto")
            produtos_ok_cfg=resultado_ok_cfg[produto_col_cfg_r].tolist()
            if produtos_ok_cfg:
                prod_sel_cfg=st.selectbox("Produto",produtos_ok_cfg,key="cfgml_prod_detalhe")
                linha_sel_cfg=resultado_cfg[resultado_cfg[produto_col_cfg_r]==prod_sel_cfg].iloc[0]
                rank_sel_cfg=linha_sel_cfg.get("rank",{}) or {}
                serie_sel_cfg=serie_mensal_produto(df_v,produto_col_cfg_r,prod_sel_cfg,data_col_cfg_r,metrica_cfg_r)
                media_sel_cfg=float(serie_sel_cfg.mean()) if len(serie_sel_cfg)>0 else 1
                rank_ord_cfg=sorted(rank_sel_cfg.items(),key=lambda x:x[1])
                if rank_ord_cfg:
                    icones_rank_cfg=["🥇","🥈","🥉","4️⃣","5️⃣","6️⃣","7️⃣","8️⃣","9️⃣","🔟"]
                    cols_rank_cfg=st.columns(max(min(len(rank_ord_cfg),4),1))
                    for i_r,(mod_r,mse_r) in enumerate(rank_ord_cfg):
                        ico_r=icones_rank_cfg[i_r] if i_r<len(icones_rank_cfg) else f"{i_r+1}º"
                        col_atual_r=cols_rank_cfg[i_r%len(cols_rank_cfg)]
                        rmse_r=float(mse_r**0.5) if mse_r<float("inf") else 0
                        erro_r=safe(rmse_r,abs(media_sel_cfg))*100
                        cls_r="g" if erro_r<5 else ("y" if erro_r<15 else "r")
                        col_atual_r.markdown(f'<div class="mc"><div class="mc-lbl">{ico_r} {mod_r}</div>'
                        f'<div class="mc-val {cls_r}" style="font-size:1rem">{erro_r:.1f}% erro</div></div>',unsafe_allow_html=True)

        sec("9️⃣ Validação Out-of-Sample (Previsto x Real)")
        if not meses_disponiveis_cfg or not escopo_valor:
            st.markdown('<div class="al-w">⚠️ Configure o Escopo (item 1) primeiro para validar.</div>',unsafe_allow_html=True)
        else:
            with st.expander("⚙️ Configurar validação",expanded=True):
                st.markdown('<div class="al-i">Escolha uma data de corte: o modelo treina SÓ com dado até essa data, prevê os meses seguintes "às cegas", e comparamos com o que realmente aconteceu (se já existir na base).</div>',unsafe_allow_html=True)
                data_corte=st.selectbox("Treinar até (data de corte)",meses_disponiveis_cfg[:-1],
                index=max(0,len(meses_disponiveis_cfg)-13),key="cfgml_data_corte")
                n_meses_valid=st.slider("Quantos meses validar depois do corte",1,12,6,key="cfgml_n_valid")
                _rodar_validacao_cfg=st.button("🔬 Rodar Validação",use_container_width=True,key="cfgml_btn_validar")

            if _rodar_validacao_cfg:
                metrica_val=next((c for c in cols_v if c.strip().lower() in ["vlr.total","vlr total","valor total","valor"]),None)
                if col_escopo_filtro and isinstance(escopo_valor,list):
                    df_escopo_val=df_v[df_v[col_escopo_filtro].astype(str).isin(escopo_valor)].copy()
                elif col_escopo_filtro:
                    df_escopo_val=df_v[df_v[col_escopo_filtro].astype(str)==str(escopo_valor)].copy()
                else:
                    df_escopo_val=df_v.copy()

                if agrupar_familia and col_desc:
                    df_escopo_val=montar_coluna_familia(df_escopo_val,col_desc)
                    produto_col_val="_FamiliaProduto"
                else:
                    produto_col_val=col_prod if col_prod else col_cat

                if produto_col_val and metrica_val and col_data_cfg:
                    df_escopo_val["_periodo_val"]=pd.to_datetime(df_escopo_val[col_data_cfg],errors="coerce",dayfirst=True).dt.to_period("M").astype(str)
                    df_treino_val=df_escopo_val[df_escopo_val["_periodo_val"]<=data_corte].copy()
                    df_treino_val[metrica_val]=parse_valor_brl(df_treino_val[metrica_val])

                    usa_exog_val = bool(meses_pico_usar) or bool(meses_promo_usar)
                    if meses_excluir_usar:
                        df_treino_val["_periodo_str_ex_val"]=pd.to_datetime(df_treino_val[col_data_cfg],errors="coerce",dayfirst=True).dt.to_period("M").astype(str)
                        df_treino_val=df_treino_val[~df_treino_val["_periodo_str_ex_val"].isin(meses_excluir_usar)]
                    ranking_val=pareto_analysis(df_treino_val,produto_col_val,metrica_val)
                    df_treino_val["_data_ml"]=pd.to_datetime(df_treino_val[col_data_cfg],errors="coerce",dayfirst=True)
                    df_treino_val["_periodo_ml"]=df_treino_val["_data_ml"].dt.to_period("M")
                    n_periodos_val=df_treino_val.dropna(subset=["_data_ml"]).groupby(produto_col_val)["_periodo_ml"].nunique()
                    produtos_hist_val=n_periodos_val[n_periodos_val>=min_hist_cfg].index
                    ranking_elegivel_val=ranking_val[ranking_val[produto_col_val].isin(produtos_hist_val)]
                    top_produtos_val=ranking_elegivel_val[produto_col_val].tolist()

                    st.markdown(f'<div class="al-i">🔎 {len(top_produtos_val)} produtos elegíveis para validação.</div>',unsafe_allow_html=True)
                    pb_val=st.progress(0)
                    texto_pb_val=st.empty()
                    linhas_val=[]
                    for idx_val,prod_v2 in enumerate(top_produtos_val):
                        texto_pb_val.caption(f"Validando {idx_val+1} de {len(top_produtos_val)}: {str(prod_v2)[:50]}")
                        serie_v2=serie_mensal_produto(df_treino_val,produto_col_val,prod_v2,col_data_cfg,metrica_val)
                        serie_v2=aplicar_correcao_precos(serie_v2, reajustes_usar)
                        ultimo_v2=float(serie_v2.iloc[-1])
                        proj_v2=None; melhor_v2=None
                        melhor_comparativo_v2,rank_v2=melhor_modelo(serie_v2,[m for m,ok in MODELOS_ML.items() if ok])
                        exog_venceu_v2=False
                        if usa_exog_val and len(serie_v2)>=12:
                            mse_exog_v2,_,_=treinar_backtest_exog(serie_v2, meses_pico_usar, meses_promo_usar)
                            if mse_exog_v2<float("inf"):
                                rank_v2["SARIMAX (c/ sazonalidade+promoção)"]=mse_exog_v2
                                mse_comparativo_v2=rank_v2.get(melhor_comparativo_v2,float("inf"))
                                if mse_exog_v2<mse_comparativo_v2:
                                    exog_venceu_v2=True
                        if exog_venceu_v2:
                            exog_hist_v2=montar_exog_calendario(serie_v2.index, meses_pico_usar, meses_promo_usar)
                            datas_fut_v2=pd.date_range(serie_v2.index[-1],periods=n_meses_valid+1,freq="MS")[1:]
                            exog_fut_v2=montar_exog_calendario(datas_fut_v2, meses_pico_usar, meses_promo_usar)
                            proj_v2,nome_v2=treinar_com_exog(serie_v2,exog_hist_v2,exog_fut_v2,n_meses_valid)
                            if proj_v2 is not None: melhor_v2=nome_v2
                        if proj_v2 is None:
                            melhor_v2=melhor_comparativo_v2
                            proj_v2=treinar(serie_v2,melhor_v2,n_meses_valid)
                        if proj_v2 is not None:
                            proj_lista_v2=proj_v2.tolist() if hasattr(proj_v2,"tolist") else list(proj_v2)
                            proj_lista_v2=[max(0.0,v) for v in proj_lista_v2]
                            linhas_val.append({produto_col_val:prod_v2,"n_periodos":len(serie_v2),
                                "modelo_escolhido":melhor_v2,"ultimo_real":ultimo_v2,
                                "previsao":[round(v,2) for v in proj_lista_v2],"status":"ok","rank":rank_v2})
                        if len(top_produtos_val)>0:
                            pb_val.progress((idx_val+1)/len(top_produtos_val))
                    pb_val.empty(); texto_pb_val.empty()
                    resultado_val=pd.DataFrame(linhas_val)

                    if resultado_val.empty or "status" not in resultado_val.columns:
                        ok_val=pd.DataFrame()
                    else:
                        ok_val=resultado_val[resultado_val["status"]=="ok"]

                    if ok_val.empty:
                        st.session_state["cfgml_df_comp_bruto"]=None
                        st.markdown('<div class="al-w">⚠️ Nenhum item teve histórico suficiente antes da data de corte para validar.</div>',unsafe_allow_html=True)
                    else:
                        linhas_comp=[]
                        for _,linha_v in ok_val.iterrows():
                            prod_v=linha_v[produto_col_val]
                            serie_completa_v=serie_mensal_produto(df_escopo_val,produto_col_val,prod_v,col_data_cfg,metrica_val)
                            corte_ts=pd.Period(data_corte).to_timestamp()
                            serie_real_pos=serie_completa_v[serie_completa_v.index>corte_ts].head(n_meses_valid)
                            proj_v=linha_v["previsao"]
                            for i_v in range(min(len(proj_v),len(serie_real_pos))):
                                real_v=float(serie_real_pos.iloc[i_v])
                                prev_v=float(proj_v[i_v])
                                erro_abs_v=abs(prev_v-real_v)
                                erro_pct_v=safe(erro_abs_v,abs(real_v))*100
                                linhas_comp.append({"Produto":prod_v,"Mes":str(serie_real_pos.index[i_v]),
                                    "Previsto":round(prev_v,2),"Real":round(real_v,2),
                                    "Erro %":round(erro_pct_v,1),"Bias":round(prev_v-real_v,2)})
                        st.session_state["cfgml_df_comp_bruto"]=pd.DataFrame(linhas_comp) if linhas_comp else None
                        st.session_state["cfgml_df_escopo_val"]=df_escopo_val
                        if st.session_state.cid and linhas_comp:
                            save_validacao_full(st.session_state.cid,
                                st.session_state["cfgml_df_comp_bruto"],df_escopo_val,
                                produto_col_val,col_data_cfg,metrica_val,
                                st.session_state.get("cfgml_filial_sel_pag"))
                        st.session_state["cfgml_produto_col_val_usado"]=produto_col_val
                        st.session_state["cfgml_col_data_val_usado"]=col_data_cfg
                        st.session_state["cfgml_metrica_val_usado"]=metrica_val
                else:
                    st.markdown('<div class="al-d">❌ Não foi possível identificar as colunas necessárias.</div>',unsafe_allow_html=True)

            df_comp_bruto=st.session_state.get("cfgml_df_comp_bruto")
            if df_comp_bruto is None and st.session_state.cid:
                _val_full=load_validacao_full(st.session_state.cid,st.session_state.get("cfgml_filial_sel_pag"))
                if _val_full is not None and _val_full.get("df_comp") is not None:
                    st.session_state["cfgml_df_comp_bruto"]=_val_full["df_comp"]
                    st.session_state["cfgml_df_escopo_val"]=_val_full["df_escopo"]
                    st.session_state["cfgml_produto_col_val_usado"]=_val_full["produto_col"]
                    st.session_state["cfgml_col_data_val_usado"]=_val_full["col_data"]
                    st.session_state["cfgml_metrica_val_usado"]=_val_full["metrica"]
                    df_comp_bruto=_val_full["df_comp"]
            if df_comp_bruto is not None and not df_comp_bruto.empty:
                produto_col_val=st.session_state.get("cfgml_produto_col_val_usado")
                col_data_val_r=st.session_state.get("cfgml_col_data_val_usado")
                metrica_val_r=st.session_state.get("cfgml_metrica_val_usado")
                df_escopo_val_r=st.session_state.get("cfgml_df_escopo_val")

                if "cfgml_limite_erro" not in st.session_state or st.session_state.get("_pg_mudou_agora"):
                    st.session_state["cfgml_limite_erro"]=20
                limite_erro_aceitavel=st.slider("🎯 Filtro de confiabilidade — erro máximo aceito por produto (%)",
                5,300,step=5,key="cfgml_limite_erro")
                if st.button("💾 Salvar este filtro para o Painel de Gestão de Estoque",key="btn_salvar_limite_erro"):
                    if st.session_state.cid:
                        save_config_ml(st.session_state.cid)
                        st.success(f"✅ Filtro de {limite_erro_aceitavel}% salvo — o Painel vai mostrar o Erro Médio com esse valor até você salvar outro.")

                erro_medio_por_produto=df_comp_bruto.groupby("Produto")["Erro %"].mean()
                produtos_dentro=erro_medio_por_produto[erro_medio_por_produto<=limite_erro_aceitavel].index
                n_total_prod_val=len(erro_medio_por_produto)
                n_dentro_val=len(produtos_dentro)

                st.markdown(f'<div class="al-i">📊 <b>{n_dentro_val} de {n_total_prod_val}</b> produtos ({n_dentro_val/n_total_prod_val*100:.0f}%) ficaram com erro médio ≤ {limite_erro_aceitavel}% — são os exibidos abaixo. Os demais ({n_total_prod_val-n_dentro_val}) tiveram erro maior e foram deixados de fora da visão detalhada, mas continuam no CSV completo para consulta.</div>',unsafe_allow_html=True)

                df_comp=df_comp_bruto[df_comp_bruto["Produto"].isin(produtos_dentro)]
                if df_comp.empty:
                    st.markdown('<div class="al-w">⚠️ Nenhum produto ficou dentro do limite escolhido. Aumente o limite acima para ver resultados.</div>',unsafe_allow_html=True)
                mape_geral=df_comp["Erro %"].mean() if not df_comp.empty else 0
                bias_geral=df_comp["Bias"].mean() if not df_comp.empty else 0
                c_v1,c_v2,c_v3=st.columns(3)
                mc(c_v1,"MAPE (Erro Médio)",f"{mape_geral:.1f}%","g" if mape_geral<15 else ("y" if mape_geral<30 else "r"))
                mc(c_v2,"Bias Médio",fmt(bias_geral),"r" if bias_geral>0 else "g",
                "Modelo superestima" if bias_geral>0 else "Modelo subestima")
                mc(c_v3,"Comparações válidas",str(len(df_comp)),"b")

                # COMPARATIVO: ML vs Média Simples
                if False:
                    with st.expander("📊 Comparativo: Motor ML vs Média Histórica Simples",expanded=True):
                        st.markdown('<div class="al-i">Veja a diferença entre prever com o Motor ML (calibrado) e simplesmente usar a média dos últimos meses — como a maioria das empresas ainda faz.</div>',unsafe_allow_html=True)
                    n_meses_media=st.slider("Período da média histórica (meses)",3,12,6,step=1,key="cfgml_n_meses_media")

                    if not df_comp.empty and df_escopo_val_r is not None:
                        linhas_comp_media=[]
                        produtos_val=df_comp["Produto"].unique()
                        for prod_v in produtos_val:
                            serie_v=serie_mensal_produto(df_escopo_val_r,produto_col_val,prod_v,col_data_val_r,metrica_val_r)
                            if len(serie_v)<n_meses_media+1: continue
                            # Pega os dados do período de validação
                            df_prod=df_comp[df_comp["Produto"]==prod_v]
                            if df_prod.empty: continue
                            n_periodos_val=len(df_prod)
                            # Calcula média dos n_meses_media anteriores ao corte
                            serie_treino=serie_v.iloc[:-(n_periodos_val)]
                            if len(serie_treino)<n_meses_media: continue
                            media_hist=serie_treino.iloc[-n_meses_media:].mean()
                            # Compara com real
                            for _,row_v in df_prod.iterrows():
                                real_v=row_v["Real"]
                                if real_v==0: continue
                                erro_media=abs(media_hist-real_v)/abs(real_v)*100
                                linhas_comp_media.append({
                                    "Produto":prod_v,
                                    "Real":real_v,
                                    "Previsto_ML":row_v["Previsto"],
                                    "Previsto_Media":media_hist,
                                    "Erro_ML":row_v["Erro %"],
                                    "Erro_Media":erro_media,
                                })

                        if linhas_comp_media:
                            df_comp_media=pd.DataFrame(linhas_comp_media)
                            mape_ml=df_comp_media["Erro_ML"].mean()
                            mape_media=df_comp_media["Erro_Media"].mean()
                            ganho=mape_media-mape_ml
                            economia_pct=ganho/mape_media*100 if mape_media>0 else 0

                            # KPIs comparativos
                            c_m1,c_m2,c_m3=st.columns(3)
                            mc(c_m1,f"MAPE Média {n_meses_media} meses",f"{mape_media:.1f}%","r")
                            mc(c_m2,"MAPE Motor ML",f"{mape_ml:.1f}%","g")
                            mc(c_m3,"Redução de Erro",f"{ganho:.1f}p.p. ({economia_pct:.0f}%)","b")

                            # Gráfico comparativo por produto
                            df_plot=df_comp_media.groupby("Produto").agg(
                                Erro_ML=("Erro_ML","mean"),
                                Erro_Media=("Erro_Media","mean")).reset_index()
                            df_plot=df_plot.sort_values("Erro_Media",ascending=False).head(30)

                            fig_comp=go.Figure()
                            fig_comp.add_trace(go.Bar(
                                x=df_plot["Produto"],y=df_plot["Erro_Media"],
                                name=f"Média {n_meses_media} meses",
                                marker=dict(color="#B91C1C",opacity=0.75,line=dict(color="white",width=0.8)),
                                hovertemplate="<b>%{x}</b><br>Erro Média: <b>%{y:.1f}%</b><extra></extra>"))
                            fig_comp.add_trace(go.Bar(
                                x=df_plot["Produto"],y=df_plot["Erro_ML"],
                                name="Motor ML (calibrado)",
                                marker=dict(color="#14243B",opacity=0.85,line=dict(color="white",width=0.8)),
                                hovertemplate="<b>%{x}</b><br>Erro ML: <b>%{y:.1f}%</b><extra></extra>"))
                            fig_comp.add_hline(y=mape_media,
                                line_dash="dot",line_color="#DC2626",line_width=1.5,opacity=0.5,
                                annotation_text=f"  Média histórica: {mape_media:.1f}%",
                                annotation_position="top left",
                                annotation_font=dict(size=10,color="#DC2626",family="Georgia, serif"))
                            fig_comp.add_hline(y=mape_ml,
                                line_dash="dot",line_color="#059669",line_width=1.5,opacity=0.5,
                                annotation_text=f"  Motor ML: {mape_ml:.1f}%",
                                annotation_position="bottom left",
                                annotation_font=dict(size=10,color="#059669",family="Georgia, serif"))
                            fig_comp.add_annotation(
                                text=f"⚡ ML reduz o erro em {ganho:.1f}p.p. ({economia_pct:.0f}% de melhoria)",
                                xref="paper",yref="paper",x=0.5,y=1.08,showarrow=False,
                                font=dict(size=12,color="#14243B",family="Georgia, serif"),
                                align="center")
                            fig_comp.update_layout(
                                title=dict(
                                    text=f"Motor ML vs Média {n_meses_media} meses — Top 30 produtos com maior erro na média",
                                    font=dict(size=14,family="Georgia, serif",color="#14243B")),
                                barmode="group",bargap=0.25,bargroupgap=0.06,
                                plot_bgcolor="white",paper_bgcolor="white",
                                font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
                                margin=dict(l=10,r=10,t=70,b=90),height=420,
                                xaxis=dict(
                                    showgrid=False,linecolor="#E5E7EB",
                                    tickangle=-45,tickfont=dict(size=8,color="#6B7280")),
                                yaxis=dict(
                                    gridcolor="#F3F4F6",linecolor="#E5E7EB",
                                    ticksuffix="%",tickfont=dict(size=9),
                                    showgrid=True,zeroline=False),
                                legend=dict(
                                    bgcolor="rgba(255,255,255,0.9)",
                                    bordercolor="#E5E7EB",borderwidth=1,
                                    orientation="h",y=-0.38,x=0.5,xanchor="center",
                                    font=dict(size=11)),
                                hovermode="x unified",
                                hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",
                                    font=dict(color="#14243B",size=11)))
                            st.plotly_chart(fig_comp,use_container_width=True)

                            pass

                            st.markdown(f'<div class="al-s">✅ O Motor ML reduziu o erro médio de <b>{mape_media:.1f}%</b> (média simples de {n_meses_media} meses) para <b>{mape_ml:.1f}%</b> — uma redução de <b>{ganho:.1f} pontos percentuais ({economia_pct:.0f}% de melhoria)</b>.</div>',unsafe_allow_html=True)

                            csv_comp=df_comp_media.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
                            st.download_button("📥 Exportar comparativo completo (CSV)",csv_comp,
                                file_name="comparativo_ml_vs_media.csv",use_container_width=True)
                        else:
                            st.markdown('<div class="al-w">⚠️ Não foi possível calcular a comparação. Verifique se há histórico suficiente antes do corte.</div>',unsafe_allow_html=True)

                csv_val=df_comp_bruto.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
                st.download_button("📥 Exportar validação COMPLETA — todos os produtos (CSV)",csv_val,file_name="validacao_out_of_sample_completa.csv",use_container_width=True)

                if not df_comp.empty:
                    sec("📊 Erro por Faixa de Valor")
                    st.markdown('<div class="al-i">O MAPE geral pode enganar quando mistura produtos de alto e baixo valor — produtos pequenos naturalmente têm % de erro maior mesmo com boa previsão em R$. Veja o erro separado por faixa.</div>',unsafe_allow_html=True)
                    valor_medio_produto=df_comp.groupby("Produto")["Real"].mean()
                    limite_baixo=valor_medio_produto.quantile(0.33)
                    limite_alto=valor_medio_produto.quantile(0.67)
                    def classificar_faixa(v):
                        if v<=limite_baixo: return "🔵 Baixo Valor"
                        if v<=limite_alto: return "🟡 Médio Valor"
                        return "🟢 Alto Valor"
                    df_comp=df_comp.copy()
                    df_comp["Faixa"]=df_comp["Produto"].map(valor_medio_produto).apply(classificar_faixa)
                    resumo_faixa=df_comp.groupby("Faixa").agg(
                        MAPE=("Erro %","mean"),
                        Bias=("Bias","mean"),
                        Produtos=("Produto","nunique"),
                        Comparacoes=("Produto","count")
                    ).reindex(["🟢 Alto Valor","🟡 Médio Valor","🔵 Baixo Valor"]).dropna()
                    cols_faixa=st.columns(len(resumo_faixa)) if len(resumo_faixa)>0 else []
                    for i_f,(faixa,row_f) in enumerate(resumo_faixa.iterrows()):
                        cls_f="g" if row_f["MAPE"]<15 else ("y" if row_f["MAPE"]<30 else "r")
                        cols_faixa[i_f].markdown(f'''<div class="mc">
                        <div class="mc-lbl">{faixa}</div>
                        <div class="mc-val {cls_f}">{row_f["MAPE"]:.1f}%</div>
                        <div class="mc-sub">{int(row_f["Produtos"])} produto(s) · {int(row_f["Comparacoes"])} comparações</div>
                        </div>''',unsafe_allow_html=True)

                with st.expander("ℹ️ Como interpretar esta validação"):
                    st.markdown("""
    <div style="color:#111827;line-height:1.6">

    Este teste treina o modelo <b>apenas com dados até a data de corte escolhida</b> — o modelo nunca vê o período seguinte. Depois, comparamos a previsão gerada com o que <b>realmente aconteceu</b> nesse período, que já está registrado na base. É um teste cego: a mesma lógica usada para validar modelos de previsão em empresas de grande porte.

    <p><b>MAPE (Erro Médio Percentual)</b> — mede o quanto a previsão errou, na média, em relação ao valor real, sem considerar direção.</p>
    <ul>
    <li>Até 10%: excelente</li>
    <li>10% a 20%: bom, confiável para a maioria das decisões</li>
    <li>20% a 30%: razoável, use com cautela e revisão manual</li>
    <li>Acima de 30%: fraco, não recomendado sem ajuste</li>
    </ul>

    <p><b>Bias (Viés Médio)</b> — mede se o modelo tende a errar sempre para o mesmo lado.</p>
    <ul>
    <li>Negativo: o modelo tende a <b>subestimar</b> (a realidade costuma vir acima do previsto) — atenção a risco de ruptura de estoque</li>
    <li>Positivo: o modelo tende a <b>superestimar</b> (a realidade costuma vir abaixo do previsto) — atenção a risco de excesso de estoque</li>
    <li>Próximo de zero: sem tendência sistemática de erro, o modelo erra "para os dois lados" de forma equilibrada</li>
    </ul>

    <p><b>Comparações válidas</b> — quantos meses reais existiam na base para comparar com a previsão nesse teste. Quanto mais meses validados, mais confiável é a conclusão sobre o MAPE e o Bias.</p>

    <p><b>Recomendação de uso:</b> utilize este teste antes de aplicar qualquer previsão a uma decisão real de compra ou estoque. Um MAPE baixo com Bias próximo de zero é o cenário ideal — indica que o modelo é preciso e não tem tendência de erro sistemático em nenhuma direção.</p>

    </div>
    """,unsafe_allow_html=True)

                if not df_comp.empty and df_escopo_val_r is not None and not df_escopo_val_r.empty:
                    sec("📈 Gráfico Comparativo — Previsto x Real, por Produto")
                    produtos_val_disp=sorted(df_comp["Produto"].unique().tolist())
                    n_val_max=min(100,len(produtos_val_disp))
                    n_val_mostrar=st.slider("Quantos gráficos mostrar",1,n_val_max,min(10,n_val_max),key="cfgml_n_graficos_val") if n_val_max>1 else n_val_max
                    produtos_mostrar_val=produtos_val_disp[:n_val_mostrar]
                    if len(produtos_val_disp)>100:
                        st.caption(f"Mostrando {len(produtos_mostrar_val)} de {len(produtos_val_disp)} produtos (limite de 100 no slider) — use o CSV para ver todos.")
                    else:
                        st.caption(f"Mostrando {len(produtos_mostrar_val)} de {len(produtos_val_disp)} produtos.")

                    for prod_val_sel in produtos_mostrar_val:
                        serie_hist_val=serie_mensal_produto(df_escopo_val_r,produto_col_val,prod_val_sel,col_data_val_r,metrica_val_r)
                        df_comp_prod=df_comp[df_comp["Produto"]==prod_val_sel].sort_values("Mes")
                        if serie_hist_val.empty or df_comp_prod.empty:
                            continue

                        fig_val=go.Figure()
                        x_hist_val=[str(p) for p in serie_hist_val.index]
                        fig_val.add_trace(go.Scatter(x=x_hist_val,y=serie_hist_val.values,name="Histórico",
                        mode="lines",line=dict(color="#14243B",width=2.3)))
                        x_val_meses=df_comp_prod["Mes"].tolist()
                        fig_val.add_trace(go.Scatter(x=x_val_meses,
                        y=df_comp_prod["Previsto"].tolist(),                           
                        name="Previsto (modelo)",mode="lines+markers",
                        line=dict(color="#A9762F",width=2.3,dash="dash"),
                        marker=dict(size=7,color="#A9762F",symbol="diamond",line=dict(color="white",width=1))))
                        fig_val.add_trace(go.Scatter(x=x_val_meses,y=df_comp_prod["Real"].tolist(),
                        name="Real (o que aconteceu)",mode="lines+markers",
                        line=dict(color="#059669",width=2.6),
                        marker=dict(size=8,color="#059669",symbol="circle",line=dict(color="white",width=1))))
                        if 'corte_ts' in dir():
                            fig_val.add_vline(x=str(corte_ts),line_dash="dash",line_color="#9CA3AF",opacity=0.6)
                            fig_val.add_annotation(x=str(corte_ts),y=1,yref="paper",yanchor="bottom",
                            text="Corte (treino termina aqui)",showarrow=False,font=dict(size=9,color="#9CA3AF"))
                                                
                        erro_prod_val=df_comp_prod["Erro %"].mean()

                        todos_valores_val=list(serie_hist_val.values)+df_comp_prod["Previsto"].tolist()+df_comp_prod["Real"].tolist()
                        v_min_val,v_max_val=min(todos_valores_val),max(todos_valores_val)
                        media_val_esc=sum(todos_valores_val)/len(todos_valores_val)
                        folga_val=max((v_max_val-v_min_val)*fator_escala_cfg, media_val_esc*0.3) if 'fator_escala_cfg' in dir() else (v_max_val-v_min_val)*1.5

                        n_total_pontos_val=len(x_hist_val)+len(x_val_meses)
                        passo_tick_val=max(1,n_total_pontos_val//10)
                        fig_val.update_layout(
                        title=dict(text=f"{str(prod_val_sel)[:55]} — Erro médio: {erro_prod_val:.1f}%",
                            font=dict(size=14,family="Georgia, serif",color="#14243B")),
                        plot_bgcolor="white",paper_bgcolor="white",font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
                        margin=dict(l=10,r=10,t=48,b=66),
                        xaxis=dict(type="category",tickangle=-40,tickfont=dict(size=9,color="#4B5563"),
                            dtick=passo_tick_val,linecolor="#E5E7EB",showgrid=False),
                        yaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=True,zeroline=False,
                            tickfont=dict(size=9),range=[max(0,v_min_val-folga_val),v_max_val+folga_val]),
                        legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.36,x=0.5,xanchor="center",font=dict(size=10)),
                        hovermode="x unified",height=400,
                        hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B")))
                        st.plotly_chart(fig_val,use_container_width=True)

            else:
                st.markdown(f'<div class="al-i">ℹ️ Ainda não existe uma Validação Out-of-Sample rodada para <b>{st.session_state.get("cfgml_filial_sel_pag")}</b>. Clique em "🔬 Rodar Validação" acima para gerar.</div>',unsafe_allow_html=True)

elif pg=="compras":
    hdr("📦 Gestão de Compras","Estoque, lead time e giro — recomendação de compra por produto, cruzando demanda prevista")
    # Seletor de Filial — fica visível acima das abas, disponível tanto em "Resultado" quanto em "Configurar"
    _df_est_top_cp=get_estoque_compras()
    _df_v_top_cp=get_vendas_df()
    _col_fil_est_top_cp=col_filial(_df_est_top_cp) if _df_est_top_cp is not None else None
    _col_fil_v_top_cp=col_filial(_df_v_top_cp) if _df_v_top_cp is not None else None
    if _col_fil_est_top_cp or _col_fil_v_top_cp:
        _filiais_disp_top_cp=sorted(set(
            (_df_est_top_cp[_col_fil_est_top_cp].dropna().astype(str).unique().tolist() if _col_fil_est_top_cp else []) +
            (_df_v_top_cp[_col_fil_v_top_cp].dropna().astype(str).unique().tolist() if _col_fil_v_top_cp else [])
        ))
        _opcoes_filial_top_cp=["(Todas as filiais)"]+_filiais_disp_top_cp

        def _on_change_filial_compras():
            st.session_state["compras_filial_sel_backup"]=st.session_state["compras_filial_sel"]
            # Lead time NÃO é limpo aqui de propósito: ele é configuração por produto/fornecedor,
            # não por filial — trocar de loja não deveria apagar o que você já configurou.
            for _k_limpar_cp in ["compras_resultado","compras_calendario","compras_morto"]:
                if _k_limpar_cp in st.session_state:
                    del st.session_state[_k_limpar_cp]

        if "compras_filial_sel_backup" not in st.session_state:
            st.session_state["compras_filial_sel_backup"]=_opcoes_filial_top_cp[0]
        if st.session_state["compras_filial_sel_backup"] not in _opcoes_filial_top_cp:
            st.session_state["compras_filial_sel_backup"]=_opcoes_filial_top_cp[0]
        st.session_state["compras_filial_sel"]=st.session_state["compras_filial_sel_backup"]

        st.selectbox("🏬 Filial",_opcoes_filial_top_cp,
          key="compras_filial_sel",on_change=_on_change_filial_compras)

    tab_resultado,tab_config=st.tabs(["📊 Resultado","⚙️ Configurar"])
    with tab_config:
        sec("Importar Estoque, Fornecedor e Lead Time")
        st.markdown('<div class="al-i">Suba um arquivo com: Produto, EstoqueAtual, CustoUnitario, Fornecedor e LeadTimeDias. O sistema tenta reconhecer as colunas automaticamente.</div>',unsafe_allow_html=True)
        arq_estoque=st.file_uploader("Selecione o arquivo (CSV ou Excel)",type=["csv","xlsx","xls"],key="compras_up_estoque")
        if arq_estoque is not None:
            senha_import_estoque=st.text_input("Senha master para confirmar a importação *",type="password",key="senha_import_estoque")
            if st.button("📤 Processar arquivo importado",use_container_width=True,key="compras_btn_processar_estoque"):
                if senha_import_estoque!=SENHA_MASTER:
                    st.error("❌ Senha master incorreta.")
                else:
                    b_est=arq_estoque.read()
                    df_estoque_raw,msg_est=ler(b_est,arq_estoque.name)
                    if df_estoque_raw is None or isinstance(df_estoque_raw,tuple):
                        st.markdown(f'<div class="al-d">❌ Não foi possível ler: {msg_est}</div>',unsafe_allow_html=True)
                    else:
                        df_estoque_raw.columns=[str(c).strip() for c in df_estoque_raw.columns]
                        col_prod_est=next((c for c in df_estoque_raw.columns if c.strip().lower() in ["produto","codigo","código","sku"]),None)
                        col_qtd_est=next((c for c in df_estoque_raw.columns if "estoqueatual" in c.strip().lower().replace(" ","") or c.strip().lower() in ["estoque","quantidade","saldo","qtd"]),None)
                        col_custo_est=next((c for c in df_estoque_raw.columns if "custo" in c.strip().lower()),None)
                        col_forn_est=next((c for c in df_estoque_raw.columns if "fornecedor" in c.strip().lower()),None)
                        col_fil_est=col_filial(df_estoque_raw)
                        if not col_prod_est or not col_qtd_est:
                            st.markdown('<div class="al-d">❌ Não encontrei as colunas de Produto e Estoque. Confira o cabeçalho do arquivo.</div>',unsafe_allow_html=True)
                        else:
                            df_est=pd.DataFrame()
                            df_est["Produto"]=df_estoque_raw[col_prod_est].astype(str).str.strip()
                            df_est["EstoqueAtual"]=parse_valor_brl(df_estoque_raw[col_qtd_est])
                            df_est["CustoUnitario"]=parse_valor_brl(df_estoque_raw[col_custo_est]) if col_custo_est else 0.0
                            df_est["Fornecedor"]=df_estoque_raw[col_forn_est].astype(str).str.strip() if col_forn_est else "Não informado"
                            if col_fil_est:
                                df_est["Filial"]=df_estoque_raw[col_fil_est].astype(str).str.strip()
                            st.session_state["compras_df_estoque"]=df_est
                            if st.session_state.cid:
                                save_estoque_compras(st.session_state.cid,df_est)
                            _msg_filial=f" · {df_est['Filial'].nunique()} filiais detectadas" if col_fil_est else ""
                            st.markdown(f'<div class="al-s">✅ {msg_est} — {len(df_est)} produtos importados e salvos ({df_est["Fornecedor"].nunique()} fornecedores{_msg_filial}).</div>',unsafe_allow_html=True)
                            st.dataframe(df_est.head(10),use_container_width=True)

        df_est_atual=get_estoque_compras()
        if df_est_atual is None:
            st.markdown('<div class="al-w">⚠️ Importe o arquivo de estoque acima para continuar.</div>',unsafe_allow_html=True)
            st.stop()
        elif arq_estoque is None:
            st.markdown(f'<div class="al-i">ℹ️ Estoque salvo já carregado ({len(df_est_atual)} produtos). Suba um novo arquivo acima para substituir.</div>',unsafe_allow_html=True)

        df_v_compras=get_vendas_df()
        if df_v_compras is None or df_v_compras.empty:
            st.markdown('<div class="al-w">⚠️ Também é necessário ter uma base de vendas importada (📥 Importar Vendas) para calcular a demanda prevista.</div>',unsafe_allow_html=True)
            st.stop()

        # Seletor de Filial agora fica acima das abas (fora do Configurar) — aqui só aplica o filtro
        # usando o que já foi escolhido lá em cima
        _col_fil_est_atual=col_filial(df_est_atual)
        _col_fil_v_compras=col_filial(df_v_compras)
        filial_sel_compras=st.session_state.get("compras_filial_sel","(Todas as filiais)")
        if filial_sel_compras!="(Todas as filiais)":
            if _col_fil_est_atual:
                df_est_atual=df_est_atual[df_est_atual[_col_fil_est_atual].astype(str)==filial_sel_compras].copy()
            if _col_fil_v_compras:
                df_v_compras=df_v_compras[df_v_compras[_col_fil_v_compras].astype(str)==filial_sel_compras].copy()

        col_prod_v=next((c for c in df_v_compras.columns if c.strip().lower() in ["produto","codigo","código","sku"]),None)
        col_data_v=next((c for c in df_v_compras.columns if c.strip().lower() in ["emissao","emissão","data"]),None)
        col_val_v=next((c for c in df_v_compras.columns if c.strip().lower() in ["vlr.total","vlr total","valor total","valor"]),None)
        st.markdown("---")
        with st.expander("Manutenção — Reclassificar Produtos"):
            st.markdown('<div class="al-i">Atualiza só a Classe ABC de cada produto (A, B, C, D, E ou SV — Sem Venda), sem refazer a previsão de demanda. Use depois de reimportar um histórico de vendas mais recente, quando quiser que a classificação reflita o dado novo sem esperar o Motor de Compras rodar de novo do zero (que pode levar horas em catálogos grandes).</div>',unsafe_allow_html=True)
            senha_reclassificar=st.text_input("Senha master para executar",type="password",key="senha_reclassificar")
            if st.button("🔄 Reclassificar Produtos (rápido, sem recalcular previsão)",use_container_width=True):
                if senha_reclassificar!=SENHA_MASTER:
                    st.markdown('<div class="al-d">❌ Senha master incorreta.</div>',unsafe_allow_html=True)
                    st.stop()
                df_res_reclass=st.session_state.get("compras_resultado")
                if df_res_reclass is None and st.session_state.cid:
                    df_res_reclass,_cal_rc,_morto_rc=load_resultado_compras(st.session_state.cid,st.session_state.get("compras_filial_sel"))
                if df_res_reclass is None or df_res_reclass.empty:
                    st.markdown('<div class="al-w">⚠️ Rode o Motor de Compras pelo menos uma vez antes de reclassificar.</div>',unsafe_allow_html=True)
                elif not (col_prod_v and col_data_v and col_val_v):
                    st.markdown('<div class="al-d">❌ Não encontrei as colunas de Produto/Data/Valor na base de Vendas.</div>',unsafe_allow_html=True)
                else:
                    with st.spinner("Reclassificando produtos..."):
                        ranking_reclass=pareto_analysis(df_v_compras,col_prod_v,col_val_v)
                        ranking_reclass=ranking_reclass.rename(columns={col_prod_v:"Produto"})
                        def _classe_5niveis_rc(pct_acum):
                            if pct_acum<=70: return "A"
                            if pct_acum<=80: return "B"
                            if pct_acum<=90: return "C"
                            if pct_acum<=97: return "D"
                            return "E"
                        ranking_reclass["classe_abc"]=ranking_reclass["pct_acumulado"].apply(_classe_5niveis_rc)
                        mapa_classe_rc=dict(zip(ranking_reclass["Produto"],ranking_reclass["classe_abc"]))
                        df_res_reclass["Classe"]=df_res_reclass["Produto"].map(mapa_classe_rc).fillna("SV")
                        st.session_state["compras_resultado"]=df_res_reclass
                        if st.session_state.cid:
                            save_resultado_compras(st.session_state.cid,df_res_reclass,
                                st.session_state.get("compras_calendario"),st.session_state.get("compras_morto"),
                                st.session_state.get("compras_filial_sel"))
                        n_sv_rc=int((df_res_reclass["Classe"]=="SV").sum())
                        st.markdown(f'<div class="al-s">✅ Reclassificação concluída — {n_sv_rc} produto(s) marcados como "SV" (sem venda no período).</div>',unsafe_allow_html=True)
        col_qtd_v=next((c for c in df_v_compras.columns if c.strip().lower() in ["quantidade","qtd"]),None)
        if not (col_prod_v and col_data_v and col_val_v):
            st.markdown('<div class="al-d">❌ A base de vendas precisa ter colunas de Produto, Data e Valor.</div>',unsafe_allow_html=True)
            st.stop()
        sec("Escopo")
    
    with tab_config:    
        st.markdown('<div class="al-i">Restrinja o cálculo a uma categoria específica, ou a produtos escolhidos manualmente — em vez de rodar sempre no catálogo inteiro.</div>',unsafe_allow_html=True)
        col_cat_v_compras=next((c for c in df_v_compras.columns if c.strip().lower() in ["categoria","segmento","grupo"]),None)

        _opcoes_escopo_compras=["🏷️ Categoria/Grupo inteiro","📦 Produto específico","🌐 Catálogo inteiro"]

        def _on_change_escopo_compras():
            st.session_state["compras_escopo_tipo_backup"]=st.session_state["compras_escopo_tipo"]

        if "compras_escopo_tipo_backup" not in st.session_state:
            st.session_state["compras_escopo_tipo_backup"]=st.session_state.get("compras_escopo_tipo",_opcoes_escopo_compras[0])
        st.session_state["compras_escopo_tipo"]=st.session_state["compras_escopo_tipo_backup"]

        escopo_tipo_compras=st.radio("Aplicar a",_opcoes_escopo_compras,
        horizontal=True,key="compras_escopo_tipo",on_change=_on_change_escopo_compras)

        produtos_no_escopo=None
        if "Categoria" in escopo_tipo_compras:
            if not col_cat_v_compras:
                st.markdown('<div class="al-w">⚠️ Esta base não tem coluna de Categoria — use "Produto específico" ou "Catálogo inteiro".</div>',unsafe_allow_html=True)
            else:
                categorias_disp_compras=sorted(df_v_compras[col_cat_v_compras].dropna().astype(str).unique().tolist())

                def _on_change_categoria_compras():
                    st.session_state["compras_categoria_sel_backup"]=st.session_state["compras_categoria_sel"]

                _bkp_cat=st.session_state.get("compras_categoria_sel_backup")
                if _bkp_cat not in categorias_disp_compras:
                    _bkp_cat=categorias_disp_compras[0] if categorias_disp_compras else None
                st.session_state["compras_categoria_sel"]=_bkp_cat
                categoria_sel_compras=st.selectbox("Categoria",categorias_disp_compras,
                key="compras_categoria_sel",on_change=_on_change_categoria_compras)
                produtos_no_escopo=set(df_v_compras[df_v_compras[col_cat_v_compras].astype(str)==categoria_sel_compras][col_prod_v].astype(str).unique())
        elif "Produto específico" in escopo_tipo_compras:
            produtos_disp_escopo=sorted(df_est_atual["Produto"].unique().tolist())
            col_busca_esc,col_limpar_esc=st.columns([4,1])
            busca_escopo=col_busca_esc.text_input("Buscar produto (busca parcial, separe por vírgula para múltiplos)",key="compras_busca_escopo",placeholder="Ex: PC000001, PC000002")
            if col_limpar_esc.button("🔄 Limpar busca",key="compras_limpar_escopo",use_container_width=True):
                st.session_state["compras_busca_escopo_limpar"]=True
                st.rerun()
            if st.session_state.pop("compras_busca_escopo_limpar",False):
                busca_escopo=""
            ja_sel_escopo=st.session_state.get("compras_produtos_escopo",[])
            if busca_escopo:
                termos_esc=[t.strip().upper() for t in busca_escopo.split(",") if t.strip()]
                produtos_filtrados_esc=[p for p in produtos_disp_escopo if any(t in p.upper() for t in termos_esc)]
            else:
                produtos_filtrados_esc=produtos_disp_escopo
            lista_final_esc=sorted(set(ja_sel_escopo+produtos_filtrados_esc[:500]))

            def _on_change_produtos_escopo():
                st.session_state["compras_produtos_escopo_backup"]=st.session_state["compras_produtos_escopo"]

            if "compras_produtos_escopo_backup" not in st.session_state:
                st.session_state["compras_produtos_escopo_backup"]=st.session_state.get("compras_produtos_escopo",[])
            _bkp_prod=[p for p in st.session_state.get("compras_produtos_escopo_backup",[]) if p in lista_final_esc]
            st.session_state["compras_produtos_escopo"]=_bkp_prod
            produtos_sel_escopo=st.multiselect(f"Produto(s) ({len(produtos_filtrados_esc)} encontrados)",
                lista_final_esc,key="compras_produtos_escopo",on_change=_on_change_produtos_escopo,
                placeholder="Selecione os produtos")
            produtos_no_escopo=set(produtos_sel_escopo) if produtos_sel_escopo else set()
        else:
            produtos_no_escopo=None

        if "Catálogo" in escopo_tipo_compras:
            _escopo_compras_texto=f"Catálogo inteiro ({len(df_est_atual)} produtos)"
        elif "Categoria" in escopo_tipo_compras and produtos_no_escopo is not None:
            _escopo_compras_texto=f"{categoria_sel_compras} ({len(produtos_no_escopo)} produtos)"
        elif produtos_no_escopo:
            _escopo_compras_texto=", ".join(sorted(produtos_no_escopo))
        else:
            _escopo_compras_texto="nenhum produto selecionado ainda"
        st.markdown(f'<div class="al-i">📌 Escopo atual: <b>{escopo_tipo_compras}</b> = <b>{_escopo_compras_texto}</b></div>',unsafe_allow_html=True)
        if st.button("💾 Salvar este Escopo",key="btn_salvar_escopo_compras"):
            if st.session_state.cid:
                save_config_compras(st.session_state.cid)
                st.markdown('<div class="al-s">✅ Escopo salvo — vai persistir mesmo após reiniciar.</div>',unsafe_allow_html=True)

        if produtos_no_escopo is not None:
            df_est_atual=df_est_atual[df_est_atual["Produto"].astype(str).isin(produtos_no_escopo)].copy()
            st.markdown(f'<div class="al-s">✅ Escopo aplicado: {len(df_est_atual)} produtos do estoque dentro do critério escolhido.</div>',unsafe_allow_html=True)
            if df_est_atual.empty:
                st.markdown('<div class="al-w">⚠️ Nenhum produto do estoque está dentro deste escopo.</div>',unsafe_allow_html=True)
                st.stop()
                sec("🚛 Cadastro de Fornecedor")
        st.markdown('<div class="al-i">Associe um fornecedor a uma categoria, produtos específicos, ou todo o catálogo — isso atualiza o estoque salvo permanentemente.</div>',unsafe_allow_html=True)

        with st.expander("➕ Associar Fornecedor"):
            nome_forn_cad=st.text_input("Nome do Fornecedor",key="compras_nome_forn_cad")
            tipo_assoc=st.radio("Aplicar a",["🏷️ Categoria/Grupo inteiro","📦 Produto específico","🌐 Catálogo inteiro"],
            horizontal=True,key="compras_tipo_assoc_forn")

            produtos_assoc=None
            if "Categoria" in tipo_assoc:
                if not col_cat_v_compras:
                    st.markdown('<div class="al-w">⚠️ Esta base não tem coluna de Categoria.</div>',unsafe_allow_html=True)
                else:
                    categorias_assoc=sorted(df_v_compras[col_cat_v_compras].dropna().astype(str).unique().tolist())
                    categoria_sel_assoc=st.selectbox("Categoria",categorias_assoc,key="compras_categoria_assoc_sel")
                    produtos_assoc=set(df_v_compras[df_v_compras[col_cat_v_compras].astype(str)==categoria_sel_assoc][col_prod_v].astype(str).unique())
            elif "Produto específico" in tipo_assoc:
                produtos_disp_assoc=sorted(df_est_atual["Produto"].unique().tolist())
                busca_assoc=st.text_input("Buscar produto",key="compras_busca_assoc")
                if busca_assoc:
                    produtos_disp_assoc=[p for p in produtos_disp_assoc if busca_assoc.upper() in p.upper()]
                produtos_sel_assoc=st.multiselect(f"Produto(s) ({len(produtos_disp_assoc)} encontrados)",
                produtos_disp_assoc[:500],key="compras_produtos_sel_assoc",placeholder="Selecione os produtos")
                produtos_assoc=set(produtos_sel_assoc)
            else:
                produtos_assoc=set(df_est_atual["Produto"].astype(str).unique())

            if st.button("💾 Salvar Associação",key="compras_salvar_assoc_forn",use_container_width=True):
                if not nome_forn_cad:
                    st.markdown('<div class="al-w">⚠️ Informe o nome do fornecedor.</div>',unsafe_allow_html=True)
                elif not produtos_assoc:
                    st.markdown('<div class="al-w">⚠️ Nenhum produto selecionado.</div>',unsafe_allow_html=True)
                else:
                    df_estoque_completo_assoc=get_estoque_compras()
                    if df_estoque_completo_assoc is not None:
                        mascara=df_estoque_completo_assoc["Produto"].astype(str).isin(produtos_assoc)
                        df_estoque_completo_assoc.loc[mascara,"Fornecedor"]=nome_forn_cad
                        st.session_state["compras_df_estoque"]=df_estoque_completo_assoc
                        if st.session_state.cid:
                            save_estoque_compras(st.session_state.cid,df_estoque_completo_assoc)
                        st.markdown(f'<div class="al-s">✅ {mascara.sum()} produto(s) associados ao fornecedor "{nome_forn_cad}" e salvos.</div>',unsafe_allow_html=True)
                        st.rerun()
                        st.markdown("---")
            df_estoque_para_excluir=get_estoque_compras()
            if df_estoque_para_excluir is not None:
                fornecedores_existentes=sorted([f for f in df_estoque_para_excluir["Fornecedor"].dropna().astype(str).unique().tolist() if f and f!="Não informado"])
                if fornecedores_existentes:
                    forn_excluir_sel=st.selectbox("Excluir fornecedor (remove a associação, produtos voltam a 'Não informado')",
                    fornecedores_existentes,key="compras_forn_excluir_sel")
                    if st.button("🗑 Excluir Fornecedor",key="compras_btn_excluir_forn",use_container_width=True):
                        mascara_excluir=df_estoque_para_excluir["Fornecedor"].astype(str)==forn_excluir_sel
                        n_afetados=int(mascara_excluir.sum())
                        df_estoque_para_excluir.loc[mascara_excluir,"Fornecedor"]="Não informado"
                        st.session_state["compras_df_estoque"]=df_estoque_para_excluir
                        if st.session_state.cid:
                            save_estoque_compras(st.session_state.cid,df_estoque_para_excluir)
                        if forn_excluir_sel in st.session_state.get("compras_leadtime_fornecedor",{}):
                            del st.session_state["compras_leadtime_fornecedor"][forn_excluir_sel]
                        st.markdown(f'<div class="al-s">✅ Fornecedor "{forn_excluir_sel}" removido — {n_afetados} produto(s) voltaram para "Não informado".</div>',unsafe_allow_html=True)
                        st.rerun()
            sec("2️⃣ Classificação Automática (Classe ABC × Volatilidade)")
        st.markdown('<div class="al-i">Cada produto é classificado por valor (Curva ABC) e por volatilidade da demanda (CV) — cruzando os dois, definimos a política de estoque ideal para cada grupo.</div>',unsafe_allow_html=True)

        with st.spinner("Classificando produtos..."):
            df_v_calc=df_v_compras.copy()
            df_v_calc[col_val_v]=parse_valor_brl(df_v_calc[col_val_v])
            ranking_abc=pareto_analysis(df_v_calc,col_prod_v,col_val_v)
            ranking_abc=ranking_abc.rename(columns={col_prod_v:"Produto"})

            df_v_calc["_periodo_cp"]=pd.to_datetime(df_v_calc[col_data_v],errors="coerce",dayfirst=True).dt.to_period("M")
            cv_por_produto={}
            for prod, g in df_v_calc.groupby(col_prod_v):
                serie=g.groupby("_periodo_cp")[col_val_v].sum()
                if len(serie)>=3 and serie.mean()>0:
                    cv_por_produto[prod]=float(serie.std()/serie.mean()*100)
            df_cv=pd.DataFrame(list(cv_por_produto.items()),columns=["Produto","CV"])

            def classe_5niveis(pct_acum):
                if pct_acum<=70: return "A"
                if pct_acum<=80: return "B"
                if pct_acum<=90: return "C"
                if pct_acum<=97: return "D"
                return "E"
            ranking_abc["classe_abc"]=ranking_abc["pct_acumulado"].apply(classe_5niveis)

            df_class=ranking_abc[["Produto","classe_abc"]].merge(df_cv,on="Produto",how="left")
            def faixa_vol(cv):
                if pd.isna(cv): return "Média"
                if cv<30: return "Baixa"
                if cv<=60: return "Média"
                return "Alta"
            df_class["Volatilidade"]=df_class["CV"].apply(faixa_vol)

        n_a=int((df_class["classe_abc"]=="A").sum())
        n_b=int((df_class["classe_abc"]=="B").sum())
        n_c=int((df_class["classe_abc"]=="C").sum())
        n_d=int((df_class["classe_abc"]=="D").sum())
        n_e=int((df_class["classe_abc"]=="E").sum())
        c_cl1,c_cl2,c_cl3,c_cl4,c_cl5=st.columns(5)
        mc(c_cl1,"Classe A",str(n_a),"g",f"{n_a/len(df_class)*100:.0f}% do catálogo")
        mc(c_cl2,"Classe B",str(n_b),"g")
        mc(c_cl3,"Classe C",str(n_c),"y")
        mc(c_cl4,"Classe D",str(n_d),"y")
        mc(c_cl5,"Classe E",str(n_e),"r")

        sec("Política de Estoque — Dias Mínimo / Máximo por Grupo")
        st.markdown('<div class="al-i">Proposta de partida, editável — ajuste conforme o perfil do negócio.</div>',unsafe_allow_html=True)

        if "compras_matriz" not in st.session_state:
            st.session_state["compras_matriz"]={"A":(15,30),"B":(20,40),"C":(30,55),"D":(40,75),"E":(55,100)}

        valores_padrao_classe={"A":(15,30),"B":(20,40),"C":(30,55),"D":(40,75),"E":(55,100)}
        cols_matriz=st.columns(5)
        for i,classe in enumerate(["A","B","C","D","E"]):
            min_atual,max_atual=st.session_state["compras_matriz"].get(classe,valores_padrao_classe[classe])
            with cols_matriz[i]:
                st.caption(f"Classe {classe}")
                novo_min=st.number_input("Dias mínimo",value=int(min_atual),min_value=1,step=1,key=f"compras_min_{classe}")
                novo_max=st.number_input("Dias máximo",value=int(max_atual),min_value=1,step=1,key=f"compras_max_{classe}")
                st.session_state["compras_matriz"][classe]=(novo_min,novo_max)

        with st.expander("🎯 Exceção de Mínimo/Máximo por Produto Específico"):
            st.markdown('<div class="al-i">Para produtos que precisam fugir da regra geral da Classe (ex: item crítico que deve ficar sempre bem estocado).</div>',unsafe_allow_html=True)
            if "compras_minmax_produto" not in st.session_state: st.session_state["compras_minmax_produto"]={}
            produtos_disp_mm=sorted(df_est_atual["Produto"].unique().tolist())
            busca_mm=st.text_input("Buscar produto",key="compras_busca_produto_mm")
            if busca_mm:
                produtos_disp_mm=[p for p in produtos_disp_mm if busca_mm.upper() in p.upper()]
            produto_sel_mm=st.selectbox(f"Produto ({len(produtos_disp_mm)} encontrados)",produtos_disp_mm[:300],key="compras_produto_sel_mm")
            atual_mm=st.session_state["compras_minmax_produto"].get(produto_sel_mm,(None,None))
            c_mm1,c_mm2,c_mm3=st.columns(3)
            novo_min_mm=c_mm1.number_input("Dias mínimo",value=int(atual_mm[0]) if atual_mm[0] else 15,min_value=1,step=1,key="compras_min_mm_input")
            novo_max_mm=c_mm2.number_input("Dias máximo",value=int(atual_mm[1]) if atual_mm[1] else 30,min_value=1,step=1,key="compras_max_mm_input")
            with c_mm3:
                st.markdown("<br>",unsafe_allow_html=True)
                if st.button("💾 Salvar exceção",key="compras_salvar_mm"):
                    st.session_state["compras_minmax_produto"][produto_sel_mm]=(novo_min_mm,novo_max_mm)
                    st.rerun()
                if produto_sel_mm in st.session_state["compras_minmax_produto"]:
                    if st.button("🗑 Remover exceção",key="compras_remover_mm"):
                        del st.session_state["compras_minmax_produto"][produto_sel_mm]
                        st.rerun()
            if st.session_state["compras_minmax_produto"]:
                st.caption(f"{len(st.session_state['compras_minmax_produto'])} produto(s) com exceção configurada: "+
                ", ".join(f"{k} ({v[0]}-{v[1]}d)" for k,v in list(st.session_state["compras_minmax_produto"].items())[:10]))

        sec("🚚 Lead Time de Entrega por Produto, Categoria e Fornecedor")
        st.markdown('<div class="al-i">Defina o prazo de entrega por produto, categoria e/ou fornecedor. O sistema usa sempre o mais específico: Produto+Fornecedor → Produto → Categoria+Fornecedor → Categoria → Fornecedor → Padrão global.</div>',unsafe_allow_html=True)

        if "compras_leadtime_produto" not in st.session_state:
            _lt_disco=load_leadtime_compras(st.session_state.cid,st.session_state.get("compras_filial_sel")) if st.session_state.cid else None
            if _lt_disco:
                for _k_lt,_v_lt in _lt_disco.items():
                    st.session_state[_k_lt]=_v_lt
                for _k_input_lt in ["compras_leadtime_catalogo_input","compras_leadtime_categoria_input",
                                    "compras_leadtime_produto_input","compras_leadtime_fornecedor_input"]:
                    if _k_input_lt in st.session_state:
                        del st.session_state[_k_input_lt]
        if "compras_leadtime_produto" not in st.session_state: st.session_state["compras_leadtime_produto"]={}
        if "compras_leadtime_categoria" not in st.session_state: st.session_state["compras_leadtime_categoria"]={}
        if "compras_leadtime_fornecedor" not in st.session_state: st.session_state["compras_leadtime_fornecedor"]={}
        if "compras_leadtime_catalogo" not in st.session_state: st.session_state["compras_leadtime_catalogo"]=15

        _opcoes_nivel_lead=["🏷️ Categoria/Grupo inteiro","📦 Produto específico","🚛 Fornecedor","🌐 Catálogo inteiro"]

        def _on_change_nivel_lead():
            st.session_state["compras_nivel_lead_backup"]=st.session_state["compras_nivel_lead"]

        if "compras_nivel_lead_backup" not in st.session_state:
            st.session_state["compras_nivel_lead_backup"]=st.session_state.get("compras_nivel_lead",_opcoes_nivel_lead[0])
        st.session_state["compras_nivel_lead"]=st.session_state["compras_nivel_lead_backup"]

        nivel_lead=st.radio("Aplicar a",_opcoes_nivel_lead,horizontal=True,
        key="compras_nivel_lead",on_change=_on_change_nivel_lead)

        fornecedores_disp_lead=sorted(df_est_atual["Fornecedor"].dropna().astype(str).unique().tolist()) if "Fornecedor" in df_est_atual.columns else []

        if "Categoria" in nivel_lead:
            if not col_cat_v_compras:
                st.markdown('<div class="al-w">⚠️ Esta base não tem coluna de Categoria.</div>',unsafe_allow_html=True)
            else:
                categorias_disp_lead=sorted(df_v_compras[col_cat_v_compras].dropna().astype(str).unique().tolist())
                cat_sel_lead=st.selectbox("Categoria",categorias_disp_lead,key="compras_categoria_lead_sel")
                forn_cat_lead=st.selectbox("Fornecedor (opcional — deixe em branco para aplicar a todos)",["(todos)"]+fornecedores_disp_lead,key="compras_forn_cat_lead")
                atual_cat=st.session_state["compras_leadtime_categoria"].get((cat_sel_lead,forn_cat_lead),15)
                _key_cat_input=f"compras_leadtime_categoria_input__{cat_sel_lead}__{forn_cat_lead}"
                st.session_state[_key_cat_input]=int(atual_cat)
                novo_cat=st.number_input("Lead time (dias)",min_value=1,step=1,key=_key_cat_input)
                if st.button("💾 Salvar",key="compras_salvar_categoria"):
                    st.session_state["compras_leadtime_categoria"][(cat_sel_lead,forn_cat_lead)]=novo_cat
                    if st.session_state.cid:
                        save_leadtime_compras(st.session_state.cid,st.session_state.get("compras_filial_sel"))
                    st.rerun()
                if st.session_state["compras_leadtime_categoria"]:
                    st.caption("Configurados: "+", ".join(f"{k[0]}/{k[1]} ({v}d)" for k,v in st.session_state["compras_leadtime_categoria"].items()))

        elif "Produto" in nivel_lead and "Fornecedor" not in nivel_lead:
            produtos_disp_lead=sorted(df_est_atual["Produto"].unique().tolist())
            busca_lead=st.text_input("Buscar produto",key="compras_busca_produto_lead")
            if busca_lead:
                produtos_disp_lead=[p for p in produtos_disp_lead if busca_lead.upper() in p.upper()]
            produtos_sel_lead=st.multiselect(f"Produto(s) ({len(produtos_disp_lead)} encontrados)",produtos_disp_lead[:500],key="compras_produtos_sel_lead")
            forn_prod_lead=st.selectbox("Fornecedor (opcional)",["(todos)"]+fornecedores_disp_lead,key="compras_forn_prod_lead")
            if "compras_leadtime_produto_input" not in st.session_state:
                st.session_state["compras_leadtime_produto_input"]=15
            novo_p=st.number_input("Lead time (dias)",min_value=1,step=1,key="compras_leadtime_produto_input")
            if st.button("💾 Salvar",key="compras_salvar_produto"):
                for p_sel in produtos_sel_lead:
                    st.session_state["compras_leadtime_produto"][(p_sel,forn_prod_lead)]=novo_p
                if st.session_state.cid:
                    save_leadtime_compras(st.session_state.cid,st.session_state.get("compras_filial_sel"))
                st.rerun()
            if st.session_state["compras_leadtime_produto"]:
                df_lt_config=pd.DataFrame([(k[0],k[1],v) for k,v in st.session_state["compras_leadtime_produto"].items()],columns=["Produto","Fornecedor","Lead Time (dias)"])
                st.dataframe(df_lt_config,use_container_width=True,height=min(300,45+35*len(df_lt_config)))

        elif "Fornecedor" in nivel_lead:
            if not fornecedores_disp_lead:
                st.markdown('<div class="al-w">⚠️ Nenhum fornecedor encontrado no estoque importado.</div>',unsafe_allow_html=True)
            else:
                forn_sel_lead=st.selectbox("Fornecedor",fornecedores_disp_lead,key="compras_forn_sel_lead")
                atual_forn=st.session_state["compras_leadtime_fornecedor"].get(forn_sel_lead,15)
                _key_forn_input=f"compras_leadtime_fornecedor_input__{forn_sel_lead}"
                st.session_state[_key_forn_input]=int(atual_forn)
                novo_forn=st.number_input("Lead time (dias)",min_value=1,step=1,key=_key_forn_input)
                if st.button("💾 Salvar",key="compras_salvar_fornecedor"):
                    st.session_state["compras_leadtime_fornecedor"][forn_sel_lead]=novo_forn
                    if st.session_state.cid:
                        save_leadtime_compras(st.session_state.cid,st.session_state.get("compras_filial_sel"))
                    st.rerun()
                if st.session_state["compras_leadtime_fornecedor"]:
                    st.caption("Fornecedores configurados: "+", ".join(f"{k} ({v}d)" for k,v in st.session_state["compras_leadtime_fornecedor"].items()))

        else:
            _key_cat_geral_input=f"compras_leadtime_catalogo_input__{st.session_state.get('compras_filial_sel','todas')}"
            st.session_state[_key_cat_geral_input]=int(st.session_state["compras_leadtime_catalogo"])
            novo_cat_geral=st.number_input("Lead time único para todo o catálogo (dias)",min_value=1,step=1,key=_key_cat_geral_input)
            st.session_state["compras_leadtime_catalogo"]=novo_cat_geral
            if st.button("💾 Salvar",key="compras_salvar_catalogo_lead"):
                if st.session_state.cid:
                    save_leadtime_compras(st.session_state.cid,st.session_state.get("compras_filial_sel"))
                st.rerun()

        lead_padrao=15

        if False:
            if not col_cat_v_compras:
                st.markdown('<div class="al-w">⚠️ Esta base não tem coluna de Categoria — use "Produto específico" ou "Catálogo inteiro".</div>',unsafe_allow_html=True)
            else:
                categorias_disp_lead=sorted(df_v_compras[col_cat_v_compras].dropna().astype(str).unique().tolist())
                categoria_sel_lead=st.selectbox("Categoria",categorias_disp_lead,key="compras_categoria_lead_sel")
                atual_cat=st.session_state["compras_leadtime_categoria"].get(categoria_sel_lead,15)
                novo_cat=st.number_input("Lead time (dias)",value=int(atual_cat),min_value=1,step=1,key="compras_leadtime_categoria_input")
                if st.button("💾 Salvar lead time desta categoria",key="compras_salvar_categoria"):
                    st.session_state["compras_leadtime_categoria"][categoria_sel_lead]=novo_cat
                    st.rerun()
                if st.session_state["compras_leadtime_categoria"]:
                    st.caption("Categorias configuradas: "+", ".join(f"{k} ({v}d)" for k,v in st.session_state["compras_leadtime_categoria"].items()))

        elif False:
            df_estoque_completo=get_estoque_compras()
            produtos_disp_lead=sorted(df_estoque_completo["Produto"].unique().tolist()) if df_estoque_completo is not None else sorted(df_est_atual["Produto"].unique().tolist())
            busca_lead=st.text_input("Buscar produto",key="compras_busca_produto_lead")
            if busca_lead:
                produtos_disp_lead=[p for p in produtos_disp_lead if busca_lead.upper() in p.upper()]
            produtos_sel_lead=st.multiselect(f"Produto(s) ({len(produtos_disp_lead)} encontrados) — pode escolher vários de uma vez",
            produtos_disp_lead[:500],key="compras_produtos_sel_lead",placeholder="Selecione um ou mais produtos")
            novo_p=st.number_input("Lead time (dias) — aplicado a todos os selecionados acima",value=15,min_value=1,step=1,key="compras_leadtime_produto_input")
            if st.button("💾 Salvar lead time para os produtos selecionados",key="compras_salvar_produto"):
                for p_sel in produtos_sel_lead:
                    st.session_state["compras_leadtime_produto"][p_sel]=novo_p
                st.rerun()

            if st.session_state["compras_leadtime_produto"]:
                st.markdown(f'<div class="al-s">✅ {len(st.session_state["compras_leadtime_produto"])} produto(s) configurado(s) até agora:</div>',unsafe_allow_html=True)
                df_lt_config=pd.DataFrame(list(st.session_state["compras_leadtime_produto"].items()),columns=["Produto","LeadTime (dias)"])
                st.dataframe(df_lt_config,use_container_width=True,height=min(300,45+35*len(df_lt_config)))
                produto_remover=st.selectbox("Remover configuração de um produto",["—"]+df_lt_config["Produto"].tolist(),key="compras_remover_lt_sel")
                if produto_remover!="—" and st.button("🗑 Remover",key="compras_remover_lt_btn"):
                    del st.session_state["compras_leadtime_produto"][produto_remover]
                    st.rerun()

        else:
            pass

        if False:
            cols_lc=st.columns(5)
            for i_c,classe_lt in enumerate(["A","B","C","D","E"]):
                with cols_lc[i_c]:
                    atual=st.session_state["compras_leadtime_classe"].get(classe_lt,15)
                    novo=st.number_input(f"Classe {classe_lt}",value=int(atual),min_value=1,step=1,key=f"leadtime_classe_{classe_lt}")
                    st.session_state["compras_leadtime_classe"][classe_lt]=novo

        elif False:
            col_desc_v=next((c for c in df_v_compras.columns if c.strip().lower() in ["descricao","descrição"]),None)
            if not col_desc_v:
                st.markdown('<div class="al-w">⚠️ Esta base não tem coluna de Descrição para identificar famílias.</div>',unsafe_allow_html=True)
            else:
                df_fam_map=df_v_compras[[col_prod_v,col_desc_v]].drop_duplicates(col_prod_v).copy()
                df_fam_map["_Familia"]=df_fam_map[col_desc_v].apply(extrair_familia_produto)
                familias_disp=sorted(df_fam_map["_Familia"].unique().tolist())
                familia_sel=st.selectbox("Família",familias_disp,key="compras_familia_sel")
                atual_f=st.session_state["compras_leadtime_familia"].get(familia_sel,15)
                novo_f=st.number_input("Lead time (dias)",value=int(atual_f),min_value=1,step=1,key="compras_leadtime_familia_input")
                if st.button("💾 Salvar lead time desta família",key="compras_salvar_familia"):
                    st.session_state["compras_leadtime_familia"][familia_sel]=novo_f
                    st.session_state["compras_fam_map"]=df_fam_map
                    st.rerun()
                if st.session_state["compras_leadtime_familia"]:
                    st.caption("Famílias configuradas: "+", ".join(f"{k} ({v}d)" for k,v in st.session_state["compras_leadtime_familia"].items()))
            sec("4️⃣ Rodar o Motor de Compras")
            st.markdown('<div class="al-i">Calcula a demanda prevista de cada produto (usando o motor de ML já validado), cruza com estoque e lead time, e gera a recomendação de compra — dentro do mesmo escopo definido no Passo 1.</div>',unsafe_allow_html=True)
        horizonte_meses=st.selectbox("Horizonte do calendário de compras",list(range(1,13)),index=5,
        format_func=lambda v:f"{v} {'mês' if v==1 else 'meses'}",key="compras_horizonte")
        

        senha_rodar_compras=st.text_input("Senha master para rodar",type="password",key="senha_rodar_compras")
        if st.button("🚀 Rodar Motor de Compras",use_container_width=True,key="compras_btn_rodar"):
            if senha_rodar_compras!=SENHA_MASTER:
                st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi executado.</div>',unsafe_allow_html=True)
                st.stop()
            if st.session_state.cid: save_config_compras(st.session_state.cid)
            produtos_com_estoque=set(df_est_atual["Produto"].astype(str))
            ranking_calc=ranking_abc[ranking_abc["Produto"].astype(str).isin(produtos_com_estoque)].copy()
            produtos_rodar=ranking_calc["Produto"].tolist()

            st.markdown(f'<div class="al-i">🔎 {len(produtos_rodar)} produtos elegíveis para cálculo.</div>',unsafe_allow_html=True)
            pb_compras=st.progress(0)
            texto_pb_compras=st.empty()
            modelos_compras=[m for m,ok in MODELOS_ML.items() if ok]
            linhas_compras=[]
            linhas_calendario=[]
            mapa_categoria_produto=dict(zip(df_v_compras[col_prod_v], df_v_compras[col_cat_v_compras])) if col_cat_v_compras else {}

            _lt_rows=[]
            for (_p_lt,_f_lt),_v_lt in st.session_state.get("compras_leadtime_produto",{}).items():
                _lt_rows.append({"Produto_ou_Categoria":_p_lt,"Fornecedor":"" if _f_lt=="(todos)" else _f_lt,"LeadTimeDias":_v_lt})
            for (_c_lt,_f_lt),_v_lt in st.session_state.get("compras_leadtime_categoria",{}).items():
                _lt_rows.append({"Produto_ou_Categoria":_c_lt,"Fornecedor":"" if _f_lt=="(todos)" else _f_lt,"LeadTimeDias":_v_lt})
            for _f_lt,_v_lt in st.session_state.get("compras_leadtime_fornecedor",{}).items():
                _lt_rows.append({"Produto_ou_Categoria":"","Fornecedor":_f_lt,"LeadTimeDias":_v_lt})
            st.session_state["compras_leadtime_tabela"]=pd.DataFrame(_lt_rows) if _lt_rows else pd.DataFrame()
            st.session_state["compras_lead_padrao"]=st.session_state.get("compras_leadtime_catalogo",15)

            # "Todas as filiais" NÃO recalcula nada — só junta os resultados já salvos de cada loja,
            # rodados individualmente antes. Isso garante que o consolidado é sempre idêntico à soma
            # exata dos resultados de cada loja, sem risco de divergência entre os dois.
            _visao_consolidada=st.session_state.get("compras_filial_sel","(Todas as filiais)")=="(Todas as filiais)"
            _cfg_ml=None
            _filiais_lista=[]
            _consolidado_pronto=False
            if _visao_consolidada and _col_fil_est_atual:
                _filiais_lista=sorted(df_est_atual[_col_fil_est_atual].dropna().astype(str).unique().tolist())
                _faltando_lojas=[]
                _resultados_lojas=[]
                _calendarios_lojas=[]
                for _fil_x in _filiais_lista:
                    _res_fil,_cal_fil,_morto_fil=load_resultado_compras(st.session_state.cid,_fil_x) if st.session_state.cid else (None,None,None)
                    if _res_fil is not None and not _res_fil.empty:
                        _resultados_lojas.append(_res_fil)
                        if _cal_fil is not None and not _cal_fil.empty:
                            _calendarios_lojas.append(_cal_fil)
                    else:
                        _faltando_lojas.append(_fil_x)
                if _faltando_lojas:
                    st.markdown(f'<div class="al-w">⚠️ Faltam rodar essas lojas antes de ver o consolidado: <b>{", ".join(_faltando_lojas)}</b>. Selecione cada uma no seletor de Filial acima e clique em "Rodar Motor de Compras" — "Todas as filiais" só junta o que já foi calculado, sem recalcular nada.</div>',unsafe_allow_html=True)
                    st.session_state["compras_resultado"]=None
                    st.session_state["compras_calendario"]=None
                else:
                    st.session_state["compras_resultado"]=pd.concat(_resultados_lojas,ignore_index=True)
                    st.session_state["compras_calendario"]=pd.concat(_calendarios_lojas,ignore_index=True) if _calendarios_lojas else pd.DataFrame()
                    # Salva o consolidado em disco também — sem isso, trocar de filial e voltar
                    # recarrega um arquivo velho (de antes dessa mudança), com números desatualizados.
                    if st.session_state.cid:
                        save_resultado_compras(st.session_state.cid,
                            st.session_state["compras_resultado"],
                            st.session_state["compras_calendario"],
                            None,"(Todas as filiais)")
                    addlog(f"Motor de Compras: consolidado de {len(_filiais_lista)} lojas ({', '.join(_filiais_lista)}), {len(st.session_state['compras_resultado'])} linhas — sem recálculo.")
                    st.markdown(f'<div class="al-s">✅ Consolidado montado a partir de {len(_filiais_lista)} lojas já calculadas — sem recalcular nada.</div>',unsafe_allow_html=True)
                _consolidado_pronto=True
                produtos_rodar=[]  # não processa nada no loop abaixo — já foi tudo carregado
            elif st.session_state.cid:
                _cfg_ml=load_cfgml_resultado(st.session_state.cid,st.session_state.get("compras_filial_sel"))

            if not _consolidado_pronto:
                if _cfg_ml is not None and not _cfg_ml.empty and "_ProdutoUnico" in _cfg_ml.columns:
                    _match_pct=len(set(str(p) for p in produtos_rodar)&set(_cfg_ml["_ProdutoUnico"].astype(str)))/len(produtos_rodar)*100 if produtos_rodar else 0
                else:
                    _match_pct=0
                if _match_pct<80:
                    st.markdown(f'<div class="al-w">⚠️ Só {_match_pct:.0f}% dos {len(produtos_rodar)} produtos batem com a previsão salva no Painel de Configuração ML. O resto será recalculado agora (mais lento).</div>',unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="al-s">✅ {_match_pct:.0f}% dos produtos batem com a previsão calibrada já salva.</div>',unsafe_allow_html=True)

            # Pré-filtra a base de vendas por loja UMA VEZ (não a cada produto) — evita refiltrar
            # a mesma base de vendas 700 vezes quando o escopo é "Todas as filiais".
            _df_v_por_filial={}
            if _visao_consolidada and _col_fil_v_compras:
                for _fil_x in _filiais_lista:
                    _df_v_por_filial[_fil_x]=df_v_calc[df_v_calc[_col_fil_v_compras].astype(str)==_fil_x]

            # Pré-processa a tabela de lead time UMA VEZ (não a cada produto/loja) — o upper()/strip()
            # nas colunas inteiras é caro repetido centenas de vezes.
            _df_lt_pre=st.session_state.get("compras_leadtime_tabela",pd.DataFrame())
            if not _df_lt_pre.empty:
                _df_lt_pre=_df_lt_pre.copy()
                _df_lt_pre["_pc"]=_df_lt_pre["Produto_ou_Categoria"].astype(str).str.strip().str.upper()
                _df_lt_pre["_forn"]=_df_lt_pre["Fornecedor"].astype(str).str.strip().str.upper()

            for idx_c,prod_c in enumerate(produtos_rodar):
                texto_pb_compras.caption(f"Calculando {idx_c+1} de {len(produtos_rodar)}: {str(prod_c)[:50]}")

                linha_est_prod=df_est_atual[df_est_atual["Produto"]==prod_c]
                if linha_est_prod.empty:
                    pb_compras.progress((idx_c+1)/len(produtos_rodar)); continue

                # Monta a lista de "unidades" a calcular pra esse produto: 1 por loja (visão consolidada,
                # cada uma com seu próprio estoque e sua própria previsão) ou 1 unidade só (visão de loja única).
                _unidades=[]
                if _visao_consolidada and _col_fil_est_atual:
                    for _fil_x in sorted(linha_est_prod[_col_fil_est_atual].dropna().astype(str).unique().tolist()):
                        _unidades.append((_fil_x,linha_est_prod[linha_est_prod[_col_fil_est_atual].astype(str)==_fil_x],_cfg_ml_por_filial.get(_fil_x)))
                else:
                    _unidades.append((None,linha_est_prod,_cfg_ml))

                for _filial_u,linha_est,_cfg_ml_u in _unidades:
                    if _visao_consolidada and _col_fil_v_compras and _filial_u is not None:
                        serie_c=serie_mensal_produto(df_v_calc[df_v_calc[_col_fil_v_compras].astype(str)==_filial_u],col_prod_v,prod_c,col_data_v,col_val_v)
                    else:
                        serie_c=serie_mensal_produto(df_v_calc,col_prod_v,prod_c,col_data_v,col_val_v)
                    if len(serie_c)<3:
                        continue
                    # Tenta usar previsão já calculada no Painel de Configuração ML (dessa loja, se consolidado)
                    demanda_mes_valor=None
                    prev_cache_c=None
                    melhor_c=None
                    fonte_previsao_c=None

                    # 1. Prioridade: Painel de Configuração ML (com sazonalidade, promoções, reajustes)
                    if _cfg_ml_u is not None and not _cfg_ml_u.empty and "_ProdutoUnico" in _cfg_ml_u.columns:
                        _row_cfg=_cfg_ml_u[_cfg_ml_u["_ProdutoUnico"]==prod_c]
                        if not _row_cfg.empty and _row_cfg.iloc[0]["status"]=="ok":
                            _prev=_row_cfg.iloc[0]["previsao"]
                            if _prev is not None and len(_prev)>0:
                                demanda_mes_valor=max(0.0,float(_prev[0]))
                                melhor_c=_row_cfg.iloc[0]["modelo_escolhido"]
                                prev_cache_c=_prev
                                fonte_previsao_c="1-Config ML (Painel)"

                    # 2. Fallback final: roda ML na hora (nível "ML por Produto Top N%" removido de propósito — não usado aqui)
                    if demanda_mes_valor is None:
                        melhor_c,_=melhor_modelo(serie_c,modelos_compras)
                        proj_c=treinar(serie_c,melhor_c,1)
                        if proj_c is None:
                            continue
                        demanda_mes_valor=max(0.0,float(proj_c.iloc[0]))
                        fonte_previsao_c="2-Recalculado agora (sem calibrações)"

                    if linha_est.empty:
                        continue
                    estoque_c=float(linha_est["EstoqueAtual"].sum())
                    custo_c=float(linha_est["CustoUnitario"].mean())
                    forn_c=linha_est["Fornecedor"].iloc[0]

                    col_qtd_v_local=next((c for c in df_v_compras.columns if c.strip().lower() in ["quantidade","qtd"]),None)
                    _df_v_preco=_df_v_por_filial.get(_filial_u,df_v_calc) if (_visao_consolidada and _filial_u is not None) else df_v_calc
                    if col_qtd_v_local:
                        serie_qtd_prod=_df_v_preco[_df_v_preco[col_prod_v]==prod_c][col_qtd_v_local]
                        serie_val_prod=_df_v_preco[_df_v_preco[col_prod_v]==prod_c][col_val_v]
                        qtd_total_prod=pd.to_numeric(serie_qtd_prod,errors="coerce").sum()
                        preco_venda_c=float(serie_val_prod.sum()/qtd_total_prod) if qtd_total_prod>0 else custo_c
                    else:
                        preco_venda_c=custo_c

                    cls_row=df_class[df_class["Produto"]==prod_c]
                    classe_c=cls_row["classe_abc"].iloc[0] if not cls_row.empty else "SV"
                    vol_c=cls_row["Volatilidade"].iloc[0] if not cls_row.empty else "Sem dados"

                    minmax_produto_map=st.session_state.get("compras_minmax_produto",{})
                    if prod_c in minmax_produto_map:
                        min_dias_c,max_dias_c=minmax_produto_map[prod_c]
                    elif classe_c=="SV":
                        min_dias_c,max_dias_c=(0,0)
                    else:
                        min_dias_c,max_dias_c=st.session_state["compras_matriz"].get(classe_c,(20,40))

                    # Lead time: Produto+Fornecedor > Produto > Categoria+Fornecedor > Categoria > Fornecedor > Padrão
                    lead_padrao_val=st.session_state.get("compras_lead_padrao",15)
                    lead_c=lead_padrao_val
                    if not _df_lt_pre.empty:
                        df_lt_=_df_lt_pre
                        prod_up=str(prod_c).strip().upper()
                        forn_up=str(forn_c).strip().upper()
                        categoria_real_c=mapa_categoria_produto.get(prod_c,"")
                        cat_up=str(categoria_real_c).strip().upper()
                        # 1. Produto + Fornecedor
                        m=df_lt_[(df_lt_["_pc"]==prod_up)&(df_lt_["_forn"]==forn_up)]
                        if not m.empty: lead_c=float(m.iloc[0]["LeadTimeDias"])
                        # 2. Produto
                        elif not df_lt_[(df_lt_["_pc"]==prod_up)&(df_lt_["_forn"]=="")].empty:
                            lead_c=float(df_lt_[(df_lt_["_pc"]==prod_up)&(df_lt_["_forn"]=="")].iloc[0]["LeadTimeDias"])
                        # 3. Categoria + Fornecedor
                        elif not df_lt_[(df_lt_["_pc"]==cat_up)&(df_lt_["_forn"]==forn_up)].empty:
                            lead_c=float(df_lt_[(df_lt_["_pc"]==cat_up)&(df_lt_["_forn"]==forn_up)].iloc[0]["LeadTimeDias"])
                        # 4. Categoria
                        elif not df_lt_[(df_lt_["_pc"]==cat_up)&(df_lt_["_forn"]=="")].empty:
                            lead_c=float(df_lt_[(df_lt_["_pc"]==cat_up)&(df_lt_["_forn"]=="")].iloc[0]["LeadTimeDias"])
                        # 5. Fornecedor
                        elif not df_lt_[(df_lt_["_pc"]=="")&(df_lt_["_forn"]==forn_up)].empty:
                            lead_c=float(df_lt_[(df_lt_["_pc"]=="")&(df_lt_["_forn"]==forn_up)].iloc[0]["LeadTimeDias"])

                    # Converte demanda em R$/mês (de venda) para unidades, usando o PREÇO DE VENDA médio do produto
                    # (não o custo — senão infla a quantidade de unidades, já que custo < preço de venda)
                    preco_ref=preco_venda_c if preco_venda_c>0 else (custo_c if custo_c>0 else 1.0)
                    demanda_unid_mes=demanda_mes_valor/preco_ref if preco_ref>0 else 0
                    demanda_unid_dia=demanda_unid_mes/30 if demanda_unid_mes else 0
                    if prev_cache_c is not None and len(prev_cache_c)>=horizonte_meses:
                        proj_horizonte=pd.Series(prev_cache_c[:horizonte_meses])
                    else:
                        proj_horizonte=treinar(serie_c,melhor_c,horizonte_meses)
                    if proj_horizonte is not None:
                        estoque_sim=estoque_c
                        hoje_cal=pd.Timestamp.now().normalize()
                        colchao_dias=max_dias_c-min_dias_c
                        total_dias_horizonte=horizonte_meses*30
                        dia_atual=0
                        while dia_atual<total_dias_horizonte:
                            mes_idx=min(int(dia_atual//30),horizonte_meses-1)
                            valor_mes_i=max(0.0,float(proj_horizonte.iloc[mes_idx]))
                            demanda_unid_mes_i=valor_mes_i/preco_ref if preco_ref>0 else 0
                            demanda_dia_i=demanda_unid_mes_i/30 if demanda_unid_mes_i else 0
                            dia_atual+=1
                            if demanda_dia_i<=0:
                                continue
                            estoque_sim-=demanda_dia_i
                            cobertura_sim=safe(max(estoque_sim,0),demanda_dia_i,999)
                            if estoque_sim<=0 or cobertura_sim<min_dias_c:
                                qtd_evento=max(0.0,(max_dias_c*demanda_dia_i)-max(estoque_sim,0))
                                if qtd_evento>0:
                                    data_evento=hoje_cal+pd.Timedelta(days=dia_atual)
                                    if dia_atual<=7: urgencia_ev="🚨 Urgente (até 7 dias)"
                                    elif dia_atual<=30: urgencia_ev="🔴 Prioritário (até 30 dias)"
                                    else: urgencia_ev="🟡 Planejado"
                                    ciclo_pedido=safe(colchao_dias,demanda_dia_i,999) if demanda_dia_i>0 else 999
                                    freq_estimada=round(365/ciclo_pedido,1) if ciclo_pedido<999 and ciclo_pedido>0 else None
                                    linhas_calendario.append({
                                        "Produto":prod_c,"Filial":_filial_u if _filial_u else st.session_state.get("compras_filial_sel"),
                                        "DataEvento":data_evento.strftime("%d/%m/%Y"),
                                        "MesHorizonte":mes_idx+1,"QtdComprar":round(qtd_evento,0),
                                        "ValorCompra":round(qtd_evento*custo_c,2),"Urgencia":urgencia_ev,
                                        "Classe":classe_c,"Fornecedor":forn_c,
                                        "CicloDias":round(ciclo_pedido,0) if ciclo_pedido<999 else None,
                                        "FrequenciaAno":freq_estimada,
                                    })
                                    estoque_sim+=qtd_evento

                    cobertura_dias=safe(estoque_c,demanda_unid_dia,999) if demanda_unid_dia>0 else 999

                    if demanda_unid_dia<=0:
                        status_c="⚪ Sem demanda prevista"
                        qtd_comprar=0.0
                    elif cobertura_dias<lead_c:
                        status_c="🚨 RUPTURA IMINENTE"
                        qtd_comprar=max(0.0,(max_dias_c*demanda_unid_dia)-estoque_c)
                    elif cobertura_dias<min_dias_c:
                        status_c="🔴 Comprar agora"
                        qtd_comprar=max(0.0,(max_dias_c*demanda_unid_dia)-estoque_c)
                    elif cobertura_dias>max_dias_c:
                        status_c="🟡 Estoque excessivo"
                        qtd_comprar=0.0
                    else:
                        status_c="🟢 OK"
                        qtd_comprar=0.0

                    giro_produto=safe(365,cobertura_dias,0) if cobertura_dias not in (None,999) and cobertura_dias>0 else 0.0
                    giro_alvo_produto=safe(365,(min_dias_c+max_dias_c)/2,0)

                    linhas_compras.append({
                        "Produto":prod_c,"Filial":_filial_u if _filial_u else st.session_state.get("compras_filial_sel"),
                        "Fornecedor":forn_c,"Classe":classe_c,"Volatilidade":vol_c,
                        "EstoqueAtual":round(estoque_c,1),"DemandaPrevMes(un)":round(demanda_unid_mes,1),
                        "CoberturaDias":round(cobertura_dias,1) if cobertura_dias<999 else None,
                        "MinDias":min_dias_c,"MaxDias":max_dias_c,"LeadTimeDias":lead_c,
                        "GiroAtual":round(giro_produto,2),"GiroAlvo":round(giro_alvo_produto,2),
                        "Status":status_c,"QtdSugerida":round(qtd_comprar,0),
                        "CustoUnitario":custo_c,"ValorSugerido":round(qtd_comprar*custo_c,2),
                        "FontePrevisao":fonte_previsao_c,"ModeloUsado":melhor_c,
                    })
                pb_compras.progress((idx_c+1)/len(produtos_rodar))
            pb_compras.empty(); texto_pb_compras.empty()
            if not _consolidado_pronto:
                st.session_state["compras_resultado"]=pd.DataFrame(linhas_compras)
                st.session_state["compras_calendario"]=pd.DataFrame(linhas_calendario)
                if st.session_state.cid:
                    save_resultado_compras(st.session_state.cid,
                        st.session_state["compras_resultado"],
                        st.session_state["compras_calendario"],
                        st.session_state.get("compras_morto"),
                        st.session_state.get("compras_filial_sel"))
                addlog(f"Motor de Compras: {len(linhas_compras)} produtos calculados")

            if st.session_state.get("compras_resultado") is not None and not st.session_state["compras_resultado"].empty and "FontePrevisao" in st.session_state["compras_resultado"].columns:
                _cont_fonte=st.session_state["compras_resultado"]["FontePrevisao"].value_counts()
                _total_fonte=int(_cont_fonte.sum())
                _partes_fonte=[]
                for _fonte_nome,_fonte_n in _cont_fonte.items():
                    _partes_fonte.append(f"{_fonte_nome}: {int(_fonte_n)} ({_fonte_n/_total_fonte*100:.0f}%)")
                st.markdown('<div class="al-i">🔎 <b>Fonte da previsão usada:</b> '+" · ".join(_partes_fonte)+'</div>',unsafe_allow_html=True)
        
    with tab_resultado:

        

        df_res_compras=st.session_state.get("compras_resultado")
        if df_res_compras is None and st.session_state.cid:
            _res_disco,_cal_disco,_morto_disco=load_resultado_compras(st.session_state.cid,st.session_state.get("compras_filial_sel"))
            if _res_disco is not None:
                st.session_state["compras_resultado"]=_res_disco
                st.session_state["compras_calendario"]=_cal_disco
                st.session_state["compras_morto"]=_morto_disco
                df_res_compras=_res_disco
        if df_res_compras is not None and not df_res_compras.empty:
            sec("⚡ Radar de Estoque — Ação Imediata")
            n_ruptura=int((df_res_compras["Status"]=="🚨 RUPTURA IMINENTE").sum())
            n_comprar=int((df_res_compras["Status"]=="🔴 Comprar agora").sum())
            n_ok=int((df_res_compras["Status"]=="🟢 OK").sum())
            n_excesso=int((df_res_compras["Status"]=="🟡 Estoque excessivo").sum())

            # Quando a visão tem mais de 1 loja, cada card conta "casos" (produto+loja), não SKUs.
            # Mostra também a contagem de SKUs únicos e quantos estão nesse status em TODAS as lojas
            # ao mesmo tempo — pra não confundir "498 casos" com "498 produtos diferentes".
            _sub_ruptura=_sub_comprar=_sub_ok=_sub_excesso=""
            if "Filial" in df_res_compras.columns and df_res_compras["Filial"].nunique()>1:
                _total_lojas_por_produto=df_res_compras.groupby("Produto")["Filial"].nunique()
                def _sku_sub(status_str):
                    _sub_df=df_res_compras[df_res_compras["Status"]==status_str]
                    if _sub_df.empty: return "0 SKUs únicos"
                    _skus=_sub_df["Produto"].nunique()
                    _status_por_produto=_sub_df.groupby("Produto")["Filial"].nunique()
                    _repetidos=int((_status_por_produto>=2).sum())
                    _em_todas=int((_status_por_produto==_total_lojas_por_produto.reindex(_status_por_produto.index)).sum())
                    return f"{_skus} SKUs únicos • {_repetidos} em 2+ lojas ({_em_todas} em todas)"
                _sub_ruptura=_sku_sub("🚨 RUPTURA IMINENTE")
                _sub_comprar=_sku_sub("🔴 Comprar agora")
                _sub_ok=_sku_sub("🟢 OK")
                _sub_excesso=_sku_sub("🟡 Estoque excessivo")

            c_r0,c_r1,c_r2,c_r3,c_r4=st.columns(5)
            mc(c_r0,"🚨 Ruptura Iminente",str(n_ruptura),"r",_sub_ruptura)
            mc(c_r1,"🔴 Comprar Agora",str(n_comprar),"r",_sub_comprar)
            mc(c_r2,"🟢 OK",str(n_ok),"g",_sub_ok)
            mc(c_r3,"🟡 Estoque Excessivo",str(n_excesso),"y",_sub_excesso)
            mc(c_r4,"💰 Total Sugerido",fmt(df_res_compras["ValorSugerido"].sum()),"b")

            # Relatório de produtos repetidos entre lojas (mesmo status em 2+ filiais) — só faz
            # sentido na visão consolidada, onde um SKU pode aparecer mais de uma vez.
            if "Filial" in df_res_compras.columns and df_res_compras["Filial"].nunique()>1:
                def _gerar_html_repetidos(_df):
                    _status_info=[("🚨 RUPTURA IMINENTE","Ruptura Iminente","#F85149"),
                                  ("🔴 Comprar agora","Comprar Agora","#F85149"),
                                  ("🟡 Estoque excessivo","Estoque Excessivo","#FFB627"),
                                  ("🟢 OK","OK","#00D4AA")]
                    _partes=[]
                    for _sval,_slabel,_scor in _status_info:
                        _sub=_df[_df["Status"]==_sval]
                        if _sub.empty: continue
                        _agg=_sub.groupby("Produto").agg(
                            Lojas=("Filial",lambda x: ", ".join(sorted(x.astype(str)))),
                            QtdLojas=("Filial","nunique"),
                            Fornecedor=("Fornecedor","first"),
                            ValorTotal=("ValorSugerido","sum"),
                        ).reset_index()
                        _agg=_agg[_agg["QtdLojas"]>=2].sort_values(["QtdLojas","ValorTotal"],ascending=[False,False])
                        if _agg.empty: continue
                        _linhas="".join(
                            f'<tr><td>{r.Produto}</td><td>{r.Fornecedor}</td><td>{r.Lojas}</td>'
                            f'<td style="text-align:center">{r.QtdLojas}</td>'
                            f'<td style="text-align:right">R$ {r.ValorTotal:,.2f}</td></tr>'
                            for r in _agg.itertuples()
                        )
                        _partes.append(f'''
                        <h2 style="color:{_scor};border-bottom:2px solid {_scor};padding-bottom:6px;margin-top:36px">
                        {_slabel} — {len(_agg)} produtos repetidos em 2 ou mais lojas</h2>
                        <table>
                        <tr><th>Produto</th><th>Fornecedor</th><th>Lojas</th><th>Qtd Lojas</th><th>Valor Sugerido Total</th></tr>
                        {_linhas}
                        </table>''')
                    _corpo="".join(_partes) if _partes else "<p>Nenhum produto repetido entre lojas no momento.</p>"
                    _filial_txt=st.session_state.get("compras_filial_sel","(Todas as filiais)")
                    _data_txt=pd.Timestamp.now().strftime("%d/%m/%Y %H:%M")
                    return f'''<!DOCTYPE html><html><head><meta charset="utf-8">
                    <title>Produtos Repetidos Entre Lojas</title>
                    <style>
                    body{{font-family:Arial,sans-serif;color:#222;padding:24px;max-width:1000px;margin:auto}}
                    h1{{color:#2176FF}}
                    table{{width:100%;border-collapse:collapse;margin-top:8px;font-size:14px}}
                    th{{background:#2176FF;color:white;padding:8px;text-align:left}}
                    td{{padding:6px 8px;border-bottom:1px solid #ddd}}
                    tr:nth-child(even){{background:#f7f7f7}}
                    .meta{{color:#666;font-size:13px;margin-bottom:20px}}
                    @media print{{ body{{padding:0}} }}
                    </style></head><body>
                    <h1>📋 Produtos Repetidos Entre Lojas</h1>
                    <div class="meta">Escopo: {_filial_txt} — Gerado em {_data_txt}<br>
                    Mostra apenas produtos com o mesmo status em 2 ou mais lojas ao mesmo tempo —
                    esses são os candidatos a decisão de rede (não resolve só transferindo estoque entre lojas).</div>
                    {_corpo}
                    </body></html>'''
                _html_repetidos=_gerar_html_repetidos(df_res_compras)
                st.download_button("📋 Baixar relatório de produtos repetidos entre lojas (HTML)",
                    data=_html_repetidos,file_name=f"produtos_repetidos_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.html",
                    mime="text/html",key="compras_btn_relatorio_repetidos")

            _tem_col_filial=("Filial" in df_res_compras.columns) and (df_res_compras["Filial"].nunique()>1)
            if _tem_col_filial:
                _col_filt1,_col_filt2,_col_filt3=st.columns(3)
            else:
                _col_filt1,_col_filt2=st.columns(2)
            with _col_filt1:
                filtro_forn=st.multiselect("Filtrar por Fornecedor",sorted(df_res_compras["Fornecedor"].unique().tolist()),key="compras_filtro_forn",placeholder="Todos os fornecedores")
            with _col_filt2:
                filtro_status=st.multiselect("Filtrar por Status",sorted(df_res_compras["Status"].unique().tolist()),key="compras_filtro_status",placeholder="Todos os status")
            filtro_filial=[]
            if _tem_col_filial:
                with _col_filt3:
                    filtro_filial=st.multiselect("Filtrar por Filial",sorted(df_res_compras["Filial"].unique().tolist()),key="compras_filtro_filial",placeholder="Todas as filiais")
            df_res_show=df_res_compras.copy()
            if filtro_forn:
                df_res_show=df_res_show[df_res_show["Fornecedor"].isin(filtro_forn)]
            if filtro_status:
                df_res_show=df_res_show[df_res_show["Status"].isin(filtro_status)]
            if filtro_filial:
                df_res_show=df_res_show[df_res_show["Filial"].isin(filtro_filial)]
            with st.expander(f"📋 Ver tabela completa ({len(df_res_show)} produtos)"):
                st.dataframe(df_res_show.sort_values("ValorSugerido",ascending=False),use_container_width=True,height=420)

            csv_compras=df_res_compras.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
            st.download_button("📥 Exportar Recomendação de Compra (CSV)",csv_compras,file_name="recomendacao_compras.csv",use_container_width=True)
            sec(f"Calendário de Compras — O Quê, Quanto e Quando")
            

            df_cal=st.session_state.get("compras_calendario")
            if df_cal is None or df_cal.empty:
                st.markdown('<div class="al-s">✅ Nenhum evento de compra necessário no horizonte escolhido, dentro dos produtos calculados.</div>',unsafe_allow_html=True)
                df_prog=pd.DataFrame()
            else:
                c_cal1,c_cal2,c_cal3=st.columns(3)
                mc(c_cal1,"📅 Eventos de Compra no Horizonte",str(len(df_cal)),"b")
                mc(c_cal2,"📦 Produtos Envolvidos",str(df_cal["Produto"].nunique()),"b")
                mc(c_cal3,"💰 Total do Horizonte",fmt(df_cal["ValorCompra"].sum()),"g")

                df_cal_show=df_cal.sort_values("MesHorizonte").copy()
                df_cal_show=df_cal_show.rename(columns={
                    "DataEvento":"Data de Compra","QtdComprar":"Quantidade","ValorCompra":"Valor",
                    "CicloDias":"Ciclo (dias)","FrequenciaAno":"Frequência (x/ano)"})

                with st.expander(f"📋 Ver calendário completo ({len(df_cal_show)} eventos)"):
                    st.dataframe(df_cal_show[["Produto","Data de Compra","Quantidade","Valor",
                    "Ciclo (dias)","Frequência (x/ano)","Classe","Fornecedor"]],use_container_width=True,height=460)

                csv_cal=df_cal.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
                st.download_button("📥 Exportar Calendário de Compras (CSV)",csv_cal,file_name=f"calendario_compras_{horizonte_meses}meses.csv",use_container_width=True)

                df_prog=df_cal.copy()

            sec("Simulador de Cenário — Impacto no Giro e no Caixa")

            df_cenario=df_res_compras[df_res_compras["CoberturaDias"].notna()].copy()
            df_cenario["CapitalAtual"]=df_cenario["EstoqueAtual"]*df_cenario["CustoUnitario"]
            df_cenario["DiasAlvo"]=(df_cenario["MinDias"]+df_cenario["MaxDias"])/2
            df_cenario["DemandaDia"]=df_cenario["DemandaPrevMes(un)"]/30
            df_cenario["CapitalIdeal"]=df_cenario["DiasAlvo"]*df_cenario["DemandaDia"]*df_cenario["CustoUnitario"]

            capital_atual_total=df_cenario["CapitalAtual"].sum()
            capital_ideal_total=df_cenario["CapitalIdeal"].sum()
            diferenca_capital=capital_atual_total-capital_ideal_total

            peso_capital_cen=df_cenario["EstoqueAtual"]*df_cenario["CustoUnitario"]
            if peso_capital_cen.sum()>0:
                giro_atual_medio=float((df_cenario["GiroAtual"]*peso_capital_cen).sum()/peso_capital_cen.sum())
                giro_alvo_medio=float((df_cenario["GiroAlvo"]*peso_capital_cen).sum()/peso_capital_cen.sum())
            else:
                giro_atual_medio=0; giro_alvo_medio=0
            giro_pct_atual=giro_atual_medio*30/365*100
            giro_pct_alvo=giro_alvo_medio*30/365*100

            c_cen1,c_cen2,c_cen3=st.columns(3)
            mc(c_cen1,"📦 Valor do Estoque Atual",fmt(capital_atual_total),"b")
            mc(c_cen2,"🎯 Valor do Estoque Ideal",fmt(capital_ideal_total),"g")
            mc(c_cen3,"💸 Capital Liberável" if diferenca_capital>0 else "📈 Capital Adicional Necessário",
            fmt(abs(diferenca_capital)),"g" if diferenca_capital>0 else "y")

            c_gir1,c_gir2=st.columns(2)
            mc(c_gir1,"🔄 Giro Médio Atual",f"{giro_pct_atual:.0f}%","b",f"{giro_atual_medio:.1f}x ao ano")
            mc(c_gir2,"🎯 Giro Médio Alvo",f"{giro_pct_alvo:.0f}%","g",f"{giro_alvo_medio:.1f}x ao ano")

            st.markdown("<br>",unsafe_allow_html=True)
            sec("📊 Giro e Capital por Classe")
            resumo_classe=[]
            for classe_res in ["A","B","C","D","E"]:
                df_cl=df_cenario[df_cenario["Classe"]==classe_res]
                if df_cl.empty: continue
                peso_cl=df_cl["CapitalAtual"]
                if peso_cl.sum()<=0: continue
                # Mesma fórmula ponderada por capital usada nos cards do topo (giro_atual_medio/giro_alvo_medio),
                # em vez de 365/média(dias) — que não é matematicamente igual a média dos giros individuais.
                giro_atual_cl=float((df_cl["GiroAtual"]*peso_cl).sum()/peso_cl.sum())
                giro_alvo_cl=float((df_cl["GiroAlvo"]*peso_cl).sum()/peso_cl.sum())
                resumo_classe.append({
                    "Classe":classe_res,
                    "Produtos":len(df_cl),
                    "Giro Atual (%)":round(giro_atual_cl*30/365*100,0),
                    "Giro Atual (x/ano)":round(giro_atual_cl,1),
                    "Giro Alvo (%)":round(giro_alvo_cl*30/365*100,0),
                    "Giro Alvo (x/ano)":round(giro_alvo_cl,1),
                    "Capital Atual":fmt(df_cl["CapitalAtual"].sum()),
                    "Capital Ideal":fmt(df_cl["CapitalIdeal"].sum()),
                })
            if resumo_classe:
                st.dataframe(pd.DataFrame(resumo_classe),use_container_width=True,hide_index=True)
                classe_raw_snapshot=[]
                for classe_res2 in ["A","B","C","D","E"]:
                    df_cl2=df_cenario[df_cenario["Classe"]==classe_res2]
                    if df_cl2.empty: continue
                    cob_cl2=df_cl2["CoberturaDias"].replace(0,np.nan).mean()
                    alvo_cl2=df_cl2["DiasAlvo"].replace(0,np.nan).mean()
                    if pd.isna(cob_cl2) or pd.isna(alvo_cl2) or cob_cl2==0 or alvo_cl2==0: continue
                    classe_raw_snapshot.append({
                        "classe":classe_res2,"produtos":len(df_cl2),
                        "giro_atual_pct":round(30/cob_cl2*100,0),"giro_atual_xano":round(365/cob_cl2,1),
                        "giro_alvo_pct":round(30/alvo_cl2*100,0),"giro_alvo_xano":round(365/alvo_cl2,1),
                        "capital_atual":float(df_cl2["CapitalAtual"].sum()),
                        "capital_ideal":float(df_cl2["CapitalIdeal"].sum()),
                    })
                st.session_state["classe_snapshot"]=classe_raw_snapshot
                if st.session_state.cid: save_snap(st.session_state.cid,"classe",st.session_state["classe_snapshot"],filial=filial_sel_compras)

            

                
                sec("Capital Parado — Estoque sem Giro")
            

            df_morto=df_res_compras.copy()
            limite_demanda_baixa=df_morto[df_morto["DemandaPrevMes(un)"]>0]["DemandaPrevMes(un)"].quantile(0.10) if (df_morto["DemandaPrevMes(un)"]>0).any() else 0
            df_morto=df_morto[(df_morto["EstoqueAtual"]>0) & (df_morto["DemandaPrevMes(un)"]<=limite_demanda_baixa)].copy()
            df_morto["CapitalParado"]=df_morto["EstoqueAtual"]*df_morto["CustoUnitario"]
            df_morto=df_morto.sort_values("CapitalParado",ascending=False)
            st.session_state["compras_morto"]=df_morto
            # Salva o Capital Morto já calculado — no momento do clique em "Rodar Motor de Compras"
            # esse cálculo ainda não tinha rodado, então o arquivo salvo ficava com dado velho/vazio
            if st.session_state.cid:
                save_resultado_compras(st.session_state.cid,None,None,df_morto,st.session_state.get("compras_filial_sel"))

            if df_morto.empty:
                st.markdown('<div class="al-s">✅ Nenhum produto identificado como capital morto, dentro dos produtos calculados.</div>',unsafe_allow_html=True)
            else:
                total_capital_morto=df_morto["CapitalParado"].sum()
                c_cm1,c_cm2=st.columns(2)
                mc(c_cm1,"📦 Produtos com Capital Parado",str(len(df_morto)),"y")
                mc(c_cm2,"💰 Capital Total Parado (baixíssimo giro)",fmt(total_capital_morto),"r")

                with st.expander(f"📋 Ver produtos com capital parado ({len(df_morto)})"):
                    st.dataframe(df_morto[["Produto","Fornecedor","Classe","EstoqueAtual","DemandaPrevMes(un)",
                    "CoberturaDias","CustoUnitario","CapitalParado"]],use_container_width=True,height=380)

                if not df_morto.empty:
                    csv_morto=df_morto.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
                    st.download_button("📥 Exportar Capital Parado (CSV)",csv_morto,file_name="capital_parado.csv",use_container_width=True)

            sec("🏆 Scorecard de Fornecedores")
            

            if df_res_compras is not None and not df_res_compras.empty and "Fornecedor" in df_res_compras.columns:
                # Compras sugeridas por fornecedor
                df_forn=df_res_compras[df_res_compras["ValorSugerido"]>0].copy()
                if not df_forn.empty:
                    score=df_forn.groupby("Fornecedor").agg(
                        Total_Compra=("ValorSugerido","sum"),
                        Qtd_SKUs=("Produto","nunique"),
                        SKUs_ClasseA=("Classe",lambda x: (x=="A").sum()),
                        Qtd_Comprar=("QtdSugerida","sum"),
                    ).reset_index().sort_values("Total_Compra",ascending=False)
                    score["Participacao_%"]=score["Total_Compra"]/score["Total_Compra"].sum()*100

                    # Capital morto por fornecedor
                    if not df_morto.empty and "Fornecedor" in df_morto.columns:
                        morto_forn=df_morto.groupby("Fornecedor")["CapitalParado"].sum().reset_index()
                        morto_forn.columns=["Fornecedor","Capital_Morto"]
                        score=score.merge(morto_forn,on="Fornecedor",how="left")
                        score["Capital_Morto"]=score["Capital_Morto"].fillna(0)
                    else:
                        score["Capital_Morto"]=0

                    # KPIs
                    total_forn=len(score)
                    forn_concentrado=score[score["Participacao_%"]>=30]
                    maior_forn=score.iloc[0]
                    c_s1,c_s2,c_s3,c_s4=st.columns(4)
                    mc(c_s1,"Fornecedores Ativos",str(total_forn),"b")
                    mc(c_s2,"Maior Fornecedor",maior_forn["Fornecedor"][:20],"b",f"{maior_forn['Participacao_%']:.1f}% do total")
                    mc(c_s3,"Fornecedores c/ >30%",str(len(forn_concentrado)),"r" if len(forn_concentrado)>0 else "g","Risco de concentração" if len(forn_concentrado)>0 else "Bem distribuído")
                    mc(c_s4,"SKUs Classe A em Compra",str(int(score["SKUs_ClasseA"].sum())),"y")

                    # Gráfico barras horizontais
                    fig_score=go.Figure()
                    fig_score.add_trace(go.Bar(
                        y=score["Fornecedor"],x=score["Total_Compra"],
                        orientation="h",name="Compras Sugeridas",
                        marker=dict(color=["#DC2626" if p>=30 else "#0F6E56" for p in score["Participacao_%"]],
                            line=dict(color="white",width=0.8)),
                        text=[f"R$ {v/1e3:.0f}k · {p:.1f}%" for v,p in zip(score["Total_Compra"],score["Participacao_%"])],
                        textposition="outside",textfont=dict(size=9,color="#374151"),cliponaxis=False,
                        hovertemplate="<b>%{y}</b><br>Compras: R$ %{x:,.0f}<extra></extra>"))
                    fig_score.update_layout(
                        title=dict(text="Compras Sugeridas por Fornecedor — Vermelho = Concentração >30%",
                            font=dict(size=13,family="Georgia, serif",color="#14243B")),
                        plot_bgcolor="white",paper_bgcolor="white",
                        font=dict(color="#6B7280",size=10),
                        margin=dict(l=10,r=120,t=44,b=20),height=max(300,len(score)*45),
                        xaxis=dict(showgrid=True,gridcolor="#F3F4F6",tickprefix="R$ ",tickfont=dict(size=9),
                        range=[0,score["Total_Compra"].max()*1.22]),
                        yaxis=dict(showgrid=False,tickfont=dict(size=10)),
                        showlegend=False,
                        hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B",size=11)))
                    st.plotly_chart(fig_score,use_container_width=True)

                    # Tabela detalhada
                    score_fmt=score.copy()
                    score_fmt["Total_Compra"]=score_fmt["Total_Compra"].apply(fmt)
                    score_fmt["Capital_Morto"]=score_fmt["Capital_Morto"].apply(fmt)
                    score_fmt["Participacao_%"]=score_fmt["Participacao_%"].apply(lambda x: f"{x:.1f}%")
                    score_fmt["Qtd_Comprar"]=score_fmt["Qtd_Comprar"].apply(lambda x: f"{int(x):,}")
                    score_fmt.columns=["Fornecedor","Total Compras","SKUs Envolvidos","SKUs Classe A","Qtd. Total","Part. %","Capital Parado"]
                    st.dataframe(score_fmt,use_container_width=True,hide_index=True)

                    # Alerta de concentração
                    if len(forn_concentrado)>0:
                        nomes=", ".join(forn_concentrado["Fornecedor"].tolist())
                        st.markdown(f'<div class="al-w">⚠️ <b>Atenção:</b> {nomes} concentra(m) mais de 30% do volume total de compras. Avalie diversificação para reduzir risco operacional.</div>',unsafe_allow_html=True)

                    csv_score=score.to_csv(sep=";",decimal=",",index=False).encode("utf-8-sig")
                    st.download_button("📥 Exportar Scorecard de Fornecedores (CSV)",csv_score,file_name="scorecard_fornecedores.csv",use_container_width=True)
                else:
                    st.markdown('<div class="al-i">Nenhuma compra sugerida para calcular o scorecard.</div>',unsafe_allow_html=True)
            else:
                st.markdown('<div class="al-w">⚠️ Rode o Motor de Compras primeiro para ver o Scorecard de Fornecedores.</div>',unsafe_allow_html=True)

            # Mesmo gráfico, agora usando o HORIZONTE de compras (6 meses, Calendário de Compras)
            # em vez do valor imediato (ValorSugerido) — participação de cada fornecedor no total
            # que vai ser gasto ao longo do horizonte, não só no que precisa comprar hoje.
            df_cal_forn=st.session_state.get("compras_calendario")
            if df_cal_forn is not None and not df_cal_forn.empty and "Fornecedor" in df_cal_forn.columns:
                st.markdown("<br>",unsafe_allow_html=True)
                
                score_h=df_cal_forn.groupby("Fornecedor").agg(
                    Total_Compra=("ValorCompra","sum"),
                    Qtd_SKUs=("Produto","nunique"),
                ).reset_index().sort_values("Total_Compra",ascending=False)
                score_h["Participacao_%"]=score_h["Total_Compra"]/score_h["Total_Compra"].sum()*100 if score_h["Total_Compra"].sum()>0 else 0

                fig_score_h=go.Figure()
                fig_score_h.add_trace(go.Bar(
                    y=score_h["Fornecedor"],x=score_h["Total_Compra"],
                    orientation="h",name="Horizonte de Compras",
                    marker=dict(color=["#FFB627" if p>=30 else "#A9762F" for p in score_h["Participacao_%"]],
                        line=dict(color="white",width=0.8)),
                    text=[f"R$ {v/1e3:.0f}k · {p:.1f}%" for v,p in zip(score_h["Total_Compra"],score_h["Participacao_%"])],
                    textposition="outside",textfont=dict(size=9,color="#374151"),cliponaxis=False,
                    hovertemplate="<b>%{y}</b><br>Horizonte: R$ %{x:,.0f}<extra></extra>"))
                fig_score_h.update_layout(
                    title=dict(text="📅 Horizonte de Compras por Fornecedor (6 meses) — Laranja = Concentração >30%",
                        font=dict(size=13,family="Georgia, serif",color="#0F766E")),
                    plot_bgcolor="white",paper_bgcolor="white",
                    font=dict(color="#6B7280",size=10),
                    margin=dict(l=10,r=120,t=44,b=20),height=max(300,len(score_h)*45),
                    xaxis=dict(showgrid=True,gridcolor="#F3F4F6",tickprefix="R$ ",tickfont=dict(size=9),
                    range=[0,score_h["Total_Compra"].max()*1.22]),
                    yaxis=dict(showgrid=False,tickfont=dict(size=10)),
                    showlegend=False,
                    hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B",size=11)))
                st.plotly_chart(fig_score_h,use_container_width=True)

            sec("Impacto no Fluxo de Caixa")
            

            if df_prog.empty:
                st.markdown('<div class="al-i">Sem compras urgentes no momento — sem impacto de caixa a projetar.</div>',unsafe_allow_html=True)
            else:
                df_fluxo_compras=df_prog.copy()
                df_fluxo_compras["DataLimite_dt"]=pd.to_datetime(df_fluxo_compras["DataEvento"],format="%d/%m/%Y")
                df_fluxo_compras["Semana"]=df_fluxo_compras["DataLimite_dt"].dt.to_period("W").apply(lambda p: p.start_time)

                resumo_semanal=df_fluxo_compras.groupby("Semana")["ValorCompra"].sum().reset_index().rename(columns={"ValorCompra":"ValorSugerido"})
                resumo_semanal=resumo_semanal.sort_values("Semana")
                resumo_semanal["Acumulado"]=resumo_semanal["ValorSugerido"].cumsum()

                total_90dias=df_fluxo_compras["ValorCompra"].sum()
                c_fx1,c_fx2,c_fx3=st.columns(3)
                mc(c_fx1,"💸 Caixa Necessário (próximas semanas)",fmt(total_90dias),"r")
                mc(c_fx2,"📅 Primeira Semana de Impacto",resumo_semanal["Semana"].iloc[0].strftime("%d/%m/%Y") if len(resumo_semanal)>0 else "—","b")
                mc(c_fx3,"📊 Semanas com Compra Programada",str(len(resumo_semanal)),"b")

                media_semanal=resumo_semanal["ValorSugerido"].mean()
                limiar_critico=media_semanal*1.5
                cores_barras=["#DC2626" if v>=limiar_critico else "#14243B" for v in resumo_semanal["ValorSugerido"]]

                # Gráfico acumulado (topo)
                fig_acum=go.Figure()
                fig_acum.add_trace(go.Scatter(
                x=resumo_semanal["Semana"].dt.strftime("%d/%m"),
                y=resumo_semanal["Acumulado"],
                name="Acumulado",mode="lines+markers",
                line=dict(color="#A9762F",width=2.8),
                marker=dict(size=8,color="#A9762F",line=dict(color="white",width=1.5)),
                fill="tozeroy",fillcolor="rgba(169,118,47,0.08)",
                hovertemplate="<b>Semana %{x}</b><br>Acumulado: R$ %{y:,.0f}<extra></extra>"))
                fig_acum.add_hline(y=media_semanal*len(resumo_semanal),
                line_dash="dot",line_color="#9CA3AF",
                annotation_text="Total médio esperado",
                annotation_font=dict(size=9,color="#9CA3AF"))
                fig_acum.update_layout(
                title=dict(text="Saída de Caixa Acumulada — Compras Programadas",
                    font=dict(size=14,family="Georgia, serif",color="#14243B")),
                plot_bgcolor="white",paper_bgcolor="white",
                font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
                margin=dict(l=10,r=10,t=44,b=40),
                xaxis=dict(showgrid=False,linecolor="#E5E7EB",tickangle=-40,tickfont=dict(size=9)),
                yaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=True,tickfont=dict(size=9),tickprefix="R$ "),
                legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.28,x=0.5,xanchor="center",font=dict(size=10)),
                hovermode="x unified",height=280,
                hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B",size=11)))
                st.plotly_chart(fig_acum,use_container_width=True)

                # Gráfico semanal (barras)
                fig_fluxo=go.Figure()
                fig_fluxo.add_trace(go.Bar(
                x=resumo_semanal["Semana"].dt.strftime("%d/%m"),
                y=resumo_semanal["ValorSugerido"],
                name="Saída prevista na semana",
                marker=dict(color=cores_barras,line=dict(color="white",width=0.8)),
                text=[fmt(v) for v in resumo_semanal["ValorSugerido"]],
                textposition="outside",textfont=dict(size=9,color="#374151"),
                hovertemplate="<b>Semana %{x}</b><br>Saída: R$ %{y:,.0f}<extra></extra>"))
                fig_fluxo.add_hline(y=media_semanal,
                line_dash="dot",line_color="#9CA3AF",
                annotation_text=f"Média: {fmt(media_semanal)}",
                annotation_position="top right",
                annotation_font=dict(size=9,color="#9CA3AF"))
                if limiar_critico<=resumo_semanal["ValorSugerido"].max():
                    fig_fluxo.add_annotation(
                        text="🔴 Semanas em vermelho = acima de 150% da média — atenção ao caixa",
                        xref="paper",yref="paper",x=0,y=-0.28,showarrow=False,
                        font=dict(size=9,color="#DC2626"),align="left")
                fig_fluxo.update_layout(
                title=dict(text="Saída de Caixa por Semana — Barras em Vermelho = Semanas Críticas",
                    font=dict(size=14,family="Georgia, serif",color="#14243B")),
                plot_bgcolor="white",paper_bgcolor="white",
                font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
                margin=dict(l=10,r=10,t=44,b=70),
                xaxis=dict(showgrid=False,linecolor="#E5E7EB",tickangle=-40,tickfont=dict(size=9)),
                yaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=True,tickfont=dict(size=9),tickprefix="R$ "),
                legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.32,x=0.5,xanchor="center",font=dict(size=10)),
                hovermode="x unified",height=340,bargap=0.25,
                hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B",size=11)))
                st.plotly_chart(fig_fluxo,use_container_width=True)

            

        sec("⚙️ Configurar Desempenho de Fornecedores")
        
        sc_salvo_cfg=load_scorecard_forn(st.session_state.cid) if st.session_state.cid else None
        if sc_salvo_cfg is not None and not sc_salvo_cfg.empty and "Fornecedor" in sc_salvo_cfg.columns:
            sc_base_cfg=sc_salvo_cfg.copy()
        else:
            sc_base_cfg=pd.DataFrame({"Fornecedor":pd.Series(dtype=str),"Prazo":pd.Series(dtype=int),
                "Qualidade":pd.Series(dtype=int),"OTIF":pd.Series(dtype=int)})
        for col_cfg,default_cfg in [("Prazo",96),("Qualidade",95),("OTIF",90)]:
            if col_cfg not in sc_base_cfg.columns: sc_base_cfg[col_cfg]=default_cfg
            sc_base_cfg[col_cfg]=pd.to_numeric(sc_base_cfg[col_cfg],errors="coerce").fillna(default_cfg).astype(int)

        sc_editado_cfg=st.data_editor(sc_base_cfg,use_container_width=True,hide_index=True,num_rows="dynamic",key="cfg_scorecard_forn_editor",
            column_config={
                "Fornecedor":st.column_config.TextColumn("Fornecedor",required=True),
                "Prazo":st.column_config.NumberColumn("Prazo (% no prazo)",min_value=0,max_value=100,step=1),
                "Qualidade":st.column_config.NumberColumn("Qualidade (% conformidade)",min_value=0,max_value=100,step=1),
                "OTIF":st.column_config.NumberColumn("OTIF (% completo e no prazo)",min_value=0,max_value=100,step=1),
            })
        senha_scorecard_forn=st.text_input("Senha master para salvar",type="password",key="senha_scorecard_forn")
        if st.button("💾 Salvar Desempenho de Fornecedores",key="cfg_salvar_scorecard_forn",use_container_width=True):
            if senha_scorecard_forn!=SENHA_MASTER:
                st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi salvo.</div>',unsafe_allow_html=True)
            else:
                sc_valido_cfg=sc_editado_cfg[sc_editado_cfg["Fornecedor"].astype(str).str.strip()!=""].copy()
                sc_valido_cfg["Fornecedor"]=sc_valido_cfg["Fornecedor"].astype(str).str.strip()
                st.session_state["gs_scorecard_forn"]=sc_valido_cfg
                if st.session_state.cid:
                    save_scorecard_forn(st.session_state.cid,sc_valido_cfg)
                    st.markdown('<div class="al-s">✅ Desempenho de fornecedores salvo — o Painel de Indicadores será atualizado automaticamente.</div>',unsafe_allow_html=True)
                else:
                    st.markdown('<div class="al-w">⚠️ Não foi possível salvar em disco — nenhum cliente selecionado (session_state.cid vazio). Os dados ficam só nesta sessão e se perdem ao reiniciar.</div>',unsafe_allow_html=True)

    st.stop()
elif pg=="fluxo_compras":
    hdr("💰 Fluxo Projetado","Saída de caixa por compras x entrada por vendas, distribuídas pelos prazos reais de pagamento e recebimento")

    df_v_ff=get_vendas_df()

    _col_fil_ff=col_filial(df_v_ff) if df_v_ff is not None else None
    if _col_fil_ff:
        _filiais_disp_ff=sorted(df_v_ff[_col_fil_ff].dropna().astype(str).unique().tolist())
        _opcoes_filial_ff=["(Todas as filiais)"]+_filiais_disp_ff

        def _on_change_filial_ff():
            st.session_state["ff_filial_sel_backup"]=st.session_state["ff_filial_sel"]
            for _k_limpar_ff in ["ff_parcelas_pagar","ff_parcelas_receber","ff_encargos_pct",
                                  "ff_saldo_inicial","ff_incluir_contas_pr","ff_config_carregada",
                                  "ff_contas_pagar_df","ff_contas_receber_df","fluxo_snapshot",
                                  "ff_expander_aberto","ff_checkbox_incluir_pr"]:
                if _k_limpar_ff in st.session_state:
                    del st.session_state[_k_limpar_ff]

        if "ff_filial_sel_backup" not in st.session_state:
            st.session_state["ff_filial_sel_backup"]=_opcoes_filial_ff[0]
        if st.session_state["ff_filial_sel_backup"] not in _opcoes_filial_ff:
            st.session_state["ff_filial_sel_backup"]=_opcoes_filial_ff[0]
        st.session_state["ff_filial_sel"]=st.session_state["ff_filial_sel_backup"]

        filial_sel_ff=st.selectbox("🏬 Filial",_opcoes_filial_ff,
          key="ff_filial_sel",on_change=_on_change_filial_ff)

        if filial_sel_ff!="(Todas as filiais)":
            df_v_ff=df_v_ff[df_v_ff[_col_fil_ff].astype(str)==filial_sel_ff].copy()

    # Só confia no calendário guardado na memória se ele for da MESMA filial que
    # está selecionada aqui agora (ex: acabou de rodar o Motor de Compras pra essa
    # mesma filial/escopo). Se for de outra filial (ou de "Todas as filiais" quando
    # você quer ver uma loja sozinha, ou vice-versa), busca do disco a versão
    # certa pra esse escopo específico — cada filial tem seu próprio arquivo salvo.
    _escopo_memoria_ff=st.session_state.get("compras_filial_sel")
    df_cal_ff=st.session_state.get("compras_calendario")
    if df_cal_ff is not None and not df_cal_ff.empty and _escopo_memoria_ff!=filial_sel_ff:
        df_cal_ff=None
    if (df_cal_ff is None or df_cal_ff.empty) and st.session_state.cid:
        _,df_cal_ff,_=load_resultado_compras(st.session_state.cid,st.session_state.get("ff_filial_sel"))

    if (df_cal_ff is None or df_cal_ff.empty) and (df_v_ff is None or df_v_ff.empty):
        st.markdown('<div class="al-w">⚠️ Rode o <b>Motor de Compras</b> (em 📦 Gestão de Compras) e/ou importe a base de <b>Vendas</b> para ver este painel.</div>',unsafe_allow_html=True)
        st.stop()

    tab_resultado,tab_config=st.tabs(["📊 Resultado","⚙️ Configurar"])
    with tab_config:    

        if "ff_expander_aberto" not in st.session_state:
            st.session_state["ff_expander_aberto"]=False
        with st.expander("⚙️ Configuração de Prazos e Encargos",expanded=st.session_state["ff_expander_aberto"]):
            st.session_state["ff_expander_aberto"]=True
            if "ff_config_carregada" not in st.session_state:
                _ff_salvo=load_fluxo_financeiro(st.session_state.cid,st.session_state.get("ff_filial_sel")) if st.session_state.cid else None
                st.session_state["ff_parcelas_pagar"]=(_ff_salvo or {}).get("ff_parcelas_pagar") or [{"pct":100,"dias":30}]
                st.session_state["ff_parcelas_receber"]=(_ff_salvo or {}).get("ff_parcelas_receber") or [{"pct":100,"dias":30}]
                st.session_state["ff_encargos_pct"]=(_ff_salvo or {}).get("ff_encargos_pct") or 0.0
                st.session_state["ff_saldo_inicial"]=(_ff_salvo or {}).get("ff_saldo_inicial") or 0.0
                st.session_state["ff_incluir_contas_pr"]=(_ff_salvo or {}).get("ff_incluir_contas_pr") or False
                st.session_state["ff_config_carregada"]=True

            c_ff_pag,c_ff_rec=st.columns(2)
            with c_ff_pag:
                st.markdown("**💸 Prazo de Pagamento a Fornecedores**")
                for idx_pp,pp in enumerate(st.session_state["ff_parcelas_pagar"]):
                    cpp1,cpp2,cpp3=st.columns([1,1,0.4])
                    pp["pct"]=cpp1.number_input("% do valor",min_value=0,max_value=100,value=int(pp["pct"]),step=5,key=f"ff_pag_pct_{idx_pp}")
                    pp["dias"]=cpp2.number_input("Dias após compra",min_value=0,max_value=365,value=int(pp["dias"]),step=15,key=f"ff_pag_dias_{idx_pp}")
                    with cpp3:
                        st.markdown("<br>",unsafe_allow_html=True)
                        if st.button("🗑",key=f"ff_pag_rm_{idx_pp}") and len(st.session_state["ff_parcelas_pagar"])>1:
                            st.session_state["ff_parcelas_pagar"].pop(idx_pp)
                            st.rerun()
                soma_pag=sum(p["pct"] for p in st.session_state["ff_parcelas_pagar"])
                if st.button("➕ Adicionar parcela",key="ff_add_pag"):
                    st.session_state["ff_parcelas_pagar"].append({"pct":0,"dias":30})
                    st.rerun()
                st.markdown(f'<div class="al-{"s" if soma_pag==100 else "d"}">{"✅" if soma_pag==100 else "⚠️"} Soma das parcelas: {soma_pag}% {"(correto)" if soma_pag==100 else "— precisa somar 100%"}</div>',unsafe_allow_html=True)

            with c_ff_rec:
                st.markdown("**💰 Prazo de Recebimento de Clientes**")
                for idx_pr,pr in enumerate(st.session_state["ff_parcelas_receber"]):
                    cpr1,cpr2,cpr3=st.columns([1,1,0.4])
                    pr["pct"]=cpr1.number_input("% do valor",min_value=0,max_value=100,value=int(pr["pct"]),step=5,key=f"ff_rec_pct_{idx_pr}")
                    pr["dias"]=cpr2.number_input("Dias após venda",min_value=0,max_value=365,value=int(pr["dias"]),step=15,key=f"ff_rec_dias_{idx_pr}")
                    with cpr3:
                        st.markdown("<br>",unsafe_allow_html=True)
                        if st.button("🗑",key=f"ff_rec_rm_{idx_pr}") and len(st.session_state["ff_parcelas_receber"])>1:
                            st.session_state["ff_parcelas_receber"].pop(idx_pr)
                            st.rerun()
                soma_rec=sum(p["pct"] for p in st.session_state["ff_parcelas_receber"])
                if st.button("➕ Adicionar parcela",key="ff_add_rec"):
                    st.session_state["ff_parcelas_receber"].append({"pct":0,"dias":30})
                    st.rerun()
                st.markdown(f'<div class="al-{"s" if soma_rec==100 else "d"}">{"✅" if soma_rec==100 else "⚠️"} Soma das parcelas: {soma_rec}% {"(correto)" if soma_rec==100 else "— precisa somar 100%"}</div>',unsafe_allow_html=True)

            c_ff_enc,c_ff_sal=st.columns(2)
            encargos_pct=c_ff_enc.number_input("📈 Outros Encargos da Empresa (%)",min_value=0.0,max_value=100.0,value=float(st.session_state["ff_encargos_pct"]),step=0.5,key="ff_encargos_input")
            saldo_inicial=c_ff_sal.number_input("🏦 Saldo Inicial de Caixa (R$)",value=float(st.session_state["ff_saldo_inicial"]),step=1000.0,key="ff_saldo_input")
            st.session_state["ff_encargos_pct"]=encargos_pct
            st.session_state["ff_saldo_inicial"]=saldo_inicial

            senha_ff_config=st.text_input("Senha master para salvar",type="password",key="senha_ff_config")
            if st.button("💾 Salvar Configuração de Prazos",key="ff_salvar_config",use_container_width=True):
                if senha_ff_config!=SENHA_MASTER:
                    st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi salvo.</div>',unsafe_allow_html=True)
                elif soma_pag!=100 or soma_rec!=100:
                    st.markdown('<div class="al-d">❌ Corrija as parcelas antes de salvar — cada lado precisa somar exatamente 100%.</div>',unsafe_allow_html=True)
                else:
                    if st.session_state.cid: save_fluxo_financeiro(st.session_state.cid,st.session_state.get("ff_filial_sel"))
                    st.markdown('<div class="al-s">✅ Configuração salva.</div>',unsafe_allow_html=True)
                    st.markdown("---")
            st.markdown("**📎 Contas a Pagar e a Receber já lançadas (opcional)**")
            

            c_up_pag,c_up_rec=st.columns(2)
            with c_up_pag:
                st.markdown("💸 Contas a Pagar")
                arq_pagar=st.file_uploader("Arquivo (CSV ou Excel)",type=["csv","xlsx","xls"],key="ff_upload_pagar")
                if arq_pagar is not None:
                    senha_import_pagar=st.text_input("Senha master *",type="password",key="senha_import_pagar")
                    if st.button("📤 Processar",use_container_width=True,key="ff_btn_processar_pagar"):
                        if senha_import_pagar!=SENHA_MASTER:
                            st.error("❌ Senha master incorreta.")
                        else:
                            try:
                                df_pagar_up=pd.read_csv(arq_pagar,sep=None,engine="python",encoding="utf-8-sig") if arq_pagar.name.endswith(".csv") else pd.read_excel(arq_pagar)
                                df_pagar_up.columns=[str(c).strip() for c in df_pagar_up.columns]
                                cols_pagar=list(df_pagar_up.columns)
                                col_venc_pag=next((c for c in cols_pagar if c.strip().lower() in ["vencimento","data","data vencimento","data de vencimento"]),None)
                                col_val_pag=next((c for c in cols_pagar if c.strip().lower() in ["valor","vlr","valor total"]),None)
                                col_cat_pag=next((c for c in cols_pagar if c.strip().lower() in ["conta","categoria","conta/categoria","natureza","fornecedor","tipo"]),None)
                                col_fil_pag=col_filial(df_pagar_up)
                                if col_venc_pag and col_val_pag:
                                    _dados_pagar_final={
                                        "Vencimento":pd.to_datetime(df_pagar_up[col_venc_pag],errors="coerce",dayfirst=True),
                                        "Valor":parse_valor_brl(df_pagar_up[col_val_pag]) if df_pagar_up[col_val_pag].dtype==object else df_pagar_up[col_val_pag],
                                        "Conta":df_pagar_up[col_cat_pag].astype(str) if col_cat_pag else "Não informado"
                                    }
                                    if col_fil_pag:
                                        _dados_pagar_final["Filial"]=df_pagar_up[col_fil_pag].astype(str).str.strip()
                                    df_pagar_final=pd.DataFrame(_dados_pagar_final).dropna(subset=["Vencimento"])
                                    if st.session_state.cid: save_contas_pr(st.session_state.cid,"pagar",df_pagar_final)
                                    st.session_state["ff_contas_pagar_df"]=df_pagar_final
                                    st.markdown(f'<div class="al-s">✅ {len(df_pagar_final)} lançamento(s) carregados — total {fmt(df_pagar_final["Valor"].sum())}</div>',unsafe_allow_html=True)
                                else:
                                    st.markdown('<div class="al-d">❌ Não encontrei as colunas Vencimento e Valor no arquivo.</div>',unsafe_allow_html=True)
                            except Exception as e:
                                st.markdown(f'<div class="al-d">❌ Erro ao ler arquivo: {e}</div>',unsafe_allow_html=True)
                _cp_salvo=st.session_state.get("ff_contas_pagar_df")
                if _cp_salvo is None and st.session_state.cid:
                    _cp_salvo=load_contas_pr(st.session_state.cid,"pagar")
                    if _cp_salvo is not None: st.session_state["ff_contas_pagar_df"]=_cp_salvo
                if _cp_salvo is not None and not _cp_salvo.empty:
                    _col_fil_cp=col_filial(_cp_salvo)
                    _filial_atual_ff=st.session_state.get("ff_filial_sel")
                    if _col_fil_cp and _filial_atual_ff and _filial_atual_ff!="(Todas as filiais)":
                        _cp_salvo=_cp_salvo[_cp_salvo[_col_fil_cp].astype(str)==_filial_atual_ff]
                    st.caption(f"📌 Já carregado: {len(_cp_salvo)} lançamento(s), total {fmt(_cp_salvo['Valor'].sum())}")

            with c_up_rec:
                st.markdown("💰 Contas a Receber")
                arq_receber=st.file_uploader("Arquivo (CSV ou Excel)",type=["csv","xlsx","xls"],key="ff_upload_receber")
                if arq_receber is not None:
                    senha_import_receber=st.text_input("Senha master *",type="password",key="senha_import_receber")
                    if st.button("📤 Processar",use_container_width=True,key="ff_btn_processar_receber"):
                        if senha_import_receber!=SENHA_MASTER:
                            st.error("❌ Senha master incorreta.")
                        else:
                            try:
                                df_receber_up=pd.read_csv(arq_receber,sep=None,engine="python",encoding="utf-8-sig") if arq_receber.name.endswith(".csv") else pd.read_excel(arq_receber)
                                df_receber_up.columns=[str(c).strip() for c in df_receber_up.columns]
                                cols_receber=list(df_receber_up.columns)
                                col_venc_rec=next((c for c in cols_receber if c.strip().lower() in ["vencimento","data","data vencimento","data de vencimento"]),None)
                                col_val_rec=next((c for c in cols_receber if c.strip().lower() in ["valor","vlr","valor total"]),None)
                                col_fil_rec=col_filial(df_receber_up)
                                if col_venc_rec and col_val_rec:
                                    _dados_receber_final={
                                        "Vencimento":pd.to_datetime(df_receber_up[col_venc_rec],errors="coerce",dayfirst=True),
                                        "Valor":parse_valor_brl(df_receber_up[col_val_rec]) if df_receber_up[col_val_rec].dtype==object else df_receber_up[col_val_rec]
                                    }
                                    if col_fil_rec:
                                        _dados_receber_final["Filial"]=df_receber_up[col_fil_rec].astype(str).str.strip()
                                    df_receber_final=pd.DataFrame(_dados_receber_final).dropna(subset=["Vencimento"])
                                    if st.session_state.cid: save_contas_pr(st.session_state.cid,"receber",df_receber_final)
                                    st.session_state["ff_contas_receber_df"]=df_receber_final
                                    st.markdown(f'<div class="al-s">✅ {len(df_receber_final)} lançamento(s) carregados — total {fmt(df_receber_final["Valor"].sum())}</div>',unsafe_allow_html=True)
                                else:
                                    st.markdown('<div class="al-d">❌ Não encontrei as colunas Vencimento e Valor no arquivo.</div>',unsafe_allow_html=True)
                            except Exception as e:
                                st.markdown(f'<div class="al-d">❌ Erro ao ler arquivo: {e}</div>',unsafe_allow_html=True)
                _cr_salvo=st.session_state.get("ff_contas_receber_df")
                if _cr_salvo is None and st.session_state.cid:
                    _cr_salvo=load_contas_pr(st.session_state.cid,"receber")
                    if _cr_salvo is not None: st.session_state["ff_contas_receber_df"]=_cr_salvo
                if _cr_salvo is not None and not _cr_salvo.empty:
                    _col_fil_cr=col_filial(_cr_salvo)
                    _filial_atual_ff2=st.session_state.get("ff_filial_sel")
                    if _col_fil_cr and _filial_atual_ff2 and _filial_atual_ff2!="(Todas as filiais)":
                        _cr_salvo=_cr_salvo[_cr_salvo[_col_fil_cr].astype(str)==_filial_atual_ff2]
                    st.caption(f"📌 Já carregado: {len(_cr_salvo)} lançamento(s), total {fmt(_cr_salvo['Valor'].sum())}")

            if "ff_checkbox_incluir_pr" not in st.session_state:
                st.session_state["ff_checkbox_incluir_pr"]=st.session_state.get("ff_incluir_contas_pr",False)

            def _on_change_incluir_contas_pr():
                st.session_state["ff_incluir_contas_pr"]=st.session_state["ff_checkbox_incluir_pr"]
                if st.session_state.cid:
                    save_fluxo_financeiro(st.session_state.cid,st.session_state.get("ff_filial_sel"))

            st.session_state["ff_incluir_contas_pr"]=st.checkbox("☑️ Incluir essas Contas a Pagar/Receber no Fluxo de Caixa Real (gráficos com prazo aplicado)",key="ff_checkbox_incluir_pr",on_change=_on_change_incluir_contas_pr)

    with tab_resultado:

        with st.expander("📊 Fluxo de Caixa Competência — Compras x Faturamento",expanded=False):
            

            if df_cal_ff is None or df_cal_ff.empty:
                st.markdown('<div class="al-i">Sem compras programadas no momento — sem gráfico a exibir.</div>',unsafe_allow_html=True)
            else:
                df_fluxo_compras_ff=df_cal_ff.copy()
                df_fluxo_compras_ff["DataLimite_dt"]=pd.to_datetime(df_fluxo_compras_ff["DataEvento"],format="%d/%m/%Y",errors="coerce")
                df_fluxo_compras_ff["Semana"]=df_fluxo_compras_ff["DataLimite_dt"].dt.to_period("W").apply(lambda p: p.start_time)

                resumo_semanal_ff=df_fluxo_compras_ff.groupby("Semana")["ValorCompra"].sum().reset_index().rename(columns={"ValorCompra":"ValorSugerido"})
                resumo_semanal_ff=resumo_semanal_ff.sort_values("Semana")
                resumo_semanal_ff["Acumulado"]=resumo_semanal_ff["ValorSugerido"].cumsum()

                resumo_semanal_ff["Faturamento"]=0.0
                totais_mensais_ff={}
                mes_base_ff=pd.Timestamp.now().to_period("M").to_timestamp()
                _cfg_ml_ff=load_cfgml_resultado(st.session_state.cid,st.session_state.get("ff_filial_sel")) if st.session_state.cid else None
                if _cfg_ml_ff is None:
                    st.markdown(f'<div class="al-w">⚠️ Não encontrei Cenário ML rodado para <b>{st.session_state.get("ff_filial_sel") or "(Todas as filiais)"}</b> — rode-o primeiro em <b>🎛️ Motor de Previsão</b> nessa Filial, ou o Fluxo Projetado vai ficar incompleto (sem a parte de demanda prevista).</div>',unsafe_allow_html=True)
                _res_ml_ff=st.session_state.get("ml_produtos_resultado")
                _fonte_prev_ff=None
                if _cfg_ml_ff is not None and not _cfg_ml_ff.empty and "previsao" in _cfg_ml_ff.columns:
                    _fonte_prev_ff=_cfg_ml_ff
                elif _res_ml_ff is not None and not _res_ml_ff.empty and "previsao" in _res_ml_ff.columns:
                    _fonte_prev_ff=_res_ml_ff

                produtos_programados_ff=set(df_cal_ff["Produto"].unique()) if df_cal_ff is not None and not df_cal_ff.empty else set()

                if _fonte_prev_ff is None:
                    st.markdown('<div class="al-w">⚠️ Nenhuma previsão de vendas encontrada ainda — rode o <b>🎛️ Motor de Previsão</b> ou o <b>🧮 ML por Produto</b> primeiro para ver o Faturamento Projetado aqui.</div>',unsafe_allow_html=True)
                elif not produtos_programados_ff:
                    st.markdown('<div class="al-i">Sem produtos na programação de compras para comparar o faturamento.</div>',unsafe_allow_html=True)
                else:
                    _col_prod_prev_ff="_ProdutoUnico" if "_ProdutoUnico" in _fonte_prev_ff.columns else None
                    totais_mensais_ff={}
                    for _,row_prev in _fonte_prev_ff.iterrows():
                        if row_prev.get("status")!="ok": continue
                        if _col_prod_prev_ff and row_prev.get(_col_prod_prev_ff) not in produtos_programados_ff: continue
                        prev_lista=row_prev.get("previsao")
                        if not prev_lista: continue
                        for i_mes,val_mes in enumerate(prev_lista):
                            totais_mensais_ff[i_mes]=totais_mensais_ff.get(i_mes,0.0)+max(0.0,float(val_mes))

                    if totais_mensais_ff:
                        linhas_fat_mensal=[]
                        for i_mes,total_mes in totais_mensais_ff.items():
                            mes_calendario=mes_base_ff+pd.DateOffset(months=i_mes+1)
                            dias_no_mes=mes_calendario.days_in_month
                            valor_dia=total_mes/dias_no_mes
                            linhas_fat_mensal.append(pd.DataFrame({
                                "_dt": pd.date_range(mes_calendario, periods=dias_no_mes, freq="D"),
                                "_valor": valor_dia}))
                        df_fat_diario_ff=pd.concat(linhas_fat_mensal,ignore_index=True)
                        df_fat_diario_ff["Semana"]=df_fat_diario_ff["_dt"].dt.to_period("W").apply(lambda p: p.start_time)
                        fat_semanal_ff=df_fat_diario_ff.groupby("Semana")["_valor"].sum()
                        resumo_semanal_ff["Faturamento"]=resumo_semanal_ff["Semana"].map(fat_semanal_ff).fillna(0.0)

                if totais_mensais_ff:
                    ultimo_mes_com_previsao=max(totais_mensais_ff.keys())
                    limite_faturamento_ff=mes_base_ff+pd.DateOffset(months=ultimo_mes_com_previsao+2)
                    n_antes=len(resumo_semanal_ff)
                    resumo_semanal_ff=resumo_semanal_ff[resumo_semanal_ff["Semana"]<=limite_faturamento_ff].reset_index(drop=True)
                    if len(resumo_semanal_ff)<n_antes:
                        st.markdown(f'<div class="al-w">⚠️ A previsão de vendas cobre menos meses que o Calendário de Compras — comparação limitada até {limite_faturamento_ff.strftime("%d/%m/%Y")} (onde os dois lados têm dado real). Rode o ML com um horizonte maior para comparar o período completo.</div>',unsafe_allow_html=True)
                resumo_semanal_ff["FaturamentoAcumulado"]=resumo_semanal_ff["Faturamento"].cumsum()

                total_90dias_ff=df_fluxo_compras_ff["ValorCompra"].sum()
                c_fx1,c_fx2,c_fx3=st.columns(3)
                mc(c_fx1,"💸 Caixa Necessário (próximas semanas)",fmt(total_90dias_ff),"r")
                mc(c_fx2,"📅 Primeira Semana de Impacto",resumo_semanal_ff["Semana"].iloc[0].strftime("%d/%m/%Y") if len(resumo_semanal_ff)>0 else "—","b")
                mc(c_fx3,"📊 Semanas com Compra Programada",str(len(resumo_semanal_ff)),"b")

                media_semanal_ff=resumo_semanal_ff["ValorSugerido"].mean()
                limiar_critico_ff=media_semanal_ff*1.5
                cores_barras_ff=["#DC2626" if v>=limiar_critico_ff else "#14243B" for v in resumo_semanal_ff["ValorSugerido"]]

                fig_acum_ff=go.Figure()
                fig_acum_ff.add_trace(go.Scatter(
                x=resumo_semanal_ff["Semana"].dt.strftime("%d/%m"),
                y=resumo_semanal_ff["Acumulado"],
                name="Compras (Acumulado)",mode="lines+markers",
                line=dict(color="#A9762F",width=2.8),
                marker=dict(size=8,color="#A9762F",line=dict(color="white",width=1.5)),
                fill="tozeroy",fillcolor="rgba(169,118,47,0.08)",
                hovertemplate="<b>Semana %{x}</b><br>Compras Acumulado: R$ %{y:,.0f}<extra></extra>"))
                fig_acum_ff.add_trace(go.Scatter(
                x=resumo_semanal_ff["Semana"].dt.strftime("%d/%m"),
                y=resumo_semanal_ff["FaturamentoAcumulado"],
                name="Faturamento (Acumulado)",mode="lines+markers",
                line=dict(color="#059669",width=2.8),
                marker=dict(size=8,color="#059669",line=dict(color="white",width=1.5)),
                fill="tozeroy",fillcolor="rgba(5,150,105,0.08)",
                hovertemplate="<b>Semana %{x}</b><br>Faturamento Acumulado: R$ %{y:,.0f}<extra></extra>"))
                fig_acum_ff.update_layout(
                title=dict(text="Saída de Caixa Acumulada (Compras) x Faturamento Acumulado",
                    font=dict(size=14,family="Georgia, serif",color="#14243B")),
                plot_bgcolor="white",paper_bgcolor="white",
                font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
                margin=dict(l=10,r=10,t=44,b=40),
                xaxis=dict(showgrid=False,linecolor="#E5E7EB",tickangle=-40,tickfont=dict(size=9)),
                yaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=True,tickfont=dict(size=9),tickprefix="R$ "),
                legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.28,x=0.5,xanchor="center",font=dict(size=10)),
                hovermode="x unified",height=280,
                hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B",size=11)))
                st.plotly_chart(fig_acum_ff,use_container_width=True,key="ff_grafico_acumulado")

                fig_fluxo_ff=go.Figure()
                fig_fluxo_ff.add_trace(go.Bar(
                x=resumo_semanal_ff["Semana"].dt.strftime("%d/%m"),
                y=resumo_semanal_ff["ValorSugerido"],
                name="Compras (Saída Prevista)",
                marker=dict(color=cores_barras_ff,line=dict(color="white",width=0.8)),
                text=[fmt(v) for v in resumo_semanal_ff["ValorSugerido"]],
                textposition="outside",textfont=dict(size=9,color="#374151"),
                hovertemplate="<b>Semana %{x}</b><br>Saída (Compras): R$ %{y:,.0f}<extra></extra>"))
                fig_fluxo_ff.add_trace(go.Bar(
                x=resumo_semanal_ff["Semana"].dt.strftime("%d/%m"),
                y=resumo_semanal_ff["Faturamento"],
                name="Faturamento (Vendas)",
                marker=dict(color="#059669",line=dict(color="white",width=0.8)),
                text=[fmt(v) for v in resumo_semanal_ff["Faturamento"]],
                textposition="outside",textfont=dict(size=9,color="#374151"),
                hovertemplate="<b>Semana %{x}</b><br>Faturamento: R$ %{y:,.0f}<extra></extra>"))
                fig_fluxo_ff.add_hline(y=media_semanal_ff,
                line_dash="dot",line_color="#9CA3AF",
                annotation_text=f"Média compras: {fmt(media_semanal_ff)}",
                annotation_position="top right",
                annotation_font=dict(size=9,color="#9CA3AF"))
                if limiar_critico_ff<=resumo_semanal_ff["ValorSugerido"].max():
                    fig_fluxo_ff.add_annotation(
                        text="🔴 Semanas em vermelho = compras acima de 150% da média — atenção ao caixa",
                        xref="paper",yref="paper",x=0,y=-0.32,showarrow=False,
                        font=dict(size=9,color="#DC2626"),align="left")
                fig_fluxo_ff.update_layout(
                title=dict(text="Saída de Caixa (Compras) x Faturamento — por Semana",
                    font=dict(size=14,family="Georgia, serif",color="#14243B")),
                plot_bgcolor="white",paper_bgcolor="white",
                font=dict(color="#6B7280",size=10,family="Segoe UI, Arial"),
                margin=dict(l=10,r=10,t=44,b=70),
                barmode="group",
                xaxis=dict(showgrid=False,linecolor="#E5E7EB",tickangle=-40,tickfont=dict(size=9)),
                yaxis=dict(gridcolor="#F3F4F6",linecolor="#E5E7EB",showgrid=True,tickfont=dict(size=9),tickprefix="R$ "),
                legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.36,x=0.5,xanchor="center",font=dict(size=10)),
                hovermode="x unified",height=340,bargap=0.25,
                hoverlabel=dict(bgcolor="white",bordercolor="#E5E7EB",font=dict(color="#14243B",size=11)))
                st.plotly_chart(fig_fluxo_ff,use_container_width=True,key="ff_grafico_semanal")

        
        sec("📊 Fluxo de Caixa Estimado — Com Configurações Aplicadas")
        

        if soma_pag!=100 or soma_rec!=100:
            st.markdown('<div class="al-w">⚠️ Ajuste as parcelas na Configuração de Prazos acima (cada lado deve somar 100%) para calcular o fluxo real.</div>',unsafe_allow_html=True)
        elif df_cal_ff is None or df_cal_ff.empty:
            st.markdown('<div class="al-i">Sem compras programadas no momento — sem fluxo real a projetar.</div>',unsafe_allow_html=True)
        else:
            eventos_saida_real=[]
            df_saida_base_real=df_cal_ff.copy()
            df_saida_base_real["_data_evt"]=pd.to_datetime(df_saida_base_real["DataEvento"],format="%d/%m/%Y",errors="coerce")
            for _,row_s in df_saida_base_real.dropna(subset=["_data_evt"]).iterrows():
                valor_evt=float(row_s["ValorCompra"])
                for parc in st.session_state["ff_parcelas_pagar"]:
                    eventos_saida_real.append({
                        "Data": row_s["_data_evt"]+pd.Timedelta(days=int(parc["dias"])),
                        "Valor": valor_evt*parc["pct"]/100})

            eventos_entrada_real=[]
            data_limite_compras=df_saida_base_real["_data_evt"].max() if not df_saida_base_real.empty else None
            if totais_mensais_ff:
                for i_mes,total_mes in totais_mensais_ff.items():
                    mes_calendario=mes_base_ff+pd.DateOffset(months=i_mes+1)
                    if data_limite_compras is not None and mes_calendario>data_limite_compras:
                        continue
                    dias_no_mes=mes_calendario.days_in_month
                    valor_dia=total_mes/dias_no_mes
                    for d in range(dias_no_mes):
                        data_dia=mes_calendario+pd.Timedelta(days=d)
                        for parc in st.session_state["ff_parcelas_receber"]:
                            valor_parcela_receber=valor_dia*parc["pct"]/100
                            data_receb=data_dia+pd.Timedelta(days=int(parc["dias"]))
                            eventos_entrada_real.append({
                                "Data": data_receb,
                                "Valor": valor_parcela_receber})
                            if encargos_pct>0:
                                eventos_saida_real.append({
                                    "Data": data_receb,
                                    "Valor": valor_parcela_receber*encargos_pct/100})
            if st.session_state.get("ff_incluir_contas_pr",False):
                hoje_ff_pr=pd.Timestamp.now().normalize()
                _filial_atual_ff3=st.session_state.get("ff_filial_sel")
                df_cp_inc=st.session_state.get("ff_contas_pagar_df")
                if df_cp_inc is not None and not df_cp_inc.empty:
                    _col_fil_cp_inc=col_filial(df_cp_inc)
                    if _col_fil_cp_inc and _filial_atual_ff3 and _filial_atual_ff3!="(Todas as filiais)":
                        df_cp_inc=df_cp_inc[df_cp_inc[_col_fil_cp_inc].astype(str)==_filial_atual_ff3]
                    for _,row_cp in df_cp_inc[df_cp_inc["Vencimento"]>=hoje_ff_pr].iterrows():
                        eventos_saida_real.append({"Data":row_cp["Vencimento"],"Valor":float(row_cp["Valor"])})
                df_cr_inc=st.session_state.get("ff_contas_receber_df")
                if df_cr_inc is not None and not df_cr_inc.empty:
                    _col_fil_cr_inc=col_filial(df_cr_inc)
                    if _col_fil_cr_inc and _filial_atual_ff3 and _filial_atual_ff3!="(Todas as filiais)":
                        df_cr_inc=df_cr_inc[df_cr_inc[_col_fil_cr_inc].astype(str)==_filial_atual_ff3]
                    for _,row_cr in df_cr_inc[df_cr_inc["Vencimento"]>=hoje_ff_pr].iterrows():
                        eventos_entrada_real.append({"Data":row_cr["Vencimento"],"Valor":float(row_cr["Valor"])})

            df_saida_ev_real=pd.DataFrame(eventos_saida_real)
            df_entrada_ev_real=pd.DataFrame(eventos_entrada_real)
            df_saida_ev_real=pd.DataFrame(eventos_saida_real)
            df_entrada_ev_real=pd.DataFrame(eventos_entrada_real)

            if df_saida_ev_real.empty and df_entrada_ev_real.empty:
                st.markdown('<div class="al-i">Sem dados suficientes para projetar o fluxo real.</div>',unsafe_allow_html=True)
            else:
                if not df_saida_ev_real.empty:
                    df_saida_ev_real["Semana"]=df_saida_ev_real["Data"].dt.to_period("W").apply(lambda p: p.start_time)
                    saida_sem_real=df_saida_ev_real.groupby("Semana")["Valor"].sum().rename("SaidaReal")
                else:
                    saida_sem_real=pd.Series(dtype=float,name="SaidaReal")

                if not df_entrada_ev_real.empty:
                    df_entrada_ev_real["Semana"]=df_entrada_ev_real["Data"].dt.to_period("W").apply(lambda p: p.start_time)
                    entrada_sem_real=df_entrada_ev_real.groupby("Semana")["Valor"].sum().rename("EntradaReal")
                else:
                    entrada_sem_real=pd.Series(dtype=float,name="EntradaReal")

                resumo_real_ff=pd.concat([entrada_sem_real,saida_sem_real],axis=1).fillna(0).sort_index().reset_index().rename(columns={"index":"Semana"})
                resumo_real_ff["SaldoSemana"]=resumo_real_ff["EntradaReal"]-resumo_real_ff["SaidaReal"]
                resumo_real_ff["SaldoAcumulado"]=saldo_inicial+resumo_real_ff["SaldoSemana"].cumsum()

                saldo_final_real=resumo_real_ff["SaldoAcumulado"].iloc[-1] if len(resumo_real_ff)>0 else saldo_inicial
                semana_critica_real=resumo_real_ff.loc[resumo_real_ff["SaldoAcumulado"].idxmin()] if len(resumo_real_ff)>0 else None

                c_kr1,c_kr2,c_kr3=st.columns(3)
                mc(c_kr1,"🏦 Saldo Final Projetado",fmt(saldo_final_real),"g" if saldo_final_real>=0 else "r")
                if semana_critica_real is not None:
                    mc(c_kr2,"⚠️ Menor Saldo Projetado",fmt(semana_critica_real["SaldoAcumulado"]),"r" if semana_critica_real["SaldoAcumulado"]<0 else "b",
                    f"semana de {semana_critica_real['Semana'].strftime('%d/%m/%Y')}")
                mc(c_kr3,"📆 Semanas Projetadas",str(len(resumo_real_ff)),"b")

                semanas_negativas_snap=resumo_real_ff[resumo_real_ff["SaldoAcumulado"]<0]
                lista_semanas_negativas=[{
                    "semana":row_sn["Semana"].strftime("%d/%m/%Y"),
                    "saldo":float(row_sn["SaldoAcumulado"])
                } for _,row_sn in semanas_negativas_snap.iterrows()]
                semanas_deficit_semanal_snap=resumo_real_ff[resumo_real_ff["SaidaReal"]>resumo_real_ff["EntradaReal"]]
                lista_semanas_deficit_semanal=[{
                    "semana":row_sd["Semana"].strftime("%d/%m/%Y"),
                    "entrada":float(row_sd["EntradaReal"]),
                    "saida":float(row_sd["SaidaReal"]),
                    "diferenca":float(row_sd["EntradaReal"]-row_sd["SaidaReal"])
                } for _,row_sd in semanas_deficit_semanal_snap.iterrows()]
                st.session_state["fluxo_snapshot"]={
                    "saldo_final_real":float(saldo_final_real),
                    "menor_saldo":float(semana_critica_real["SaldoAcumulado"]) if semana_critica_real is not None else None,
                    "semana_critica":semana_critica_real["Semana"].strftime("%d/%m/%Y") if semana_critica_real is not None else None,
                    "semanas_negativas":lista_semanas_negativas,
                    "n_semanas_negativas":len(lista_semanas_negativas),
                    "semanas_deficit_semanal":lista_semanas_deficit_semanal,
                    "n_semanas_deficit_semanal":len(lista_semanas_deficit_semanal),
                    "total_entradas":float(resumo_real_ff["EntradaReal"].sum()),
                    "total_saidas":float(resumo_real_ff["SaidaReal"].sum()),
                    "n_semanas":len(resumo_real_ff),
                    "periodo_inicio":resumo_real_ff["Semana"].min().strftime("%d/%m/%Y") if len(resumo_real_ff)>0 else None,
                    "periodo_fim":resumo_real_ff["Semana"].max().strftime("%d/%m/%Y") if len(resumo_real_ff)>0 else None,
                    "incluiu_contas_pr":st.session_state.get("ff_incluir_contas_pr",False),
                    "gerado_em":datetime.now().strftime("%Y-%m-%d %H:%M"),
                }
                if st.session_state.cid: save_snap(st.session_state.cid,"fluxo",st.session_state["fluxo_snapshot"],filial=st.session_state.get('ff_filial_sel'))

                

                fig_real1=go.Figure()
                fig_real1.add_trace(go.Scatter(x=resumo_real_ff["Semana"].dt.strftime("%d/%m"),y=resumo_real_ff["EntradaReal"].cumsum(),
                    name="Entrada Real Acumulada",mode="lines+markers",line=dict(color="#00A86B",width=3.5),marker=dict(size=8,line=dict(color="white",width=1.5))))
                fig_real1.add_trace(go.Scatter(x=resumo_real_ff["Semana"].dt.strftime("%d/%m"),y=resumo_real_ff["SaidaReal"].cumsum(),
                    name="Saída Real Acumulada",mode="lines+markers",line=dict(color="#E11D2E",width=3.5),marker=dict(size=8,line=dict(color="white",width=1.5))))
                fig_real1.add_trace(go.Scatter(x=resumo_real_ff["Semana"].dt.strftime("%d/%m"),y=resumo_real_ff["SaldoAcumulado"],
                    name="Saldo de Caixa Acumulado",mode="lines+markers",line=dict(color="#F0A500",width=4.5),marker=dict(size=10,color="#F0A500",line=dict(color="#14243B",width=1.5)),
                    fill="tozeroy",fillcolor="rgba(240,165,0,0.12)"))
                fig_real1.add_hline(y=0,line_dash="dash",line_color="#14243B",line_width=1.5)
                fig_real1.update_layout(
                    title=dict(text="🔥 Fluxo de Caixa Estimado — Entradas x Saídas x Saldo (com prazos aplicados)",
                        font=dict(size=15,family="Georgia, serif",color="#14243B")),
                    plot_bgcolor="#FFFDF8",paper_bgcolor="#FFFDF8",
                    font=dict(color="#374151",size=10,family="Segoe UI, Arial"),
                    margin=dict(l=10,r=10,t=48,b=40),
                    xaxis=dict(showgrid=False,linecolor="#D1D5DB",tickangle=-40,tickfont=dict(size=9)),
                    yaxis=dict(gridcolor="#F0EAD8",linecolor="#D1D5DB",showgrid=True,tickfont=dict(size=9),tickprefix="R$ "),
                    legend=dict(bgcolor="rgba(0,0,0,0)",orientation="h",y=-0.3,x=0.5,xanchor="center",font=dict(size=10,color="#14243B")),
                    hovermode="x unified",height=340,
                    hoverlabel=dict(bgcolor="#14243B",bordercolor="#F0A500",font=dict(color="white",size=11)))
                st.plotly_chart(fig_real1,use_container_width=True,key="ff_grafico_real_acumulado")

                with st.expander("🔍 Conferir os números semana a semana (tabela crua)"):
                    df_conferencia=resumo_real_ff.copy()
                    df_conferencia["Semana"]=df_conferencia["Semana"].dt.strftime("%d/%m/%Y")
                    df_conferencia=df_conferencia.rename(columns={
                        "EntradaReal":"Entrada da Semana","SaidaReal":"Saída da Semana",
                        "SaldoSemana":"Diferença da Semana (Entrada-Saída)","SaldoAcumulado":"Saldo Acumulado"})
                    st.dataframe(df_conferencia,use_container_width=True,height=min(400,45+35*len(df_conferencia)))

                    n_semanas_vermelhas=int((resumo_real_ff["SaidaReal"]>resumo_real_ff["EntradaReal"]).sum())
                    n_semanas_verdes=int((resumo_real_ff["EntradaReal"]>=resumo_real_ff["SaidaReal"]).sum())
                    soma_deficit_vermelhas=float(resumo_real_ff.loc[resumo_real_ff["SaidaReal"]>resumo_real_ff["EntradaReal"],"SaldoSemana"].sum())
                    soma_superavit_verdes=float(resumo_real_ff.loc[resumo_real_ff["EntradaReal"]>=resumo_real_ff["SaidaReal"],"SaldoSemana"].sum())
                    st.markdown(f'''<div style="background:#F8F5EE;border-radius:8px;padding:12px 16px;margin-top:10px;font-size:.82rem">
                        <b>Resumo de conferência:</b><br>
                        🔴 {n_semanas_vermelhas} semana(s) com saída maior — soma do déficit: <b style="color:#DC2626">{fmt(soma_deficit_vermelhas)}</b><br>
                        🟢 {n_semanas_verdes} semana(s) com entrada maior ou igual — soma do superávit: <b style="color:#059669">{fmt(soma_superavit_verdes)}</b><br>
                        ➡️ Soma total (déficit + superávit) = <b>{fmt(soma_deficit_vermelhas+soma_superavit_verdes)}</b> — deve bater com o Saldo Acumulado da última semana da tabela acima.
                    </div>''',unsafe_allow_html=True)

                semanas_criticas_mask=resumo_real_ff["SaidaReal"]>resumo_real_ff["EntradaReal"]

                fig_real2=go.Figure()

                for idx_sc in resumo_real_ff.index[semanas_criticas_mask]:
                    fig_real2.add_vrect(x0=idx_sc-0.5,x1=idx_sc+0.5,fillcolor="rgba(225,29,46,0.06)",line_width=0,layer="below")

                fig_real2.add_trace(go.Bar(x=resumo_real_ff["Semana"].dt.strftime("%d/%m"),y=resumo_real_ff["EntradaReal"],
                    name="💚 Entrada Real (Recebimentos)",
                    marker=dict(color="#00C878",line=dict(color="#FFFFFF",width=1.2),opacity=0.92),
                    text=[fmt(v) if v>0 else "" for v in resumo_real_ff["EntradaReal"]],
                    textposition="outside",textfont=dict(size=9,color="#00874F",family="Segoe UI, Arial"),
                    hovertemplate="<b>Semana de %{x}</b><br>Entrada: R$ %{y:,.0f}<extra></extra>"))

                fig_real2.add_trace(go.Bar(x=resumo_real_ff["Semana"].dt.strftime("%d/%m"),y=resumo_real_ff["SaidaReal"],
                    name="🔴 Saída Real (Pagamentos)",
                    marker=dict(color="#F0142C",line=dict(color="#FFFFFF",width=1.2),opacity=0.92),
                    text=[fmt(v) for v in resumo_real_ff["SaidaReal"]],
                    textposition="outside",textfont=dict(size=9,color="#B00020",family="Segoe UI, Arial"),
                    hovertemplate="<b>Semana de %{x}</b><br>Saída: R$ %{y:,.0f}<extra></extra>"))

                fig_real2.update_layout(
                    title=dict(text="🔥 Entradas x Saídas Estimadas por Semana",
                        font=dict(size=18,family="Georgia, serif",color="#14243B"),
                        subtitle=dict(text="Faixas em vermelho claro = semanas onde a saída supera a entrada",
                            font=dict(size=11,color="#9CA3AF"))),
                    plot_bgcolor="#FFFEFB",paper_bgcolor="#FFFEFB",
                    font=dict(color="#374151",size=10,family="Segoe UI, Arial"),
                    margin=dict(l=10,r=10,t=70,b=70),
                    barmode="group",bargap=0.35,bargroupgap=0.1,
                    xaxis=dict(showgrid=False,linecolor="#E5E0D0",tickangle=-40,tickfont=dict(size=9,color="#6B7280")),
                    yaxis=dict(gridcolor="#F2ECD9",linecolor="#E5E0D0",showgrid=True,zeroline=True,zerolinecolor="#D1CBB8",zerolinewidth=1.5,tickfont=dict(size=9),tickprefix="R$ "),
                    legend=dict(bgcolor="rgba(255,255,255,0.7)",bordercolor="#F0A500",borderwidth=1,orientation="h",y=-0.34,x=0.5,xanchor="center",font=dict(size=10,color="#14243B")),
                    hovermode="x unified",height=380,
                    hoverlabel=dict(bgcolor="#14243B",bordercolor="#F0A500",font=dict(color="white",size=11,family="Segoe UI, Arial")))
                st.plotly_chart(fig_real2,use_container_width=True,key="ff_grafico_real_semanal")
                with st.expander("🔍 Ver o total de Entradas e Saídas deste gráfico"):
                    total_entrada_real=float(resumo_real_ff["EntradaReal"].sum())
                    total_saida_real=float(resumo_real_ff["SaidaReal"].sum())
                    diferenca_total_real=total_entrada_real-total_saida_real
                    st.markdown(f'''<div style="background:#F8F5EE;border-radius:8px;padding:12px 16px;font-size:.82rem">
                        🟢 Total de Entradas (todas as semanas somadas): <b style="color:#059669">{fmt(total_entrada_real)}</b><br>
                        🔴 Total de Saídas (todas as semanas somadas): <b style="color:#DC2626">{fmt(total_saida_real)}</b><br>
                        ➡️ Diferença (Entrada − Saída): <b style="color:{"#059669" if diferenca_total_real>=0 else "#DC2626"}">{fmt(diferenca_total_real)}</b><br>
                        <span style="color:#888">Esse valor de Diferença deve bater com o Saldo Acumulado da última semana no gráfico de linhas acima (mais o Saldo Inicial configurado).</span>
                    </div>''',unsafe_allow_html=True)
elif pg=="parecer_ia":
    hdr("🧭 Parecer Estratégico","Análise executiva com IA, cruzando as 3 filiais — Painel de Indicadores, Fluxo de Caixa, Configuração de ML e Fornecedores")

    _df_raw_par=get_df_raw_bruto()
    _col_fil_par=col_filial(_df_raw_par) if _df_raw_par is not None else None
    FILIAIS_PARECER=sorted(v for v in _df_raw_par[_col_fil_par].dropna().astype(str).unique().tolist() if v!="(Todas as filiais)") if _col_fil_par else []

    dados_par={}
    if st.session_state.cid and FILIAIS_PARECER:
        for _fil_p in FILIAIS_PARECER:
            dados_par[_fil_p]={
                "painel":load_snap(st.session_state.cid,"painel",_fil_p),
                "fluxo":load_snap(st.session_state.cid,"fluxo",_fil_p),
                "classe":load_snap(st.session_state.cid,"classe",_fil_p),
                "mlconfig":load_snap(st.session_state.cid,"mlconfig",_fil_p),
            }

    snap_pareto=st.session_state.get("pareto_resultado_atual")
    if snap_pareto is None and st.session_state.cid:
        _spr=load_pareto_snap(st.session_state.cid)
        if _spr is not None:
            st.session_state["pareto_resultado_atual"]=_spr
            snap_pareto=_spr
            _spv=load_snap(st.session_state.cid,"pareto_visao")
            if _spv: st.session_state["pareto_visao"]=_spv.get("visao","")
    pareto_visao_usada=st.session_state.get("pareto_visao","")

    sc_forn_global=load_scorecard_forn(st.session_state.cid) if st.session_state.cid else None
    if sc_forn_global is not None: st.session_state["gs_scorecard_forn"]=sc_forn_global

    sec("✅ Checklist — dado necessário por filial")
    if not FILIAIS_PARECER:
        st.markdown('<div class="al-w">⚠️ Nenhuma filial encontrada — importe os dados primeiro.</div>',unsafe_allow_html=True)
        st.stop()

    faltando_par=[]
    linhas_check=[]
    for _fil_p in FILIAIS_PARECER:
        _ok_painel="✅" if dados_par[_fil_p]["painel"] else "❌"
        _ok_fluxo="✅" if dados_par[_fil_p]["fluxo"] else "❌"
        _ok_classe="✅" if dados_par[_fil_p]["classe"] else "⚪ opcional"
        _ok_mlconfig="✅" if dados_par[_fil_p]["mlconfig"] else "⚪ opcional"
        linhas_check.append({"Filial":_fil_p,"Painel de Indicadores":_ok_painel,"Fluxo Comercial Projetado":_ok_fluxo,
                              "Simulador de Cenário":_ok_classe,"Config. ML":_ok_mlconfig})
        if not dados_par[_fil_p]["painel"]:
            faltando_par.append(f"{_fil_p}: abra o 📊 Dashboard Executivo")
        if not dados_par[_fil_p]["fluxo"]:
            faltando_par.append(f"{_fil_p}: abra o 💰 Fluxo de Caixa Comercial Projetado")
    st.dataframe(pd.DataFrame(linhas_check),use_container_width=True,hide_index=True)

    if faltando_par:
        st.markdown('<div class="al-w">⚠️ Ainda falta dado obrigatório pra gerar o parecer:<br>'+"<br>".join(faltando_par)+'</div>',unsafe_allow_html=True)

    pronto_par=len(faltando_par)==0

    if pronto_par:
        senha_gerar_parecer_com=st.text_input("Senha master para gerar",type="password",key="senha_gerar_parecer_com")
        if st.button("🚀 Gerar Parecer Estratégico",key="btn_gerar_parecer",use_container_width=True):
            if senha_gerar_parecer_com!=SENHA_MASTER:
                st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi executado.</div>',unsafe_allow_html=True)
                st.stop()
            if not st.session_state.api_key:
                st.markdown('<div class="al-d">❌ Nenhuma API Key configurada — vá em ⚙️ Configurações e cole sua chave da OpenAI ou Claude.</div>',unsafe_allow_html=True)
            else:
                with st.spinner("Analisando as 3 filiais e cruzando os dados..."):
                    forn_txt=""
                    sc_forn=sc_forn_global
                    if sc_forn is not None and not sc_forn.empty:
                        linhas_forn=[]
                        for _,r_f in sc_forn.iterrows():
                            linhas_forn.append(f"- {r_f['Fornecedor']}: Prazo {r_f['Prazo']}%, Qualidade {r_f['Qualidade']}%, OTIF {r_f['OTIF']}%")
                        forn_txt="\n".join(linhas_forn)
                    else:
                        forn_txt="Nenhum fornecedor com desempenho (Prazo/Qualidade/OTIF) cadastrado ainda."

                    pareto_txt="Não disponível — gere a Curva de Pareto."
                    if snap_pareto is not None and "classe_abc" in snap_pareto.columns:
                        n_total_p=len(snap_pareto)
                        n_a_p=int((snap_pareto["classe_abc"]=="A").sum())
                        n_b_p=int((snap_pareto["classe_abc"]=="B").sum())
                        n_c_p=int((snap_pareto["classe_abc"]=="C").sum())
                        pareto_txt=(f"Visão: {pareto_visao_usada} | Classe A (80% do valor): {n_a_p} de {n_total_p} itens "
                                    f"({n_a_p/n_total_p*100:.1f}% dos itens) | Classe B: {n_b_p} | Classe C: {n_c_p}")

                    # Monta um bloco de texto por filial, incluindo fornecedores calculados do Motor de Compras salvo
                    blocos_filiais=[]
                    for _fil_p in FILIAIS_PARECER:
                        _sp=dados_par[_fil_p]["painel"]
                        _sf=dados_par[_fil_p]["fluxo"]
                        _sc=dados_par[_fil_p]["classe"]
                        _sm=dados_par[_fil_p]["mlconfig"]

                        _radar_txt=(f"Ruptura Iminente {_sp.get('n_rup_im',0)} | Comprar Agora {_sp.get('n_comprar',0)} | "
                                    f"OK {_sp.get('n_ok',0)} | Estoque Excessivo {_sp.get('n_excesso',0)} | "
                                    f"Total Sugerido Compra Imediata {fmt(_sp.get('total_sugerido_imediato',0))}")

                        if _sp.get('cap_parado',0)>0:
                            _cap_parado_txt=f"{fmt(_sp['cap_parado'])} em {_sp.get('n_parado',0)} produto(s) com demanda próxima de zero"
                        else:
                            _cap_parado_txt="R$ 0 (nenhum produto com demanda tão próxima de zero a ponto de ser considerado morto)"

                        _classe_txt="não disponível nesta filial"
                        if _sc:
                            _linhas_c=[]
                            for c_item in _sc:
                                _liber=c_item["capital_atual"]-c_item["capital_ideal"]
                                _linhas_c.append(f"Classe {c_item['classe']} ({c_item['produtos']} produtos): Giro {c_item['giro_atual_xano']}x/ano "
                                                 f"(alvo {c_item['giro_alvo_xano']}x/ano), Capital {fmt(c_item['capital_atual'])} "
                                                 f"({'liberável' if _liber>=0 else 'necessário adicional'} de {fmt(abs(_liber))})")
                            _classe_txt=" | ".join(_linhas_c)

                        _mlconfig_txt="não disponível nesta filial"
                        if _sm:
                            _configs_ativas_txt=", ".join(_sm['configs_ativas']) if _sm.get('configs_ativas') else "padrão estatístico natural"
                            _mlconfig_txt=f"Escopo: {_sm.get('escopo_valor','—')} | Calibrações ativas: {_configs_ativas_txt}"

                        _sem_neg=_sf.get('semanas_negativas',[]) if _sf else []
                        if _sem_neg:
                            _pior_sem=min(_sem_neg,key=lambda s:s['saldo'])
                            _caixa_txt=(f"{len(_sem_neg)} semana(s) com SALDO ACUMULADO negativo, a mais crítica é {_pior_sem['semana']} "
                                        f"com {fmt(_pior_sem['saldo'])}")
                        elif _sf and _sf.get('menor_saldo') is not None and _sf['menor_saldo']<0:
                            _caixa_txt=f"menor saldo acumulado do horizonte é {fmt(_sf['menor_saldo'])} na semana de {_sf.get('semana_critica','—')}"
                        else:
                            _caixa_txt="nenhuma semana com saldo ACUMULADO negativo no horizonte projetado"

                        _sem_defc=_sf.get('semanas_deficit_semanal',[]) if _sf else []
                        if _sem_defc:
                            _pior_defc=min(_sem_defc,key=lambda s:s['diferenca'])
                            _caixa_txt+=(f". ALÉM DISSO: {len(_sem_defc)} semana(s) individuais tiveram SAÍDA maior que ENTRADA "
                                        f"naquela semana específica (mesmo o saldo acumulado total podendo estar positivo) — "
                                        f"a pior foi {_pior_defc['semana']}, com entrada de {fmt(_pior_defc['entrada'])} "
                                        f"contra saída de {fmt(_pior_defc['saida'])} (déficit de {fmt(abs(_pior_defc['diferenca']))})"
                                        f". Cite essas semanas de déficit semanal no parecer como um alerta à parte — "
                                        f"são semanas de atenção mesmo que o caixa acumulado não tenha ficado negativo.")

                        # Fornecedores desta filial, calculado ao vivo do Motor de Compras salvo (participação + capital parado)
                        _forn_txt_p="sem dado de fornecedor calculado para esta filial"
                        try:
                            _res_p,_cal_p,_morto_p=load_resultado_compras(st.session_state.cid,_fil_p)
                            if _res_p is not None and not _res_p.empty and "Fornecedor" in _res_p.columns and "ValorSugerido" in _res_p.columns:
                                _df_forn_p=_res_p[_res_p["ValorSugerido"]>0].copy()
                                if not _df_forn_p.empty:
                                    _score_p=_df_forn_p.groupby("Fornecedor").agg(
                                        Total=("ValorSugerido","sum"),SKUs=("Produto","nunique")).reset_index()
                                    _score_p["Part_pct"]=_score_p["Total"]/_score_p["Total"].sum()*100
                                    _cap_forn_p=pd.Series(dtype=float)
                                    if _morto_p is not None and not _morto_p.empty and "Fornecedor" in _morto_p.columns and "CapitalParado" in _morto_p.columns:
                                        _cap_forn_p=_morto_p.groupby("Fornecedor")["CapitalParado"].sum()
                                    _linhas_f=[]
                                    for _,r_f in _score_p.sort_values("Total",ascending=False).iterrows():
                                        _cp_f=_cap_forn_p.get(r_f["Fornecedor"],0)
                                        _linhas_f.append(f"{r_f['Fornecedor']}: {fmt(r_f['Total'])} sugerido ({r_f['Part_pct']:.1f}% do total desta filial), "
                                                        f"{int(r_f['SKUs'])} SKUs"+(f", capital parado {fmt(_cp_f)}" if _cp_f>0 else ""))
                                    _forn_txt_p="; ".join(_linhas_f) if _linhas_f else "nenhuma compra sugerida no momento"
                        except Exception:
                            pass

                        blocos_filiais.append(f"""
=== FILIAL: {_fil_p} ===
Score Executivo: {_sp['score_final']}/100 ({_sp['score_lbl']})
Valor do Estoque: {fmt(_sp['val_estoque'])} | Capital na Política Ideal: {fmt(_sp['cap_ideal'])} | Capital Liberável (excesso, giro normal): {fmt(_sp['cap_liberavel'])}
Capital Parado (sem giro / candidatos a liquidação): {_cap_parado_txt}
Ruptura: {_sp['pct_ruptura']:.1f}% dos produtos | Cobertura média: {_sp['cob_media']:.0f} dias | Giro médio: {_sp['giro_ano']}x/ano | Lead Time médio: {_sp['lt_medio']:.0f} dias
MAPE do modelo de previsão (validação out-of-sample): {f"{_sp['mape_ml']:.1f}%" if _sp.get('mape_ml') else "não validado ainda"}
Radar de Estoque — ação imediata: {_radar_txt}
Giro e Capital por Classe ABC/D/E: {_classe_txt}
Configuração do modelo de ML usado: {_mlconfig_txt}
Fluxo de Caixa Comercial Projetado: saldo final {fmt(_sf['saldo_final_real']) if _sf else '—'} em {_sf.get('periodo_fim','—') if _sf else '—'}; {_caixa_txt}
Fornecedores desta filial (participação e capital parado, calculado das compras sugeridas): {_forn_txt_p}
""")

                    bloco_todas_filiais="\n".join(blocos_filiais)

                    prompt_parecer=f"""Você é um consultor de Controladoria/FP&A sênior, analisando a operação de estoque e caixa de uma empresa com MÚLTIPLAS FILIAIS ({', '.join(FILIAIS_PARECER)}).

SUA TAREFA PRINCIPAL não é resumir cada filial isoladamente — é ENCONTRAR CORRELAÇÕES E PADRÕES entre elas. Especificamente, para cada ponto analisado, avalie:
- Esse padrão se repete nas 3 filiais (sugerindo causa sistêmica — ex: o modelo de previsão, uma política de estoque mal calibrada) ou aparece só em 1 (sugerindo causa local — ex: fornecedor específico daquela região, característica de demanda local)?
- Existe um fornecedor que aparece com concentração relevante em mais de uma filial? Se sim, isso é um risco de dependência AINDA MAIOR do que pareceria olhando cada filial isolada, porque o problema desse fornecedor com uma filial pode se repetir nas outras.
- Onde está concentrado o MAIOR capital liberável somando as 3 filiais — isso deve guiar a priorização das ações.
- As semanas críticas de caixa de filiais diferentes caem nas mesmas semanas do calendário (sugerindo um padrão sazonal da empresa toda) ou em semanas diferentes (sugerindo que cada filial tem ciclo de compra próprio)?

IMPORTANTE SOBRE DATAS: sempre que citar saldo, semana crítica ou projeção, diga a qual filial e a qual período se refere.

REFERÊNCIAS DE MERCADO PARA CLASSIFICAR CADA MÉTRICA (diga explicitamente onde cada número se encaixa):
- MAPE: <10% Excelente | 10-20% Bom | 20-30% Aceitável | >30% Baixa confiabilidade
- Cobertura de estoque (dias): <15 Enxuta | 15-45 Saudável | 45-90 Alta | >90 Excessiva
- Giro de estoque (x/ano): <4x Baixo | 4-8x Moderado | 8-15x Bom | >15x Excelente
- Lead Time médio: até 15d Ótimo | 16-30d Aceitável | >30d Elevado
- Ruptura (%): <5% Excelente | 5-15% Aceitável | 15-30% Atenção | >30% Crítico
- Fornecedores (Prazo/Qualidade/OTIF): ≥93% Ótimo | 85-92% Atenção | <85% Crítico
- Capital Parado (% do Estoque Total): <5% Saudável | 5-15% Atenção | >15% Crítico
- Concentração de fornecedor (participação nas compras): <20% Saudável | 20-30% Atenção | >30% Risco de dependência

GLOSSÁRIO — explique estes conceitos de forma educativa quando relevante:
Score Executivo: nota de 0-100 resumindo a saúde do estoque, considerando Ruptura (fator mais determinante, por afetar receita diretamente), Giro, Cobertura, Capital Parado, MAPE e Lead Time.
MAPE e Validação às Cegas (Out-of-Sample): o modelo é treinado só com dado até uma data de corte, "prevê" os meses seguintes sem ver o que aconteceu, e comparamos com a realidade já conhecida — garante que o erro medido reflete uma previsão real, sem "colar".
Como o motor de previsão escolhe modelo: para cada produto, testa vários modelos estatísticos e de Machine Learning diferentes (cada um captura tendência/sazonalidade/ruído de um jeito), e usa automaticamente o que teve menor erro para aquele produto especificamente — produtos diferentes podem usar modelos diferentes.
Calibrações disponíveis no sistema (capacidades, independente de estarem ativas nesta rodada): Sazonalidade (marcar meses de pico conhecidos), Promoções/Eventos (isolar esse efeito da sazonalidade normal), Exclusão de Outliers, Correção de Preço/Reajustes, Fatores Condicionantes de Mercado (câmbio, crise de fornecimento).

DADOS POR FILIAL:
{bloco_todas_filiais}

CURVA DE PARETO (concentração de vendas — empresa toda, classificação INDEPENDENTE do Giro/Capital por Classe acima; não misture os percentuais de uma com a outra):
{pareto_txt}

DESEMPENHO DE FORNECEDORES (Prazo/Qualidade/OTIF — avaliação manual, vale para todas as filiais igualmente):
{forn_txt}

TAREFA:
Escreva um parecer executivo em português, rico, profissional e ESPECÍFICO (não genérico), para um Coordenador/Gerente de Controladoria que supervisiona as {len(FILIAIS_PARECER)} filiais. Estruture assim:

**1. Diagnóstico Geral e Comparativo entre Filiais** — explique em 1-2 frases o que é o Score Executivo (métrica composta considerando Ruptura, Giro, Cobertura, Capital Parado, MAPE e Lead Time; Ruptura é o fator mais determinante por impactar receita diretamente). Depois cite o Score de CADA UMA das {len(FILIAIS_PARECER)} filiais individualmente (nunca resuma como "varia entre X e Y" sem nomear cada uma). Só afirme que a situação é "semelhante" ou "muito diferente" entre filiais se os números que você já vai citar nas seções seguintes realmente sustentarem essa afirmação — não conclua isso antes de checar.

**2. Radar de Estoque — Ação Imediata, por Filial** — para cada filial, cite: ruptura iminente/comprar agora (total sugerido), quantidade de produtos em Estoque Excessivo, e o valor de Capital Parado (sem giro). Esses 3 números (ruptura, estoque excessivo, capital parado) estão nos dados de cada filial abaixo — cite todos os 3 aqui, mesmo que só formalmente, porque a Seção 7 vai precisar reusar alguns deles. Se uma filial está bem pior que as outras em algum desses pontos, destaque isso com os números lado a lado.

**3. Estoque e Capital por Curva — Onde Focar Primeiro** — para CADA filial, cite o capital liberável (nunca omita nenhuma, mesmo que o valor seja pequeno ou zero) E o giro por classe (não só capital — cite pelo menos a Classe A de cada filial: giro atual vs alvo). Cruzando as filiais, diga onde está o maior capital liberável no total, para priorizar a ação. Você pode citar a Curva de Pareto como contexto adicional, mas NUNCA aplique o percentual dela sobre a Classe A do Giro/Capital — são bases diferentes.

**4. Confiabilidade da Previsão** — explique brevemente a validação às cegas e como o motor escolhe modelo por produto. Apresente as calibrações disponíveis como capacidade educativa, depois diga quais estavam ativas em cada filial nesta rodada (sem inventar nenhuma que não foi informada).

**5. Situação de Caixa — Padrão entre Filiais** — para cada filial com semana negativa, você recebeu a lista completa, cronológica, de todas as semanas negativas com seus saldos. NÃO liste as semanas uma por uma no texto — em vez disso, identifique e descreva o PADRÃO: a situação piora progressivamente até um ponto e depois melhora? Há um pico isolado destoante do resto? O problema é concentrado em um período específico (ex: um trimestre) ou espalhado ao longo de todo o horizonte? Cite a semana mais crítica como ponto de referência, mas a narrativa deve ser sobre a tendência, não sobre cada dado individual. Só afirme que as semanas críticas "coincidem" ou "não coincidem" entre filiais se houver pelo menos 2 filiais com semana crítica pra comparar — se só 1 filial tiver, diga apenas que só ela apresentou criticidade, sem tentar comparar padrão.

**6. Fornecedores — Concentração e Risco Cruzado** — identifique fornecedores que aparecem com participação relevante em mais de uma filial (risco de dependência combinado). VERIFICAÇÃO OBRIGATÓRIA: para cada fornecedor citado, compare o valor de compra sugerida com o capital parado que ele já tem — se um fornecedor tem capital parado relevante (produtos dele já parados sem giro) E ao mesmo tempo está recebendo uma sugestão de compra significativa, isso é um alerta específico que precisa ser dito explicitamente (ex: "Fornecedor X tem R$Y parado sem giro, mas ainda assim está recebendo sugestão de compra de R$Z — vale revisar se essa compra nova é realmente necessária antes de aumentar ainda mais o estoque parado desse fornecedor"). Não deixe essa checagem implícita — se ela se aplicar a algum fornecedor, escreva a frase de alerta. Traga também prazo/qualidade/OTIF quando relevante.

**7. Ações Recomendadas** — 6 a 9 ações práticas, priorizadas (mais urgente primeiro), cada uma dizendo A QUAL FILIAL (ou "todas") se aplica. REGRA OBRIGATÓRIA: toda ação deve referenciar um número, filial ou achado JÁ CITADO em uma das seções 1-6 acima — nunca introduza aqui um valor que não apareceu antes no texto. REGRA ANTI-REDUNDÂNCIA: cada ação deve tratar de um PROBLEMA diferente das outras — antes de escrever a lista final, revise se duas ações não estão recomendando a mesma coisa pra mesma filial só com números diferentes (ex: uma ação sobre "capital parado da Loja X" e outra sobre "capital liberável da Loja X" são a mesma recomendação disfarçada, a menos que sejam explicitamente duas ações distintas com objetivos diferentes) — se isso acontecer, funda as duas em uma ação só, mais completa, em vez de listar separadamente.

ANTES DE FINALIZAR, REVISE VOCÊ MESMO O TEXTO E CORRIJA SE PRECISAR:
- Todo número usado na Seção 7 apareceu em alguma seção anterior?
- Toda comparação entre filiais ("semelhante", "diferente", "coincide") tem pelo menos 2 valores reais sendo comparados?
- Toda filial citada tem seu nome formatado como "Loja Centro", "Loja Norte", "Loja Sul" (nunca "loja_centro" ou variações em minúsculo/underscore)?
- Alguma conclusão da Seção 1 contradiz um número citado depois, nas Seções 2-6? Se sim, ajuste a Seção 1 antes de responder.

Seja específico usando os números e filiais fornecidos. Não invente dados que não foram fornecidos — se algo estiver "não disponível", mencione brevemente que a análise ficaria mais completa com esse dado, sem inventar números."""

                    try:
                        texto_parecer=""
                        _debug_resp=None
                        if not st.session_state.api_key.startswith("sk-ant-") and OPENAI_OK:
                            client_ia=OpenAI(api_key=st.session_state.api_key)
                            resp_ia=client_ia.chat.completions.create(model="gpt-4o",max_tokens=6000,temperature=0.3,
                                messages=[{"role":"system","content":"Você é um consultor de Controladoria/FP&A sênior, direto, específico e rico em conexões entre os dados — especialmente entre filiais diferentes. Nunca omita seções ou resuma o conteúdo para economizar espaço — seja completo em todas as 7 seções pedidas."},
                                          {"role":"user","content":prompt_parecer}])
                            texto_parecer=resp_ia.choices[0].message.content.strip()
                            _debug_resp=f"finish_reason={resp_ia.choices[0].finish_reason}"
                        elif ANTHROPIC_OK:
                            client_ia=anthropic.Anthropic(api_key=st.session_state.api_key)
                            resp_ia=client_ia.messages.create(model="claude-sonnet-4-6",max_tokens=8000,
                                system="Você é um consultor de Controladoria/FP&A sênior, direto, específico e rico em conexões entre os dados — especialmente entre filiais diferentes.",
                                messages=[{"role":"user","content":prompt_parecer}])
                            _tipos_blocos=[getattr(b,"type","?") for b in resp_ia.content]
                            _debug_resp=f"stop_reason={resp_ia.stop_reason} | blocos={_tipos_blocos} | qtd_blocos={len(resp_ia.content)}"
                            for _bloco in resp_ia.content:
                                if getattr(_bloco,"type",None)=="text":
                                    texto_parecer=_bloco.text.strip()
                                    break

                        if texto_parecer:
                            st.session_state["parecer_atual"]=texto_parecer
                            _score_medio=sum(dados_par[f]["painel"]["score_final"] for f in FILIAIS_PARECER)/len(FILIAIS_PARECER)
                            _saldo_total=sum((dados_par[f]["fluxo"] or {}).get("saldo_final_real",0) for f in FILIAIS_PARECER)
                            parecer_salvo={
                                "data":datetime.now().strftime("%Y-%m-%d %H:%M"),
                                "score":round(_score_medio,1),
                                "saldo_final":_saldo_total,
                                "filiais":FILIAIS_PARECER,
                                "texto":texto_parecer,
                            }
                            if st.session_state.cid: save_parecer_ia(st.session_state.cid,parecer_salvo,tipo="comercial")
                        else:
                            _erro_import=f"OPENAI_OK={OPENAI_OK} (erro: {OPENAI_ERR if not OPENAI_OK and 'OPENAI_ERR' in dir() else 'n/a'}) | ANTHROPIC_OK={ANTHROPIC_OK} (erro: {ANTHROPIC_ERR if not ANTHROPIC_OK and 'ANTHROPIC_ERR' in dir() else 'n/a'})"
                            st.markdown(f'<div class="al-d">❌ Não foi possível gerar o parecer — resposta vazia da IA.<br>'
                                        f'<span style="font-size:.75rem;opacity:.8">DEBUG: {_debug_resp} | tamanho do prompt: {len(prompt_parecer)} caracteres<br>{_erro_import}</span></div>',unsafe_allow_html=True)
                    except Exception as e:
                        st.markdown(f'<div class="al-d">❌ Erro ao chamar a IA: {e}</div>',unsafe_allow_html=True)
    else:
        st.button("🚀 Gerar Parecer Estratégico",key="btn_gerar_parecer_desabilitado",use_container_width=True,disabled=True)

    if st.session_state.get("parecer_atual"):
        sec("📋 Parecer Gerado")
        texto_editado_atual=st.text_area("Edite o parecer se quiser ajustar algo:",
            value=st.session_state["parecer_atual"],height=420,key="edicao_parecer_atual")
        senha_salvar_atual=st.text_input("Senha master para salvar",type="password",key="senha_salvar_atual")
        if st.button("💾 Salvar edição",key="btn_salvar_edicao_atual",use_container_width=True):
            if senha_salvar_atual!=SENHA_MASTER:
                st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi salvo.</div>',unsafe_allow_html=True)
            else:
                st.session_state["parecer_atual"]=texto_editado_atual
                hist_atual=load_parecer_ia_historico(st.session_state.cid) if st.session_state.cid else []
                if hist_atual:
                    update_parecer_ia(st.session_state.cid,0,texto_editado_atual)
                st.markdown('<div class="al-s">✅ Edição salva.</div>',unsafe_allow_html=True)
        _p2_html_com=load_cli(st.session_state.cid) if st.session_state.cid else None
        _html_com=gerar_html_parecer("Parecer Estratégico — Módulo Comercial",
            "Motor de Compras, Estoque, ML de Demanda e Fornecedores",
            st.session_state["parecer_atual"],_p2_html_com.get("nome","") if _p2_html_com else "",
            cor_faixa="#0F6E56",cor_clara="#9FE1CB")
        st.download_button("📄 Baixar HTML (imprimir como PDF)",_html_com,
            file_name="parecer_comercial.html",mime="text/html",
            use_container_width=True,key="btn_html_comercial")

    sec("🕓 Histórico de Pareceres — Módulo Comercial")
    historico_parecer=load_parecer_ia_historico(st.session_state.cid,tipo="comercial") if st.session_state.cid else []
    if not historico_parecer:
        st.markdown('<div class="al-i">Nenhum parecer gerado ainda.</div>',unsafe_allow_html=True)
    else:
        for i_h,item_h in enumerate(historico_parecer):
            with st.expander(f"📅 {item_h['data']} — Score {item_h['score']}/100 · Saldo Final {fmt(item_h['saldo_final'])}"):
                texto_editado_hist=st.text_area("Texto (editável):",value=item_h["texto"],
                    height=380,key=f"edicao_parecer_hist_{i_h}")
                senha_hist_com=st.text_input("Senha master para salvar/apagar",type="password",key=f"senha_hist_com_{i_h}")
                c_hist1,c_hist2=st.columns(2)
                if c_hist1.button("💾 Salvar edição",key=f"salvar_parecer_hist_{i_h}",use_container_width=True):
                    if senha_hist_com!=SENHA_MASTER:
                        st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi salvo.</div>',unsafe_allow_html=True)
                    else:
                        if st.session_state.cid:
                            update_parecer_ia(st.session_state.cid,i_h,texto_editado_hist,tipo="comercial")
                        st.markdown('<div class="al-s">✅ Edição salva.</div>',unsafe_allow_html=True)
                if c_hist2.button("🗑️ Apagar este parecer",key=f"del_parecer_{i_h}",use_container_width=True):
                    if senha_hist_com!=SENHA_MASTER:
                        st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi apagado.</div>',unsafe_allow_html=True)
                    else:
                        if st.session_state.cid:
                            delete_parecer_ia(st.session_state.cid,i_h,tipo="comercial")
                        st.rerun()
                _p2_html_hist=load_cli(st.session_state.cid) if st.session_state.cid else None
                _html_hist=gerar_html_parecer("Parecer Estratégico — Módulo Comercial",
                    "Motor de Compras, Estoque, ML de Demanda e Fornecedores",
                    item_h["texto"],_p2_html_hist.get("nome","") if _p2_html_hist else "",
                    cor_faixa="#0F6E56",cor_clara="#9FE1CB")
                st.download_button("📄 Baixar HTML (imprimir como PDF)",_html_hist,
                    file_name="parecer_comercial.html",mime="text/html",
                    use_container_width=True,key=f"btn_html_hist_{i_h}")

    st.divider()
    sec("💼 Parecer Financeiro — Módulo Financeiro")

    FILIAIS_FIN=FILIAIS_PARECER if 'FILIAIS_PARECER' in dir() and FILIAIS_PARECER else (
        sorted(v for v in _df_raw_par[_col_fil_par].dropna().astype(str).unique().tolist() if v!="(Todas as filiais)") if _col_fil_par else [])

    dados_fin={}
    fin_consolidado=None
    if st.session_state.cid and FILIAIS_FIN:
        for _fil_f in FILIAIS_FIN:
            dados_fin[_fil_f]=load_snap(st.session_state.cid,"financeiro",_fil_f)
        fin_consolidado=load_snap(st.session_state.cid,"financeiro","(Todas as filiais)")

    sec("✅ Checklist — dado necessário (Financeiro)")
    if not FILIAIS_FIN:
        st.markdown('<div class="al-w">⚠️ Nenhuma filial encontrada — importe os dados primeiro.</div>',unsafe_allow_html=True)
    else:
        faltando_fin=[]
        linhas_check_fin=[]
        for _fil_f in FILIAIS_FIN:
            _ok_fin="✅" if dados_fin.get(_fil_f) else "❌"
            linhas_check_fin.append({"Filial":_fil_f,"Alertas e Diagnóstico (DRE)":_ok_fin})
            if not dados_fin.get(_fil_f):
                faltando_fin.append(f"{_fil_f}: abra 🚨 Alertas e Diagnóstico com essa filial selecionada")
        linhas_check_fin.append({"Filial":"(Todas as filiais) — Liquidez/Kanitz/ROE","Alertas e Diagnóstico (DRE)":"✅" if fin_consolidado else "❌"})
        if not fin_consolidado:
            faltando_fin.append('"(Todas as filiais)": abra 🚨 Alertas e Diagnóstico com "(Todas as filiais)" selecionada')
        st.dataframe(pd.DataFrame(linhas_check_fin),use_container_width=True,hide_index=True)
        if faltando_fin:
            st.markdown('<div class="al-w">⚠️ Ainda falta dado obrigatório:<br>'+"<br>".join(faltando_fin)+'</div>',unsafe_allow_html=True)
        pronto_fin=len(faltando_fin)==0

        if pronto_fin:
            senha_gerar_parecer_fin=st.text_input("Senha master para gerar",type="password",key="senha_gerar_parecer_fin")
            if st.button("🚀 Gerar Parecer Financeiro",key="btn_gerar_parecer_fin",use_container_width=True):
                if senha_gerar_parecer_fin!=SENHA_MASTER:
                    st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi executado.</div>',unsafe_allow_html=True)
                    st.stop()
                if not st.session_state.api_key:
                    st.markdown('<div class="al-d">❌ Nenhuma API Key configurada.</div>',unsafe_allow_html=True)
                else:
                    with st.spinner("Analisando o Módulo Financeiro..."):
                        blocos_fin=[]
                        for _fil_f in FILIAIS_FIN:
                            _df_f=dados_fin[_fil_f]
                            _alertas_txt=" | ".join(f"[{a[0]}] {a[1]}"+(f" ({a[2]})" if a[2] else "") for a in _df_f.get("alertas",[]))
                            blocos_fin.append(f"""
=== FILIAL: {_fil_f} (período: {_df_f.get('periodo','—')}) ===
Receita Bruta: {fmt(_df_f['receita_bruta'])} | Receita Líquida: {fmt(_df_f['receita_liquida'])}
Lucro Bruto: {fmt(_df_f.get('lucro_bruto',0))} | Lucro Líquido: {fmt(_df_f.get('lucro_liquido',0))} | EBITDA: {fmt(_df_f.get('ebitda_rs',0))}
Margem Bruta: {_df_f['margem_bruta_pct']:.1f}% | Margem Líquida: {_df_f['margem_liquida_pct']:.1f}% | Margem Contribuição: {_df_f['margem_contrib_pct']:.1f}% | EBITDA: {_df_f['ebitda_pct']:.1f}%
PMR: {_df_f['pmr']:.0f}d | PMP: {_df_f['pmp']:.0f}d | PME: {_df_f['pme']:.0f}d | Ciclo de Caixa: {_df_f['ciclo_caixa']:.0f}d | Giro de Estoque: {_df_f.get('giro_estoque',0):.1f}x
Score de Saúde: {_df_f['score_saude']:.0f}/100
Alertas já classificados (referências de mercado já aplicadas): {_alertas_txt if _alertas_txt else 'nenhum alerta'}
""")
                        bloco_fin_todas="\n".join(blocos_fin)

                        _alertas_cons_txt=" | ".join(f"[{a[0]}] {a[1]}"+(f" ({a[2]})" if a[2] else "") for a in (fin_consolidado.get("alertas",[]) if fin_consolidado else []))
                        bloco_consolidado=f"""
=== EMPRESA TODA — "(Todas as filiais)" (período: {fin_consolidado.get('periodo','—') if fin_consolidado else '—'}) ===
Estas métricas usam contas do Balanço Patrimonial (Capital Social, Empréstimos, Patrimônio Líquido) que NÃO são segregadas por filial — por isso só existem no nível consolidado, nunca compare-as entre as lojas.
Liquidez Corrente: {fin_consolidado['liquidez_corrente']:.2f}x | Liquidez Imediata: {fin_consolidado['liquidez_imediata']:.2f}x
Kanitz: {fin_consolidado['kanitz']:.2f} | ROE: {fin_consolidado['roe']:.1f}% | ICD: {fin_consolidado.get('icd',0):.1f}%
Ativo Total: {fmt(fin_consolidado.get('ativo_total',0))} | Passivo Total: {fmt(fin_consolidado.get('passivo_total',0))} | Patrimônio Líquido: {fmt(fin_consolidado.get('patrimonio_liquido',0))}
Alertas já classificados: {_alertas_cons_txt if _alertas_cons_txt else 'nenhum alerta'}
""" if fin_consolidado else "Não disponível."

                        prompt_fin=f"""Você é um consultor de Controladoria/FP&A sênior, analisando a saúde financeira de uma empresa com múltiplas filiais ({', '.join(FILIAIS_FIN)}), a partir da DRE e do Balanço Patrimonial.

REGRA CRÍTICA SOBRE ESCOPO: Receita, Margens, Ciclo de Caixa (PMR/PMP/PME) e Score de Saúde são calculados por filial e PODEM ser comparados entre elas. Liquidez, Kanitz e ROE vêm do Balanço Patrimonial, que tem contas corporativas (Capital Social, Empréstimos, Patrimônio Líquido) não segregadas por loja — por isso essas 3 métricas SÓ existem no nível "(Todas as filiais)" e NUNCA devem ser atribuídas ou comparadas entre lojas individualmente. Não invente uma "Liquidez da Loja Centro", por exemplo — isso não existe nos dados.

Os alertas fornecidos abaixo JÁ FORAM CLASSIFICADOS usando faixas de referência de mercado padrão (ex: Margem Líquida <5% é atenção, Liquidez Corrente <1x é crítico, Kanitz <-3 é zona insolvente) — você não precisa reclassificar, só usar essa classificação já pronta na sua análise.

DADOS POR FILIAL (Receita, Margens, Ciclo de Caixa, Score):
{bloco_fin_todas}

DADOS CONSOLIDADOS (Liquidez, Kanitz, ROE — só nível empresa toda):
{bloco_consolidado}

TAREFA:
Escreva um parecer executivo em português, rico, específico e SEM repetir a mesma recomendação disfarçada duas vezes. Estruture assim:

**1. Diagnóstico Geral** — Score de Saúde de cada filial, e a situação de Liquidez/Kanitz/ROE da empresa como um todo (nível consolidado, nunca por loja).

**2. Rentabilidade e Margens, por Filial** — compare Margem Bruta/Líquida/Contribuição/EBITDA (%) entre as filiais, citando também os valores em R$ (Lucro Bruto, Lucro Líquido, EBITDA) quando ajudar a dar escala à comparação. Aponte qual filial está com a rentabilidade mais fraca e cite os alertas relacionados já classificados. ANÁLISE OBRIGATÓRIA: calcule e compare a taxa de conversão de Lucro Bruto para Lucro Líquido (Lucro Líquido ÷ Lucro Bruto) entre as filiais — uma filial com Lucro Bruto alto mas conversão baixa (ex: perde muito entre bruto e líquido) tem um problema de ESTRUTURA DE DESPESA, não de venda; uma filial com Lucro Bruto baixo mas conversão saudável tem o problema na ORIGEM (venda/CMV), não na despesa. Diga explicitamente qual dos dois casos cada filial se encaixa.

**3. Ciclo Operacional e Prazos, por Filial** — compare PMR/PMP/PME/Ciclo de Caixa e Giro de Estoque entre as filiais. Cite o que um ciclo de caixa maior ou giro de estoque menor significam em termos de capital de giro.

**4. Saúde Patrimonial da Empresa** — Liquidez Corrente/Imediata, Kanitz, ROE e ICD, no nível consolidado, junto com Ativo Total, Passivo Total e Patrimônio Líquido para dar contexto de tamanho. Explique brevemente o que o Kanitz representa (zona de insolvência/penumbra/solvência).

**5. Ações Recomendadas** — 5 a 8 ações práticas, priorizadas, cada uma dizendo a qual filial se aplica (ou "empresa toda" para as ações relacionadas a Liquidez/Kanitz/ROE). REGRA: toda ação deve referenciar um número já citado nas seções 1-4 acima. REGRA ANTI-REDUNDÂNCIA: cada ação trata de um problema diferente — não recomende a mesma coisa duas vezes com números diferentes. REGRA DE CONSISTÊNCIA ENTRE FILIAIS: se duas ou mais filiais têm o MESMO valor (ou valor muito próximo) num indicador problemático, todas elas devem receber o mesmo tratamento na recomendação — ou você inclui todas as filiais afetadas na mesma ação, ou explica explicitamente por que só uma delas está sendo priorizada apesar do número ser igual nas outras. Nunca destaque só 1 filial silenciosamente quando o dado mostra que outra(s) têm o problema igual. REGRA DE COERÊNCIA DE CAUSA: se a Seção 2 já concluiu que a causa da rentabilidade fraca de uma filial é ESTRUTURA DE DESPESA (conversão Bruto→Líquido baixa) ou é ORIGEM/VENDA (Margem Bruta baixa), nenhuma ação sobre essa mesma filial pode sugerir investigar a causa oposta — a ação precisa ser consistente com o diagnóstico já dado, nunca contradizê-lo. A taxa de conversão Bruto→Líquido é o sinal PRINCIPAL de causa — "essa filial tem a Margem Bruta mais baixa entre as 3" não é motivo suficiente para recomendar ação de origem/venda se a diferença for pequena (poucos pontos percentuais) e a conversão já tiver apontado estrutura de despesa como causa principal; só recomende ação de origem/venda se a Margem Bruta estiver em faixa realmente baixa em termos absolutos (abaixo de 20%, conforme referência de mercado), não apenas relativamente mais baixa que as outras filiais.

Seja específico. Não invente dados que não foram fornecidos."""

                        try:
                            texto_fin=""
                            if not st.session_state.api_key.startswith("sk-ant-") and OPENAI_OK:
                                client_fin=OpenAI(api_key=st.session_state.api_key)
                                resp_fin=client_fin.chat.completions.create(model="gpt-4o",max_tokens=5000,temperature=0.3,
                                    messages=[{"role":"system","content":"Você é um consultor de Controladoria/FP&A sênior, direto e específico. Nunca atribua Liquidez, Kanitz ou ROE a uma filial individual — essas métricas só existem no nível consolidado."},
                                              {"role":"user","content":prompt_fin}])
                                texto_fin=resp_fin.choices[0].message.content.strip()
                            elif ANTHROPIC_OK:
                                client_fin=anthropic.Anthropic(api_key=st.session_state.api_key)
                                resp_fin=client_fin.messages.create(model="claude-sonnet-4-6",max_tokens=3500,
                                    system="Você é um consultor de Controladoria/FP&A sênior, direto e específico. Nunca atribua Liquidez, Kanitz ou ROE a uma filial individual — essas métricas só existem no nível consolidado.",
                                    messages=[{"role":"user","content":prompt_fin}])
                                texto_fin=resp_fin.content[0].text.strip()

                            if texto_fin:
                                st.session_state["parecer_fin_atual"]=texto_fin
                                parecer_fin_salvo={
                                    "data":datetime.now().strftime("%Y-%m-%d %H:%M"),
                                    "score":round(sum(dados_fin[f]["score_saude"] for f in FILIAIS_FIN)/len(FILIAIS_FIN),1),
                                    "saldo_final":fin_consolidado.get("liquidez_corrente",0) if fin_consolidado else 0,
                                    "filiais":FILIAIS_FIN,
                                    "texto":texto_fin,
                                }
                                if st.session_state.cid: save_parecer_ia(st.session_state.cid,parecer_fin_salvo,tipo="financeiro")
                            else:
                                st.markdown('<div class="al-d">❌ Resposta vazia da IA.</div>',unsafe_allow_html=True)
                        except Exception as e:
                            st.markdown(f'<div class="al-d">❌ Erro ao chamar a IA: {e}</div>',unsafe_allow_html=True)
        else:
            st.button("🚀 Gerar Parecer Financeiro",key="btn_gerar_parecer_fin_desabilitado",use_container_width=True,disabled=True)

    if st.session_state.get("parecer_fin_atual"):
        sec("📋 Parecer Financeiro Gerado")
        texto_editado_fin=st.text_area("Edite o parecer se quiser ajustar algo:",
            value=st.session_state["parecer_fin_atual"],height=420,key="edicao_parecer_fin_atual")
        senha_salvar_fin=st.text_input("Senha master para salvar",type="password",key="senha_salvar_fin")
        if st.button("💾 Salvar edição",key="btn_salvar_edicao_fin_atual",use_container_width=True):
            if senha_salvar_fin!=SENHA_MASTER:
                st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi salvo.</div>',unsafe_allow_html=True)
            else:
                st.session_state["parecer_fin_atual"]=texto_editado_fin
                hist_fin_atual=load_parecer_ia_historico(st.session_state.cid,tipo="financeiro") if st.session_state.cid else []
                if hist_fin_atual:
                    update_parecer_ia(st.session_state.cid,0,texto_editado_fin,tipo="financeiro")
                st.markdown('<div class="al-s">✅ Edição salva.</div>',unsafe_allow_html=True)
        _p2_html_fin=load_cli(st.session_state.cid) if st.session_state.cid else None
        _html_fin=gerar_html_parecer("Parecer Estratégico — Módulo Financeiro",
            "DRE, Margens, Liquidez e Ciclo de Caixa",
            st.session_state["parecer_fin_atual"],_p2_html_fin.get("nome","") if _p2_html_fin else "",
            cor_faixa="#0F6E56",cor_clara="#9FE1CB")
        st.download_button("📄 Baixar HTML (imprimir como PDF)",_html_fin,
            file_name="parecer_financeiro.html",mime="text/html",
            use_container_width=True,key="btn_html_financeiro")

    sec("🕓 Histórico de Pareceres — Módulo Financeiro")
    historico_parecer_fin=load_parecer_ia_historico(st.session_state.cid,tipo="financeiro") if st.session_state.cid else []
    if not historico_parecer_fin:
        st.markdown('<div class="al-i">Nenhum parecer financeiro gerado ainda.</div>',unsafe_allow_html=True)
    else:
        for i_hf,item_hf in enumerate(historico_parecer_fin):
            with st.expander(f"📅 {item_hf['data']} — Score médio {item_hf['score']}/100"):
                texto_editado_histf=st.text_area("Texto (editável):",value=item_hf["texto"],
                    height=380,key=f"edicao_parecer_histf_{i_hf}")
                senha_hist_fin=st.text_input("Senha master para salvar/apagar",type="password",key=f"senha_hist_fin_{i_hf}")
                c_histf1,c_histf2=st.columns(2)
                if c_histf1.button("💾 Salvar edição",key=f"salvar_parecer_histf_{i_hf}",use_container_width=True):
                    if senha_hist_fin!=SENHA_MASTER:
                        st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi salvo.</div>',unsafe_allow_html=True)
                    else:
                        if st.session_state.cid:
                            update_parecer_ia(st.session_state.cid,i_hf,texto_editado_histf,tipo="financeiro")
                        st.markdown('<div class="al-s">✅ Edição salva.</div>',unsafe_allow_html=True)
                if c_histf2.button("🗑️ Apagar este parecer",key=f"del_parecer_fin_{i_hf}",use_container_width=True):
                    if senha_hist_fin!=SENHA_MASTER:
                        st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi apagado.</div>',unsafe_allow_html=True)
                    else:
                        if st.session_state.cid:
                            delete_parecer_ia(st.session_state.cid,i_hf,tipo="financeiro")
                        st.rerun()
                _p2_html_histf=load_cli(st.session_state.cid) if st.session_state.cid else None
                _html_histf=gerar_html_parecer("Parecer Estratégico — Módulo Financeiro",
                    "DRE, Margens, Liquidez e Ciclo de Caixa",
                    item_hf["texto"],_p2_html_histf.get("nome","") if _p2_html_histf else "",
                    cor_faixa="#0F6E56",cor_clara="#9FE1CB")
                st.download_button("📄 Baixar HTML (imprimir como PDF)",_html_histf,
                    file_name="parecer_financeiro.html",mime="text/html",
                    use_container_width=True,key=f"btn_html_histf_{i_hf}")

    st.divider()
    sec("🎯 Parecer Consolidado Final — Comercial + Financeiro")
    

    _hist_com_final=load_parecer_ia_historico(st.session_state.cid,tipo="comercial") if st.session_state.cid else []
    _hist_fin_final=load_parecer_ia_historico(st.session_state.cid,tipo="financeiro") if st.session_state.cid else []
    _texto_com_final=_hist_com_final[0]["texto"] if _hist_com_final else st.session_state.get("parecer_atual")
    _texto_fin_final=_hist_fin_final[0]["texto"] if _hist_fin_final else st.session_state.get("parecer_fin_atual")

    _falta_final=[]
    if not _texto_com_final: _falta_final.append("Gere o Parecer Comercial primeiro (acima nesta mesma página).")
    if not _texto_fin_final: _falta_final.append("Gere o Parecer Financeiro primeiro (acima nesta mesma página).")

    if _falta_final:
        st.markdown('<div class="al-w">⚠️ '+"<br>".join(_falta_final)+'</div>',unsafe_allow_html=True)
    else:
        senha_gerar_parecer_fnl=st.text_input("Senha master para gerar",type="password",key="senha_gerar_parecer_fnl")
        if st.button("🚀 Gerar Parecer Consolidado Final",key="btn_gerar_parecer_final",use_container_width=True):
            if senha_gerar_parecer_fnl!=SENHA_MASTER:
                st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi executado.</div>',unsafe_allow_html=True)
                st.stop()
            if not st.session_state.api_key:
                st.markdown('<div class="al-d">❌ Nenhuma API Key configurada.</div>',unsafe_allow_html=True)
            else:
                with st.spinner("Cruzando os dois relatórios..."):
                    prompt_final=f"""Você é um consultor sênior de Controladoria/FP&A. Você já recebeu, prontos e validados, dois relatórios separados sobre a mesma empresa: um do MÓDULO COMERCIAL (Motor de Compras, Estoque, ML de demanda, Fornecedores) e outro do MÓDULO FINANCEIRO (DRE, Margens, Liquidez, Ciclo de Caixa).

SUA TAREFA não é repetir os dois relatórios um atrás do outro — é ENCONTRAR CONEXÕES REAIS entre eles. Por exemplo (só exemplos, use os achados reais dos relatórios abaixo, não invente):
- Capital parado em estoque (Comercial) pode explicar parte de um Ciclo de Caixa elevado (Financeiro).
- Ruptura de estoque numa filial (Comercial) pode explicar uma Receita ou Margem mais fraca nessa mesma filial (Financeiro).
- Concentração de compra num fornecedor (Comercial) pode se conectar com uma semana de caixa mais apertada (Financeiro).
- Uma filial com bom desempenho nos dois relatórios reforça que ela é referência; uma filial fraca nos dois reforça prioridade.

REGRA: só afirme uma conexão entre os dois relatórios se os dados de AMBOS realmente sustentarem ela — não force uma correlação que não existe nos textos. Se não houver conexão clara entre um achado comercial e um financeiro, tudo bem trazê-lo isolado, mas priorize as conexões reais que você encontrar.

RELATÓRIO COMERCIAL (já gerado, validado):
{_texto_com_final}

RELATÓRIO FINANCEIRO (já gerado, validado):
{_texto_fin_final}

TAREFA — estruture o parecer consolidado assim:

**1. Visão Executiva Consolidada** — 4-6 frases resumindo a saúde geral da empresa, cruzando as duas fontes, para um leitor que só vai ler esse resumo.

**2. Correlações Encontradas entre Comercial e Financeiro** — liste as conexões reais e específicas que você encontrou entre os dois relatórios (mínimo 3, máximo 6). Cada uma deve citar o achado comercial e o achado financeiro que se conectam, e explicar a relação.

**3. Panorama por Filial** — para cada filial, uma síntese curta cruzando a situação comercial e financeira dela juntas (é a loja com melhor ou pior desempenho geral? Por quê, considerando os dois lados?).

**4. Prioridades Finais da Empresa** — 6 a 10 ações finais, priorizadas (mais urgente primeiro), fundindo as recomendações dos dois relatórios originais. Remova duplicatas, junte ações relacionadas dos dois lados numa só quando fizer sentido, e diga a qual filial (ou "empresa toda") cada uma se aplica. Toda ação deve ter origem rastreável em um dos dois relatórios acima — não invente ação nova sem base neles.

Seja específico, use os números e filiais já citados nos dois relatórios. Não invente dado novo."""

                    try:
                        texto_final=""
                        if not st.session_state.api_key.startswith("sk-ant-") and OPENAI_OK:
                            client_final=OpenAI(api_key=st.session_state.api_key)
                            resp_final=client_final.chat.completions.create(model="gpt-4o",max_tokens=6000,temperature=0.3,
                                messages=[{"role":"system","content":"Você é um consultor de Controladoria/FP&A sênior, especialista em conectar achados operacionais e financeiros. Nunca force uma correlação que os dados não sustentam."},
                                          {"role":"user","content":prompt_final}])
                            texto_final=resp_final.choices[0].message.content.strip()
                        elif ANTHROPIC_OK:
                            client_final=anthropic.Anthropic(api_key=st.session_state.api_key)
                            resp_final=client_final.messages.create(model="claude-sonnet-4-6",max_tokens=4500,
                                system="Você é um consultor de Controladoria/FP&A sênior, especialista em conectar achados operacionais e financeiros. Nunca force uma correlação que os dados não sustentam.",
                                messages=[{"role":"user","content":prompt_final}])
                            texto_final=resp_final.content[0].text.strip()

                        if texto_final:
                            st.session_state["parecer_final_atual"]=texto_final
                            parecer_final_salvo={
                                "data":datetime.now().strftime("%Y-%m-%d %H:%M"),
                                "score":0,"saldo_final":0,
                                "texto":texto_final,
                            }
                            if st.session_state.cid: save_parecer_ia(st.session_state.cid,parecer_final_salvo,tipo="consolidado")
                        else:
                            st.markdown('<div class="al-d">❌ Resposta vazia da IA.</div>',unsafe_allow_html=True)
                    except Exception as e:
                        st.markdown(f'<div class="al-d">❌ Erro ao chamar a IA: {e}</div>',unsafe_allow_html=True)

    if st.session_state.get("parecer_final_atual"):
        sec("📋 Parecer Consolidado Final Gerado")
        texto_editado_final=st.text_area("Edite o parecer se quiser ajustar algo:",
            value=st.session_state["parecer_final_atual"],height=460,key="edicao_parecer_final_atual")
        senha_salvar_final=st.text_input("Senha master para salvar",type="password",key="senha_salvar_final")
        if st.button("💾 Salvar edição",key="btn_salvar_edicao_final_atual",use_container_width=True):
            if senha_salvar_final!=SENHA_MASTER:
                st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi salvo.</div>',unsafe_allow_html=True)
            else:
                st.session_state["parecer_final_atual"]=texto_editado_final
                hist_final_atual=load_parecer_ia_historico(st.session_state.cid,tipo="consolidado") if st.session_state.cid else []
                if hist_final_atual:
                    update_parecer_ia(st.session_state.cid,0,texto_editado_final,tipo="consolidado")
                st.markdown('<div class="al-s">✅ Edição salva.</div>',unsafe_allow_html=True)
        _p2_html_final=load_cli(st.session_state.cid) if st.session_state.cid else None
        _html_final=gerar_html_parecer("Parecer Consolidado Final",
            "Cruzamento entre os Módulos Comercial e Financeiro",
            st.session_state["parecer_final_atual"],_p2_html_final.get("nome","") if _p2_html_final else "",
            cor_faixa="#0F6E56",cor_clara="#9FE1CB")
        st.download_button("📄 Baixar HTML (imprimir como PDF)",_html_final,
            file_name="parecer_consolidado_final.html",mime="text/html",
            use_container_width=True,key="btn_html_consolidado_atual")

    sec("🕓 Histórico de Pareceres — Consolidado Final")
    historico_parecer_final=load_parecer_ia_historico(st.session_state.cid,tipo="consolidado") if st.session_state.cid else []
    if not historico_parecer_final:
        st.markdown('<div class="al-i">Nenhum parecer consolidado final gerado ainda.</div>',unsafe_allow_html=True)
    else:
        for i_hff,item_hff in enumerate(historico_parecer_final):
            with st.expander(f"📅 {item_hff['data']}"):
                texto_editado_histff=st.text_area("Texto (editável):",value=item_hff["texto"],
                    height=420,key=f"edicao_parecer_histff_{i_hff}")
                senha_hist_final=st.text_input("Senha master para salvar/apagar",type="password",key=f"senha_hist_final_{i_hff}")
                c_histff1,c_histff2=st.columns(2)
                if c_histff1.button("💾 Salvar edição",key=f"salvar_parecer_histff_{i_hff}",use_container_width=True):
                    if senha_hist_final!=SENHA_MASTER:
                        st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi salvo.</div>',unsafe_allow_html=True)
                    else:
                        if st.session_state.cid:
                            update_parecer_ia(st.session_state.cid,i_hff,texto_editado_histff,tipo="consolidado")
                        st.markdown('<div class="al-s">✅ Edição salva.</div>',unsafe_allow_html=True)
                if c_histff2.button("🗑️ Apagar este parecer",key=f"del_parecer_final_{i_hff}",use_container_width=True):
                    if senha_hist_final!=SENHA_MASTER:
                        st.markdown('<div class="al-d">❌ Senha master incorreta — nada foi apagado.</div>',unsafe_allow_html=True)
                    else:
                        if st.session_state.cid:
                            delete_parecer_ia(st.session_state.cid,i_hff,tipo="consolidado")
                        st.rerun()
                _p2_html_histff=load_cli(st.session_state.cid) if st.session_state.cid else None
                _html_histff=gerar_html_parecer("Parecer Consolidado Final",
                    "Cruzamento entre os Módulos Comercial e Financeiro",
                    item_hff["texto"],_p2_html_histff.get("nome","") if _p2_html_histff else "",
                    cor_faixa="#0F6E56",cor_clara="#9FE1CB")
                st.download_button("📄 Baixar HTML (imprimir como PDF)",_html_histff,
                    file_name="parecer_consolidado_final.html",mime="text/html",
                    use_container_width=True,key=f"btn_html_histff_{i_hff}")

        
                
# ── EXPORTAR ────────────────────────────────────────
elif pg=="exportar":
    hdr("💾 Exportar","CSV, Excel multi-abas e JSON para Power BI")
    _df_raw_exp=get_df_raw_bruto()
    _col_fil_exp=col_filial(_df_raw_exp) if _df_raw_exp is not None else None
    filial_sel_exp=None
    if _col_fil_exp:
        _filiais_disp_exp=sorted(v for v in _df_raw_exp[_col_fil_exp].dropna().astype(str).unique().tolist() if v!="(Todas as filiais)")
        _opcoes_filial_exp=["(Todas as filiais)"]+_filiais_disp_exp

        def _on_change_filial_exp():
            st.session_state["exp_filial_sel_backup"]=st.session_state["exp_filial_sel"]

        if "exp_filial_sel_backup" not in st.session_state:
            st.session_state["exp_filial_sel_backup"]=_opcoes_filial_exp[0]
        if st.session_state["exp_filial_sel_backup"] not in _opcoes_filial_exp:
            st.session_state["exp_filial_sel_backup"]=_opcoes_filial_exp[0]
        st.session_state["exp_filial_sel"]=st.session_state["exp_filial_sel_backup"]

        filial_sel_exp=st.selectbox("🏬 Filial",_opcoes_filial_exp,
          key="exp_filial_sel",on_change=_on_change_filial_exp)

    df=get_df_filial(filial_sel_exp)
    if df is None: no_data(); st.stop()

    # Garante que a projeção exportada é da MESMA filial selecionada aqui
    if st.session_state.get("projecoes_filial_atual")!=filial_sel_exp:
        st.session_state.projecoes={}
        st.session_state["projecoes_filial_atual"]=filial_sel_exp
        if st.session_state.cid:
            _proj_disco_exp=load_projecoes_ml(st.session_state.cid,filial_sel_exp)
            if _proj_disco_exp:
                st.session_state.projecoes=_proj_disco_exp

    p2=load_cli(st.session_state.cid) if st.session_state.cid else {}
    nid=gid(p2.get("nome","dados") if p2 else "dados")
    df_exp=df.copy()
    if "Data" in df_exp.columns:
        df_exp["Data"]=pd.to_datetime(df_exp["Data"],errors="coerce").dt.strftime("%d/%m/%Y")
    if st.session_state.projecoes:
        st.markdown(f'<div class="al-i">📊 {len(st.session_state.projecoes)} campos de projeção incluídos (Filial: {filial_sel_exp or "(Todas as filiais)"})</div>',unsafe_allow_html=True)
        for c2_,pe in st.session_state.projecoes.items():
            vals_e=pe["valores"]
            serie_e=pd.Series([np.nan]*(len(df_exp)-len(vals_e))+vals_e)
            df_exp[f"{c2_} (Proj.)"]=serie_e.values[:len(df_exp)]
    st.markdown(f'<div class="al-s">✅ <b>{len(df_exp)}</b> períodos · <b>{df_exp.shape[1]}</b> campos</div>',unsafe_allow_html=True)
    c1,c2,c3=st.columns(3)
    with c1:
        csv=df_exp.to_csv(sep=";",decimal=",",index=False,encoding="utf-8-sig").encode("utf-8-sig")
        st.download_button("📥 CSV (Power BI)",csv,file_name=f"{nid}_analytics.csv",mime="text/csv",use_container_width=True)
    with c2:
        buf=BytesIO()
        with pd.ExcelWriter(buf,engine="openpyxl") as w:
            df_exp.to_excel(w,index=False,sheet_name="Dados")
            dre_c=[c for c in ["Ano","mês","receita bruta de vendas","receita líquida","lucro bruto","lucro líquido","EBITDA","margem líquida %","EBITDA %"] if c in df_exp.columns]
            if dre_c: df_exp[dre_c].to_excel(w,index=False,sheet_name="DRE")
            bal_c=[c for c in ["Ano","mês","ativo total","ativo circ","pass total","pass circ","PL","liquidez corrente"] if c in df_exp.columns]
            if bal_c: df_exp[bal_c].to_excel(w,index=False,sheet_name="Balanço")
        buf.seek(0)
        st.download_button("📥 Excel (3 abas)",buf.getvalue(),file_name=f"{nid}_analytics.xlsx",
          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
    with c3:
        st.download_button("📥 JSON",df_exp.to_json(orient="records",force_ascii=False,indent=2).encode(),
          file_name=f"{nid}_analytics.json",mime="application/json",use_container_width=True)
    sec("👁️ Preview")
    cm2=cm_(df_exp); ca2=ca_(df_exp)
    prev=[c for c in [ca2,cm2,"receita bruta de vendas","lucro bruto","lucro líquido","EBITDA %","kanitz","score_risco"] if c and c in df_exp.columns]
    st.dataframe(df_exp[prev] if prev else df_exp,use_container_width=True,height=380)
    if st.session_state.log:
        st.divider()
        with st.expander("📋 Log"):
            for l in st.session_state.log[:30]: st.caption(l)

elif pg=="gestao_estoque":
    st.markdown("""<style>
.ge-kpi-box{background:#fff;border:1px solid #EBEBEB;border-radius:10px;padding:7px 6px;text-align:center}
.block-container{padding-top:1rem!important}
.ge-kpi-lbl{font-size:.57rem;color:#999;text-transform:uppercase;letter-spacing:.05em;margin-bottom:3px;line-height:1.3}
.ge-kpi-val{font-family:Georgia,serif;font-size:1.05rem;font-weight:700;color:#14243B;line-height:1.1}
.ge-kpi-sub{font-size:.60rem;margin-top:2px}
.ge-kpi-sub.up{color:#059669}.ge-kpi-sub.dn{color:#DC2626}.ge-kpi-sub.nt{color:#aaa}
.ge-sec{font-family:Georgia,serif;font-size:.78rem;font-weight:700;color:#14243B;
  padding:5px 10px;border-left:3px solid #F0A500;background:#FFFBF0;
  border-radius:0 5px 5px 0;margin-bottom:8px}
.ge-mini{border-radius:8px;padding:8px 10px;text-align:center;margin-bottom:6px}
.ge-mini-lbl{font-size:.58rem;text-transform:uppercase;letter-spacing:.05em;margin-bottom:3px;opacity:.8}
.ge-mini-val{font-family:Georgia,serif;font-size:.95rem;font-weight:700;line-height:1}
.ge-mini-sub{font-size:.58rem;margin-top:2px;opacity:.7}
.ge-radar{display:grid;grid-template-columns:1fr 1fr;gap:5px}
.ge-radar-item{border-radius:7px;padding:7px 5px;text-align:center}
.ge-radar-n{font-family:Georgia,serif;font-size:1.1rem;font-weight:800;line-height:1}
.ge-radar-l{font-size:.58rem;text-transform:uppercase;letter-spacing:.03em;margin-top:2px;opacity:.8}
.ge-alert-list{list-style:none;padding:0;margin:0}
.ge-alert-list li{display:flex;align-items:flex-start;gap:6px;padding:5px 0;
  border-bottom:1px solid #F5F5F5;font-size:.72rem;color:#444;line-height:1.4}
.ge-alert-list li:last-child{border-bottom:none}
.ge-table{width:100%;border-collapse:collapse;font-size:.68rem}
.ge-table th{background:#F8F5EE;color:#666;padding:4px 6px;text-align:center;font-weight:600;border-bottom:1px solid #EBEBEB}
.ge-table th:first-child{text-align:left}
.ge-table td{padding:4px 6px;text-align:center;border-bottom:1px solid #F5F5F5;color:#333}
.ge-table td:first-child{text-align:left;font-weight:500;font-size:.65rem}
.ge-pct{font-weight:700;font-size:.72rem}
.ge-pct.g{color:#059669}.ge-pct.y{color:#D97706}.ge-pct.r{color:#DC2626}
.ge-cal-grid{display:grid;grid-template-columns:1fr 1fr;gap:5px}
.ge-cal-item{background:#F8F5EE;border-radius:7px;padding:7px 6px;text-align:center}
.ge-cal-n{font-family:Georgia,serif;font-size:.95rem;font-weight:700;color:#14243B;line-height:1}
.ge-cal-l{font-size:.57rem;color:#888;text-transform:uppercase;letter-spacing:.03em;margin-top:2px}
</style>""", unsafe_allow_html=True)

    # HEADER (score será inserido após cálculo)
    _header_placeholder = st.empty()

    # Seletor de Filial — decide qual resultado salvo (por filial) carregar do disco.
    # O Motor de Compras já salva um arquivo separado por filial; aqui só falta escolher qual ler.
    _df_v_ge=get_vendas_df()
    _col_fil_ge=col_filial(_df_v_ge) if _df_v_ge is not None else None
    filial_sel_ge=None
    if _col_fil_ge:
        _filiais_disp_ge=sorted(v for v in _df_v_ge[_col_fil_ge].dropna().astype(str).unique().tolist() if v!="(Todas as filiais)")
        _opcoes_filial_ge=["(Todas as filiais)"]+_filiais_disp_ge

        def _on_change_filial_ge():
            st.session_state["ge_filial_sel_backup"]=st.session_state["ge_filial_sel"]

        if "ge_filial_sel_backup" not in st.session_state:
            st.session_state["ge_filial_sel_backup"]=_opcoes_filial_ge[0]
        if st.session_state["ge_filial_sel_backup"] not in _opcoes_filial_ge:
            st.session_state["ge_filial_sel_backup"]=_opcoes_filial_ge[0]
        st.session_state["ge_filial_sel"]=st.session_state["ge_filial_sel_backup"]

        filial_sel_ge=st.selectbox("🏬 Filial",_opcoes_filial_ge,
          key="ge_filial_sel",on_change=_on_change_filial_ge)

    # DADOS — sempre carregados do disco pra filial selecionada (não usa cache de sessão,
    # que poderia ser de outra filial rodada por último no Motor de Compras)
    df_est      = st.session_state.get("compras_df_estoque")
    df_prog_gs  = st.session_state.get("compras_prog")
    df_comp_val=None
    if st.session_state.cid:
        _val_full=load_validacao_full(st.session_state.cid,filial_sel_ge)
        if _val_full is not None and _val_full.get("df_comp") is not None:
            df_comp_val=_val_full["df_comp"]

    df_res=df_cal_gs=df_morto_gs=None
    if st.session_state.cid:
        df_res,df_cal_gs,df_morto_gs=load_resultado_compras(st.session_state.cid,filial_sel_ge)

    if df_res is None or df_res.empty:
        st.markdown('<div class="al-w">⚠️ Rode o <b>Motor de Compras</b> primeiro para essa filial, para popular este painel.</div>', unsafe_allow_html=True)
        st.stop()

    def fv(v):
        if abs(v)>=1e6: return f"R$ {v/1e6:.2f}mi"
        if abs(v)>=1e3: return f"R$ {v/1e3:.1f}k"
        return f"R$ {v:.0f}"

    # KPIs
    val_estoque   = float((df_res["EstoqueAtual"]*df_res["CustoUnitario"]).sum()) if "EstoqueAtual" in df_res.columns else 0
    # Giro ponderado por capital — mais representativo financeiramente
    if "CoberturaDias" in df_res.columns and "CustoUnitario" in df_res.columns and "EstoqueAtual" in df_res.columns:
        df_giro_v=df_res[df_res["CoberturaDias"].notna()&(df_res["CoberturaDias"]<999)&(df_res["CoberturaDias"]>0)].copy()
        if len(df_giro_v)>0:
            giro_ano=float((df_giro_v["GiroAtual"]*df_giro_v["EstoqueAtual"]*df_giro_v["CustoUnitario"]).sum()/(df_giro_v["EstoqueAtual"]*df_giro_v["CustoUnitario"]).sum())
            giro_ano=round(giro_ano,1)
        else:
            giro_ano=0
    else:
        cobertura_media_pnl = float(df_res["CoberturaDias"].replace(0,np.nan).mean()) if "CoberturaDias" in df_res.columns else 0
        giro_ano   = round(365/cobertura_media_pnl,1) if cobertura_media_pnl>0 else 0
    if "GiroAlvo" in df_res.columns and "EstoqueAtual" in df_res.columns and "CustoUnitario" in df_res.columns:
        peso_capital_alvo = df_res["EstoqueAtual"]*df_res["CustoUnitario"]
        giro_alvo_ano = round(float((df_res["GiroAlvo"]*peso_capital_alvo).sum()/peso_capital_alvo.sum()),1) if peso_capital_alvo.sum()>0 else 0
    else:
        giro_alvo_ano = 0
    # Cobertura Média ponderada por capital, mesma metodologia do Giro Médio acima —
    # média simples de dias favorece demais produtos com cobertura muito alta (poucos itens, valor baixo).
    if "df_giro_v" in dir() and len(df_giro_v)>0:
        _peso_cob=df_giro_v["EstoqueAtual"]*df_giro_v["CustoUnitario"]
        cob_media=float((df_giro_v["CoberturaDias"]*_peso_cob).sum()/_peso_cob.sum()) if _peso_cob.sum()>0 else 0
    else:
        cob_media = float(df_res["CoberturaDias"].dropna().mean()) if "CoberturaDias" in df_res.columns else 0
    n_rup_im  = int(len(df_res[df_res["Status"]=="🚨 RUPTURA IMINENTE"])) if "Status" in df_res.columns else 0
    n_comprar = int(len(df_res[df_res["Status"]=="🔴 Comprar agora"])) if "Status" in df_res.columns else 0
    n_ok      = int(len(df_res[df_res["Status"]=="🟢 OK"])) if "Status" in df_res.columns else 0
    n_excesso = int(len(df_res[df_res["Status"]=="🟡 Estoque excessivo"])) if "Status" in df_res.columns else 0
    n_ruptura = n_rup_im + n_comprar
    pct_ruptura = round(n_ruptura/len(df_res)*100,1) if len(df_res)>0 else 0
    cap_parado=0; n_parado=0
    if df_morto_gs is not None and not df_morto_gs.empty and "CapitalParado" in df_morto_gs.columns:
        cap_parado=float(df_morto_gs["CapitalParado"].sum()); n_parado=len(df_morto_gs)
    lt_medio      = float(df_res["LeadTimeDias"].mean()) if "LeadTimeDias" in df_res.columns else 0
    if "MinDias" in df_res.columns and "MaxDias" in df_res.columns and "DemandaPrevMes(un)" in df_res.columns:
        dias_alvo_prod   = (df_res["MinDias"]+df_res["MaxDias"])/2
        demanda_dia_prod = df_res["DemandaPrevMes(un)"]/30
        cap_ideal = float((dias_alvo_prod*demanda_dia_prod*df_res["CustoUnitario"]).sum())
    else:
        cap_ideal = 0
    cap_liberavel = max(0, val_estoque-cap_ideal)
    # Calendário
    n_eventos=0; n_prods_cal=0; total_horizon=0; caixa_sem=0; n_semanas=0; primeira_sem="—"
    periodo_cal="—"
    if df_cal_gs is not None and not df_cal_gs.empty:
        n_eventos=len(df_cal_gs)
        n_prods_cal=df_cal_gs["Produto"].nunique() if "Produto" in df_cal_gs.columns else 0
        total_horizon=float(df_cal_gs["ValorCompra"].sum()) if "ValorCompra" in df_cal_gs.columns else 0
        if "DataEvento" in df_cal_gs.columns:
            try:
                datas_cal=pd.to_datetime(df_cal_gs["DataEvento"],format="%d/%m/%Y")
                periodo_cal=f"{datas_cal.min().strftime('%d/%m/%y')} a {datas_cal.max().strftime('%d/%m/%y')}"
            except: pass
    if df_prog_gs is not None and not df_prog_gs.empty and "ValorCompra" in df_prog_gs.columns:
        caixa_sem=float(df_prog_gs["ValorCompra"].sum())
        if "DataEvento" in df_prog_gs.columns:
            try:
                df_prog_gs["_dt"]=pd.to_datetime(df_prog_gs["DataEvento"],format="%d/%m/%Y")
                df_prog_gs["_sem"]=df_prog_gs["_dt"].dt.to_period("W")
                n_semanas=df_prog_gs["_sem"].nunique()
                primeira_sem=df_prog_gs["_dt"].min().strftime("%d/%m/%Y")
            except: pass
    # MAPE — usa o filtro de confiabilidade salvo pelo botão "💾 Salvar" no Motor de Previsão.
    # Lê sempre do disco (nunca da sessão), pra não depender de sincronia entre páginas.
    _limite_erro_ge=20
    if st.session_state.cid:
        _cfg_ml_disco_ge=load_config_ml(st.session_state.cid)
        if _cfg_ml_disco_ge and "cfgml_limite_erro" in _cfg_ml_disco_ge:
            _limite_erro_ge=_cfg_ml_disco_ge["cfgml_limite_erro"]

    mape_ml=None
    if df_comp_val is not None and not df_comp_val.empty and "Erro %" in df_comp_val.columns:
        if "Produto" in df_comp_val.columns:
            _erro_medio_prod_ge=df_comp_val.groupby("Produto")["Erro %"].mean()
            _produtos_dentro_ge=_erro_medio_prod_ge[_erro_medio_prod_ge<=_limite_erro_ge].index
            _df_comp_filtrado_ge=df_comp_val[df_comp_val["Produto"].isin(_produtos_dentro_ge)]
            if not _df_comp_filtrado_ge.empty:
                mape_ml=round(float(_df_comp_filtrado_ge["Erro %"].mean()),1)
        if mape_ml is None:
            mape_ml=round(float(df_comp_val["Erro %"].mean()),1)        

            


    # ── SCORE DE SAÚDE DO ESTOQUE ───────────────────────────────────────
    score_giro     = min(100, max(0, round((giro_ano/giro_alvo_ano)*100))) if giro_alvo_ano>0 else 50
    score_cobertura= min(100, max(0, round((1-(abs(cob_media-37)/37))*100)))
    score_ruptura  = min(100, max(0, round(100-(pct_ruptura*3))))
    score_capital  = min(100, max(0, round((1-cap_parado/val_estoque)*100))) if val_estoque>0 else 100
    score_lead     = min(100, max(0, round(100-max(0,(lt_medio-7)*3)))) if lt_medio>0 else 70
    score_ml       = min(100, max(0, round((1-mape_ml/100)*100))) if mape_ml else 70
    score_final    = max(0, min(100, round(score_giro*0.15 + score_cobertura*0.15 + score_ruptura*0.35 + score_capital*0.15 + score_lead*0.05 + score_ml*0.15)))

    if score_final>=85:   score_cor="#059669"; score_lbl="Excelente"; score_stars="★★★★★"
    elif score_final>=70: score_cor="#059669"; score_lbl="Bom";       score_stars="★★★★☆"
    elif score_final>=55: score_cor="#D97706"; score_lbl="Regular";   score_stars="★★★☆☆"
    elif score_final>=40: score_cor="#DC2626"; score_lbl="Atenção";   score_stars="★★☆☆☆"
    else:                 score_cor="#7f1d1d"; score_lbl="Crítico";   score_stars="★☆☆☆☆"

    total_sug_snap=float(df_res[df_res["ValorSugerido"]>0]["ValorSugerido"].sum()) if "ValorSugerido" in df_res.columns else 0
    st.session_state["painel_snapshot"]={
        "score_final":score_final,"score_lbl":score_lbl,
        "val_estoque":val_estoque,"cap_ideal":cap_ideal,"cap_liberavel":cap_liberavel,
        "cap_parado":cap_parado,"n_parado":n_parado,
        "pct_ruptura":pct_ruptura,"n_rup_im":n_rup_im,"n_comprar":n_comprar,
        "n_ok":n_ok,"n_excesso":n_excesso,"total_sugerido_imediato":total_sug_snap,
        "cob_media":cob_media,"lt_medio":lt_medio,"giro_ano":giro_ano,
        "mape_ml":mape_ml,
        "gerado_em":datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    if st.session_state.cid: save_snap(st.session_state.cid,"painel",st.session_state["painel_snapshot"],filial=filial_sel_ge)

    # Renderizar header com score
    _header_placeholder.markdown(f'''<div style="background:linear-gradient(135deg,#0F6E56 0%,#085041 100%);
      border-radius:14px;padding:22px 32px;margin-bottom:4px;box-shadow:0 4px 14px rgba(0,0,0,.12);
      display:flex;align-items:center;justify-content:space-between;filter:invert(1) hue-rotate(180deg)">
      <div style="display:flex;align-items:center;gap:10px">
        <div style="background:#9FE1CB;border-radius:7px;width:34px;height:34px;display:flex;align-items:center;justify-content:center;font-size:18px">📦</div>
        <div>
          <div style="font-family:Georgia,serif;font-size:1.05rem;font-weight:600;color:#fff">PAINEL EXECUTIVO | GESTÃO INTELIGENTE DE ESTOQUES</div>
          <div style="color:#9FE1CB;font-size:.78rem;margin-top:4px;letter-spacing:.02em">Da previsão de demanda ao planejamento de compras.</div>
        </div>
      </div>
      <div style="display:flex;align-items:center;gap:10px">
        <div style="text-align:right">
          <div style="font-size:.6rem;color:#C9A876;text-transform:uppercase;letter-spacing:.08em">Score NetExame</div>
          <div style="font-size:.85rem;color:{score_cor}">{score_stars}</div>
        </div>
        <div style="display:flex;align-items:baseline;gap:3px;background:{score_cor}26;border:1px solid {score_cor}55;
          border-radius:12px;padding:4px 14px">
          <div style="font-family:Georgia,serif;font-size:2.2rem;font-weight:800;color:{score_cor};line-height:1">{score_final}</div>
          <div style="font-size:.75rem;color:{score_cor};opacity:.85">/100</div>
        </div>
        <div style="font-size:.75rem;font-weight:700;color:{score_cor};background:{score_cor}22;padding:3px 10px;border-radius:20px">{score_lbl}</div>
      </div>
    </div>''', unsafe_allow_html=True)

    st.markdown(f"""<div title="Ruptura {pct_ruptura:.1f}% → {max(0,round(100-(pct_ruptura*3)))}/100 | Giro {giro_ano}x → {score_giro}/100 | Cobertura {cob_media:.0f}d → {score_cobertura}/100 | Capital {fv(cap_parado)} → {score_capital}/100 | ML {f'{mape_ml:.1f}%' if mape_ml else '—'} → {score_ml}/100 | Lead Time {lt_medio:.0f}d → {score_lead}/100"
      style="display:inline-block;cursor:help"></div>""", unsafe_allow_html=True)

    # ── LINHA 1: KPIs ──────────────────────────────────────────────────
    k1,k2,k3,k4,k5,k6,k7,k8=st.columns(8)

    # k2 — Giro
    _giro_status = "🟢 Acima do alvo" if giro_ano>=giro_alvo_ano else "🔴 Abaixo do alvo"
    _giro_cor = "#059669" if giro_ano>=giro_alvo_ano else "#DC2626"
    k2.markdown(f'''<div class="ge-kpi-box">
        <div style="font-size:16px;margin-bottom:3px">🔄</div>
        <div class="ge-kpi-lbl">Giro Médio do Estoque</div>
        <div class="ge-kpi-val">{giro_ano}x</div>
        <div class="ge-kpi-sub nt">alvo: {giro_alvo_ano}x</div>
        <div class="ge-kpi-sub" style="color:{_giro_cor};font-size:.62rem">{_giro_status}</div>
    </div>''', unsafe_allow_html=True)

    # k3 — Cobertura
    _cob_status = "🟢 Ideal" if 25<=cob_media<=40 else ("🔴 Acima" if cob_media>40 else "⚠️ Baixa")
    _cob_cor = "#059669" if 25<=cob_media<=40 else "#DC2626"
    k3.markdown(f'''<div class="ge-kpi-box">
        <div style="font-size:16px;margin-bottom:3px">📅</div>
        <div class="ge-kpi-lbl">Cobertura Média</div>
        <div class="ge-kpi-val">{cob_media:.0f} dias</div>
        <div class="ge-kpi-sub nt">meta: 25–40 dias</div>
        <div class="ge-kpi-sub" style="color:{_cob_cor};font-size:.62rem">{_cob_status}</div>
    </div>''', unsafe_allow_html=True)

    # k4 — Ruptura
    _rup_status = "🟢 Controlada" if pct_ruptura<=5 else ("⚠️ Atenção" if pct_ruptura<=15 else "🔴 Crítica")
    _rup_cor = "#059669" if pct_ruptura<=5 else ("#D97706" if pct_ruptura<=15 else "#DC2626")
    k4.markdown(f'''<div class="ge-kpi-box">
        <div style="font-size:16px;margin-bottom:3px">⚠️</div>
        <div class="ge-kpi-lbl">Ruptura</div>
        <div class="ge-kpi-val" style="color:{_rup_cor}">{pct_ruptura:.1f}%</div>
        <div class="ge-kpi-sub nt">{n_ruptura} produtos</div>
        <div class="ge-kpi-sub" style="color:{_rup_cor};font-size:.62rem">{_rup_status}</div>
    </div>''', unsafe_allow_html=True)

    # k5 — Capital Parado
    _cap_status = "🟢 Sem capital parado" if cap_parado==0 else "🔴 Capital imobilizado"
    _cap_cor = "#059669" if cap_parado==0 else "#DC2626"
    k5.markdown(f'''<div class="ge-kpi-box">
        <div style="font-size:16px;margin-bottom:3px">💰</div>
        <div class="ge-kpi-lbl">Capital Parado</div>
        <div class="ge-kpi-val">{fv(cap_parado)}</div>
        <div class="ge-kpi-sub nt">{n_parado} itens sem giro</div>
        <div class="ge-kpi-sub" style="color:{_cap_cor};font-size:.62rem">{_cap_status}</div>
    </div>''', unsafe_allow_html=True)

    # k6 — Erro Médio ML
    _precisao = round(100-mape_ml,1) if mape_ml else None
    _ml_status = "🟢 Excelente" if mape_ml and mape_ml<10 else ("⚠️ Aceitável" if mape_ml and mape_ml<20 else "🔴 Alto erro")
    _ml_cor = "#059669" if mape_ml and mape_ml<10 else ("#D97706" if mape_ml and mape_ml<20 else "#DC2626")
    k6.markdown(f'''<div class="ge-kpi-box">
        <div style="font-size:16px;margin-bottom:3px">📈</div>
        <div class="ge-kpi-lbl">Erro Médio</div>
        <div class="ge-kpi-val">{f"{mape_ml:.1f}%" if mape_ml else "—"}</div>
        <div class="ge-kpi-sub nt">precisão: {f"{_precisao:.1f}%" if _precisao else "—"}</div>
        <div class="ge-kpi-sub" style="color:{_ml_cor};font-size:.62rem">{_ml_status if mape_ml else "sem dados"}</div>
    </div>''', unsafe_allow_html=True)

    # k7 — Lead Time
    _lt_status = "🟢 Ótimo" if lt_medio<=10 else ("⚠️ Aceitável" if lt_medio<=20 else "🔴 Elevado")
    _lt_cor = "#059669" if lt_medio<=10 else ("#D97706" if lt_medio<=20 else "#DC2626")
    k7.markdown(f'''<div class="ge-kpi-box">
        <div style="font-size:16px;margin-bottom:3px">🚚</div>
        <div class="ge-kpi-lbl">Lead Time Médio</div>
        <div class="ge-kpi-val">{lt_medio:.0f} dias</div>
        <div class="ge-kpi-sub nt">meta: ≤ 10 dias</div>
        <div class="ge-kpi-sub" style="color:{_lt_cor};font-size:.62rem">{_lt_status}</div>
    </div>''', unsafe_allow_html=True)

    # k8 — Caixa p/ Compras
    k8.markdown(f'''<div class="ge-kpi-box">
        <div style="font-size:16px;margin-bottom:3px">💸</div>
        <div class="ge-kpi-lbl">Planejamneto de Compras</div>
        <div class="ge-kpi-val">{fv(total_horizon)}</div>
        <div class="ge-kpi-sub nt">horizonte: {periodo_cal}</div>
    </div>''', unsafe_allow_html=True)

    status_est = "🟢 Meta" if val_estoque <= cap_ideal*1.1 else "🔴 Acima"
    cor_est = "#059669" if val_estoque <= cap_ideal*1.1 else "#DC2626"
    k1.markdown(f'''<div class="ge-kpi-box">
        <div style="font-size:16px;margin-bottom:3px">📦</div>
        <div class="ge-kpi-lbl">Valor do Estoque</div>
        <div class="ge-kpi-val">{fv(val_estoque)}</div>
        <div class="ge-kpi-sub nt">alvo: {fv(cap_ideal)}</div>
        <div class="ge-kpi-sub" style="color:{cor_est};font-size:.62rem">{status_est}</div>
    </div>''', unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # ── LINHA 2: ABC + GIRO/CAPITAL POR CLASSE + RADAR + COBERTURA ─────
    g1,g2,g3=st.columns([1,1.1,1])

    with g1:
        st.markdown('<div class="ge-sec">🔵 Curva ABC</div>', unsafe_allow_html=True)
        if "Classe" in df_res.columns and "EstoqueAtual" in df_res.columns and "CustoUnitario" in df_res.columns:
            df_res["_CapitalClasse"]=df_res["EstoqueAtual"]*df_res["CustoUnitario"]
            abc=df_res.groupby("Classe")["_CapitalClasse"].sum().reset_index()
            abc.columns=["Classe","Capital"]
            abc["Pct"]=abc["Capital"]/abc["Capital"].sum()*100 if abc["Capital"].sum()>0 else 0
            mapa_cores_classe={"A":"#F0A500","B":"#14243B","C":"#6b7280","D":"#9ca3af","E":"#d1d5db","SV":"#DC2626"}
            fig_abc=go.Figure(go.Pie(
                labels=abc["Classe"],values=abc["Pct"],
                marker=dict(colors=[mapa_cores_classe.get(cl,"#cccccc") for cl in abc["Classe"]],
                    line=dict(color="white",width=2)),
                hole=0.45,textinfo="label+percent",textfont=dict(size=9),
                customdata=abc["Capital"],
                hovertemplate="<b>Classe %{label}</b><br>%{value:.1f}% do estoque<br>%{customdata:,.0f}<extra></extra>"))
            fig_abc.update_layout(plot_bgcolor="white",paper_bgcolor="white",
                margin=dict(l=0,r=0,t=0,b=0),height=140,showlegend=True,
                legend=dict(orientation="v",x=0.72,y=0.5,font=dict(size=8),bgcolor="rgba(0,0,0,0)"))
            st.plotly_chart(fig_abc,use_container_width=True,key="ge_abc")

    with g2:
        st.markdown('<div class="ge-sec">💰 Giro e Capital por Classe</div>', unsafe_allow_html=True)
        if "Classe" in df_res.columns and "EstoqueAtual" in df_res.columns:
            df_res["_CapitalLinha"]=df_res["EstoqueAtual"]*df_res["CustoUnitario"]
            def _giro_pond_ge(grupo):
                peso=grupo["_CapitalLinha"]
                if peso.sum()<=0 or "GiroAtual" not in grupo.columns: return 0.0
                return round(float((grupo["GiroAtual"]*peso).sum()/peso.sum()),1)
            gc=df_res.groupby("Classe")["_CapitalLinha"].sum().reset_index()
            gc.columns=["Classe","Capital"]
            gc_giro=df_res.groupby("Classe").apply(_giro_pond_ge).reset_index()
            gc_giro.columns=["Classe","GiroAno"]
            gc=gc.merge(gc_giro,on="Classe")
            gc=gc.sort_values("Capital",ascending=False)
            fig_gc=go.Figure()
            fig_gc.add_trace(go.Bar(
                x=gc["Classe"],y=gc["Capital"],name="Capital (R$)",
                marker=dict(color="#F0A500",line=dict(color="white",width=0.8)),
                text=[fv(v) for v in gc["Capital"]],textposition="outside",textfont=dict(size=8)))
            fig_gc.add_trace(go.Scatter(
                x=gc["Classe"],y=gc["GiroAno"],name="Giro (x/ano)",
                mode="lines+markers",yaxis="y2",
                line=dict(color="#14243B",width=2),marker=dict(size=6,color="#14243B")))
            fig_gc.update_layout(plot_bgcolor="white",paper_bgcolor="white",
                margin=dict(l=0,r=0,t=0,b=0),height=140,
                xaxis=dict(showgrid=False,tickfont=dict(size=9)),
                yaxis=dict(gridcolor="#F5F5F5",tickfont=dict(size=8),showticklabels=False),
                yaxis2=dict(overlaying="y",side="right",showgrid=False,tickfont=dict(size=8),ticksuffix="x"),
                legend=dict(orientation="h",y=-0.3,x=0.5,xanchor="center",font=dict(size=8),bgcolor="rgba(0,0,0,0)"),
                hovermode="x unified")
            st.plotly_chart(fig_gc,use_container_width=True,key="ge_gc")

    with g3:
        st.markdown('<div class="ge-sec">⚡ Radar de Estoque</div>', unsafe_allow_html=True)
        total_sug=float(df_res[df_res["ValorSugerido"]>0]["ValorSugerido"].sum()) if "ValorSugerido" in df_res.columns else 0
        st.markdown(f'''<div class="ge-radar">
            <div class="ge-radar-item" style="background:#FEF2F2;color:#DC2626">
                <div class="ge-radar-n">{n_rup_im}</div>
                <div class="ge-radar-l">Ruptura Iminente</div>
            </div>
            <div class="ge-radar-item" style="background:#FEF2F2;color:#DC2626">
                <div class="ge-radar-n">{n_comprar}</div>
                <div class="ge-radar-l">Comprar Agora</div>
            </div>
            <div class="ge-radar-item" style="background:#ECFDF5;color:#059669">
                <div class="ge-radar-n">{n_ok}</div>
                <div class="ge-radar-l">OK</div>
            </div>
            <div class="ge-radar-item" style="background:#FFFBEB;color:#D97706">
                <div class="ge-radar-n">{n_excesso}</div>
                <div class="ge-radar-l">Estoque Excessivo</div>
            </div>
        </div>
        <div style="background:#F8F5EE;border-radius:7px;padding:7px 10px;margin-top:6px;text-align:center">
            <div style="font-size:.6rem;color:#888;text-transform:uppercase;letter-spacing:.05em">Total Sugerido Imediato</div>
            <div style="font-family:Georgia,serif;font-size:1rem;font-weight:700;color:#14243B">{fv(total_sug)}</div>
        </div>''', unsafe_allow_html=True)

    

    st.markdown("<div style='height:0px'></div>", unsafe_allow_html=True)

    # ── LINHA 3: FORNECEDORES + COBERTURA ──────────────────────────────
    f1,f_cob=st.columns([1,1])

    with f1:
        st.markdown('<div class="ge-sec">🏆 Desempenho de Fornecedores</div>', unsafe_allow_html=True)
        if "gs_scorecard_forn" not in st.session_state:
            sc_persist=load_scorecard_forn(st.session_state.cid) if st.session_state.cid else None
            st.session_state["gs_scorecard_forn"]=sc_persist if sc_persist is not None else pd.DataFrame({"Fornecedor":[],"Prazo":[],"Qualidade":[],"OTIF":[]})
        sc=st.session_state["gs_scorecard_forn"]
        def cor(v): return "g" if v>=93 else ("y" if v>=85 else "r")
        def circulo(v):
            c=cor(v)
            cor_hex="#059669" if c=="g" else ("#D97706" if c=="y" else "#DC2626")
            return f'<span style="display:inline-flex;align-items:center;gap:5px;font-weight:700;color:{cor_hex};font-size:.78rem"><span style="width:14px;height:14px;border-radius:50%;border:2px solid {cor_hex};display:inline-block"></span>{v}%</span>'
        rows=""
        for _,row in sc.iterrows():
            p=int(row["Prazo"]); q=int(row["Qualidade"]); o=int(row["OTIF"])
            rows+=f"""<tr>
                <td style='padding:3px 6px;font-size:.65rem;font-weight:500;color:#14243B'>
                    🏭 {str(row['Fornecedor'])}
                </td>
                <td style='padding:3px 6px;text-align:center'>{circulo(p)}</td>
                <td style='padding:3px 6px;text-align:center'>{circulo(q)}</td>
                <td style='padding:3px 6px;text-align:center'>{circulo(o)}</td>
            </tr>"""
        st.markdown(f'''<table style="width:100%;border-collapse:collapse;font-size:.68rem">
            <thead><tr style="background:#F8F5EE">
                <th style="padding:4px 6px;text-align:left;color:#666;font-weight:600;border-bottom:2px solid #F0A500;font-size:.62rem;white-space:nowrap">Fornecedor</th>
                <th style="padding:4px 6px;text-align:center;color:#666;font-weight:600;border-bottom:2px solid #F0A500;font-size:.62rem">Prazo</th>
                <th style="padding:4px 6px;text-align:center;color:#666;font-weight:600;border-bottom:2px solid #F0A500;font-size:.62rem">Qualidade</th>
                <th style="padding:4px 6px;text-align:center;color:#666;font-weight:600;border-bottom:2px solid #F0A500;font-size:.62rem">OTIF</th>
            </tr></thead>
            <tbody>{rows}</tbody>
        </table>''',unsafe_allow_html=True)

    with f_cob:
        st.markdown('<div class="ge-sec">📦 Cobertura por Categoria</div>', unsafe_allow_html=True)
        if "CoberturaDias" in df_res.columns:
            if df_est is not None and "Categoria" in df_est.columns:
                df_rc=df_res.merge(df_est[["Produto","Categoria"]],on="Produto",how="left"); cat_c="Categoria"
            else:
                df_rc=df_res.copy(); cat_c="Classe"
            df_rc["_CapitalLinha"]=df_rc["EstoqueAtual"]*df_rc["CustoUnitario"]
            def _cob_pond_ge(grupo):
                peso=grupo["_CapitalLinha"]
                if peso.sum()<=0: return np.nan
                return float((grupo["CoberturaDias"]*peso).sum()/peso.sum())
            cob_cat=df_rc.groupby(cat_c).apply(_cob_pond_ge).reset_index()
            cob_cat.columns=["Cat","Dias"]
            cob_cat=cob_cat.dropna()
            cob_cat=cob_cat.sort_values("Dias",ascending=False).head(6)
            fig_cob2=go.Figure(go.Bar(
                y=cob_cat["Cat"],x=cob_cat["Dias"],orientation="h",
                marker=dict(color="#F0A500",line=dict(color="white",width=0.8)),
                text=[f"{v:.0f}d" for v in cob_cat["Dias"]],
                textposition="outside",textfont=dict(size=9)))
            fig_cob2.update_layout(plot_bgcolor="white",paper_bgcolor="white",
                margin=dict(l=0,r=40,t=0,b=0),height=max(90,min(200,30+22*len(cob_cat))),
                xaxis=dict(showgrid=True,gridcolor="#F5F5F5",tickfont=dict(size=8)),
                yaxis=dict(showgrid=False,tickfont=dict(size=8)),showlegend=False)
            st.plotly_chart(fig_cob2,use_container_width=True,key="ge_cob2")

    st.markdown('<div style="background:#F8F5EE;border-top:1px solid #EBEBEB;padding:5px 16px;border-radius:0 0 10px 10px;font-size:.60rem;color:#bbb;margin-top:6px">ℹ️ Dados calculados pelo Motor de Compras e ML — rode novamente para atualizar.</div>',unsafe_allow_html=True)
